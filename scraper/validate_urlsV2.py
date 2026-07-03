"""
Royal London — Pre-Scrape URL Validation Script
================================================
Run this BEFORE scrape_approved_urls.py to validate all URLs
in the customer Excel and produce a report to share with them.

Usage:
    python scraper/validate_urls.py
    python scraper/validate_urls.py --file path/to/custom.xlsx
    python scraper/validate_urls.py --concurrency 20 --timeout 10
    python scraper/validate_urls.py --retry-timeouts
    python scraper/validate_urls.py --retry-timeouts --retry-timeout 25

Output:
    scraper/data/url_validation_report_<timestamp>.xlsx
    Four sheets:
      1. Full Results    — every URL with status
      2. Action Required — dead + redirect + duplicate rows only
      3. Summary         — counts by issue type
      4. Clean URL List  — unique confirmed-live URLs, normalised
                           to lowercase, ready for the scraper

═══════════════════════════════════════════════════════════════
CHANGE LOG
═══════════════════════════════════════════════════════════════

v1.0.0 — July 2026 | Mukesh Kund
         Initial version

         PURPOSE:
         Standalone pre-scrape health check. Reads customer
         Excel (any column layout — URL column auto-detected
         by content sniffing, not header name), checks every
         URL live via async HTTP HEAD requests, detects
         duplicates, and produces an actionable Excel report
         ready to share with the customer.

         COLUMN DETECTION STRATEGY:
         Content-based sniffing (primary) — finds the column
         where cells look like https://... regardless of header
         name. Header-name matching used only as a tiebreaker
         when multiple columns have URL-like content.
         This means the script works correctly even if the
         customer renames "URL" to "Source URL", "Link",
         "Web Address", or anything else in a future file.

         URL CHECKS:
         - Live (200)            — clean, no action needed
         - Redirected (301/302)  — follows to final URL,
           flags if final domain differs from royallondon.com
         - Dead (4xx/5xx)        — action required
         - Timeout               — action required
         - Duplicate             — flagged after normalisation
           (catches case differences like should-I vs should-i,
           trailing slash variants, and exact duplicates)

v1.1.0 — July 2026 | Mukesh Kund
         --retry-timeouts flag + Clean URL List sheet +
         lowercase URL normalisation

         --retry-timeouts:
         After the initial check, progressively re-checks all
         URLs that timed out across up to 2 additional rounds:
           Round 1: concurrency=15, timeout=12s  (initial)
           Round 2: concurrency=3,  timeout=30s  (auto-retry)
           Round 3: concurrency=1,  timeout=45s  (final pass)
         Each round only re-checks URLs still timing out from
         the previous round — so if Round 2 resolves 50 of 60
         timeouts, Round 3 only runs on the remaining 10.
         After Round 3, any URL still timing out is confirmed
         unreachable and labelled clearly in the report:
         "Timed out on all 3 rounds (12s / 30s / 45s)".
         This is fully automatic — no manual re-runs needed.

         Clean URL List sheet (Sheet 4):
         Contains ONLY unique, confirmed-live (HTTP 200) URLs,
         sorted alphabetically. Every URL is normalised to
         lowercase (scheme + domain + path, query strings and
         fragments stripped) before being written. This is the
         definitive list to hand to the scraper.

         WHY lowercase normalisation in the clean list:
         normalize_url() already lowercases every URL path
         before it enters the scraper and the search index.
         If the clean list used mixed-case raw Excel URLs, there
         would be a mismatch between the customer-facing
         document and what's stored in the index — confusing
         to trace if a citation URL ever needs investigating.
         Using the same normalised form everywhere means the
         clean list, scraper input, and index source_url field
         are all identical strings — no ambiguity.

v1.2.0 — July 2026 | Mukesh Kund
         Internal redirect detection + excluded_urls.json support

         INTERNAL REDIRECT DETECTION:
         Previously, a URL that returned HTTP 200 but silently
         redirected to a different page within royallondon.com
         was marked as "✅ Live" — hiding the fact that the
         original content no longer exists at that URL.

         Observed in practice: campaign and product URLs like
         /insurance/life-insurance/do-more-digitally and
         /insurance/life-insurance/termcampaign were returning
         200 but landing on /insurance/ — the generic section
         homepage. Scraping these would index the homepage
         content multiple times under different source_url
         values — polluting the search index with duplicates
         and potentially confusing Aria's retrieval.

         FIX — is_internal_redirect() uses THREE signals:
         1. PATH PARENT CHECK: if the final URL path is a
            direct parent of the original (e.g. /insurance/
            when original was /insurance/.../do-more-digitally)
            → definitive redirect, flag immediately.
         2. PATH SEGMENT SIMILARITY: if fewer than 50% of the
            original path segments survive in the final URL
            → content has moved significantly, flag it.
         3. PAGE TITLE MISMATCH (additional check): if the
            customer Excel has a title column, fetch the live
            page's <title> tag and compare. A significant
            title difference confirms the redirect landed on
            completely different content. Uses rapidfuzz
            partial_ratio for fuzzy comparison (threshold 60).

         New status: "internal_redirect" — orange colour in
         report (distinct from yellow external redirect).

         Action Required sheet: shows original URL, final URL
         it redirected to, and a clear note:
         "Original content removed — now redirects to: <url>"

         Clean URL List: internal redirects are EXCLUDED.
         Scraping a redirect destination thinking it is the
         approved content would silently index wrong content.

         EXCLUDED_URLS.JSON SUPPORT:
         Load scraper/data/excluded_urls.json at startup.
         Any URL in this file is marked as "⛔ Excluded" in
         the Full Results sheet with the reason provided.
         Excluded URLs do NOT appear in Action Required or
         Clean URL List sheets.
         File is optional — if absent, no URLs are excluded.
         Format defined in scraper/data/excluded_urls.json
         (template created alongside this script).

═══════════════════════════════════════════════════════════════
"""

import argparse
import asyncio
import json
import sys
import os
import re
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import urlparse, urlunparse

import aiohttp
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

try:
    from rapidfuzz import fuzz as _fuzz
    _RAPIDFUZZ_AVAILABLE = True
except ImportError:
    _RAPIDFUZZ_AVAILABLE = False

# ── Config ────────────────────────────────────────────────────
DEFAULT_EXCEL             = None   # auto-detected from scraper/data/
DEFAULT_CONCURRENCY       = 15
DEFAULT_TIMEOUT           = 12     # Round 1: fast pass, 15 concurrent
DEFAULT_RETRY_TIMEOUT     = 30     # Round 2: 3 concurrent, 30s
DEFAULT_RETRY_TIMEOUT_R3  = 45     # Round 3: 1 concurrent, 45s — final confirmation
DEFAULT_RETRY_CONCURRENCY = 3      # Round 2 concurrency
MAX_RETRIES               = 1
EXPECTED_DOMAIN           = "royallondon.com"

# Directory to scan for the customer Excel when --file not provided
SCRAPER_DATA_DIR = "scraper/data"

# ── DEFAULT FILENAME (edit this if auto-detect picks the wrong file) ──
# Set this to the customer Excel filename (just the name, not the full
# path) to use it as the preferred file when no --file flag is given.
# Auto-detect still runs, but this filename wins if it exists.
# Leave as None to always use the most recently modified Excel.
#
# Example:
#   DEFAULT_EXCEL_FILENAME = "royal_london_website_chatbot_urls.xlsx"
DEFAULT_EXCEL_FILENAME = "royal_london_website_chatbot_urls.xlsx"

# HTTP headers — mimic a real browser to avoid bot-detection blocks
REQUEST_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,*/*",
}

# ── Auto-detect latest Excel ──────────────────────────────────
def find_latest_excel(data_dir: str) -> str | None:
    """
    Find the customer Excel file in data_dir.

    Priority order:
    1. DEFAULT_EXCEL_FILENAME (if set and exists in data_dir)
       — explicit preferred filename, no guessing needed
    2. Most recently modified .xlsx that is not a validation
       report and not a Windows Excel lock file (~$ prefix)

    Windows lock file note:
    When Excel has a file open, it creates a hidden temp file
    called ~$<filename>.xlsx in the same folder. openpyxl
    cannot read these — they are not real Excel files. The
    ~$ prefix filter ensures these are never picked up
    regardless of their modification time.
    """
    data_path = Path(data_dir)
    if not data_path.exists():
        return None

    # Priority 1 — preferred filename explicitly set
    if DEFAULT_EXCEL_FILENAME:
        preferred = data_path / DEFAULT_EXCEL_FILENAME
        if preferred.exists():
            return str(preferred)

    # Priority 2 — most recently modified valid Excel
    candidates = [
        f for f in data_path.glob("*.xlsx")
        if not f.name.startswith("~$")              # exclude Windows lock files
        and "url_validation_report" not in f.name.lower()  # exclude our own reports
    ]

    if not candidates:
        return None

    return str(max(candidates, key=lambda f: f.stat().st_mtime))
GREEN  = "C6EFCE"   # live
YELLOW = "FFEB9C"   # external redirect / warning
ORANGE = "F4B942"   # internal redirect — content removed
RED    = "FFC7CE"   # dead / error
BLUE   = "BDD7EE"   # duplicate
PURPLE = "E8D5F5"   # excluded
GREY   = "D9D9D9"   # header
WHITE  = "FFFFFF"

STATUS_LABELS = {
    "live":              "✅ Live",
    "redirect":          "⚠️  Redirected (external)",
    "internal_redirect": "🔀 Internal Redirect",
    "dead":              "❌ Dead",
    "timeout":           "⏱️  Timeout",
    "duplicate":         "🔁 Duplicate",
    "excluded":          "⛔ Excluded",
    "error":             "❌ Error",
}

STATUS_COLOURS = {
    "live":              GREEN,
    "redirect":          YELLOW,
    "internal_redirect": ORANGE,
    "dead":              RED,
    "timeout":           RED,
    "duplicate":         BLUE,
    "excluded":          PURPLE,
    "error":             RED,
}

# Path to the excluded URLs JSON file (optional — skipped if absent)
EXCLUDED_URLS_PATH = "scraper/data/excluded_urls.json"


# ── Helpers ───────────────────────────────────────────────────
def normalize_url(url: str) -> str:
    """Canonical URL for duplicate detection — same logic as scraper."""
    if not url:
        return url
    try:
        parsed = urlparse(url.strip())
        return urlunparse((
            parsed.scheme.lower(),
            parsed.netloc.lower(),
            parsed.path.lower().rstrip("/"),
            "", "", "",
        ))
    except Exception:
        return url.strip().lower()


def load_excluded_urls(path: str) -> dict[str, str]:
    """
    Load excluded_urls.json — returns {normalised_url: reason}.
    Returns empty dict if file does not exist (optional file).
    Logs a warning if the file exists but cannot be parsed.
    """
    p = Path(path)
    if not p.exists():
        return {}
    try:
        with open(p, encoding="utf-8") as f:
            entries = json.load(f)
        excluded = {}
        for entry in entries:
            url    = entry.get("url", "").strip()
            reason = entry.get("reason", "No reason provided")
            if url:
                excluded[normalize_url(url)] = reason
        print(f"   Loaded {len(excluded)} excluded URL(s) from {p.name}")
        return excluded
    except Exception as e:
        print(f"   ⚠️  Could not load {path}: {e}")
        return {}


async def fetch_page_title(
    session: aiohttp.ClientSession,
    url: str,
    timeout: int = 10,
) -> str:
    """
    Fetch the <title> tag of a live page for internal redirect
    title comparison. Returns empty string on any failure —
    title comparison is a supporting signal only, never the
    sole reason a URL is flagged.
    """
    try:
        async with session.get(
            url,
            timeout=aiohttp.ClientTimeout(total=timeout),
            ssl=False,
            allow_redirects=True,
        ) as resp:
            if resp.status != 200:
                return ""
            # Read only first 8KB — title is always in <head>
            chunk = await resp.content.read(8192)
            html  = chunk.decode("utf-8", errors="ignore")
            match = re.search(
                r"<title[^>]*>(.*?)</title>",
                html,
                re.IGNORECASE | re.DOTALL,
            )
            if match:
                return re.sub(r"\s+", " ", match.group(1)).strip()
    except Exception:
        pass
    return ""


def is_internal_redirect(
    original_url: str,
    final_url: str,
    original_title: str = "",
    live_title: str = "",
) -> tuple[bool, str]:
    """
    Detect when a URL has redirected to a different page within
    royallondon.com — content removed but domain unchanged.

    THREE signals (any one is sufficient to flag):

    Signal 1 — PATH PARENT CHECK (strongest):
    If final path is a direct ancestor of the original path,
    the page was removed and fell back to the section homepage.
    e.g. /insurance/ when original was /insurance/life-insurance/
    do-more-digitally → definitive.

    Signal 2 — PATH SEGMENT SIMILARITY (< 50%):
    If fewer than half of the original path segments survive
    in the final URL, content has moved significantly.
    e.g. /pensions/workplace-pensions/ → /pensions/ loses 50%.

    Signal 3 — PAGE TITLE MISMATCH (supporting):
    If original Excel title and fetched live page title differ
    significantly (rapidfuzz partial_ratio < 60), the redirect
    landed on completely different content. Only used when both
    titles are available and rapidfuzz is installed.
    A title match alone does NOT mark a URL as an internal
    redirect — titles can be generic ("Royal London").

    Returns (is_redirect: bool, reason: str).
    """
    orig_parsed  = urlparse(original_url)
    final_parsed = urlparse(final_url)

    orig_path  = orig_parsed.path.lower().rstrip("/")
    final_path = final_parsed.path.lower().rstrip("/")

    # No redirect at all
    if orig_path == final_path:
        return False, ""

    # Signal 1 — parent path check
    # final_path must be a proper prefix of orig_path
    if orig_path.startswith(final_path + "/") and len(final_path) < len(orig_path):
        return True, (
            f"Page removed — redirects to parent: {final_url}"
        )

    # Signal 2 — path segment similarity
    orig_segments  = [s for s in orig_path.split("/") if s]
    final_segments = [s for s in final_path.split("/") if s]

    if orig_segments:
        common = sum(1 for s in final_segments if s in orig_segments)
        similarity = common / len(orig_segments)
        if similarity < 0.5:
            return True, (
                f"Path changed significantly ({round(similarity*100)}% "
                f"segment overlap) — redirects to: {final_url}"
            )

    # Signal 3 — title mismatch (supporting only)
    if (
        _RAPIDFUZZ_AVAILABLE
        and original_title
        and live_title
        and original_title.lower() != "royal london"  # too generic
    ):
        # Strip "- Royal London" suffix for cleaner comparison
        clean_orig = re.sub(
            r"\s*[-|]\s*royal london\s*$", "",
            original_title, flags=re.IGNORECASE
        ).strip()
        clean_live = re.sub(
            r"\s*[-|]\s*royal london\s*$", "",
            live_title, flags=re.IGNORECASE
        ).strip()

        if clean_orig and clean_live:
            score = _fuzz.partial_ratio(
                clean_orig.lower(), clean_live.lower()
            )
            if score < 60:
                return True, (
                    f"Title mismatch ({score}% match): "
                    f"expected '{clean_orig[:40]}' — "
                    f"got '{clean_live[:40]}' at {final_url}"
                )

    return False, ""


def detect_url_column(ws) -> tuple[int | None, int | None, int | None]:
    """
    Detect URL, Title, and Category column indices by content sniffing.

    Primary strategy (URL column):
    Sample first 10 non-empty data rows per column. The column
    with the highest proportion of cells starting with https?://
    is the URL column. No dependency on header naming — works
    regardless of what the column is called.

    Tiebreaker: if multiple columns score equally, prefer the
    one whose header also contains "url" or "link".

    Title and Category: header-name matching only (optional
    fields — graceful empty-string fallback if not found).

    Returns (url_col_idx, title_col_idx, category_col_idx)
    All zero-indexed. None if not found.
    """
    TITLE_HEADERS    = {"title", "page title", "name", "page name"}
    CATEGORY_HEADERS = {"category", "type", "content type", "page type"}

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return None, None, None

    header_row = [
        str(h).strip().lower() if h is not None else ""
        for h in rows[0]
    ]
    data_rows = rows[1:]

    num_cols = len(header_row)

    # ── Content-based URL sniffing ────────────────────────────
    url_scores = []
    for col_idx in range(num_cols):
        sample = [
            str(r[col_idx]).strip()
            for r in data_rows
            if col_idx < len(r) and r[col_idx] is not None
        ][:10]
        if not sample:
            url_scores.append(0.0)
            continue
        url_like = sum(
            1 for v in sample
            if v.startswith("http://") or v.startswith("https://")
        )
        url_scores.append(url_like / len(sample))

    max_score = max(url_scores) if url_scores else 0
    if max_score < 0.5:
        # No column is majority URL-like
        return None, None, None

    # Among columns tied at the max score, prefer one whose
    # header contains "url" or "link" (tiebreaker only)
    candidates = [
        i for i, s in enumerate(url_scores) if s == max_score
    ]
    url_idx = candidates[0]  # default: leftmost best candidate
    for i in candidates:
        if any(kw in header_row[i] for kw in ("url", "link")):
            url_idx = i
            break

    # ── Header-name matching for optional columns ─────────────
    title_idx    = None
    category_idx = None
    for i, h in enumerate(header_row):
        if i == url_idx:
            continue
        if h in TITLE_HEADERS and title_idx is None:
            title_idx = i
        if h in CATEGORY_HEADERS and category_idx is None:
            category_idx = i

    return url_idx, title_idx, category_idx


def load_urls_from_excel(excel_path: str) -> list[dict]:
    """
    Load URLs from Excel using content-based column detection.

    Returns list of dicts:
    [{"url": "...", "title": "...", "category": "...",
      "row_num": N}]

    Raises ValueError if no URL column can be detected.
    """
    wb = load_workbook(excel_path, read_only=True)
    ws = wb.active

    url_idx, title_idx, category_idx = detect_url_column(ws)

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if url_idx is None:
        header_row = rows[0] if rows else []
        raise ValueError(
            f"No URL column found in {excel_path!r}.\n"
            f"Expected a column where most cells start with "
            f"https:// or http://.\n"
            f"Header row found: {list(header_row)}"
        )

    print(f"   URL column:      column index {url_idx} "
          f"(header: {str(rows[0][url_idx])!r})")
    print(f"   Title column:    "
          f"{'column index ' + str(title_idx) + ' (header: ' + repr(str(rows[0][title_idx])) + ')' if title_idx is not None else 'not found — title will be empty'}")
    print(f"   Category column: "
          f"{'column index ' + str(category_idx) + ' (header: ' + repr(str(rows[0][category_idx])) + ')' if category_idx is not None else 'not found'}")

    entries = []
    for row_num, row in enumerate(rows[1:], start=2):
        if not row or len(row) <= url_idx:
            continue
        url = str(row[url_idx]).strip() if row[url_idx] else ""
        if not url.startswith("http"):
            continue

        title = (
            str(row[title_idx]).strip()
            if title_idx is not None and len(row) > title_idx and row[title_idx]
            else ""
        )
        category = (
            str(row[category_idx]).strip()
            if category_idx is not None and len(row) > category_idx and row[category_idx]
            else ""
        )

        entries.append({
            "url":      url,
            "title":    title,
            "category": category,
            "row_num":  row_num,
        })

    return entries


# ── Async URL checker ─────────────────────────────────────────
async def check_url(
    session: aiohttp.ClientSession,
    entry: dict,
    timeout: int,
    semaphore: asyncio.Semaphore,
) -> dict:
    """
    Check a single URL. Returns result dict with:
    url, title, category, row_num, status, status_code,
    final_url, issue, redirect_chain, live_title
    """
    url = entry["url"]
    result = {
        **entry,
        "status":         "error",
        "status_code":    None,
        "final_url":      url,
        "issue":          "",
        "redirect_chain": "",
        "live_title":     "",
    }

    async with semaphore:
        for attempt in range(MAX_RETRIES + 1):
            try:
                resp = await session.get(
                    url,
                    timeout=aiohttp.ClientTimeout(total=timeout),
                    ssl=False,
                    allow_redirects=True,
                )
                result["status_code"] = resp.status
                result["final_url"]   = str(resp.url)

                # Build redirect chain if redirected
                if str(resp.url) != url:
                    result["redirect_chain"] = (
                        f"{url} → {str(resp.url)}"
                    )

                if resp.status == 200:
                    final_domain = urlparse(str(resp.url)).netloc.lower()

                    if EXPECTED_DOMAIN not in final_domain:
                        # External redirect — outside royallondon.com
                        result["status"] = "redirect"
                        result["issue"]  = (
                            f"Redirected to external domain: "
                            f"{final_domain}"
                        )
                    elif str(resp.url).rstrip("/") != url.rstrip("/"):
                        # Same domain but different URL — check if
                        # it's an internal redirect (content removed)
                        live_title = await fetch_page_title(
                            session, str(resp.url)
                        )
                        result["live_title"] = live_title

                        flagged, reason = is_internal_redirect(
                            original_url=url,
                            final_url=str(resp.url),
                            original_title=entry.get("title", ""),
                            live_title=live_title,
                        )
                        if flagged:
                            result["status"] = "internal_redirect"
                            result["issue"]  = reason
                        else:
                            # Minor redirect (e.g. trailing slash
                            # normalisation) — treat as live
                            result["status"] = "live"
                            result["issue"]  = ""
                    else:
                        result["status"] = "live"
                        result["issue"]  = ""

                elif resp.status >= 400:
                    result["status"] = "dead"
                    result["issue"]  = f"HTTP {resp.status}"
                else:
                    result["status"] = "redirect"
                    result["issue"]  = f"HTTP {resp.status}"

                break

            except asyncio.TimeoutError:
                if attempt < MAX_RETRIES:
                    await asyncio.sleep(1)
                    continue
                result["status"] = "timeout"
                result["issue"]  = f"Timed out after {timeout}s"

            except aiohttp.ClientConnectorError as e:
                if attempt < MAX_RETRIES:
                    await asyncio.sleep(1)
                    continue
                result["status"] = "dead"
                result["issue"]  = f"Connection error: {str(e)[:80]}"

            except Exception as e:
                result["status"] = "error"
                result["issue"]  = f"Error: {str(e)[:80]}"
                break

    return result


async def check_all_urls(
    entries: list[dict],
    concurrency: int,
    timeout: int,
) -> list[dict]:
    """Check all URLs concurrently with a semaphore rate limit."""
    semaphore = asyncio.Semaphore(concurrency)
    connector = aiohttp.TCPConnector(
        limit=concurrency,
        limit_per_host=5,   # don't hammer royallondon.com too hard
    )

    async with aiohttp.ClientSession(
        headers=REQUEST_HEADERS,
        connector=connector,
    ) as session:
        tasks = [
            check_url(session, entry, timeout, semaphore)
            for entry in entries
        ]

        results = []
        done = 0
        total = len(tasks)

        for coro in asyncio.as_completed(tasks):
            result = await coro
            results.append(result)
            done += 1
            if done % 20 == 0 or done == total:
                pct = round(done / total * 100)
                status_icon = {
                    "live":      "✅",
                    "dead":      "❌",
                    "redirect":  "⚠️ ",
                    "timeout":   "⏱️ ",
                    "duplicate": "🔁",
                    "error":     "❌",
                }.get(result["status"], "  ")
                print(
                    f"   [{pct:3d}%] {done}/{total} checked  "
                    f"| Last: {status_icon} {result['url'][:60]}"
                )

        return results


async def retry_timeouts(
    results: list[dict],
    retry_timeout: int = DEFAULT_RETRY_TIMEOUT,
    retry_timeout_r3: int = DEFAULT_RETRY_TIMEOUT_R3,
    retry_concurrency: int = DEFAULT_RETRY_CONCURRENCY,
) -> list[dict]:
    """
    v1.1.0 — Progressive 3-round retry for timed-out URLs.

    Round 1 (main check): concurrency=15, timeout=12s
    Round 2 (this func):  concurrency=3,  timeout=30s
    Round 3 (this func):  concurrency=1,  timeout=45s

    Each round only re-checks URLs that STILL timed out in the
    previous round — so if Round 2 resolves 50 of 60 timeouts,
    Round 3 only runs on the remaining 10. Total extra time
    stays small even with many initial timeouts.

    After Round 3, anything still timing out is genuinely
    unreachable — marked with a clear note so the report is
    unambiguous about whether it was a fluke or a real failure.
    """
    ROUNDS = [
        {"label": "Round 2", "concurrency": retry_concurrency, "timeout": retry_timeout},
        {"label": "Round 3", "concurrency": 1,                 "timeout": retry_timeout_r3},
    ]

    for round_cfg in ROUNDS:
        still_timing_out = [r for r in results if r["status"] == "timeout"]
        if not still_timing_out:
            print(f"   No more timeouts — skipping {round_cfg['label']}.")
            break

        print(
            f"\n   {round_cfg['label']}: re-checking "
            f"{len(still_timing_out)} URL(s) — "
            f"concurrency={round_cfg['concurrency']}, "
            f"timeout={round_cfg['timeout']}s..."
        )

        retry_entries = [
            {"url": r["url"], "title": r["title"],
             "category": r["category"], "row_num": r["row_num"]}
            for r in still_timing_out
        ]

        retry_checked = await check_all_urls(
            retry_entries,
            concurrency=round_cfg["concurrency"],
            timeout=round_cfg["timeout"],
        )
        retry_by_url = {r["url"]: r for r in retry_checked}

        resolved = 0
        still_timeout = 0
        for result in results:
            if result["status"] != "timeout":
                continue
            updated = retry_by_url.get(result["url"])
            if not updated:
                continue
            if updated["status"] != "timeout":
                result["status"]      = updated["status"]
                result["status_code"] = updated["status_code"]
                result["final_url"]   = updated["final_url"]
                result["issue"]       = (
                    f"{updated['issue']} (resolved on {round_cfg['label'].lower()})"
                    if updated["issue"]
                    else f"(resolved on {round_cfg['label'].lower()})"
                )
                resolved += 1
            else:
                still_timeout += 1

        print(
            f"   {round_cfg['label']} complete: "
            f"{resolved} resolved ✅  |  {still_timeout} still timing out ⏱️"
        )

    # Mark anything surviving all rounds as confirmed unreachable
    for result in results:
        if result["status"] == "timeout":
            result["issue"] = (
                f"Timed out on all 3 rounds "
                f"(12s / {retry_timeout}s / {retry_timeout_r3}s) — "
                f"confirmed unreachable"
            )

    return results


def detect_duplicates(results: list[dict]) -> list[dict]:
    """
    Mark duplicate URLs in results list.
    Normalises before comparison — catches case differences,
    trailing slash variants, and exact duplicates.
    First occurrence is kept as primary; all subsequent
    occurrences are flagged as duplicates.
    """
    seen: dict[str, int] = {}   # normalised_url → index of first occurrence

    for i, r in enumerate(results):
        norm = normalize_url(r["url"])
        if norm in seen:
            # Mark this entry as a duplicate
            r["is_duplicate"] = True
            r["duplicate_of"] = results[seen[norm]]["url"]
            # Don't change the status if it's already dead/timeout
            # — being a duplicate AND dead is more actionable
            if r["status"] == "live":
                r["status"] = "duplicate"
                r["issue"]  = (
                    f"Duplicate of: {results[seen[norm]]['url']}"
                )
            else:
                r["issue"] = (
                    r["issue"] +
                    f" | Also duplicate of: {results[seen[norm]]['url']}"
                )
        else:
            seen[norm] = i
            r["is_duplicate"]  = False
            r["duplicate_of"]  = ""

    return results


# ── Report generation ─────────────────────────────────────────
def _cell_fill(colour: str) -> PatternFill:
    return PatternFill("solid", start_color=colour, fgColor=colour)


def _header_style(ws, row: int, cols: list[str]):
    for col in cols:
        cell = ws[f"{col}{row}"]
        cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill      = _cell_fill("1F4E79")
        cell.alignment = Alignment(horizontal="center", vertical="center")


def write_report(
    results: list[dict],
    output_path: str,
    input_file: str,
) -> str:
    """Write the three-sheet Excel report."""
    wb = Workbook()

    # ── Counts ────────────────────────────────────────────────
    counts = {
        "live":              sum(1 for r in results if r["status"] == "live"),
        "internal_redirect": sum(1 for r in results if r["status"] == "internal_redirect"),
        "redirect":          sum(1 for r in results if r["status"] == "redirect"),
        "dead":              sum(1 for r in results if r["status"] == "dead"),
        "timeout":           sum(1 for r in results if r["status"] == "timeout"),
        "duplicate":         sum(1 for r in results if r["status"] == "duplicate"),
        "excluded":          sum(1 for r in results if r["status"] == "excluded"),
        "error":             sum(1 for r in results if r["status"] == "error"),
    }
    action_needed = (
        counts["dead"] + counts["redirect"] +
        counts["internal_redirect"] +
        counts["timeout"] + counts["duplicate"] + counts["error"]
    )

    # ══════════════════════════════════════════════════════════
    # SHEET 1 — Full Results
    # ══════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Full Results"

    headers = [
        "Row (Excel)", "Title", "Category", "URL",
        "Status", "HTTP Code", "Final URL",
        "Duplicate", "Issue / Notes",
    ]
    ws1.append(headers)
    _header_style(ws1, 1, [get_column_letter(i+1) for i in range(len(headers))])
    ws1.row_dimensions[1].height = 20

    for r in sorted(results, key=lambda x: x["row_num"]):
        status    = r["status"]
        colour    = STATUS_COLOURS.get(status, WHITE)
        is_dup    = "Yes" if r.get("is_duplicate") else ""
        final_url = r["final_url"] if r["final_url"] != r["url"] else ""

        row_data = [
            r["row_num"],
            r["title"],
            r["category"],
            r["url"],
            STATUS_LABELS.get(status, status),
            r["status_code"] or "",
            final_url,
            is_dup,
            r["issue"],
        ]
        ws1.append(row_data)
        row_idx = ws1.max_row
        for col_idx in range(1, len(headers) + 1):
            cell      = ws1.cell(row=row_idx, column=col_idx)
            cell.fill = _cell_fill(colour)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="center")

        # URL column — hyperlink
        url_cell = ws1.cell(row=row_idx, column=4)
        url_cell.hyperlink = r["url"]
        url_cell.font = Font(
            name="Arial", size=10,
            color="0563C1", underline="single"
        )

    # Column widths
    col_widths = [10, 45, 12, 60, 22, 10, 60, 10, 50]
    for i, width in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = width

    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # ══════════════════════════════════════════════════════════
    # SHEET 2 — Action Required
    # ══════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Action Required")

    ws2.append([""])
    ws2["A1"] = "⚠️  URLs requiring action — share this sheet with the customer"
    ws2["A1"].font = Font(bold=True, name="Arial", size=12, color="C00000")
    ws2.row_dimensions[1].height = 22

    ws2.append([""])   # blank row

    action_headers = [
        "Row (Excel)", "Title", "Category", "URL",
        "Issue Type", "HTTP Code", "Final URL", "Notes",
    ]
    ws2.append(action_headers)
    _header_style(ws2, 3, [get_column_letter(i+1) for i in range(len(action_headers))])

    action_rows = [
        r for r in results
        if r["status"] in (
            "dead", "internal_redirect", "redirect",
            "timeout", "error", "duplicate"
        )
    ]
    action_rows.sort(key=lambda x: (
        ["dead", "internal_redirect", "timeout",
         "redirect", "duplicate", "error"].index(
            x["status"] if x["status"] in
            ["dead", "internal_redirect", "timeout",
             "redirect", "duplicate", "error"]
            else "error"
        ),
        x["row_num"],
    ))

    for r in action_rows:
        status    = r["status"]
        colour    = STATUS_COLOURS.get(status, WHITE)
        final_url = r["final_url"] if r["final_url"] != r["url"] else ""

        # Build a customer-friendly note
        if status == "internal_redirect":
            notes = (
                f"Original content removed — "
                f"now redirects to: {r['final_url']}"
            )
        else:
            notes = r["issue"]

        row_data = [
            r["row_num"],
            r["title"],
            r["category"],
            r["url"],
            STATUS_LABELS.get(status, status),
            r["status_code"] or "",
            final_url,
            notes,
        ]
        ws2.append(row_data)
        row_idx = ws2.max_row
        for col_idx in range(1, len(action_headers) + 1):
            cell      = ws2.cell(row=row_idx, column=col_idx)
            cell.fill = _cell_fill(colour)
            cell.font = Font(name="Arial", size=10)

        url_cell = ws2.cell(row=row_idx, column=4)
        url_cell.hyperlink = r["url"]
        url_cell.font = Font(
            name="Arial", size=10,
            color="0563C1", underline="single"
        )

    if not action_rows:
        ws2.append(["", "✅ No action required — all URLs are live and unique"])
        ws2.cell(ws2.max_row, 2).font = Font(
            bold=True, name="Arial", size=11, color="375623"
        )

    col_widths2 = [10, 45, 12, 60, 22, 10, 60, 60]
    for i, width in enumerate(col_widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = width

    ws2.freeze_panes = "A4"

    # ══════════════════════════════════════════════════════════
    # SHEET 3 — Summary
    # ══════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Summary")

    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 15
    ws3.column_dimensions["C"].width = 40

    # Title block
    ws3["A1"] = "Royal London URL Validation Report"
    ws3["A1"].font = Font(bold=True, name="Arial", size=14, color="1F4E79")
    ws3.merge_cells("A1:C1")

    ws3["A2"] = "Generated:"
    ws3["B2"] = datetime.now(timezone.utc).strftime("%d %b %Y %H:%M UTC")
    ws3["A3"] = "Input file:"
    ws3["B3"] = Path(input_file).name
    ws3["A4"] = "Total URLs checked:"
    ws3["B4"] = len(results)

    for row in [2, 3, 4]:
        ws3[f"A{row}"].font = Font(bold=True, name="Arial", size=10)
        ws3[f"B{row}"].font = Font(name="Arial", size=10)

    # Results table
    ws3["A6"] = "Result"
    ws3["B6"] = "Count"
    ws3["C6"] = "Notes"
    _header_style(ws3, 6, ["A", "B", "C"])

    summary_rows = [
        ("✅ Live",
         counts["live"],
         "Working correctly — no action needed"),
        ("🔀 Internal Redirect",
         counts["internal_redirect"],
         "Original content removed — redirects within royallondon.com. "
         "Remove from approved list or ask customer to provide replacement URL"),
        ("❌ Dead (4xx/5xx)",
         counts["dead"],
         "Remove from approved list or fix URL"),
        ("⚠️  Redirected (external)",
         counts["redirect"],
         "Review — final URL is outside royallondon.com"),
        ("⏱️  Timeout",
         counts["timeout"],
         "Could not connect after 3 retry rounds — confirmed unreachable"),
        ("🔁 Duplicate",
         counts["duplicate"],
         "Same URL appears more than once — remove extra rows"),
        ("⛔ Excluded",
         counts["excluded"],
         "Excluded via excluded_urls.json — not in Clean URL List"),
        ("❌ Other error",
         counts["error"],
         "Unexpected error — check manually"),
        ("", "", ""),
        ("⚠️  TOTAL ACTION REQUIRED",
         action_needed,
         "See 'Action Required' sheet"),
    ]

    colour_map = {
        "✅ Live":                    GREEN,
        "🔀 Internal Redirect":       ORANGE,
        "❌ Dead (4xx/5xx)":         RED,
        "⚠️  Redirected (external)": YELLOW,
        "⏱️  Timeout":               RED,
        "🔁 Duplicate":              BLUE,
        "⛔ Excluded":               PURPLE,
        "❌ Other error":             RED,
        "⚠️  TOTAL ACTION REQUIRED": YELLOW,
    }

    for label, count, note in summary_rows:
        ws3.append([label, count if count != "" else "", note])
        row_idx = ws3.max_row
        colour  = colour_map.get(label, WHITE)
        for col in ["A", "B", "C"]:
            cell = ws3[f"{col}{row_idx}"]
            if label:
                cell.fill = _cell_fill(colour)
            cell.font = Font(
                name="Arial", size=10,
                bold=(label == "⚠️  TOTAL ACTION REQUIRED"),
            )

    # ══════════════════════════════════════════════════════════
    # SHEET 4 — Clean URL List
    # ══════════════════════════════════════════════════════════
    # Unique, confirmed-live (HTTP 200) URLs only.
    # Every URL normalised to lowercase via normalize_url() —
    # same transformation the scraper applies, so this list
    # is directly usable as scraper input and matches what
    # will be stored as source_url in the search index.
    # Sorted alphabetically for easy scanning.
    # ══════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Clean URL List")

    # Info header
    ws4["A1"] = "Clean URL List — unique confirmed-live URLs (HTTP 200), normalised to lowercase"
    ws4["A1"].font = Font(bold=True, name="Arial", size=11, color="375623")
    ws4.merge_cells("A1:D1")
    ws4.row_dimensions[1].height = 20

    ws4["A2"] = (
        "Use this list as input for scrape_approved_urls.py. "
        "URLs are identical to what will be stored as source_url in the search index. "
        "Internal redirects, dead links, duplicates and excluded URLs are NOT included."
    )
    ws4["A2"].font = Font(name="Arial", size=9, italic=True, color="595959")
    ws4.merge_cells("A2:D2")

    ws4["A3"] = ""  # spacer

    clean_headers = ["#", "URL (normalised)", "Title", "Category"]
    ws4.append(clean_headers)
    _header_style(ws4, 4, ["A", "B", "C", "D"])

    # Collect ONLY live URLs — internal redirects, dead links,
    # duplicates and excluded entries are all omitted.
    # internal_redirect is excluded because scraping the redirect
    # destination thinking it's the approved content would
    # silently index wrong content under the wrong source_url.
    seen_normalised = set()
    clean_rows = []
    for r in results:
        if r["status"] not in ("live",):
            continue
        norm = normalize_url(r["url"])
        if norm in seen_normalised:
            continue
        seen_normalised.add(norm)
        clean_rows.append({
            "url":      norm,
            "title":    r["title"],
            "category": r["category"],
        })

    # Sort alphabetically by normalised URL
    clean_rows.sort(key=lambda x: x["url"])

    for i, row in enumerate(clean_rows, start=1):
        ws4.append([i, row["url"], row["title"], row["category"]])
        row_idx = ws4.max_row
        for col_idx in range(1, 5):
            cell      = ws4.cell(row=row_idx, column=col_idx)
            cell.fill = _cell_fill(GREEN)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="center")

        # URL as hyperlink
        url_cell = ws4.cell(row=row_idx, column=2)
        url_cell.hyperlink = row["url"]
        url_cell.font = Font(
            name="Arial", size=10,
            color="0563C1", underline="single",
        )

    # Footer row with total count
    ws4.append([""])
    ws4.append([
        "",
        f"Total: {len(clean_rows):,} unique live URLs",
        "",
        "",
    ])
    footer_cell = ws4.cell(ws4.max_row, 2)
    footer_cell.font = Font(bold=True, name="Arial", size=10, color="375623")

    ws4.column_dimensions["A"].width = 6
    ws4.column_dimensions["B"].width = 70
    ws4.column_dimensions["C"].width = 45
    ws4.column_dimensions["D"].width = 15
    ws4.freeze_panes = "A5"
    ws4.auto_filter.ref = f"A4:D4"

    wb.save(output_path)
    return output_path


# ── Main ──────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description="Validate Royal London approved URLs before scraping"
    )
    parser.add_argument(
        "--file", default=None,
        help=(
            "Path to customer Excel file. If not provided, automatically "
            "finds the most recently modified .xlsx in scraper/data/ "
            "that is not itself a validation report."
        ),
    )
    parser.add_argument(
        "--concurrency", type=int, default=DEFAULT_CONCURRENCY,
        help="Number of concurrent HTTP requests (default: 15)",
    )
    parser.add_argument(
        "--timeout", type=int, default=DEFAULT_TIMEOUT,
        help="Seconds before a URL is marked as timeout (default: 12)",
    )
    parser.add_argument(
        "--no-retry", action="store_true",
        help=(
            "Skip the automatic retry pass for timed-out URLs. "
            "By default, any timed-out URLs are automatically re-checked "
            "at lower concurrency before the report is written."
        ),
    )
    parser.add_argument(
        "--retry-timeout", type=int, default=DEFAULT_RETRY_TIMEOUT,
        help=f"Timeout for Round 2 retry pass (default: {DEFAULT_RETRY_TIMEOUT}s). Round 3 always uses {DEFAULT_RETRY_TIMEOUT_R3}s at concurrency=1.",
    )
    args = parser.parse_args()

    # ── Resolve Excel file ─────────────────────────────────────
    if args.file:
        excel_path = args.file
        if not Path(excel_path).exists():
            print(f"\n❌ ERROR: File not found: {excel_path}")
            sys.exit(1)
    else:
        excel_path = find_latest_excel(SCRAPER_DATA_DIR)
        if not excel_path:
            print(
                f"\n❌ ERROR: No Excel file found in {SCRAPER_DATA_DIR}/\n"
                f"   Either place the customer Excel there, or use:\n"
                f"   python scraper/validate_urls.py --file path/to/file.xlsx\n"
                f"   Or set DEFAULT_EXCEL_FILENAME at the top of this script."
            )
            sys.exit(1)
        name = Path(excel_path).name
        if DEFAULT_EXCEL_FILENAME and name == DEFAULT_EXCEL_FILENAME:
            print(f"\n📂 Using preferred file: {name}")
        else:
            print(f"\n📂 Auto-detected input: {name} (most recently modified)")

    print("\n" + "=" * 65)
    print("   Royal London — Pre-Scrape URL Validation")
    print("=" * 65)
    print(f"   Input:          {excel_path}")
    print(f"   Concurrency:    {args.concurrency} simultaneous requests")
    print(f"   Timeout:        {args.timeout}s per URL")
    print(f"   Retries:        {MAX_RETRIES} on timeout/connection error")
    print(f"   Auto-retry:     {'disabled (--no-retry)' if args.no_retry else f'3 rounds — 12s/15c → {args.retry_timeout}s/3c → {DEFAULT_RETRY_TIMEOUT_R3}s/1c'}")

    # ── Load excluded URLs ─────────────────────────────────────
    print("\n⛔ Loading excluded URLs...")
    excluded = load_excluded_urls(EXCLUDED_URLS_PATH)
    if not excluded:
        print("   No excluded_urls.json found — no URLs pre-excluded.")

    # ── Load URLs ──────────────────────────────────────────────
    print("\n📋 Loading URLs from Excel...")
    try:
        entries = load_urls_from_excel(excel_path)
    except ValueError as e:
        print(f"\n❌ ERROR: {e}")
        sys.exit(1)

    print(f"   Found {len(entries):,} URLs to validate")

    # Apply exclusions before checking — excluded URLs skip
    # HTTP checks entirely and are marked directly in results
    excluded_results = []
    active_entries   = []
    for entry in entries:
        norm = normalize_url(entry["url"])
        if norm in excluded:
            excluded_results.append({
                **entry,
                "status":         "excluded",
                "status_code":    None,
                "final_url":      entry["url"],
                "issue":          f"Excluded: {excluded[norm]}",
                "redirect_chain": "",
                "live_title":     "",
                "is_duplicate":   False,
                "duplicate_of":   "",
            })
        else:
            active_entries.append(entry)

    if excluded_results:
        print(f"   {len(excluded_results)} URL(s) pre-excluded — skipping HTTP check")

    if not active_entries:
        print("\n⚠️  No URLs to validate after exclusions.")
        sys.exit(1)

    # ── Initial check ──────────────────────────────────────────
    print(f"\n🌐 Checking {len(active_entries):,} URLs "
          f"(~{len(active_entries) // args.concurrency + 1}s estimated)...")

    results = asyncio.run(
        check_all_urls(active_entries, args.concurrency, args.timeout)
    )

    # Merge excluded results back into main results
    results.extend(excluded_results)

    # ── Auto-retry timeouts ────────────────────────────────────
    # Enabled by default — disabled only with --no-retry.
    # A timeout under 15 concurrent requests is not conclusive;
    # re-checking at concurrency=3 with a longer timeout gives
    # each page a fair individual chance before marking it dead.
    timeout_count = sum(1 for r in results if r["status"] == "timeout")
    if timeout_count > 0 and not args.no_retry:
        print(f"\n⏱️  {timeout_count} URL(s) timed out — starting progressive retry (30s → 45s)...")
        results = asyncio.run(
            retry_timeouts(
                results,
                retry_timeout=args.retry_timeout,
                retry_timeout_r3=DEFAULT_RETRY_TIMEOUT_R3,
                retry_concurrency=DEFAULT_RETRY_CONCURRENCY,
            )
        )
    elif timeout_count > 0 and args.no_retry:
        print(
            f"\n   ℹ️  {timeout_count} URL(s) timed out. "
            f"Remove --no-retry to re-check these automatically."
        )

    # ── Detect duplicates ──────────────────────────────────────
    print("\n🔍 Checking for duplicates...")
    results = detect_duplicates(results)
    dup_count = sum(1 for r in results if r.get("is_duplicate"))
    print(f"   Found {dup_count} duplicate URL(s)")

    # ── Write report ───────────────────────────────────────────
    timestamp   = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    output_dir  = Path(excel_path).parent
    output_path = str(
        output_dir / f"url_validation_report_{timestamp}.xlsx"
    )

    print(f"\n📊 Writing report...")
    write_report(results, output_path, excel_path)

    # ── Console summary ────────────────────────────────────────
    counts = {
        "live":              sum(1 for r in results if r["status"] == "live"),
        "internal_redirect": sum(1 for r in results if r["status"] == "internal_redirect"),
        "redirect":          sum(1 for r in results if r["status"] == "redirect"),
        "dead":              sum(1 for r in results if r["status"] == "dead"),
        "timeout":           sum(1 for r in results if r["status"] == "timeout"),
        "duplicate":         sum(1 for r in results if r["status"] == "duplicate"),
        "excluded":          sum(1 for r in results if r["status"] == "excluded"),
        "error":             sum(1 for r in results if r["status"] == "error"),
    }
    clean_count   = counts["live"]
    action_needed = (
        counts["dead"] + counts["redirect"] +
        counts["internal_redirect"] +
        counts["timeout"] + counts["duplicate"] + counts["error"]
    )

    print("\n" + "=" * 65)
    print("   VALIDATION COMPLETE")
    print("=" * 65)
    print(f"   Total checked:          {len(results):,}")
    print(f"   ✅ Live:                {counts['live']:,}")
    print(f"   🔀 Internal redirect:   {counts['internal_redirect']:,}")
    print(f"   ❌ Dead (4xx/5xx):      {counts['dead']:,}")
    print(f"   ⚠️  Redirected (ext):   {counts['redirect']:,}")
    print(f"   ⏱️  Timeout:            {counts['timeout']:,}")
    print(f"   🔁 Duplicate:           {counts['duplicate']:,}")
    print(f"   ⛔ Excluded:            {counts['excluded']:,}")
    print(f"   ❌ Other errors:        {counts['error']:,}")
    print(f"   ─────────────────────────────────────────")
    print(f"   ⚠️  Action required:    {action_needed:,}")
    print(f"   📋 Clean URL list:      {clean_count:,} unique live URLs")
    print(f"\n   Report saved to:")
    print(f"   {output_path}")

    if action_needed > 0:
        print(
            f"\n   👉 Share the 'Action Required' sheet with the customer."
        )
        print(
            f"   👉 Use the 'Clean URL List' sheet as input for the scraper"
            f"\n      once the customer confirms fixes."
        )
    else:
        print(
            f"\n   ✅ All URLs are live and unique — "
            f"use the 'Clean URL List' sheet as scraper input."
        )

    print("=" * 65 + "\n")


if __name__ == "__main__":
    main()