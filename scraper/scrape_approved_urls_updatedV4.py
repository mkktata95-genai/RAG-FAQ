"""
Royal London FAQ — Web Scraper v4 (Customer Approved URLs Only)
════════════════════════════════════════════════════════════
Scrapes customer-approved Royal London pages and saves structured
JSON for chunk_and_index_hqaV3.py to index into Azure AI Search.

Pipeline:
  1. Load approved URLs from Excel (header-based column detection —
     no hardcoded positions, works with any column layout)
  2. Scrape each URL with crawl4ai (main content only)
  3. Clean content (deduplicate, strip nav/footer boilerplate)
  4. Extract rich metadata (video, content type, product category,
     description, thumbnail, publish date — from HTML, zero extra calls)
  5. Save JSON → local file or Azure Blob Storage (production)

Input:
  Any Excel file with a "URL" column (header-based detection).
  Title and status columns optional — detected automatically.

Output:
  scraper/data/royal_london_faq_approved_<timestamp>.json
  (or Azure Blob Storage when AZURE_STORAGE_CONNECTION is set)

Output fields per page:
  url, title, section, audience, content, scraped_at,
  content_length, content_hash, has_video, content_type,
  product_category, description, thumbnail_url, publish_date,
  collection_name, read_time_mins

═══════════════════════════════════════════════════════════════
LOCAL USAGE
═══════════════════════════════════════════════════════════════

    # Standard scrape (reads APPROVED_EXCEL from .env / constant)
    python scraper/scrape_approved_urls_updatedV2.py

    # Custom Excel file
    python scraper/scrape_approved_urls_updatedV2.py \
        --file scraper/data/your_approved_urls.xlsx

    # Dry run — validate Excel + URL detection, no scraping
    python scraper/scrape_approved_urls_updatedV2.py --dry-run

    # Dry run on custom file
    python scraper/scrape_approved_urls_updatedV2.py \
        --file scraper/data/your_approved_urls.xlsx --dry-run

═══════════════════════════════════════════════════════════════
PRODUCTION — AZURE CONTAINER APPS JOB
═══════════════════════════════════════════════════════════════
Andy: Container Apps Job (aria-scraper-job) runs this script.
      Required env vars (set in Key Vault, not in code):

    AZURE_STORAGE_CONNECTION   — Blob Storage connection string.
                                  Not set = local mode (local dev).
                                  Set = uploads JSON to Blob Storage.
    BLOB_CONTAINER_NAME        — Blob container (default: scraper-data)
    BLOB_SCRAPED_FILENAME      — Output blob filename. Must match
                                  BLOB_SCRAPED_FILENAME in
                                  chunk_and_index_hqaV3.py so the
                                  indexer reads the correct file.
                                  Default: royal_london_faq_latest.json
    APPROVED_EXCEL_PATH        — Override the default Excel path
                                  (optional — for production URL list)

    # ADO pipeline / manual trigger:
    az containerapp job start \
        --name aria-scraper-job \
        --resource-group <rg> \
        --env-vars AZURE_STORAGE_CONNECTION=<conn> DRY_RUN=false

    # In the Container Apps Job entrypoint script:
    dry_run     = os.getenv("DRY_RUN", "false").lower() == "true"
    excel_path  = os.getenv("APPROVED_EXCEL_PATH", None)
    result      = run_scraper(excel_path=excel_path, dry_run=dry_run)

    # Run order (both jobs must be in same resource group):
    # Job 1: aria-scraper-job     → produces Blob JSON
    # Job 2: aria-indexer-job     → reads Blob JSON, indexes to AI Search
    # Both triggered by ADO pipeline in sequence after URL list update.

═══════════════════════════════════════════════════════════════
PROGRAMMATIC (DEVOPS / TESTING)
═══════════════════════════════════════════════════════════════

    from scraper.scrape_approved_urls_updatedV2 import run_scraper

    result = run_scraper()                          # default Excel
    result = run_scraper(excel_path="custom.xlsx") # custom file
    result = run_scraper(dry_run=True)             # dry run

    # Result dict:
    # {
    #   "success":       bool,
    #   "pages_scraped": int,
    #   "pages_failed":  int,
    #   "output_path":   str,  # local path or blob filename
    #   "dry_run":       bool,
    #   "error":         str,  # empty if success
    # }

═══════════════════════════════════════════════════════════════
CHANGE LOG
═══════════════════════════════════════════════════════════════

v1.0.0 — Initial version
         crawl4ai scraping from customer Excel file.
         Saves output JSON to scraper/data/ locally.

v2.0.0 — June 2026 | Mukesh Kund
         Production readiness: Blob Storage + entry point

         PRODUCTION GAP (pre v2.0.0):
         - Scraper saved output JSON to local disk only.
         - Azure Function App has no persistent local disk.
         - Output JSON must go to Azure Blob Storage so
           chunk_and_index.py can read it from a separate
           Function App invocation.
         - Excel URL source is developer-only. Production
           URL source (Blob JSON, SharePoint, CMS API) to
           be agreed with brand/marketing team and DevOps.

         save_scraped_pages() [NEW]:
         - Abstracts local vs Blob Storage output.
         - Local: saves to scraper/data/ as before.
         - Production: uploads to Azure Blob Storage.
         - Switched by AZURE_STORAGE_CONNECTION env var.
         - Not set → local mode, zero behaviour change.

         load_url_source() [NEW]:
         - Abstracts local Excel vs Blob Storage URL list.
         - Local: reads Excel file as before.
         - Production: reads JSON from Blob Storage.
         - Switched by AZURE_STORAGE_CONNECTION env var.
         - TODO: production URL source format to be agreed
           with brand/marketing team and DevOps.

         run_scraper() [NEW]:
         - Clean entry point for DevOps / Function App.
         - Returns structured result dict with stats.
         - TODO (DevOps): wrap in Function App trigger.

v3.0.0 — June 2026 | Mukesh Kund
         Rich metadata extraction — video detection, content type,
         product category, audience, publish date, thumbnail

         WHY:
         - Royal London pages contain videos (webinars, guides,
           product explainers) on ANY page type — not just /webinars/.
           Videos appear on pension pages, insurance pages, tools,
           existing customer pages etc.
         - UI/UX team will need metadata to render rich citations
           (video cards, product badges, publish dates, thumbnails)
           without requiring a re-index.
         - Principle: index once, serve many use cases. All metadata
           extracted at scrape time from HTML already fetched by
           crawl4ai — zero extra HTTP calls, zero extra cost.

         VIDEO DETECTION STRATEGY:
         - Royal London uses a proprietary/JS-rendered video player.
           The video embed URL is NOT reliably extractable from HTML.
         - Decision: source_url IS the video reference. UI team links
           to the page — the video is on that page.
         - Detection uses three independent signals (any one = True):
             Signal 1: URL pattern (/webinars/, /videos/, /video/)
             Signal 2: meta-Collection_name contains "webinar"/"video"
             Signal 3: rendered HTML contains video player CSS classes
                       or data attributes (fallback, less reliable)
         - has_video field: True/False stored per page and per chunk.
           UI team uses this to render a "📹 Watch video" indicator.

         METADATA EXTRACTION — extract_page_metadata() [NEW]:
         - Parses result.html (already fetched by crawl4ai, no extra
           HTTP call) using BeautifulSoup to extract:
             has_video        — bool: page contains video content
             content_type     — webinar/guide/article/faq/tool/news
             product_category — pensions/insurance/isa/retirement/general
             audience         — customer/adviser/employer (from URL)
             description      — from meta-description or og:description
             thumbnail_url    — from meta-teaser_image or og:image
             publish_date     — from meta-st-publish-date (ISO format)
             collection_name  — from meta-Collection_name (e.g. "Pension webinar")
             read_time_mins   — estimated from word count (200 wpm)

         OUTPUT FORMAT UPDATED:
         - Old: url, title, section, audience, content, scraped_at, content_length
         - New: + has_video, content_type, product_category, description,
                  thumbnail_url, publish_date, collection_name, read_time_mins
         - All new fields have safe defaults — if extraction fails,
           defaults are used and scraping continues normally.
         - chunk_and_index.py passes all new fields through to index.

         audience FIELD FIX:
         - Was hardcoded to "customer" regardless of URL.
         - Now derived from URL: adviser.royallondon.com → "adviser",
           employer.royallondon.com → "employer", else → "customer".
         - Matches the audience derivation in extract_page_metadata().

v4.0.0 — July 2026 | Mukesh Kund
         Production hardening: dotenv import fix, URL normalisation,
         content_hash field.

         CRITICAL BUG FIX — dotenv import missing:
         find_dotenv() and load_dotenv() were called at module level
         but neither was imported — NameError on startup.
         Fixed: added from dotenv import load_dotenv, find_dotenv.
         Also added override=True so .env always beats shell vars.

         URL NORMALISATION — normalize_url() helper:
         load_approved_pages() only stripped trailing slash + query.
         No path lowercasing meant should-I vs should-i survived
         deduplication as two separate URLs — double-indexing same
         content. normalize_url() lowercases path only (domain kept).
         Applied in load_approved_pages() (dedup key + stored URL)
         and in scrape_page() output (defence-in-depth for redirects).

         CONTENT HASH:
         SHA-256 hash of page content added to scrape output.
         Used by content_freshness.py (Sprint 2) to detect changed
         pages without re-scraping all 350 URLs.

v4.1.0 — July 2026 | Mukesh Kund
         Live HTTP status check at scrape time.

         ROOT CAUSE:
         The Excel "status" column reflects verification-time status —
         recorded when the Excel was last built, potentially weeks
         before a scrape run. scrape_page() only checked
         result.success from crawl4ai — but success=True just means
         the browser loaded a page; a 404 error page still
         "successfully" loads. result.status_code was never checked.

         FIX:
         scrape_page() now reads result.status_code after the
         crawl4ai fetch. status_code >= 400 → page dropped and
         logged as scrape_http_error with the actual code.
         status_code is None (crawl4ai occasionally doesn't
         populate it) → fails open, does not reject, so working
         pages are never dropped over a missing field.

v4.2.0 — July 2026 | Mukesh Kund
         Header-based column detection — no hardcoded positions.

         ROOT CAUSE:
         load_approved_pages() assumed fixed column layout
         (title=col B, url=col C, status=col E) from the internal
         verification file. A customer-supplied Excel is not
         guaranteed to match that layout. A URL-only file (2 columns)
         would skip EVERY row due to `len(row) < 5` guard —
         silently returning zero pages with no error.

         FIX — detect columns by HEADER NAME (row 1):
         URL_HEADERS    = {"url", "page url", "link", ...}
         TITLE_HEADERS  = {"title", "page title", "name"}
         STATUS_HEADERS = {"status", "status code", "http status"}
         All matched case-insensitively. URL column REQUIRED —
         raises ValueError with actual headers if not found.
         Title and status are optional.

         STATUS HANDLING:
         Only skips on unambiguous dead signals (HTTP >= 400,
         or words like dead/broken/removed). Blank, "200", "OK",
         "Live", or unrecognised values are KEPT — ambiguity
         defers to "keep it, let the live scrape decide".

v4.3.0 — July 2026 | Mukesh Kund
         Excel Category column used as primary content_type source.
         derive_section() function added (was missing — NameError).

         BUG FIX — derive_section() never defined:
         scrape_page() called derive_section(url) but the function
         was missing from the codebase — would crash with NameError
         on every scrape. Added derive_section() using SECTION_MAP.

         Excel Category as primary content_type:
         - Customer supplies Category per URL (Brand/Guidance/Other/
           Product/Tool) — more authoritative than URL-pattern.
         - CATEGORY_HEADERS set added to column detection.
         - _EXCEL_CATEGORY_MAP: Brand/Other/Product → article,
           Guidance → guide, Tool → tool.
         - URL-pattern still wins for high-signal types (webinar,
           video, faq, news) — a "Product" page on /webinars/ is
           still correctly typed as "webinar".
         - Fallback: Category absent or unrecognised → derive_content_type().
         - No changes to chunk_and_index_hqaV3.py required.

v4.4.0 — July 2026 | Mukesh Kund
         Multi-state dropdown scraping + 4 bug fixes.

         FEATURE — Multi-state dropdown scraping:
         Pages like /tell-us-about-a-bereavement/ render different
         phone numbers and contact details per dropdown selection
         (e.g. Scottish Provident → 0345 646 2096, Royal London/
         Bright Grey → 0345 646 2108). With wait_until="domcontentloaded"
         and a single arun() call the scraper only captured the
         default/unselected state — all per-policy contact numbers
         were missing from the index.

         FIX — _detect_routing_dropdowns() + _scrape_dropdown_states():
         After the initial page load, the scraper checks the rendered
         HTML for <select> elements with more than one non-placeholder
         option. If found, it iterates each option by injecting JS to
         fire 'input' + 'change' events (same technique validated in
         manual Playwright testing), waits 1.5 s for the DOM to settle,
         then captures the updated body text. A line-by-line diff
         against the default state extracts ONLY the content that
         changed (phone numbers, addresses, form links) — shared
         static content is not duplicated. Each option produces a
         separate page_data entry with a synthetic URL
         (base_url + #policy=<option_value>) so the indexer treats
         them as distinct documents and chunk_and_index_hqaV3.py does
         not deduplicate them. The base page (default state) is still
         indexed as the primary document.

         scrape_page() return type widened to list[dict] | dict | None:
         - Standard page → dict (unchanged behaviour)
         - Dropdown page → list[dict] (base + one entry per option)
         - Failure → None (unchanged behaviour)
         Both _run() (run_scraper) and main() batch loops updated to
         flatten list returns correctly. content_length guard for
         dropdown entries lowered to 20 chars (phone number + label
         is valid content even at short length).

         BUG FIX #1 — asyncio.run() crashes in async contexts:
         run_scraper() used asyncio.run(_run()) which raises
         RuntimeError when called from an already-running event loop
         (Azure Functions, Jupyter, FastAPI). Fixed with
         nest_asyncio.apply() + asyncio.get_event_loop().run_until_complete()
         pattern, with graceful fallback to asyncio.run() when
         nest_asyncio is not installed (plain script usage unchanged).

         BUG FIX #2 — read_time_mins stored as str not int:
         extract_page_metadata() stored read_time_mins as str(read_time)
         — e.g. "5" — making downstream arithmetic (sum, average) fail
         silently. Fixed: stored as int. Default also changed from
         "5" (str) to 5 (int).

         BUG FIX #3 — traceback and asyncio imported inside functions:
         traceback was imported inside run_scraper(); asyncio was
         imported twice (module level + inside run_scraper()). Both
         moved to top-level imports.

         BUG FIX #4 — save_scraped_pages() saves empty JSON silently:
         If all URLs failed, save_scraped_pages() wrote [] to disk
         with no warning — next indexer run would silently wipe the
         index. Fixed: early-exit with log.error and return "" when
         results list is empty. Callers check for empty output_path
         and surface the failure in the result dict.

v4.5.0 — July 2026 | Mukesh Kund
         Playwright-based dropdown scraping + base page deduplication fix.

         PROBLEM WITH v4.4.0 DROPDOWN APPROACH:
         _scrape_dropdown_states() called crawler.arun() a SECOND time
         to detect dropdowns — this reloaded the page with
         wait_until="networkidle" causing 30s timeouts on Royal London
         contact pages (background XHR activity never fully settles).
         Result: dropdown_detect_failed warning on most pages, and the
         bereavement page #policy= variants were never captured.

         ADDITIONALLY: The base page crawl4ai scrape on dropdown pages
         (e.g. /tell-us-about-a-bereavement/) captured ALL option
         content rendered in the DOM simultaneously (Royal London
         renders hidden panels for every option at page load). This
         caused the content to repeat 3x in the scraped output — the
         existing remove_duplicate_content() only catches H1/H2-level
         duplication, not inline paragraph repetition.

         FIX 1 — Playwright for dropdown pages:
         After crawl4ai scrapes the base page, BeautifulSoup checks
         result.html for <select> elements with >1 non-placeholder
         option. If found, a Playwright browser handles the dropdown
         interaction — single page load (wait_until="networkidle",
         timeout=45000ms), then per-option JS event injection +
         1.5s DOM wait + body text diff. This exactly mirrors the
         proven standalone Playwright script (crawler.py v0.1.0)
         that correctly captured all 13 bereavement policy options.
         Playwright runs in a thread pool executor to avoid blocking
         the crawl4ai asyncio event loop.

         FIX 2 — Base page content truncation for dropdown pages:
         When a page has routing dropdowns, the crawl4ai base page
         content is truncated at the first dropdown-related marker
         (select, dropdown, policy, please select etc.) to remove
         the repeated option content that Royal London renders inline.
         This gives a clean intro paragraph as the base page — exactly
         matching what the proven Playwright script captured as the
         default state (234 words, clean intro only).

         FIX 3 — Page timeout 30s → 45s for dropdown detection:
         All CrawlerRunConfig page_timeout values inside dropdown
         handling raised from 30000 → 45000ms to match the Playwright
         script's proven timeout value.

         NEW FUNCTIONS:
         - _has_routing_dropdowns_in_html(): BeautifulSoup check on
           already-fetched HTML — zero extra network call.
         - _scrape_dropdown_states_playwright(): Playwright thread
           that mirrors crawler.py _scrape_multi_state_page() exactly.
         - _truncate_base_content_at_dropdown(): strips repeated option
           content from base page markdown before storing.

         REMOVED:
         - _JS_DETECT_DROPDOWNS, _JS_SELECT_OPTION, _JS_GET_BODY_TEXT
           constants (crawl4ai JS injection approach — replaced by
           Playwright).
         - _scrape_dropdown_states() async function (replaced by
           _scrape_dropdown_states_playwright()).

═══════════════════════════════════════════════════════════════
"""

import argparse
import asyncio
import concurrent.futures
import hashlib
import os
import json
import re
import sys
import time
import traceback
from pathlib import Path
from datetime import datetime, timezone

# nest_asyncio allows asyncio.run()-equivalent to work inside an already-running
# event loop (Azure Functions, FastAPI, Jupyter). Optional — plain script use
# works without it via the except ImportError fallback in run_scraper().
try:
    import nest_asyncio as _nest_asyncio
    _NEST_ASYNCIO_AVAILABLE = True
except ImportError:
    _NEST_ASYNCIO_AVAILABLE = False

import structlog
from bs4 import BeautifulSoup       # v3.0.0: meta tag + video detection
from openpyxl import load_workbook
from crawl4ai import AsyncWebCrawler, BrowserConfig, CrawlerRunConfig
from crawl4ai.markdown_generation_strategy import DefaultMarkdownGenerator
from crawl4ai.content_filter_strategy import PruningContentFilter

# v4.0.0 FIX: find_dotenv / load_dotenv were called but never imported
# — caused NameError on startup. Added the import.
# override=True ensures .env always wins over shell environment
# variables — prevents silent misconfiguration in DevOps pipelines.
from dotenv import load_dotenv, find_dotenv

_dotenv_path = find_dotenv(usecwd=False)
load_dotenv(_dotenv_path, override=True)
log = structlog.get_logger()

# ── Config ────────────────────────────────────────────
APPROVED_EXCEL = "scraper/data/royal_london_verification_retried.xlsx"
BATCH_SIZE = 5
BATCH_DELAY_SECONDS = 2

# ── Production: Azure Blob Storage ─────────────────────────────
# TODO (DevOps): Set these in Azure Key Vault before go-live.
#
# AZURE_STORAGE_CONNECTION — Blob Storage connection string.
#   Not set (default) → local mode, saves to scraper/data/.
#   Set in production → uploads JSON to Blob Storage instead.
#
# BLOB_CONTAINER_NAME — container name in Blob Storage.
#   DevOps creates this in the Azure Storage Account.
#   Default: "scraper-data"
#
# BLOB_SCRAPED_FILENAME — filename written to Blob Storage.
#   chunk_and_index.py reads this exact filename.
#   Both scripts must agree on the same filename.
#   Default: "royal_london_faq_latest.json"
#
# TODO: Production URL source format TBD with client.
#   Currently reads approved URLs from Excel (dev-only workflow).
#   Production options (agree with brand/marketing + DevOps):
#     Option A: JSON file in Blob Storage (content team uploads)
#     Option B: SharePoint list via Microsoft Graph API
#     Option C: CMS API endpoint from Royal London website team
#   Until decided, Excel path works for local + manual prod runs.
BLOB_STORAGE_CONNECTION = os.getenv("AZURE_STORAGE_CONNECTION", "")
BLOB_CONTAINER_NAME     = os.getenv("BLOB_CONTAINER_NAME", "scraper-data")
BLOB_SCRAPED_FILENAME   = os.getenv(
    "BLOB_SCRAPED_FILENAME",
    "royal_london_faq_latest.json",
)

# Section mapping from first URL path segment
SECTION_MAP = {
    "existing-customers":       "Existing Customers",
    "insurance":                "Insurance",
    "pensions":                 "Pensions",
    "guides-tools":             "Guides and Tools",
    "retirement-planning":      "Retirement Planning",
    "isa":                      "ISA",
    "profitshare":              "ProfitShare",
    "about-us":                 "About Us",
    "find-a-financial-adviser": "Find a Financial Adviser",
    "accessibility":            "Accessibility",
    "informational-pages":      "Information",
}


def derive_section(url: str) -> str:
    """
    Derive section label from first URL path segment.

    Uses SECTION_MAP — first segment of the path after the domain.
    Falls back to "General" if no segment matches.

    Examples:
        .../pensions/...            → "Pensions"
        .../existing-customers/...  → "Existing Customers"
        .../isa/...                 → "ISA"
        .../unknown/...             → "General"
    """
    try:
        path = url.split("://", 1)[-1]          # strip scheme
        path = path.split("/", 1)[-1]            # strip domain
        first_segment = path.split("/")[0].lower()
        return SECTION_MAP.get(first_segment, "General")
    except Exception:
        return "General"


# ── Step 0: Metadata extraction (v3.0.0) ────────────────────

# CSS class names and data attributes that Royal London uses
# for their video player components. Detected in rendered HTML
# as a fallback signal when meta tags don't confirm video.
# These were identified from inspecting Royal London page source.
# Update this list if new player classes are discovered.
VIDEO_CSS_SIGNALS = [
    "video-player",
    "webinar-player",
    "brightcove-player",
    "bc-player",
    "vjs-tech",          # Video.js (common enterprise player)
    "kaltura-player",
    "jwplayer",
    "data-video-id",
    "data-webinar-id",
    "data-brightcove",
]

# URL path segments that reliably indicate video content.
# Any segment present in the page URL → has_video = True.
VIDEO_URL_SIGNALS = [
    "/webinars/",
    "/videos/",
    "/video/",
    "/webinar/",
]

# meta-Collection_name values that indicate video pages.
# crawl4ai renders these as text in the HTML head.
# Case-insensitive matching used.
VIDEO_COLLECTION_SIGNALS = [
    "webinar",
    "video",
    "podcast",          # audio/video content
]

# Product category keyword mapping.
# URL path segments → product_category value.
# Order matters — first match wins.
PRODUCT_CATEGORY_MAP = [
    ("/pension",              "pensions"),
    ("/retirement",           "retirement"),
    ("/life-insurance",       "life_insurance"),
    ("/life-cover",           "life_insurance"),
    ("/whole-of-life",        "life_insurance"),
    ("/income-protection",    "income_protection"),
    ("/critical-illness",     "critical_illness"),
    ("/illness-income",       "income_protection"),
    ("/isa",                  "isa"),
    ("/investments",          "investments"),
    ("/investment",           "investments"),
    ("/fund",                 "investments"),
    ("/funeral",              "funeral"),
    ("/profitshare",          "profitshare"),
    ("/financial-adviser",    "financial_advice"),
    ("/find-a-financial",     "financial_advice"),
    ("/about-us",             "corporate"),
    ("/media",                "corporate"),
    ("/existing-customers",   "customer_support"),
]

# Content type mapping from URL path segments.
# Order matters — more specific patterns first.
CONTENT_TYPE_MAP = [
    ("/webinars/",            "webinar"),
    ("/videos/",              "video"),
    ("/video/",               "video"),
    ("/guides-tools/",        "guide"),
    ("/pension-calculator",   "tool"),
    ("/retirement-planner",   "tool"),
    ("/lump-sum-calculator",  "tool"),
    ("/risk-profiler",        "tool"),
    ("/calculator",           "tool"),
    ("/planner",              "tool"),
    ("/existing-customers/",  "faq"),
    ("/help-and-support/",    "faq"),
    ("/pensions-explained",   "faq"),
    ("/about-us/",            "corporate"),
    ("/media/",               "news"),
    ("/press-release",        "news"),
    ("/news/",                "news"),
    ("/agm/",                 "corporate"),
]


def derive_content_type(url: str) -> str:
    """
    Derive content_type from URL path pattern.

    Uses CONTENT_TYPE_MAP — first match wins (most specific first).
    Falls back to "article" if no pattern matches.

    Returns one of:
        webinar, video, guide, tool, faq, corporate, news, article
    """
    url_lower = url.lower()
    for pattern, content_type in CONTENT_TYPE_MAP:
        if pattern in url_lower:
            return content_type
    return "article"


def derive_product_category(url: str) -> str:
    """
    Derive product_category from URL path pattern.

    Uses PRODUCT_CATEGORY_MAP — first match wins.
    Falls back to "general" if no pattern matches.

    Returns one of:
        pensions, retirement, life_insurance, income_protection,
        critical_illness, isa, investments, funeral, profitshare,
        financial_advice, corporate, customer_support, general
    """
    url_lower = url.lower()
    for pattern, category in PRODUCT_CATEGORY_MAP:
        if pattern in url_lower:
            return category
    return "general"


def derive_audience_from_url(url: str) -> str:
    """
    Derive audience from URL domain/path.

    v3.0.0 FIX: was hardcoded to "customer" in scrape_page().
    Now properly derived from URL:
        adviser.royallondon.com  → "adviser"
        employer.royallondon.com → "employer"
        /adviser/ in path        → "adviser"
        /employer/ in path       → "employer"
        everything else          → "customer"
    """
    url_lower = url.lower()
    if "adviser.royallondon.com" in url_lower or "/adviser/" in url_lower:
        return "adviser"
    if "employer.royallondon.com" in url_lower or "/employer/" in url_lower:
        return "employer"
    return "customer"


def detect_video_from_html(html: str, url: str) -> bool:
    """
    Detect whether a page contains video content.

    Uses three independent signals — any one True = has_video:

    Signal 1 — URL pattern (most reliable):
        URL contains /webinars/, /videos/, /video/, /webinar/

    Signal 2 — Meta tag (very reliable):
        meta-Collection_name contains "webinar" or "video"
        Parsed from HTML <head> — always present, not JS-dependent.

    Signal 3 — HTML class/attribute (fallback):
        Rendered HTML contains known video player CSS class names
        or data attributes (VIDEO_CSS_SIGNALS list).
        Less reliable as Royal London's player is JS-rendered and
        may not appear at domcontentloaded — but included as belt-
        and-suspenders.

    Args:
        html: Raw HTML string from crawl4ai result.html
        url:  Page URL for Signal 1 check

    Returns:
        bool — True if any signal indicates video content
    """
    url_lower = url.lower()

    # Signal 1: URL pattern
    for pattern in VIDEO_URL_SIGNALS:
        if pattern in url_lower:
            return True

    # Signal 2 + 3: parse HTML
    if not html:
        return False

    try:
        soup = BeautifulSoup(html, "html.parser")

        # Signal 2: meta-Collection_name tag
        # Royal London sets this in the page head as:
        # <meta name="Collection_name" content="Pension webinar">
        collection_meta = (
            soup.find("meta", attrs={"name": "Collection_name"}) or
            soup.find("meta", attrs={"name": "collection_name"}) or
            soup.find("meta", property="Collection_name")
        )
        if collection_meta:
            collection_val = (
                collection_meta.get("content", "") or ""
            ).lower()
            for signal in VIDEO_COLLECTION_SIGNALS:
                if signal in collection_val:
                    return True

        # Also check og:type = "video" or similar
        og_type = soup.find("meta", property="og:type")
        if og_type and "video" in (og_type.get("content", "") or "").lower():
            return True

        # Signal 3: HTML class/attribute scan
        # Check rendered HTML (not just head) for video player signals
        html_lower = html.lower()
        for signal in VIDEO_CSS_SIGNALS:
            if signal in html_lower:
                return True

    except Exception as e:
        log.warning(
            "video_detection_parse_error",
            url=url,
            error=str(e),
        )

    return False


def extract_page_metadata(html: str, url: str) -> dict:
    """
    Extract rich metadata from page HTML for index enrichment.

    Parses result.html (already fetched by crawl4ai — zero extra
    HTTP calls, zero extra cost) using BeautifulSoup.

    All extractions are defensive — any failure returns safe
    defaults so scraping continues normally.

    Args:
        html: Raw HTML string from crawl4ai result.html
        url:  Page URL for URL-based derivations

    Returns dict with keys:
        has_video        (bool) — page contains video content
        content_type     (str)  — webinar/guide/article/faq/tool/news
        product_category (str)  — pensions/insurance/isa/etc
        audience         (str)  — customer/adviser/employer
        description      (str)  — page description for UI previews
        thumbnail_url    (str)  — teaser/og image URL
        publish_date     (str)  — ISO date string or ""
        collection_name  (str)  — e.g. "Pension webinar"
        read_time_mins   (str)  — estimated reading time

    All string fields default to "" if not found.
    has_video defaults to False if detection fails.
    """
    # Safe defaults — used if any extraction fails
    metadata = {
        "has_video":        False,
        "content_type":     derive_content_type(url),
        "product_category": derive_product_category(url),
        "audience":         derive_audience_from_url(url),
        "description":      "",
        "thumbnail_url":    "",
        "publish_date":     "",
        "collection_name":  "",
        "read_time_mins":   5,    # default 5 min if calculation fails (int, not str)
    }

    # Video detection — uses URL + HTML signals
    metadata["has_video"] = detect_video_from_html(html, url)

    if not html:
        # No HTML available — return URL-derived defaults
        return metadata

    try:
        soup = BeautifulSoup(html, "html.parser")

        # ── Description ──────────────────────────────────────
        # Priority: meta-description > og:description > st-description
        # UI team uses this for citation preview tooltips.
        for attr, key in [
            ({"name": "description"}, "content"),
            ({"property": "og:description"}, "content"),
            ({"name": "st-description"}, "content"),
        ]:
            tag = soup.find("meta", attrs=attr)
            if tag and tag.get(key, "").strip():
                metadata["description"] = tag[key].strip()[:300]
                break

        # ── Thumbnail URL ─────────────────────────────────────
        # Priority: meta-teaser_image > og:image
        # meta-teaser_image is Royal London's custom teaser image
        # (350x200px, page-specific). og:image is usually the RL logo.
        # UI team can use teaser_image for rich citation cards.
        teaser = soup.find("meta", attrs={"name": "teaser_image"})
        if teaser and teaser.get("content", "").strip():
            metadata["thumbnail_url"] = teaser["content"].strip()
        else:
            og_image = soup.find("meta", property="og:image")
            if og_image and og_image.get("content", "").strip():
                # Only store if it's Royal London hosted (not generic logo)
                img_url = og_image["content"].strip()
                if "rl-logo-meta-image" not in img_url:
                    metadata["thumbnail_url"] = img_url

        # ── Publish date ──────────────────────────────────────
        # Royal London sets meta-st-publish-date as human-readable
        # e.g. "13 March 2024". Parse to ISO format for indexing.
        pub_date_tag = soup.find(
            "meta", attrs={"name": "st-publish-date"}
        )
        if pub_date_tag and pub_date_tag.get("content", "").strip():
            raw_date = pub_date_tag["content"].strip()
            try:
                # Parse "13 March 2024" → "2024-03-13"
                parsed = datetime.strptime(raw_date, "%d %B %Y")
                metadata["publish_date"] = parsed.strftime("%Y-%m-%d")
            except ValueError:
                # If format differs, store raw value
                metadata["publish_date"] = raw_date[:20]

        # ── Collection name ───────────────────────────────────
        # e.g. "Pension webinar", "Life insurance guide"
        # UI team can use this for content category badges.
        collection_tag = soup.find(
            "meta", attrs={"name": "Collection_name"}
        )
        if collection_tag and collection_tag.get("content", "").strip():
            metadata["collection_name"] = (
                collection_tag["content"].strip()[:100]
            )

        # ── Read time estimate ────────────────────────────────
        # Estimate from visible text word count at 200 wpm.
        # For webinar pages: transcript is ~5,000-8,000 words →
        # 25-40 min. For articles: ~500-1,500 words → 2-7 min.
        # UI team can display "8 min read" or "30 min webinar".
        body_text = soup.get_text(separator=" ", strip=True)
        word_count = len(body_text.split())
        read_time = max(1, round(word_count / 200))
        metadata["read_time_mins"] = int(read_time)  # v4.4.0: int not str

    except Exception as e:
        log.warning(
            "metadata_extraction_error",
            url=url,
            error=str(e),
            note="Using safe defaults for this page",
        )

    return metadata


# ── URL normalisation helper ──────────────────────────────────
def normalize_url(url: str) -> str:
    """
    Canonical URL form for deduplication and index consistency.

    Strips trailing slash, query string, fragment, and lowercases
    the path. Domain is left as-is (already lowercase in practice).

    WHY lowercase the path:
    Customer Excel contained both:
        .../should-I-consolidate-my-pensions
        .../should-i-consolidate-my-pensions
    Without path lowercasing these survive deduplication as two
    separate entries and are scraped + indexed twice — duplicate
    content with different chunk_ids causing retrieval noise.
    """
    url = url.strip()
    url = url.split("?")[0].split("#")[0]   # strip query + fragment
    url = url.rstrip("/")                    # strip trailing slash
    if "://" in url:
        scheme_host, _, path = url.partition("://")
        domain_end = path.find("/")
        if domain_end == -1:
            return f"{scheme_host}://{path}"
        domain = path[:domain_end]
        rest   = path[domain_end:].lower()  # lowercase path only
        return f"{scheme_host}://{domain}{rest}"
    return url.lower()


# ── Excel column detection + URL loading ─────────────────────
def load_approved_pages(excel_path: str) -> list[dict]:
    """
    Reads customer Excel, returns list of dicts:
    [{"url": "...", "title": "..."}]

    Deduplicates by normalized URL (keeps first occurrence's title).

    v4.2.0 — COLUMN DETECTION BY HEADER NAME, NOT FIXED POSITION:
    Previous versions assumed a fixed column layout (title at
    column B, url at column C, status at column E) copied from
    the internal verification file
    (royal_london_verification_retried.xlsx), which included a
    status column from an earlier internal check pass. A real
    customer-supplied Excel is NOT guaranteed to have that
    layout, or even a status column at all — the customer
    typically supplies ONLY a URL column, sometimes with a title
    column, and no status information whatsoever.

    Previous behaviour on a URL-only or URL+title file:
    `if not row or len(row) < 5: continue` would skip EVERY row
    (fewer than 5 columns), silently producing zero pages with
    no error — the worst possible failure mode.

    Column detection (case-insensitive, header = row 1):
      URL column    (REQUIRED): "url", "page url", "link",
                                 "webpage", "web page", "web url"
      Title column  (optional): "title", "page title", "name"
      Status column (optional): "status", "status code",
                                 "http status"

    If no URL column is found, raises ValueError listing the
    headers actually present — an immediate, clear failure
    instead of a silent empty result.

    STATUS HANDLING — Excel status is a pre-filter hint, NEVER
    the final authority:
    - If a status column exists: rows are skipped ONLY when the
      value is an unambiguous dead signal — a numeric HTTP code
      >= 400, or a word like "dead"/"broken"/"404"/"removed"/
      "gone"/"not found". Anything else (a plain "200", "OK",
      "Live", blank, or unrecognised text) is KEPT — customer
      status notation isn't guaranteed to be a numeric HTTP code,
      so being lenient avoids wrongly discarding a working URL
      over an ambiguous cell value.
    - If NO status column exists at all: every row with a valid
      URL is kept as a candidate.
    - Either way, the LIVE HTTP status_code check in scrape_page()
      (v4.1.0) is what actually determines whether a URL is live
      at scrape time — this function only avoids wasting scrape
      effort on links already known-dead when that information
      happens to be available in the Excel.
    """
    wb = load_workbook(excel_path, read_only=True)
    ws = wb.active

    # ── Detect columns by header name (row 1) ──────────────────
    # Sets of recognised header names (case-insensitive).
    # Adding new synonyms here is all that's needed if the
    # customer changes their column heading in future.
    URL_HEADERS      = {"url", "page url", "link", "webpage", "web page", "web url"}
    TITLE_HEADERS    = {"title", "page title", "name"}
    STATUS_HEADERS   = {"status", "status code", "http status"}
    CATEGORY_HEADERS = {"category", "content category", "page category", "type"}

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    headers    = [
        str(h).strip().lower() if h is not None else ""
        for h in header_row
    ]

    def find_col(candidates: set) -> int | None:
        for idx, h in enumerate(headers):
            if h in candidates:
                return idx
        return None

    url_idx      = find_col(URL_HEADERS)
    title_idx    = find_col(TITLE_HEADERS)
    status_idx   = find_col(STATUS_HEADERS)
    category_idx = find_col(CATEGORY_HEADERS)

    if url_idx is None:
        wb.close()
        raise ValueError(
            f"load_approved_pages: no URL column found in "
            f"{excel_path!r}. Expected a header matching one of "
            f"{sorted(URL_HEADERS)} (case-insensitive). "
            f"Headers actually found: {header_row!r}"
        )

    log.info(
        "approved_pages_columns_detected",
        url_column=headers[url_idx],
        title_column=headers[title_idx] if title_idx is not None else None,
        status_column=headers[status_idx] if status_idx is not None else None,
        category_column=headers[category_idx] if category_idx is not None else None,
        has_status_column=status_idx is not None,
        has_category_column=category_idx is not None,
    )

    def _is_dead_status(value) -> bool:
        """
        Lenient dead-link detector for an OPTIONAL status column.
        Only returns True for unambiguous dead signals — see
        load_approved_pages() docstring for rationale.
        Never used to REQUIRE a "200" — absence or ambiguity
        always means "keep it, let the live scrape decide".
        """
        if value is None:
            return False
        s = str(value).strip().lower()
        if not s:
            return False
        try:
            code = int(s)
            return code >= 400
        except ValueError:
            pass
        dead_words = {"dead", "broken", "404", "removed", "gone", "not found"}
        return any(w in s for w in dead_words)

    seen           = set()
    pages          = []
    total_rows     = 0
    skipped_status = 0
    duplicates     = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) <= url_idx:
            continue

        url = str(row[url_idx]).strip() if row[url_idx] else ""
        if not url.startswith("http"):
            continue

        raw_title = (
            str(row[title_idx]).strip()
            if title_idx is not None and len(row) > title_idx and row[title_idx]
            else ""
        )
        status_value = (
            row[status_idx]
            if status_idx is not None and len(row) > status_idx
            else None
        )
        excel_category = (
            str(row[category_idx]).strip().lower()
            if category_idx is not None and len(row) > category_idx and row[category_idx]
            else ""
        )

        total_rows += 1

        if _is_dead_status(status_value):
            skipped_status += 1
            log.debug(
                "url_skipped_dead_status",
                url=url,
                status_value=status_value,
            )
            continue

        normalized = normalize_url(url)

        if normalized in seen:
            duplicates.append(url)
            log.debug(
                "url_duplicate_skipped",
                raw_url=url,
                normalized=normalized,
            )
            continue

        seen.add(normalized)

        # Clean title — strip " - Royal London" suffix variants
        title = raw_title
        for suffix in [
            " - Royal London", " | Royal London",
            "- Royal London",  "| Royal London",
        ]:
            if title.endswith(suffix):
                title = title[: -len(suffix)].strip()
                break

        pages.append({"url": normalized, "title": title, "excel_category": excel_category})

    wb.close()

    # Audit log — production canary for Excel data quality
    log.info(
        "approved_pages_loaded",
        total_rows=total_rows,
        skipped_dead_status=skipped_status,
        duplicates_removed=len(duplicates),
        unique_pages=len(pages),
        had_status_column=status_idx is not None,
        had_title_column=title_idx is not None,
        had_category_column=category_idx is not None,
    )
    if duplicates:
        log.warning(
            "duplicate_urls_in_excel",
            count=len(duplicates),
            examples=duplicates[:5],
        )
    if not pages:
        log.warning(
            "approved_pages_empty",
            note=(
                "No pages loaded — check that the URL column was "
                "detected correctly and rows contain http(s) URLs."
            ),
        )

    return pages



# ── Step 2: Content cleaning ───────────────────────────
def remove_duplicate_content(content: str) -> str:
    """
    Remove exact duplicate of article content.
    crawl4ai sometimes scrapes the page twice.

    Finds H1/H2 headings — if the second one starts after
    40% of content AND has >60% word overlap with the first
    chunk, it's a true duplicate → cut it.
    """
    h1_pattern = re.compile(r'^#{1,2}\s+\S', re.MULTILINE)
    matches = list(h1_pattern.finditer(content))

    if len(matches) < 2:
        return content

    second_pos = matches[1].start()
    content_len = len(content)

    if second_pos < content_len * 0.40:
        return content

    first_chunk = content[:second_pos].strip()
    second_chunk = content[second_pos:].strip()

    first_words = set(first_chunk.split()[:200])
    second_words = set(second_chunk.split()[:200])

    if not first_words:
        return content

    overlap_pct = len(first_words & second_words) / len(first_words)

    if overlap_pct > 0.60:
        log.info(
            "duplicate_removed",
            overlap_pct=round(overlap_pct, 2),
            chars_removed=len(second_chunk),
        )
        return first_chunk

    return content


def clean_content(content: str) -> str:
    """
    Safe content cleaning — removes only pure UI noise.
    No meaningful content removed.

    NOTE: External (non-royallondon) URL stripping is NOT
    done here — chunk_and_index.py handles that separately
    at chunk time.
    """

    # Step 1: Remove duplicate article copy
    content = remove_duplicate_content(content)

    # Step 2: Remove breadcrumb navigation
    # "1. [ Home ](url) >" or last item "5. Section Name"
    content = re.sub(
        r'^\s*\d+\.\s*\[.*?\]\(.*?\)\s*>\s*$',
        '', content, flags=re.MULTILINE,
    )
    content = re.sub(
        r'^\s*\d+\.\s*\[.*?\]\(.*?\)\s*$',
        '', content, flags=re.MULTILINE,
    )
    content = re.sub(
        r'^\s*\d+\.\s+[A-Z][^\n]{3,60}$',
        '', content, flags=re.MULTILINE,
    )

    # Step 3: Remove social share sections
    # "Share" heading + following empty bullets
    content = re.sub(
        r'Share\s*\n(\s*\*\s*(\[?\s*\]?\([^\)]*\))?\s*\n)+',
        '', content,
    )
    content = re.sub(
        r'^\s*\*\s*\[?\s*\]?\(\s*[^\)]{0,10}\)\s*$',
        '', content, flags=re.MULTILINE,
    )
    content = re.sub(
        r'^Share\s*$',
        '', content, flags=re.MULTILINE,
    )

    # Step 4: Remove Twitter share links
    content = re.sub(
        r'\[?\s*\]?\(https://twitter\.com/intent/tweet[^\)]*\)\s*',
        '', content,
    )

    # Step 5: Remove other social media URLs
    content = re.sub(
        r'https://www\.(facebook|instagram|linkedin|x|youtube|twitter)\.com/\S+',
        '', content,
    )

    # Step 6: Remove empty markdown links
    content = re.sub(r'\[\s*\]\(\s*\)', '', content)
    content = re.sub(
        r'^\s*\*\s*\[\s*\]\s*$', '', content, flags=re.MULTILINE,
    )

    # Step 7: Remove Previous/Next Item labels
    content = re.sub(
        r'^(Previous Item|Next Item)\s*$',
        '', content, flags=re.MULTILINE | re.IGNORECASE,
    )

    # Step 8: Remove footer boilerplate
    content = re.sub(
        r'Your browser is not supported\..*?×\s*',
        '', content, flags=re.DOTALL,
    )
    content = re.sub(
        r'#{1,3}\s*Connect with us.*$',
        '', content, flags=re.DOTALL | re.MULTILINE,
    )
    content = re.sub(
        r'#{1,3}\s*Products and services.*$',
        '', content, flags=re.DOTALL | re.MULTILINE,
    )
    content = re.sub(
        r'#{1,3}\s*About Royal London.*$',
        '', content, flags=re.DOTALL | re.MULTILINE,
    )
    content = re.sub(
        r'#{1,3}\s*Useful links.*$',
        '', content, flags=re.DOTALL | re.MULTILINE,
    )
    content = re.sub(
        r'\*\*The Royal London Mutual Insurance.*$',
        '', content, flags=re.DOTALL | re.MULTILINE,
    )
    content = re.sub(
        r'©\s*Royal London \d{4}.*$',
        '', content, flags=re.DOTALL | re.MULTILINE,
    )
    content = re.sub(r'\[Back to top\].*?\n', '', content)

    # Step 9: Normalize whitespace
    content = re.sub(r'\n{3,}', '\n\n', content)
    content = re.sub(r'[ \t]+\n', '\n', content)
    content = re.sub(r'\n[ \t]+\n', '\n\n', content)

    return content.strip()


# ── Excel category → content_type mapping (v4.3.0) ────
# Maps customer-supplied Category values to content_type.
# Excel wins over URL-pattern detection when present.
# URL-pattern detection is the fallback when Excel category
# is blank or unrecognised.
_EXCEL_CATEGORY_MAP = {
    "brand":    "article",    # brand/marketing pages → article
    "guidance": "guide",      # guidance content → guide
    "other":    "article",    # catch-all → article
    "product":  "article",    # product pages → article
    "tool":     "tool",       # tools/calculators → tool
}

def map_excel_category_to_content_type(excel_category: str, url: str) -> str:
    """
    v4.3.0 — Map customer Excel Category to content_type.

    Priority: Excel category (primary) → URL-pattern (fallback).

    Args:
        excel_category: lowercased value from Excel Category column
                        (e.g. "brand", "guidance", "product", "tool", "other")
        url:            page URL — used as fallback via derive_content_type()

    Returns:
        content_type string: guide / tool / webinar / video /
                             faq / news / corporate / article
    """
    if excel_category:
        mapped = _EXCEL_CATEGORY_MAP.get(excel_category)
        if mapped:
            # URL-pattern can still upgrade: a "product" page
            # on /webinars/ should still be "webinar".
            url_type = derive_content_type(url)
            if url_type in ("webinar", "video", "tool", "faq", "news"):
                return url_type
            return mapped
    # Fallback: URL-pattern only
    return derive_content_type(url)


# ── Step 2b: Multi-state dropdown helpers (v4.5.0) ─────────
#
# Replaced crawl4ai JS-injection approach (v4.4.0) with Playwright.
# Root cause of v4.4.0 failure: a second crawler.arun() call with
# wait_until="networkidle" timed out on Royal London contact pages.
# Playwright loads the page ONCE, detects dropdowns from the live DOM,
# then iterates options via JS event injection — matching the proven
# standalone crawler.py script (v0.1.0) exactly.

# Dropdown option text that indicates a placeholder (skip these)
_DROPDOWN_PLACEHOLDERS = {
    "select...", "select", "please select", "--", "choose...",
    "choose", "please choose",
}


def _has_routing_dropdowns_in_html(html: str) -> bool:
    """
    Check rendered HTML for routing <select> elements.

    Uses BeautifulSoup on already-fetched crawl4ai HTML — zero
    extra network call. Returns True only if a <select> has more
    than one non-placeholder <option>.
    """
    if not html:
        return False
    try:
        soup = BeautifulSoup(html, "html.parser")
        for select in soup.find_all("select"):
            valid_opts = [
                o for o in select.find_all("option")
                if o.get_text(strip=True).lower() not in _DROPDOWN_PLACEHOLDERS
                and o.get_text(strip=True)
            ]
            if len(valid_opts) > 1:
                return True
        return False
    except Exception:
        return False


def _truncate_base_content_at_dropdown(content: str) -> str:
    """
    For dropdown pages, truncate base page content at the first
    dropdown-related marker.

    Royal London renders all dropdown option content in the DOM
    simultaneously (hidden panels shown via JS). crawl4ai captures
    everything, causing the contact details to appear 3x in the
    markdown output (once per visible render pass).

    We keep only the intro paragraph — the part before the dropdown
    widget appears. This matches exactly what the Playwright script
    captured as the default state (clean intro, no repeated options).

    Markers that indicate the dropdown/contact panel has started:
    These appear in the markdown just before the repeated content.
    """
    DROPDOWN_MARKERS = [
        "please select an option",
        "please select",
        "select an option",
        "getting in touch",
        "get in touch",
        "to show you how best",
        "we need to know what kind of policy",
        "don't worry if you don't know",
        "choose from the list below",
        "select from the list",
    ]

    content_lower = content.lower()
    earliest_cut  = len(content)

    for marker in DROPDOWN_MARKERS:
        idx = content_lower.find(marker)
        if idx != -1 and idx < earliest_cut:
            earliest_cut = idx

    if earliest_cut < len(content):
        truncated = content[:earliest_cut].strip()
        # Only truncate if we still have meaningful content
        if len(truncated) >= 100:
            log.info(
                "base_content_truncated_at_dropdown",
                original_chars=len(content),
                truncated_chars=len(truncated),
            )
            return truncated

    return content


def _scrape_dropdown_states_playwright(
    url:            str,
    base_title:     str,
    base_page_data: dict,
) -> list[dict]:
    """
    Scrape per-option content from a routing dropdown page using Playwright.

    Runs synchronously — called via asyncio thread pool executor from
    scrape_page() to avoid blocking the crawl4ai event loop.

    Mirrors _scrape_multi_state_page() from crawler.py v0.1.0 exactly:
    1. Single page load (networkidle, 45s timeout)
    2. Detect <select> elements from live DOM
    3. For each option: JS event injection → 1.5s wait → body text
    4. Line-by-line diff against default state → only changed lines
    5. Build page_data dict with synthetic #policy= URL

    Returns list of page_data dicts (one per option with changed content).
    Empty list on any failure — base page is still used.
    """
    try:
        from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
    except ImportError:
        log.error(
            "playwright_not_installed",
            note="pip install playwright && playwright install chromium",
        )
        return []

    results: list[dict] = []

    try:
        with sync_playwright() as pw:
            browser = pw.chromium.launch(headless=True)
            try:
                page = browser.new_page()

                # Block images/fonts — speed up load, no impact on text content
                page.route(
                    "**/*.{png,jpg,jpeg,gif,svg,woff,woff2,ttf,eot}",
                    lambda route: route.abort(),
                )

                log.info("playwright_navigating", url=url)
                page.goto(url, wait_until="networkidle", timeout=45000)

                # Try to wait for main content area
                try:
                    page.wait_for_selector(
                        "main, article, [role='main']",
                        timeout=10000,
                    )
                except PWTimeout:
                    pass  # Continue anyway — main content may still be present

                # Detect routing dropdowns from live DOM
                selects = page.query_selector_all("select")
                routing_dropdowns = []

                for select in selects:
                    options     = select.query_selector_all("option")
                    valid_opts  = []
                    for opt in options:
                        text  = opt.inner_text().strip()
                        value = opt.get_attribute("value") or ""
                        if text and text.lower() not in _DROPDOWN_PLACEHOLDERS:
                            valid_opts.append({"value": value, "text": text})

                    if len(valid_opts) > 1:
                        routing_dropdowns.append({
                            "select_element": select,
                            "options":        valid_opts,
                        })

                if not routing_dropdowns:
                    log.info("playwright_no_dropdowns_found", url=url)
                    return []

                log.info(
                    "playwright_dropdowns_detected",
                    url=url,
                    dropdown_count=len(routing_dropdowns),
                )

                # Capture default body text for diffing
                default_raw_text = page.inner_text("body")
                default_lines    = {
                    line.strip()
                    for line in default_raw_text.split("\n")
                    if line.strip()
                }

                # Iterate each dropdown × each option
                for dropdown in routing_dropdowns:
                    select_el = dropdown["select_element"]
                    options   = dropdown["options"]

                    for option in options:
                        opt_value = option["value"]
                        opt_text  = option["text"]

                        try:
                            # Fire JS input + change events on the select element
                            # Mirrors crawler.py evaluate() call exactly
                            select_el.evaluate(
                                """(el, optText) => {
                                    const options = Array.from(el.options);
                                    const target = options.find(
                                        o => o.text.trim() === optText
                                    );
                                    if (target) {
                                        el.value = target.value;
                                        el.dispatchEvent(
                                            new Event('input', { bubbles: true })
                                        );
                                        el.dispatchEvent(
                                            new Event('change', { bubbles: true })
                                        );
                                    }
                                }""",
                                opt_text,
                            )

                            # Wait 1.5s for DOM to update — same as crawler.py
                            time.sleep(1.5)

                            new_raw_text = page.inner_text("body")

                            # Line-by-line diff — only lines NOT in default state
                            new_lines     = [
                                line.strip()
                                for line in new_raw_text.split("\n")
                                if line.strip()
                            ]
                            changed_lines = [
                                line for line in new_lines
                                if line not in default_lines
                            ]
                            dynamic_content = "\n".join(changed_lines)

                            if not dynamic_content or len(dynamic_content.strip()) < 20:
                                log.warning(
                                    "playwright_option_no_change",
                                    url=url,
                                    option=opt_text,
                                )
                                continue

                            # Synthetic URL — mirrors crawler.py state_url pattern
                            safe_value = opt_value if opt_value else opt_text
                            state_url  = f"{url}#policy={safe_value}"
                            content    = dynamic_content.strip()

                            results.append({
                                "url":              state_url,
                                "title":            f"{base_title} — {opt_text}",
                                "section":          base_page_data["section"],
                                "content":          content,
                                "scraped_at":       datetime.now(timezone.utc).isoformat(),
                                "content_length":   len(content),
                                "content_hash":     hashlib.sha256(
                                    content.encode("utf-8")
                                ).hexdigest(),
                                "audience":         base_page_data["audience"],
                                "has_video":        base_page_data["has_video"],
                                "content_type":     base_page_data["content_type"],
                                "product_category": base_page_data["product_category"],
                                "description":      base_page_data["description"],
                                "thumbnail_url":    base_page_data["thumbnail_url"],
                                "publish_date":     base_page_data["publish_date"],
                                "collection_name":  base_page_data["collection_name"],
                                "read_time_mins":   max(1, len(content.split()) // 200),
                                "dropdown_state":   opt_text,
                                "dropdown_value":   opt_value or "",
                            })

                            log.info(
                                "playwright_option_scraped",
                                url=state_url,
                                option=opt_text,
                                chars=len(content),
                            )

                        except Exception as e:
                            log.warning(
                                "playwright_option_error",
                                url=url,
                                option=opt_text,
                                error=str(e),
                            )
                            continue

            finally:
                browser.close()

    except Exception as e:
        log.error("playwright_dropdown_scrape_error", url=url, error=str(e))

    return results


# ── Step 3: Scrape a single page ───────────────────────
async def scrape_page(
    crawler: AsyncWebCrawler,
    page_info: dict,
    index: int,
    total: int,
) -> "list[dict] | dict | None":
    """
    Scrape a single page and return cleaned page data.

    Return type (v4.4.0):
        dict        — standard page (unchanged behaviour)
        list[dict]  — multi-state dropdown page: first entry is the
                      base page, subsequent entries are per-option states
        None        — scrape failed (unchanged behaviour)
    """
    url            = page_info["url"]
    title          = page_info["title"]
    excel_category = page_info.get("excel_category", "")

    log.info("scraping_page", url=url, index=index, total=total)

    try:
        run_config = CrawlerRunConfig(
            css_selector=(
                "main, article, .content, #content, "
                ".page-content, .main-content, [role='main']"
            ),
            markdown_generator=DefaultMarkdownGenerator(
                content_filter=PruningContentFilter(
                    threshold=0.45,
                    threshold_type="fixed",
                ),
                options={
                    "ignore_links": False,
                    "ignore_images": True,
                    "skip_internal_links": True,
                }
            ),
            wait_until="domcontentloaded",
            page_timeout=30000,
            verbose=False,
            excluded_tags=[
                "nav", "header", "footer", "aside",
                "script", "style", "noscript",
            ],
        )

        result = await crawler.arun(url=url, config=run_config)

        if not result.success:
            log.error(
                "scrape_failed",
                url=url,
                error=result.error_message,
            )
            return None

        page_content = result.markdown.raw_markdown

        if not page_content or len(page_content.strip()) < 100:
            log.warning(
                "content_too_short",
                url=url,
                length=len(page_content or ""),
            )
            return None

        page_content = clean_content(page_content)

        if len(page_content.strip()) < 50:
            log.warning("content_too_short_after_cleaning", url=url)
            return None

        # v3.0.0 — extract rich metadata from raw HTML.
        # result.html is the full rendered HTML already fetched
        # by crawl4ai — no extra HTTP call needed.
        # Falls back to safe defaults if HTML unavailable.
        raw_html = getattr(result, "html", "") or ""
        metadata = extract_page_metadata(raw_html, url)

        # v4.0.0: re-normalise at output — defence for redirects
        url = normalize_url(url)

        page_data = {
            # ── Core fields ────────────────────────────────────
            "url":            url,
            "title":          title,
            "section":        derive_section(url),
            "content":        page_content.strip(),
            "scraped_at":     datetime.now(timezone.utc).isoformat(),
            "content_length": len(page_content.strip()),
            # v4.0.0: SHA-256 for freshness detection (content_freshness.py)
            "content_hash":   hashlib.sha256(
                page_content.strip().encode("utf-8")
            ).hexdigest(),

            # ── Enrichment fields (v3.0.0) ────────────────────
            # All extracted from HTML already fetched by crawl4ai.
            # Zero extra HTTP calls. Safe defaults if extraction fails.
            "audience":         metadata["audience"],
            "has_video":        metadata["has_video"],
            # v4.3.0: Excel Category takes priority; URL-pattern is fallback.
            "content_type":     map_excel_category_to_content_type(excel_category, url),
            "product_category": metadata["product_category"],
            "description":      metadata["description"],
            "thumbnail_url":    metadata["thumbnail_url"],
            "publish_date":     metadata["publish_date"],
            "collection_name":  metadata["collection_name"],
            "read_time_mins":   metadata["read_time_mins"],
        }

        log.info(
            "scrape_success",
            url=url,
            index=index,
            total=total,
            content_length=page_data["content_length"],
            has_video=metadata["has_video"],
            content_type=page_data["content_type"],
            product_category=metadata["product_category"],
            excel_category=excel_category or "none",
        )

        # v4.5.0 — detect routing dropdowns from already-fetched HTML.
        # BeautifulSoup check on result.html — zero extra network call.
        # If routing <select> elements found, use Playwright (in thread
        # pool executor) to scrape each option state — mirrors proven
        # crawler.py v0.1.0 approach that captured all 13 bereavement
        # policy options correctly.
        raw_html = getattr(result, "html", "") or ""

        if _has_routing_dropdowns_in_html(raw_html):
            # Truncate base page content at dropdown marker —
            # removes repeated option content Royal London renders
            # inline (all hidden panels in DOM simultaneously).
            truncated_content = _truncate_base_content_at_dropdown(
                page_data["content"]
            )
            page_data["content"]        = truncated_content
            page_data["content_length"] = len(truncated_content)
            page_data["content_hash"]   = hashlib.sha256(
                truncated_content.encode("utf-8")
            ).hexdigest()

            log.info("dropdown_page_detected_via_html", url=url)

            # Run Playwright in thread pool — non-blocking for asyncio
            try:
                loop     = asyncio.get_event_loop()
                executor = concurrent.futures.ThreadPoolExecutor(max_workers=1)
                dropdown_states = await loop.run_in_executor(
                    executor,
                    _scrape_dropdown_states_playwright,
                    url,
                    page_data["title"],
                    page_data,
                )
            except Exception as _de:
                log.warning(
                    "playwright_dropdown_skipped",
                    url=url,
                    error=str(_de),
                )
                dropdown_states = []

            if dropdown_states:
                log.info(
                    "multi_state_page_scraped",
                    url=url,
                    option_count=len(dropdown_states),
                )
                return [page_data] + dropdown_states

        return page_data

    except Exception as e:
        log.error("scrape_error", url=url, error=str(e))
        return None



def load_url_source(excel_path: str) -> list[dict]:
    """
    Load approved URLs from local Excel or Blob Storage.

    Local development (AZURE_STORAGE_CONNECTION not set):
        Reads from Excel file at excel_path — same as v1.0.0.
        No behaviour change for local development workflow.

    Production (AZURE_STORAGE_CONNECTION set):
        Reads from Azure Blob Storage JSON file instead.
        JSON format: [{"url": "...", "title": "..."}, ...]
        DevOps uploads this JSON to Blob Storage; brand/marketing
        team maintains the URL list in agreed format.

    TODO: Production URL source format to be agreed with client.
    Until decided, this falls back to Excel for all environments.
    """
    # ── Production: read from Blob Storage ───────────────────
    # TODO (DevOps): uncomment when AZURE_STORAGE_CONNECTION
    # is configured in Key Vault and URL source format agreed.
    #
    # if BLOB_STORAGE_CONNECTION:
    #     from azure.storage.blob import BlobServiceClient
    #     blob_service = BlobServiceClient.from_connection_string(
    #         BLOB_STORAGE_CONNECTION
    #     )
    #     blob_client = blob_service.get_blob_client(
    #         container=BLOB_CONTAINER_NAME,
    #         blob="approved_urls.json",   # DevOps agrees this name
    #     )
    #     data  = blob_client.download_blob().readall()
    #     pages = json.loads(data)
    #     log.info(
    #         "url_source_loaded_from_blob",
    #         container=BLOB_CONTAINER_NAME,
    #         count=len(pages),
    #     )
    #     return pages   # [{"url": ..., "title": ...}]

    # ── Local development: read from Excel ───────────────────
    # Current behaviour — unchanged from v1.0.0.
    return load_approved_pages(excel_path)


def save_scraped_pages(
    results: list[dict],
    output_file: "Path",
) -> str:
    """
    Save scraped pages to local file or Azure Blob Storage.

    Local development (AZURE_STORAGE_CONNECTION not set):
        Saves JSON to scraper/data/ as before.
        Returns local file path string.

    Production (AZURE_STORAGE_CONNECTION set):
        Uploads JSON to Azure Blob Storage.
        chunk_and_index.py reads from the same Blob container.
        Returns blob filename string.

    Args:
        results:     List of scraped page dicts
        output_file: Local Path object (used in local mode,
                     filename used as blob name in production)

    Returns:
        str — local file path or blob filename
    """
    # ── Production: upload to Blob Storage ───────────────────
    # TODO (DevOps): uncomment when AZURE_STORAGE_CONNECTION
    # is configured in Key Vault.
    #
    # if BLOB_STORAGE_CONNECTION:
    #     from azure.storage.blob import BlobServiceClient
    #     blob_service = BlobServiceClient.from_connection_string(
    #         BLOB_STORAGE_CONNECTION
    #     )
    #     blob_name   = BLOB_SCRAPED_FILENAME
    #     blob_client = blob_service.get_blob_client(
    #         container=BLOB_CONTAINER_NAME,
    #         blob=blob_name,
    #     )
    #     data = json.dumps(results, ensure_ascii=False, indent=2)
    #     blob_client.upload_blob(data, overwrite=True)
    #     log.info(
    #         "scraped_pages_saved_to_blob",
    #         container=BLOB_CONTAINER_NAME,
    #         blob=blob_name,
    #         pages=len(results),
    #     )
    #     print(f"   ✅ Uploaded to Blob Storage: {blob_name}")
    #     return blob_name

    # v4.4.0 BUG FIX: guard against empty results — writing [] silently
    # would cause the next indexer run to wipe the index with no warning.
    if not results:
        log.error(
            "save_scraped_pages_empty",
            note=(
                "No pages to save — all URLs failed to scrape. "
                "Output file NOT written. Check scrape errors above."
            ),
        )
        return ""

    # ── Local development: save to local file ────────────────
    output_file.parent.mkdir(parents=True, exist_ok=True)
    with open(output_file, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    log.info(
        "scraped_pages_saved_locally",
        file=str(output_file),
        pages=len(results),
    )
    return str(output_file)


def run_scraper(
    excel_path: str | None = None,
    dry_run: bool = False,
) -> dict:
    """
    Programmatic entry point for the scraping pipeline.
    Called by DevOps / Azure Container Apps Job.

    Args:
        excel_path: Path to approved URLs Excel file.
                    If None, uses APPROVED_EXCEL constant.
                    Production: override via APPROVED_EXCEL_PATH env var.
        dry_run:    If True, detect columns + count URLs but do NOT
                    scrape or save anything. Use to validate the Excel
                    file is correctly parsed before a production run.

    Returns:
        dict with keys:
            success        (bool) — True if completed without error
            pages_scraped  (int)  — pages scraped (0 if dry_run)
            pages_failed   (int)  — pages that failed to scrape
            output_path    (str)  — local file path or blob filename
            dry_run        (bool) — whether this was a dry run
            error          (str)  — error message if success=False

    TODO (DevOps — Sprint 2):
    Wrap in Azure Container Apps Job entrypoint script.
    Container Apps Job: aria-scraper-job

        # In job entrypoint (e.g. entrypoint.py):
        import os
        from scraper.scrape_approved_urls_updatedV2 import run_scraper

        dry_run    = os.getenv("DRY_RUN", "false").lower() == "true"
        excel_path = os.getenv("APPROVED_EXCEL_PATH", None)
        result     = run_scraper(excel_path=excel_path, dry_run=dry_run)

        if not result["success"]:
            raise RuntimeError(f"Scrape failed: {result['error']}")

        print(f"Scraped {result['pages_scraped']} pages → {result['output_path']}")

    ADO pipeline trigger (runs aria-scraper-job, then aria-indexer-job):

        az containerapp job start \
            --name aria-scraper-job \
            --resource-group <rg> \
            --env-vars DRY_RUN=false

    Required env vars in Key Vault (see production section in module docstring).
    """
    result = {
        "success":       False,
        "pages_scraped": 0,
        "pages_failed":  0,
        "output_path":   "",
        "dry_run":       dry_run,
        "error":         "",
    }

    try:
        excel = excel_path or os.getenv("APPROVED_EXCEL_PATH") or APPROVED_EXCEL

        if dry_run:
            # Dry run: detect columns + count URLs, no scraping
            pages = load_url_source(excel)
            print(f"\n✅ DRY RUN COMPLETE — no scraping performed.")
            print(f"   Excel file:       {excel}")
            print(f"   URLs detected:    {len(pages)}")
            print(f"\n   Remove --dry-run or set DRY_RUN=false to scrape for real.")
            result["success"]       = True
            result["pages_scraped"] = 0
            result["output_path"]   = ""
            return result

        async def _run():
            pages_to_scrape = load_url_source(excel)
            browser_config  = BrowserConfig(
                browser_type="chromium",
                headless=True,
                verbose=False,
            )
            scraped      = []
            failed_urls  = []
            total        = len(pages_to_scrape)

            log.info(
                "scraper_pipeline_started",
                total_urls=total,
                excel=excel,
                blob_storage=bool(BLOB_STORAGE_CONNECTION),
            )

            async with AsyncWebCrawler(config=browser_config) as crawler:
                for batch_start in range(0, total, BATCH_SIZE):
                    batch  = pages_to_scrape[
                        batch_start:batch_start + BATCH_SIZE
                    ]
                    tasks  = [
                        scrape_page(
                            crawler, page_info,
                            batch_start + i + 1, total,
                        )
                        for i, page_info in enumerate(batch)
                    ]
                    results = await asyncio.gather(*tasks)
                    for page_info, r in zip(batch, results):
                        if r is None:
                            failed_urls.append(page_info["url"])
                        elif isinstance(r, list):
                            # v4.4.0: multi-state dropdown page —
                            # flatten all option entries into scraped.
                            # Minimum content_length 20 chars (a phone
                            # number + label is valid at short length).
                            scraped.extend(
                                entry for entry in r
                                if entry.get("content_length", 0) >= 20
                            )
                        else:
                            scraped.append(r)
                    if batch_start + BATCH_SIZE < total:
                        await asyncio.sleep(BATCH_DELAY_SECONDS)

            return scraped, failed_urls

        # v4.4.0 BUG FIX: asyncio.run() raises RuntimeError when called
        # from an already-running event loop (Azure Functions, FastAPI,
        # Jupyter). Use nest_asyncio when available; fall back to plain
        # asyncio.run() for normal script execution.
        if _NEST_ASYNCIO_AVAILABLE:
            _nest_asyncio.apply()
            loop = asyncio.get_event_loop()
            scraped, failed = loop.run_until_complete(_run())
        else:
            # nest_asyncio not installed — plain script context, safe to use.
            # Install nest_asyncio if calling from async host environments.
            scraped, failed = asyncio.run(_run())

        timestamp   = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        output_file = (
            Path("scraper/data") /
            f"royal_london_faq_approved_{timestamp}.json"
        )
        output_path = save_scraped_pages(scraped, output_file)

        # v4.4.0: save_scraped_pages returns "" when results are empty
        if not output_path:
            result["error"] = (
                "All URLs failed to scrape — output file not written. "
                "Check scrape errors in the log."
            )
            result["pages_failed"] = len(failed)
            return result

        log.info(
            "scraper_pipeline_complete",
            pages_scraped=len(scraped),
            pages_failed=len(failed),
            output_path=output_path,
        )

        result["success"]       = True
        result["pages_scraped"] = len(scraped)
        result["pages_failed"]  = len(failed)
        result["output_path"]   = output_path
        return result

    except Exception as e:
        result["error"] = str(e)
        log.error(
            "scraper_pipeline_error",
            error=str(e),
            traceback=traceback.format_exc(),
        )
        return result

# ── Main ────────────────────────────────────────────────
async def main():
    parser = argparse.ArgumentParser(
        description="RLG FAQ Scraper — scrapes customer-approved URLs"
    )
    parser.add_argument(
        "--file", default=None,
        help="Path to approved URLs Excel file. "
             "Default: APPROVED_EXCEL constant.",
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Detect Excel columns + count URLs, do NOT scrape. "
             "Validates the Excel file is parsed correctly.",
    )
    args = parser.parse_args()

    print("\n" + "=" * 60)
    print("RLG FAQ SCRAPER — Customer Approved URLs Only")
    print("=" * 60)

    excel_path = Path(
        args.file
        or os.getenv("APPROVED_EXCEL_PATH")
        or APPROVED_EXCEL
    )
    if not excel_path.exists():
        print(f"\nERROR: Excel file not found: {excel_path}")
        sys.exit(1)

    pages_to_scrape = load_url_source(str(excel_path))
    print(f"\nExcel file:       {excel_path}")
    print(f"Approved URLs:    {len(pages_to_scrape)}")

    if args.dry_run:
        print("\n✅ DRY RUN COMPLETE — no scraping performed.")
        print("   Run without --dry-run to scrape for real.")
        return

    # ── Scrape in batches ───────────────────────────────
    browser_config = BrowserConfig(
        browser_type="chromium",
        headless=True,
        verbose=False,
    )

    results = []
    failed_urls = []
    total = len(pages_to_scrape)

    async with AsyncWebCrawler(config=browser_config) as crawler:
        for batch_start in range(0, total, BATCH_SIZE):
            batch = pages_to_scrape[batch_start:batch_start + BATCH_SIZE]
            batch_num = batch_start // BATCH_SIZE + 1
            total_batches = (total + BATCH_SIZE - 1) // BATCH_SIZE

            print(f"\nBatch {batch_num}/{total_batches}")

            tasks = [
                scrape_page(crawler, page_info, batch_start + i + 1, total)
                for i, page_info in enumerate(batch)
            ]
            batch_results = await asyncio.gather(*tasks)

            for page_info, result in zip(batch, batch_results):
                if result is None:
                    failed_urls.append(page_info["url"])
                elif isinstance(result, list):
                    # v4.4.0: multi-state dropdown page — flatten entries.
                    results.extend(
                        entry for entry in result
                        if entry.get("content_length", 0) >= 20
                    )
                else:
                    results.append(result)

            if batch_start + BATCH_SIZE < total:
                await asyncio.sleep(BATCH_DELAY_SECONDS)

    # ── Save output ─────────────────────────────────────
    # v2.0.0: save_scraped_pages() abstracts local file vs
    # Blob Storage. Local: saves to scraper/data/ as before.
    # Production: uploads to Azure Blob Storage when
    # AZURE_STORAGE_CONNECTION env var is set.
    timestamp   = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    output_file = (
        Path("scraper/data") /
        f"royal_london_faq_approved_{timestamp}.json"
    )
    output_path = save_scraped_pages(results, output_file)

    # ── Summary ───────────────────────────────────────────
    print("\n" + "=" * 60)
    print("SCRAPE SUMMARY")
    print("=" * 60)
    print(f"Approved URLs:    {total}")
    print(f"Scraped success:  {len(results)}")
    print(f"Failed:           {len(failed_urls)}")

    if failed_urls:
        print("\nFailed URLs:")
        for url in failed_urls:
            print(f"  - {url}")

    if results:
        total_chars = sum(r["content_length"] for r in results)
        lengths     = [r["content_length"] for r in results]
        print(f"\nTotal content:    {total_chars:,} chars")
        print(f"Shortest page:    {min(lengths):,} chars")
        print(f"Longest page:     {max(lengths):,} chars")
        print(f"Average page:     {total_chars // len(results):,} chars")

        # v3.0.0 — enrichment summary
        video_pages = sum(1 for r in results if r.get("has_video"))
        type_counts = {}
        cat_counts  = {}
        for r in results:
            ct = r.get("content_type", "article")
            pc = r.get("product_category", "general")
            type_counts[ct] = type_counts.get(ct, 0) + 1
            cat_counts[pc]  = cat_counts.get(pc, 0) + 1

        print(f"\nPages with video: {video_pages}")
        print("\nContent types:")
        for ct, count in sorted(type_counts.items(), key=lambda x: -x[1]):
            print(f"  {ct:<20} {count}")
        print("\nProduct categories:")
        for pc, count in sorted(cat_counts.items(), key=lambda x: -x[1]):
            print(f"  {pc:<25} {count}")

    print(f"\nSaved to: {output_path}")
    print("=" * 60)
    print(f"\n👉 Next step: index the scraped content:")
    print(f"   python scraper/chunk_and_index_hqaV3.py --full --file {output_path}")
    print(f"   python scraper/chunk_and_index_hqaV3.py --full --no-hqa --file {output_path}  # baseline")


if __name__ == "__main__":
    asyncio.run(main())