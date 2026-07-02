"""
Royal London — Pre-Scrape URL Validation Script
================================================
Run this BEFORE scrape_approved_urls.py to validate all URLs
in the customer Excel and produce a report to share with them.

Usage:
    python scraper/validate_urls.py
    python scraper/validate_urls.py --file path/to/custom.xlsx
    python scraper/validate_urls.py --concurrency 20 --timeout 10

Output:
    scraper/data/url_validation_report_<timestamp>.xlsx
    Three sheets:
      1. Full Results   — every URL with status
      2. Action Required — dead + redirect + duplicate rows only
      3. Summary        — counts by issue type

═══════════════════════════════════════════════════════════════
CHANGE LOG
═══════════════════════════════════════════════════════════════

v1.0.0 — July 2026 | Mukesh Kund
         Initial version

         PURPOSE:
         Standalone pre-scrape health check. Reads customer
         Excel (any column layout — URL column auto-detected
         by content sniffing, not header name), checks every
         URL live via async HTTP HEAD requests, detects
         duplicates, and produces an actionable Excel report
         ready to share with the customer.

         COLUMN DETECTION STRATEGY:
         Content-based sniffing (primary) — finds the column
         where cells look like https://... regardless of header
         name. Header-name matching used only as a tiebreaker
         when multiple columns have URL-like content.
         This means the script works correctly even if the
         customer renames "URL" to "Source URL", "Link",
         "Web Address", or anything else in a future file.

         URL CHECKS:
         - Live (200)            — clean, no action needed
         - Redirected (301/302)  — follows to final URL,
           flags if final domain differs from royallondon.com
         - Dead (4xx/5xx)        — action required
         - Timeout               — action required
         - Duplicate             — flagged after normalisation
           (catches case differences like should-I vs should-i,
           trailing slash variants, and exact duplicates)

═══════════════════════════════════════════════════════════════
"""

import argparse
import asyncio
import sys
import os
import re
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import urlparse, urlunparse

import aiohttp
from openpyxl import load_workbook, Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, PatternFill
)
from openpyxl.utils import get_column_letter

# ── Config ────────────────────────────────────────────────────
DEFAULT_EXCEL       = "scraper/data/royal_london_verification_retried.xlsx"
DEFAULT_CONCURRENCY = 15
DEFAULT_TIMEOUT     = 12       # seconds per request
MAX_RETRIES         = 1        # retry once on timeout/connection error
EXPECTED_DOMAIN     = "royallondon.com"

# HTTP headers — mimic a real browser to avoid bot-detection blocks
REQUEST_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,*/*",
}

# ── Colour palette ────────────────────────────────────────────
GREEN  = "C6EFCE"   # live
YELLOW = "FFEB9C"   # redirect / warning
RED    = "FFC7CE"   # dead / error
BLUE   = "BDD7EE"   # duplicate
GREY   = "D9D9D9"   # header
WHITE  = "FFFFFF"

STATUS_LABELS = {
    "live":       "✅ Live",
    "redirect":   "⚠️  Redirected (external)",
    "dead":       "❌ Dead",
    "timeout":    "⏱️  Timeout",
    "duplicate":  "🔁 Duplicate",
    "error":      "❌ Error",
}

STATUS_COLOURS = {
    "live":       GREEN,
    "redirect":   YELLOW,
    "dead":       RED,
    "timeout":    RED,
    "duplicate":  BLUE,
    "error":      RED,
}


# ── Helpers ───────────────────────────────────────────────────
def normalize_url(url: str) -> str:
    """Canonical URL for duplicate detection — same logic as scraper."""
    if not url:
        return url
    try:
        parsed = urlparse(url.strip())
        return urlunparse((
            parsed.scheme.lower(),
            parsed.netloc.lower(),
            parsed.path.lower().rstrip("/"),
            "", "", "",
        ))
    except Exception:
        return url.strip().lower()


def detect_url_column(ws) -> tuple[int | None, int | None, int | None]:
    """
    Detect URL, Title, and Category column indices by content sniffing.

    Primary strategy (URL column):
    Sample first 10 non-empty data rows per column. The column
    with the highest proportion of cells starting with https?://
    is the URL column. No dependency on header naming — works
    regardless of what the column is called.

    Tiebreaker: if multiple columns score equally, prefer the
    one whose header also contains "url" or "link".

    Title and Category: header-name matching only (optional
    fields — graceful empty-string fallback if not found).

    Returns (url_col_idx, title_col_idx, category_col_idx)
    All zero-indexed. None if not found.
    """
    TITLE_HEADERS    = {"title", "page title", "name", "page name"}
    CATEGORY_HEADERS = {"category", "type", "content type", "page type"}

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return None, None, None

    header_row = [
        str(h).strip().lower() if h is not None else ""
        for h in rows[0]
    ]
    data_rows = rows[1:]

    num_cols = len(header_row)

    # ── Content-based URL sniffing ────────────────────────────
    url_scores = []
    for col_idx in range(num_cols):
        sample = [
            str(r[col_idx]).strip()
            for r in data_rows
            if col_idx < len(r) and r[col_idx] is not None
        ][:10]
        if not sample:
            url_scores.append(0.0)
            continue
        url_like = sum(
            1 for v in sample
            if v.startswith("http://") or v.startswith("https://")
        )
        url_scores.append(url_like / len(sample))

    max_score = max(url_scores) if url_scores else 0
    if max_score < 0.5:
        # No column is majority URL-like
        return None, None, None

    # Among columns tied at the max score, prefer one whose
    # header contains "url" or "link" (tiebreaker only)
    candidates = [
        i for i, s in enumerate(url_scores) if s == max_score
    ]
    url_idx = candidates[0]  # default: leftmost best candidate
    for i in candidates:
        if any(kw in header_row[i] for kw in ("url", "link")):
            url_idx = i
            break

    # ── Header-name matching for optional columns ─────────────
    title_idx    = None
    category_idx = None
    for i, h in enumerate(header_row):
        if i == url_idx:
            continue
        if h in TITLE_HEADERS and title_idx is None:
            title_idx = i
        if h in CATEGORY_HEADERS and category_idx is None:
            category_idx = i

    return url_idx, title_idx, category_idx


def load_urls_from_excel(excel_path: str) -> list[dict]:
    """
    Load URLs from Excel using content-based column detection.

    Returns list of dicts:
    [{"url": "...", "title": "...", "category": "...",
      "row_num": N}]

    Raises ValueError if no URL column can be detected.
    """
    wb = load_workbook(excel_path, read_only=True)
    ws = wb.active

    url_idx, title_idx, category_idx = detect_url_column(ws)

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if url_idx is None:
        header_row = rows[0] if rows else []
        raise ValueError(
            f"No URL column found in {excel_path!r}.\n"
            f"Expected a column where most cells start with "
            f"https:// or http://.\n"
            f"Header row found: {list(header_row)}"
        )

    print(f"   URL column:      column index {url_idx} "
          f"(header: {str(rows[0][url_idx])!r})")
    print(f"   Title column:    "
          f"{'column index ' + str(title_idx) + ' (header: ' + repr(str(rows[0][title_idx])) + ')' if title_idx is not None else 'not found — title will be empty'}")
    print(f"   Category column: "
          f"{'column index ' + str(category_idx) + ' (header: ' + repr(str(rows[0][category_idx])) + ')' if category_idx is not None else 'not found'}")

    entries = []
    for row_num, row in enumerate(rows[1:], start=2):
        if not row or len(row) <= url_idx:
            continue
        url = str(row[url_idx]).strip() if row[url_idx] else ""
        if not url.startswith("http"):
            continue

        title = (
            str(row[title_idx]).strip()
            if title_idx is not None and len(row) > title_idx and row[title_idx]
            else ""
        )
        category = (
            str(row[category_idx]).strip()
            if category_idx is not None and len(row) > category_idx and row[category_idx]
            else ""
        )

        entries.append({
            "url":      url,
            "title":    title,
            "category": category,
            "row_num":  row_num,
        })

    return entries


# ── Async URL checker ─────────────────────────────────────────
async def check_url(
    session: aiohttp.ClientSession,
    entry: dict,
    timeout: int,
    semaphore: asyncio.Semaphore,
) -> dict:
    """
    Check a single URL. Returns result dict with:
    url, title, category, row_num, status, status_code,
    final_url, issue, redirect_chain
    """
    url = entry["url"]
    result = {
        **entry,
        "status":         "error",
        "status_code":    None,
        "final_url":      url,
        "issue":          "",
        "redirect_chain": "",
    }

    async with semaphore:
        for attempt in range(MAX_RETRIES + 1):
            try:
                resp = await session.get(
                    url,
                    timeout=aiohttp.ClientTimeout(total=timeout),
                    ssl=False,           # don't fail on cert issues
                    allow_redirects=True,
                )
                result["status_code"] = resp.status
                result["final_url"]   = str(resp.url)

                # Build redirect chain if redirected
                if str(resp.url) != url:
                    result["redirect_chain"] = (
                        f"{url} → {str(resp.url)}"
                    )

                if resp.status == 200:
                    # Check if we ended up on the right domain
                    final_domain = urlparse(str(resp.url)).netloc.lower()
                    if EXPECTED_DOMAIN not in final_domain:
                        result["status"] = "redirect"
                        result["issue"]  = (
                            f"Redirected to external domain: "
                            f"{final_domain}"
                        )
                    else:
                        result["status"] = "live"
                        result["issue"]  = ""
                elif resp.status >= 400:
                    result["status"] = "dead"
                    result["issue"]  = f"HTTP {resp.status}"
                else:
                    # 3xx that wasn't followed (shouldn't happen
                    # with allow_redirects=True, but be safe)
                    result["status"] = "redirect"
                    result["issue"]  = f"HTTP {resp.status}"

                break  # success — don't retry

            except asyncio.TimeoutError:
                if attempt < MAX_RETRIES:
                    await asyncio.sleep(1)
                    continue
                result["status"] = "timeout"
                result["issue"]  = f"Timed out after {timeout}s"

            except aiohttp.ClientConnectorError as e:
                if attempt < MAX_RETRIES:
                    await asyncio.sleep(1)
                    continue
                result["status"] = "dead"
                result["issue"]  = f"Connection error: {str(e)[:80]}"

            except Exception as e:
                result["status"] = "error"
                result["issue"]  = f"Error: {str(e)[:80]}"
                break

    return result


async def check_all_urls(
    entries: list[dict],
    concurrency: int,
    timeout: int,
) -> list[dict]:
    """Check all URLs concurrently with a semaphore rate limit."""
    semaphore = asyncio.Semaphore(concurrency)
    connector = aiohttp.TCPConnector(
        limit=concurrency,
        limit_per_host=5,   # don't hammer royallondon.com too hard
    )

    async with aiohttp.ClientSession(
        headers=REQUEST_HEADERS,
        connector=connector,
    ) as session:
        tasks = [
            check_url(session, entry, timeout, semaphore)
            for entry in entries
        ]

        results = []
        done = 0
        total = len(tasks)

        for coro in asyncio.as_completed(tasks):
            result = await coro
            results.append(result)
            done += 1
            if done % 20 == 0 or done == total:
                pct = round(done / total * 100)
                status_icon = {
                    "live":      "✅",
                    "dead":      "❌",
                    "redirect":  "⚠️ ",
                    "timeout":   "⏱️ ",
                    "duplicate": "🔁",
                    "error":     "❌",
                }.get(result["status"], "  ")
                print(
                    f"   [{pct:3d}%] {done}/{total} checked  "
                    f"| Last: {status_icon} {result['url'][:60]}"
                )

        return results


def detect_duplicates(results: list[dict]) -> list[dict]:
    """
    Mark duplicate URLs in results list.
    Normalises before comparison — catches case differences,
    trailing slash variants, and exact duplicates.
    First occurrence is kept as primary; all subsequent
    occurrences are flagged as duplicates.
    """
    seen: dict[str, int] = {}   # normalised_url → index of first occurrence

    for i, r in enumerate(results):
        norm = normalize_url(r["url"])
        if norm in seen:
            # Mark this entry as a duplicate
            r["is_duplicate"] = True
            r["duplicate_of"] = results[seen[norm]]["url"]
            # Don't change the status if it's already dead/timeout
            # — being a duplicate AND dead is more actionable
            if r["status"] == "live":
                r["status"] = "duplicate"
                r["issue"]  = (
                    f"Duplicate of: {results[seen[norm]]['url']}"
                )
            else:
                r["issue"] = (
                    r["issue"] +
                    f" | Also duplicate of: {results[seen[norm]]['url']}"
                )
        else:
            seen[norm] = i
            r["is_duplicate"]  = False
            r["duplicate_of"]  = ""

    return results


# ── Report generation ─────────────────────────────────────────
def _cell_fill(colour: str) -> PatternFill:
    return PatternFill("solid", start_color=colour, fgColor=colour)


def _header_style(ws, row: int, cols: list[str]):
    for col in cols:
        cell = ws[f"{col}{row}"]
        cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill      = _cell_fill("1F4E79")
        cell.alignment = Alignment(horizontal="center", vertical="center")


def write_report(
    results: list[dict],
    output_path: str,
    input_file: str,
) -> str:
    """Write the three-sheet Excel report."""
    wb = Workbook()

    # ── Counts ────────────────────────────────────────────────
    counts = {
        "live":      sum(1 for r in results if r["status"] == "live"),
        "redirect":  sum(1 for r in results if r["status"] == "redirect"),
        "dead":      sum(1 for r in results if r["status"] == "dead"),
        "timeout":   sum(1 for r in results if r["status"] == "timeout"),
        "duplicate": sum(1 for r in results if r["status"] == "duplicate"),
        "error":     sum(1 for r in results if r["status"] == "error"),
    }
    action_needed = (
        counts["dead"] + counts["redirect"] +
        counts["timeout"] + counts["duplicate"] + counts["error"]
    )

    # ══════════════════════════════════════════════════════════
    # SHEET 1 — Full Results
    # ══════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Full Results"

    headers = [
        "Row (Excel)", "Title", "Category", "URL",
        "Status", "HTTP Code", "Final URL",
        "Duplicate", "Issue / Notes",
    ]
    ws1.append(headers)
    _header_style(ws1, 1, [get_column_letter(i+1) for i in range(len(headers))])
    ws1.row_dimensions[1].height = 20

    for r in sorted(results, key=lambda x: x["row_num"]):
        status    = r["status"]
        colour    = STATUS_COLOURS.get(status, WHITE)
        is_dup    = "Yes" if r.get("is_duplicate") else ""
        final_url = r["final_url"] if r["final_url"] != r["url"] else ""

        row_data = [
            r["row_num"],
            r["title"],
            r["category"],
            r["url"],
            STATUS_LABELS.get(status, status),
            r["status_code"] or "",
            final_url,
            is_dup,
            r["issue"],
        ]
        ws1.append(row_data)
        row_idx = ws1.max_row
        for col_idx in range(1, len(headers) + 1):
            cell      = ws1.cell(row=row_idx, column=col_idx)
            cell.fill = _cell_fill(colour)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(vertical="center")

        # URL column — hyperlink
        url_cell = ws1.cell(row=row_idx, column=4)
        url_cell.hyperlink = r["url"]
        url_cell.font = Font(
            name="Arial", size=10,
            color="0563C1", underline="single"
        )

    # Column widths
    col_widths = [10, 45, 12, 60, 22, 10, 60, 10, 50]
    for i, width in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = width

    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    # ══════════════════════════════════════════════════════════
    # SHEET 2 — Action Required
    # ══════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Action Required")

    ws2.append([""])
    ws2["A1"] = "⚠️  URLs requiring action — share this sheet with the customer"
    ws2["A1"].font = Font(bold=True, name="Arial", size=12, color="C00000")
    ws2.row_dimensions[1].height = 22

    ws2.append([""])   # blank row

    action_headers = [
        "Row (Excel)", "Title", "Category", "URL",
        "Issue Type", "HTTP Code", "Final URL", "Notes",
    ]
    ws2.append(action_headers)
    _header_style(ws2, 3, [get_column_letter(i+1) for i in range(len(action_headers))])

    action_rows = [
        r for r in results
        if r["status"] in ("dead", "redirect", "timeout", "error", "duplicate")
    ]
    action_rows.sort(key=lambda x: (
        # Sort order: dead first, then timeout, redirect, duplicate, error
        ["dead", "timeout", "redirect", "duplicate", "error"].index(
            r["status"] if r["status"] in
            ["dead", "timeout", "redirect", "duplicate", "error"]
            else "error"
        ),
        x["row_num"],
    ))

    for r in action_rows:
        status    = r["status"]
        colour    = STATUS_COLOURS.get(status, WHITE)
        final_url = r["final_url"] if r["final_url"] != r["url"] else ""

        row_data = [
            r["row_num"],
            r["title"],
            r["category"],
            r["url"],
            STATUS_LABELS.get(status, status),
            r["status_code"] or "",
            final_url,
            r["issue"],
        ]
        ws2.append(row_data)
        row_idx = ws2.max_row
        for col_idx in range(1, len(action_headers) + 1):
            cell      = ws2.cell(row=row_idx, column=col_idx)
            cell.fill = _cell_fill(colour)
            cell.font = Font(name="Arial", size=10)

        url_cell = ws2.cell(row=row_idx, column=4)
        url_cell.hyperlink = r["url"]
        url_cell.font = Font(
            name="Arial", size=10,
            color="0563C1", underline="single"
        )

    if not action_rows:
        ws2.append(["", "✅ No action required — all URLs are live and unique"])
        ws2.cell(ws2.max_row, 2).font = Font(
            bold=True, name="Arial", size=11, color="375623"
        )

    col_widths2 = [10, 45, 12, 60, 22, 10, 60, 60]
    for i, width in enumerate(col_widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = width

    ws2.freeze_panes = "A4"

    # ══════════════════════════════════════════════════════════
    # SHEET 3 — Summary
    # ══════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Summary")

    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 15
    ws3.column_dimensions["C"].width = 40

    # Title block
    ws3["A1"] = "Royal London URL Validation Report"
    ws3["A1"].font = Font(bold=True, name="Arial", size=14, color="1F4E79")
    ws3.merge_cells("A1:C1")

    ws3["A2"] = "Generated:"
    ws3["B2"] = datetime.now(timezone.utc).strftime("%d %b %Y %H:%M UTC")
    ws3["A3"] = "Input file:"
    ws3["B3"] = Path(input_file).name
    ws3["A4"] = "Total URLs checked:"
    ws3["B4"] = len(results)

    for row in [2, 3, 4]:
        ws3[f"A{row}"].font = Font(bold=True, name="Arial", size=10)
        ws3[f"B{row}"].font = Font(name="Arial", size=10)

    # Results table
    ws3["A6"] = "Result"
    ws3["B6"] = "Count"
    ws3["C6"] = "Notes"
    _header_style(ws3, 6, ["A", "B", "C"])

    summary_rows = [
        ("✅ Live",                      counts["live"],
         "Working correctly — no action needed"),
        ("❌ Dead (4xx/5xx)",            counts["dead"],
         "Remove from approved list or fix URL"),
        ("⚠️  Redirected (external)",    counts["redirect"],
         "Review — final URL is outside royallondon.com"),
        ("⏱️  Timeout",                  counts["timeout"],
         "Could not connect — check manually"),
        ("🔁 Duplicate",                 counts["duplicate"],
         "Same URL appears more than once — remove extra rows"),
        ("❌ Other error",               counts["error"],
         "Unexpected error — check manually"),
        ("", "", ""),
        ("⚠️  TOTAL ACTION REQUIRED",    action_needed,
         "See 'Action Required' sheet"),
    ]

    colour_map = {
        "✅ Live":                    GREEN,
        "❌ Dead (4xx/5xx)":         RED,
        "⚠️  Redirected (external)": YELLOW,
        "⏱️  Timeout":               RED,
        "🔁 Duplicate":              BLUE,
        "❌ Other error":             RED,
        "⚠️  TOTAL ACTION REQUIRED": YELLOW,
    }

    for label, count, note in summary_rows:
        ws3.append([label, count if count != "" else "", note])
        row_idx = ws3.max_row
        colour  = colour_map.get(label, WHITE)
        for col in ["A", "B", "C"]:
            cell = ws3[f"{col}{row_idx}"]
            if label:
                cell.fill = _cell_fill(colour)
            cell.font = Font(
                name="Arial", size=10,
                bold=(label == "⚠️  TOTAL ACTION REQUIRED"),
            )

    wb.save(output_path)
    return output_path


# ── Main ──────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description="Validate Royal London approved URLs before scraping"
    )
    parser.add_argument(
        "--file", default=DEFAULT_EXCEL,
        help="Path to customer Excel file",
    )
    parser.add_argument(
        "--concurrency", type=int, default=DEFAULT_CONCURRENCY,
        help="Number of concurrent HTTP requests (default: 15)",
    )
    parser.add_argument(
        "--timeout", type=int, default=DEFAULT_TIMEOUT,
        help="Seconds before a URL is marked as timeout (default: 12)",
    )
    args = parser.parse_args()

    excel_path = args.file
    if not Path(excel_path).exists():
        print(f"\n❌ ERROR: File not found: {excel_path}")
        sys.exit(1)

    print("\n" + "=" * 65)
    print("   Royal London — Pre-Scrape URL Validation")
    print("=" * 65)
    print(f"   Input:       {excel_path}")
    print(f"   Concurrency: {args.concurrency} simultaneous requests")
    print(f"   Timeout:     {args.timeout}s per URL")
    print(f"   Retries:     {MAX_RETRIES} on timeout/connection error")

    # ── Load URLs ──────────────────────────────────────────────
    print("\n📋 Loading URLs from Excel...")
    try:
        entries = load_urls_from_excel(excel_path)
    except ValueError as e:
        print(f"\n❌ ERROR: {e}")
        sys.exit(1)

    print(f"   Found {len(entries):,} URLs to validate")

    if not entries:
        print("\n⚠️  No URLs found. Check the Excel file format.")
        sys.exit(1)

    # ── Check URLs ─────────────────────────────────────────────
    print(f"\n🌐 Checking {len(entries):,} URLs "
          f"(~{len(entries) // args.concurrency + 1}s estimated)...")

    results = asyncio.run(
        check_all_urls(entries, args.concurrency, args.timeout)
    )

    # ── Detect duplicates ──────────────────────────────────────
    print("\n🔍 Checking for duplicates...")
    results = detect_duplicates(results)
    dup_count = sum(1 for r in results if r.get("is_duplicate"))
    print(f"   Found {dup_count} duplicate URL(s)")

    # ── Write report ───────────────────────────────────────────
    timestamp   = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    output_dir  = Path(excel_path).parent
    output_path = str(
        output_dir / f"url_validation_report_{timestamp}.xlsx"
    )

    print(f"\n📊 Writing report...")
    write_report(results, output_path, excel_path)

    # ── Console summary ────────────────────────────────────────
    counts = {
        "live":      sum(1 for r in results if r["status"] == "live"),
        "redirect":  sum(1 for r in results if r["status"] == "redirect"),
        "dead":      sum(1 for r in results if r["status"] == "dead"),
        "timeout":   sum(1 for r in results if r["status"] == "timeout"),
        "duplicate": sum(1 for r in results if r["status"] == "duplicate"),
        "error":     sum(1 for r in results if r["status"] == "error"),
    }
    action_needed = (
        counts["dead"] + counts["redirect"] +
        counts["timeout"] + counts["duplicate"] + counts["error"]
    )

    print("\n" + "=" * 65)
    print("   VALIDATION COMPLETE")
    print("=" * 65)
    print(f"   Total checked:         {len(results):,}")
    print(f"   ✅ Live:               {counts['live']:,}")
    print(f"   ❌ Dead (4xx/5xx):     {counts['dead']:,}")
    print(f"   ⚠️  Redirected (ext):  {counts['redirect']:,}")
    print(f"   ⏱️  Timeout:           {counts['timeout']:,}")
    print(f"   🔁 Duplicate:          {counts['duplicate']:,}")
    print(f"   ❌ Other errors:       {counts['error']:,}")
    print(f"   ─────────────────────────────")
    print(f"   ⚠️  Action required:   {action_needed:,}")
    print(f"\n   Report saved to:")
    print(f"   {output_path}")

    if action_needed > 0:
        print(
            f"\n   👉 Open the report and share the "
            f"'Action Required' sheet with the customer."
        )
    else:
        print("\n   ✅ All URLs are live and unique — ready to scrape!")

    print("=" * 65 + "\n")


if __name__ == "__main__":
    main()