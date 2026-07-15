"""
Royal London ARIA — Dropdown Detection Diagnostic
==================================================
Scans all approved URLs and detects which pages have <select>
elements, classifying them by type so we know exactly what we
are dealing with before updating the scraper strategy.

Outputs:
  - Console summary table
  - Excel report: scraper/data/dropdown_detection_<timestamp>.xlsx

Usage:
    python scraper/detect_dropdowns.py --file scraper/data/Approved_URLs.xlsx

No index or cache changes — read-only diagnostic tool.
"""

import argparse
import asyncio
import os
import re
import sys
import time
from datetime import datetime, timezone
from pathlib import Path

# ── Imports ───────────────────────────────────────────────────
from dotenv import find_dotenv, load_dotenv
load_dotenv(find_dotenv(usecwd=False), override=True)

import structlog
log = structlog.get_logger()

# ── CDP constants (mirrors scraper v4) ────────────────────────
_CDP_PORT    = int(os.getenv("PLAYWRIGHT_CDP_PORT", "9222"))
_CDP_URL     = f"http://localhost:{_CDP_PORT}"
_CHROME_PROC = None

PLAYWRIGHT_EXECUTABLE_PATH = os.getenv(
    "PLAYWRIGHT_EXECUTABLE_PATH",
    r"C:\Program Files\Google\Chrome\Application\chrome.exe",
)

# ── Placeholder options to ignore ─────────────────────────────
PLACEHOLDERS = {
    "select...", "select", "please select", "--", "choose...",
    "choose", "please choose", "all", "-- select an option --",
    "- select -", "select an option", "select option",
}

# ── Contact signal patterns ────────────────────────────────────
PHONE_RE    = re.compile(r'0\d{3,4}\s?\d{3,4}\s?\d{3,4}')
POSTCODE_RE = re.compile(r'[A-Z]{1,2}\d{1,2}\s?\d[A-Z]{2}')
CONTACT_KEYWORDS = [
    "call us", "write to us", "lines are open", "fill in our",
    "tell us someone", "online form", "monday to friday",
    "excluding bank holidays",
]

# ── Content filter signals ─────────────────────────────────────
FILTER_KEYWORDS = [
    "filter by", "category", "years", "months", "newest", "oldest",
    "sort by", "results", "page 1 of",
]

# ── Form field signals ─────────────────────────────────────────
FORM_KEYWORDS = [
    "mr", "mrs", "ms", "miss", "other", "title",
    "first name", "last name", "date of birth",
]


def _is_contact_content(text: str) -> bool:
    t = text.lower()
    return (
        bool(PHONE_RE.search(t)) or
        bool(POSTCODE_RE.search(text)) or
        any(kw in t for kw in CONTACT_KEYWORDS)
    )


def classify_dropdown(options: list[str], page_url: str) -> str:
    """
    Classify a dropdown type from its option labels and page URL.

    Returns one of:
      contact_routing  — bereavement/contact pages with policy options
      content_filter   — Category/Years/Months/sort controls
      form_field       — Mr/Mrs/Ms title selects and similar
      navigation       — tabs switching between sibling pages
      unknown          — can't determine without JS execution
    """
    url_lower   = page_url.lower()
    opts_lower  = [o.lower() for o in options]
    opts_joined = " ".join(opts_lower)

    # Form field detection
    form_hits = sum(1 for kw in FORM_KEYWORDS if kw in opts_joined)
    if form_hits >= 2:
        return "form_field"

    # Content filter detection
    filter_hits = sum(1 for kw in FILTER_KEYWORDS if kw in opts_joined)
    if filter_hits >= 1:
        return "content_filter"

    # URL-based hints
    if any(p in url_lower for p in ["/life-events", "/retirement-guidance",
                                     "/guides-tools", "/news", "/blog"]):
        return "content_filter"

    # Contact routing hints
    if any(p in url_lower for p in ["/contact", "/bereavement",
                                     "/make-a-claim", "/help-and-support"]):
        return "contact_routing"

    return "unknown"


def start_chrome_cdp() -> bool:
    global _CHROME_PROC
    import socket, subprocess, sys as _sys, time as _t

    if not PLAYWRIGHT_EXECUTABLE_PATH or not os.path.exists(PLAYWRIGHT_EXECUTABLE_PATH):
        return False

    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        s.settimeout(1)
        if s.connect_ex(("localhost", _CDP_PORT)) == 0:
            return True

    try:
        flags = 0
        if _sys.platform == "win32":
            flags = subprocess.CREATE_NO_WINDOW
        _CHROME_PROC = subprocess.Popen(
            [PLAYWRIGHT_EXECUTABLE_PATH,
             f"--remote-debugging-port={_CDP_PORT}",
             "--headless=new", "--no-sandbox",
             "--disable-dev-shm-usage", "--disable-gpu",
             "--no-first-run", "--no-default-browser-check"],
            stdin=subprocess.DEVNULL,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            creationflags=flags,
        )
        for _ in range(15):
            _t.sleep(0.5)
            with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
                s.settimeout(0.5)
                if s.connect_ex(("localhost", _CDP_PORT)) == 0:
                    return True
        return False
    except Exception as e:
        print(f"CDP start failed: {e}")
        return False


def stop_chrome_cdp():
    global _CHROME_PROC
    if _CHROME_PROC:
        import sys as _sys
        try: _CHROME_PROC.terminate()
        except Exception: pass
        try: _CHROME_PROC.wait(timeout=5)
        except Exception: pass
        for pipe in [_CHROME_PROC.stdin, _CHROME_PROC.stdout, _CHROME_PROC.stderr]:
            if pipe:
                try: pipe.close()
                except Exception: pass
        _CHROME_PROC = None


def load_urls_from_excel(file_path: str) -> list[dict]:
    from openpyxl import load_workbook
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    URL_HEADERS   = {"url", "page url", "link", "webpage", "web url"}
    TITLE_HEADERS = {"title", "page title", "name"}

    url_idx = title_idx = None
    entries = []

    for row in ws.iter_rows(values_only=True):
        if url_idx is None:
            for i, cell in enumerate(row):
                if cell and str(cell).strip().lower() in URL_HEADERS:
                    url_idx = i
                elif cell and str(cell).strip().lower() in TITLE_HEADERS:
                    title_idx = i
            continue

        url = str(row[url_idx]).strip() if row[url_idx] else ""
        if not url or not url.startswith("http"):
            continue
        title = str(row[title_idx]).strip() if (title_idx is not None and row[title_idx]) else ""
        entries.append({"url": url, "title": title})

    wb.close()
    return entries


async def detect_dropdowns_on_page(page, url: str, title: str) -> dict:
    """
    Navigate to URL and detect all <select> elements.
    Returns result dict for this URL.
    """
    result = {
        "url":              url,
        "title":            title,
        "has_dropdown":     False,
        "dropdown_count":   0,
        "total_options":    0,
        "dropdown_type":    "",
        "option_samples":   "",
        "notes":            "",
    }

    try:
        response = await page.goto(
            url,
            wait_until="domcontentloaded",
            timeout=30000,
        )

        if not response or response.status >= 400:
            result["notes"] = f"HTTP {response.status if response else 'no response'}"
            return result

        # Check for redirect
        final_url = page.url
        if final_url.rstrip("/") != url.rstrip("/"):
            result["notes"] = f"Redirected → {final_url[:80]}"
            return result

        # Detect all <select> elements
        select_data = await page.evaluate("""
        () => {
            const selects = Array.from(document.querySelectorAll('select'));
            return selects.map(sel => ({
                options: Array.from(sel.options).map(o => o.text.trim()).filter(t => t)
            }));
        }
        """)

        valid_selects = []
        for sel in select_data:
            opts = [
                o for o in sel["options"]
                if o.lower() not in {
                    "select...", "select", "please select", "--",
                    "choose...", "choose", "please choose",
                    "-- select an option --", "- select -",
                    "select an option", "all",
                }
            ]
            if len(opts) > 1:
                valid_selects.append(opts)

        if not valid_selects:
            result["notes"] = "No meaningful dropdowns"
            return result

        result["has_dropdown"]   = True
        result["dropdown_count"] = len(valid_selects)
        result["total_options"]  = sum(len(s) for s in valid_selects)

        # Classify using first dropdown's options
        first_opts       = valid_selects[0]
        dropdown_type    = classify_dropdown(first_opts, url)
        result["dropdown_type"]  = dropdown_type
        result["option_samples"] = " | ".join(first_opts[:5])
        if len(first_opts) > 5:
            result["option_samples"] += f" ... (+{len(first_opts)-5} more)"

    except Exception as e:
        result["notes"] = f"Error: {type(e).__name__}: {str(e)[:80]}"

    return result


async def run_detection(entries: list[dict]) -> list[dict]:
    """Run detection across all URLs using Playwright."""
    try:
        from playwright.async_api import async_playwright
    except ImportError:
        print("❌ playwright not installed. pip install playwright")
        sys.exit(1)

    results = []
    total   = len(entries)

    use_cdp = start_chrome_cdp()

    async with async_playwright() as pw:
        if use_cdp:
            print(f"   CDP mode — connecting to Chrome on port {_CDP_PORT}")
            browser = await pw.chromium.connect_over_cdp(_CDP_URL)
        else:
            print("   Playwright mode — launching own browser")
            browser = await pw.chromium.launch(headless=True)

        context = await browser.new_context(
            user_agent="ARIA-DropdownDetector/1.0 (internal diagnostic)"
        )
        page    = await context.new_page()

        # Block images/fonts for speed
        await page.route(
            "**/*.{png,jpg,jpeg,gif,svg,woff,woff2,ttf,eot}",
            lambda route: route.abort()
        )

        try:
            for i, entry in enumerate(entries, 1):
                url   = entry["url"]
                title = entry.get("title", "")
                print(f"  [{i:3d}/{total}] {url[:80]}", end="\r")

                result = await detect_dropdowns_on_page(page, url, title)
                results.append(result)

                # Small delay to avoid hammering CDN
                await asyncio.sleep(0.3)

        finally:
            await context.close()
            if not use_cdp:
                await browser.close()

    stop_chrome_cdp()
    return results


def build_report(results: list[dict], output_path: Path) -> None:
    """Build Excel report with 3 sheets."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl not installed — skipping Excel report")
        return

    COLOURS = {
        "contact_routing": "C8E6C9",   # green
        "content_filter":  "FFF9C4",   # yellow
        "form_field":      "FFE0B2",   # orange
        "navigation":      "E1BEE7",   # purple
        "unknown":         "B3E5FC",   # blue
        "no_dropdown":     "F5F5F5",   # grey
        "header":          "2C3E50",   # dark
    }

    wb  = Workbook()

    # ── Sheet 1: All URLs ──────────────────────────────────────
    ws1 = wb.active
    ws1.title = "All URLs"
    headers = ["URL", "Title", "Has Dropdown", "Count",
               "Total Options", "Type", "Option Samples", "Notes"]
    ws1.append(headers)
    _style_header(ws1, COLOURS["header"])

    for r in results:
        colour = COLOURS.get(r["dropdown_type"] if r["has_dropdown"] else "no_dropdown", "FFFFFF")
        ws1.append([
            r["url"], r["title"],
            "YES" if r["has_dropdown"] else "No",
            r["dropdown_count"], r["total_options"],
            r["dropdown_type"], r["option_samples"], r["notes"],
        ])
        fill = PatternFill("solid", fgColor=colour)
        for cell in ws1[ws1.max_row]:
            cell.fill = fill

    _auto_width(ws1)

    # ── Sheet 2: Dropdown URLs Only ────────────────────────────
    ws2 = wb.create_sheet("Dropdown URLs")
    ws2.append(headers)
    _style_header(ws2, COLOURS["header"])
    for r in results:
        if r["has_dropdown"]:
            colour = COLOURS.get(r["dropdown_type"], "FFFFFF")
            ws2.append([
                r["url"], r["title"], "YES",
                r["dropdown_count"], r["total_options"],
                r["dropdown_type"], r["option_samples"], r["notes"],
            ])
            fill = PatternFill("solid", fgColor=colour)
            for cell in ws2[ws2.max_row]:
                cell.fill = fill
    _auto_width(ws2)

    # ── Sheet 3: Summary ───────────────────────────────────────
    ws3 = wb.create_sheet("Summary")
    dropdown_urls = [r for r in results if r["has_dropdown"]]
    type_counts   = {}
    for r in dropdown_urls:
        t = r["dropdown_type"] or "unknown"
        type_counts[t] = type_counts.get(t, 0) + 1

    ws3.append(["ARIA Dropdown Detection Report", ""])
    ws3.append(["Run At", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")])
    ws3.append(["Total Approved URLs", len(results)])
    ws3.append(["URLs with dropdowns", len(dropdown_urls)])
    ws3.append(["URLs without dropdowns", len(results) - len(dropdown_urls)])
    ws3.append(["", ""])
    ws3.append(["DROPDOWN TYPE BREAKDOWN", ""])
    for t, count in sorted(type_counts.items()):
        ws3.append([t, count])
    ws3.append(["", ""])
    ws3.append(["COLOUR KEY", ""])
    ws3.append(["contact_routing — GREEN",  "These are the ones we SHOULD scrape dropdown states for"])
    ws3.append(["content_filter  — YELLOW", "Category/Years/Months/sort filters — skip dropdown states"])
    ws3.append(["form_field      — ORANGE", "Mr/Mrs/Ms title selects — skip dropdown states"])
    ws3.append(["navigation      — PURPLE", "Tab controls switching between sibling pages — skip"])
    ws3.append(["unknown         — BLUE",   "Cannot classify without JS — review manually"])
    _auto_width(ws3)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))


def _style_header(ws, bg_hex: str):
    fill = PatternFill("solid", fgColor=bg_hex)
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def _auto_width(ws, max_w: int = 80):
    from openpyxl.utils import get_column_letter
    for col in ws.columns:
        max_len    = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, max_w)


def main():
    parser = argparse.ArgumentParser(
        description="Detect which approved Royal London URLs have dropdown elements."
    )
    parser.add_argument(
        "--file", required=True,
        help="Path to approved URLs Excel file."
    )
    parser.add_argument(
        "--output", default=None,
        help="Output Excel path. Default: scraper/data/dropdown_detection_<timestamp>.xlsx"
    )
    args = parser.parse_args()

    if not os.path.exists(args.file):
        print(f"❌ File not found: {args.file}")
        sys.exit(1)

    ts          = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = Path(args.output) if args.output else \
                  Path(f"scraper/data/dropdown_detection_{ts}.xlsx")

    print("\n" + "=" * 65)
    print("   ARIA Dropdown Detection Diagnostic")
    print("=" * 65)
    print(f"   Input:  {args.file}")
    print(f"   Output: {output_path}")
    print("=" * 65 + "\n")

    print("📋 Loading approved URLs...")
    entries = load_urls_from_excel(args.file)
    print(f"   {len(entries):,} URLs loaded\n")

    print(f"🔍 Scanning {len(entries):,} URLs for dropdowns...")
    results = asyncio.run(run_detection(entries))
    print(f"\n   Scan complete — {len(results):,} URLs processed\n")

    # Console summary
    dropdown_urls = [r for r in results if r["has_dropdown"]]
    type_counts: dict[str, list] = {}
    for r in dropdown_urls:
        t = r["dropdown_type"] or "unknown"
        type_counts.setdefault(t, []).append(r)

    print("=" * 65)
    print("   RESULTS")
    print("=" * 65)
    print(f"   Total URLs scanned       : {len(results):,}")
    print(f"   URLs WITH dropdowns      : {len(dropdown_urls):,}")
    print(f"   URLs without dropdowns   : {len(results) - len(dropdown_urls):,}")
    print()
    print("   BREAKDOWN BY TYPE:")
    for t, urls in sorted(type_counts.items()):
        print(f"   {t:<20}: {len(urls):3d} URLs")
        for r in urls[:3]:
            print(f"     → {r['url'][:70]}")
        if len(urls) > 3:
            print(f"       ... and {len(urls)-3} more")
    print("=" * 65)

    print(f"\n📊 Building Excel report...")
    build_report(results, output_path)
    print(f"   Saved: {output_path}\n")


if __name__ == "__main__":
    import sys as _sys
    if _sys.platform == "win32":
        import warnings as _warnings
        _warnings.filterwarnings("ignore", message=".*I/O operation on closed pipe.*", category=ResourceWarning)
        _warnings.filterwarnings("ignore", message=".*unclosed transport.*", category=ResourceWarning)
        _orig = _sys.unraisablehook
        def _hook(u):
            if "I/O operation on closed pipe" in str(u.exc_value): return
            _orig(u)
        _sys.unraisablehook = _hook

    main()