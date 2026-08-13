"""
Royal London Digital Assistance — Content Freshness Manager (V1)
====================================================
Nightly job that compares the approved URL Excel (from Azure Blob
Storage) against the live Azure AI Search index, detects content
changes using SHA-256 hashing, and keeps BOTH indexes + Redis cache
in sync with what is actually on the Royal London website.

FULLY SELF-CONTAINED — no imports from scrape_approved_urls_updatedV5.py,
chunk_and_index_hqaV5.py, or element_detector.py. All scraping, HQA,
chunking (including element-aware chunking, v1.7.0), embedding, and
index management logic lives here — deliberate design principle
(v1.4.0), confirmed unchanged when element-aware chunking was added.

DUAL-INDEX AUTO MODE:
Every apply run automatically updates BOTH indexes:
  rlg-faq-index-v5          — full HQA + title_questions
  rlg-faq-index-v5-baseline — no HQA (title as retrieval signal only)
No flags needed. Consistency guaranteed — a changed page always gets
full HQA treatment in main index and baseline treatment in baseline index.
Eliminates the mismatch that would occur if freshness only updated one.

DECISION LOG:
  - Excel in Azure Blob is the SINGLE source of truth for approved URLs.
  - Internal redirects treated as removed — new URL must be in Excel.
  - External redirects always removed.
  - 404 / 5xx URLs removed. Re-added automatically if they recover.
  - Cache invalidation is TARGETED — only affected keys flushed.
  - content_hash (SHA-256) is the change detection signal.
  - HQA generated for ALL delta-indexed pages — no quality mismatch
    between full-index and freshness-indexed chunks.
  - Scrape once → chunk once → HQA augment → index to BOTH indexes.
  - Dropdown pages (#state=): all variants deleted when base URL changes.

TWO MODES:
  --mode report  (default / safe)
      Full scan, produce Excel report, NO writes to index or cache.
  --mode apply   (nightly job)
      Full scan + execute all changes to BOTH indexes + cache.

═══════════════════════════════════════════════════════════════
LOCAL USAGE (VDI)
═══════════════════════════════════════════════════════════════

    # Report only — safe, no writes (uses local Excel)
    python scraper/content_freshnessV1.py --mode report --file scraper/data/Approved_URLs.xlsx

    # Apply mode — local Excel
    python scraper/content_freshnessV1.py --mode apply --file scraper/data/Approved_URLs.xlsx

    # Report — production (downloads Excel from Blob)
    python scraper/content_freshnessV1.py --mode report

    # Apply — production (nightly job)
    python scraper/content_freshnessV1.py --mode apply

    # Dry run — validate config + connectivity, no writes
    python scraper/content_freshnessV1.py --mode apply --dry-run

═══════════════════════════════════════════════════════════════
PRODUCTION — AZURE CONTAINER APPS JOB (DevOps)
═══════════════════════════════════════════════════════════════

# TODO (DevOps): Create Container Apps Job: digital-assistance-freshness-job
# Schedule: nightly at 02:00 UTC (after CMS publishing windows)
# Image: same image as digital-assistance-scraper-job (crawl4ai + Playwright)
#
# ── DOCKERFILE ────────────────────────────────────────────────
#
#   FROM python:3.11-slim
#   RUN apt-get update && apt-get install -y \
#       wget gnupg ca-certificates fonts-liberation \
#       libasound2 libatk-bridge2.0-0 libdrm2 libxkbcommon0 \
#       libxcomposite1 libxdamage1 libxfixes3 libxrandr2 \
#       libgbm1 libnss3 libnspr4 libdbus-1-3 libatspi2.0-0 \
#       && rm -rf /var/lib/apt/lists/*
#   COPY requirements.txt .
#   RUN pip install -r requirements.txt
#   RUN playwright install chromium --with-deps
#   # DO NOT set PLAYWRIGHT_EXECUTABLE_PATH in container.
#   # Leave unset — CDP mode auto-skipped on Linux.
#   COPY . .
#   CMD ["python", "scraper/content_freshnessV1.py", "--mode", "apply"]
#
# ── AZURE MANAGED IDENTITY ────────────────────────────────────
#
#   Resource                    Role
#   ─────────────────────────── ────────────────────────────────
#   Azure AI Search             Search Index Data Contributor
#   Azure OpenAI                Cognitive Services OpenAI User
#   Azure Blob Storage          Storage Blob Data Contributor
#   Azure Key Vault             Key Vault Secrets User
#
#   DefaultAzureCredential picks up Managed Identity automatically.
#   TODO (DevOps): assign identity to digital-assistance-freshness-job + grant roles.
#
# ── AZURE KEY VAULT — required secrets ───────────────────────
#
#   Secret Name                              Value
#   ──────────────────────────────────────── ────────────────────
#   AZURE-STORAGE-CONNECTION                 Blob Storage conn string
#   AZURE-SEARCH-ENDPOINT                    https://<n>.search.windows.net
#   AZURE-SEARCH-INDEX-NAME                  rlg-faq-index-v5
#   AZURE-SEARCH-BASELINE-INDEX-NAME         rlg-faq-index-v5-baseline
#   AZURE-OPENAI-ENDPOINT                    https://<n>.openai.azure.com
#   AZURE-OPENAI-EMBEDDING-DEPLOYMENT        text-embedding-3-large
#   AZURE-OPENAI-EMBEDDING-DIMENSIONS        1536
#   AZURE-OPENAI-DEPLOYMENT-HQA              gpt-4o-mini
#   REDIS-URL                                rediss://<host>:6380
#   BLOB-CONTAINER-NAME                      scraper-data
#   BLOB-APPROVED-EXCEL-NAME                 approved-urls/Approved_URLs.xlsx
#   FRESHNESS-MODE                           apply
#   BLOB-ARCHIVE-PREFIX                      freshness/archive/
#   CHUNK-SIZE                               1600
#   CHUNK-OVERLAP                            200
#   EMBEDDING-BATCH-SIZE                     50
#   HQA-QUESTIONS-FIRST-CHUNK               8
#   HQA-QUESTIONS-OTHER-CHUNKS              5
#   TITLE-QUESTIONS-COUNT                    3
#   TITLE-QUESTIONS-MAX-WORDS               12
#   MAX-COLLISION-THRESHOLD                  3
#
#   DO NOT add PLAYWRIGHT-EXECUTABLE-PATH to Key Vault.
#   In production the container installs chromium via Dockerfile.
#   CDP mode auto-skipped when path doesn't exist (Linux).
#
# ── BLOB STORAGE PATHS ────────────────────────────────────────
#
#   scraper-data/
#     approved-urls/Approved_URLs.xlsx    <- dedicated team updates
#     freshness/content_hashes.json       <- written after each apply run
#     freshness/reports/freshness_report_apply_<ts>.xlsx
#     freshness/manifests/run_manifest_<ts>.json
#
# ── CONTAINER APPS JOB TRIGGER ───────────────────────────────
#
#   # Nightly schedule:
#   Schedule: "0 2 * * *"  (02:00 UTC daily)
#
#   # Manual trigger:
#   az containerapp job start \
#       --name digital-assistance-freshness-job \
#       --resource-group <rg>
#
# ── FIRST RUN CHECKLIST ───────────────────────────────────────
#
#   1. chunk_and_index_hqaV5.py --full completes for BOTH indexes
#      (main + baseline) — builds v5 schema with all version fields
#      (including element_type, v5.10.0)
#   2. Set AZURE-SEARCH-INDEX-NAME -> rlg-faq-index-v5 in Key Vault
#   3. Set AZURE-SEARCH-BASELINE-INDEX-NAME -> rlg-faq-index-v5-baseline
#   4. Run freshness in report mode to verify scan:
#      python scraper/content_freshnessV1.py --mode report
#   5. Review Excel report — confirm URLs, hashes, dropdown states
#   6. Enable nightly Container Apps Job schedule

CHANGE LOG
==========

v1.0.0 — July 2026 | Mukesh Kund
         Initial production version. Two-mode design (report/apply).
         Targeted Redis cache invalidation. Dropdown page support.
         SHA-256 content hashing. Azure Blob state storage.

v1.1.0 — July 2026 | Mukesh Kund
         Atomic chunking for dropdown state pages — aligned with
         chunk_and_index_hqaV4.py v5.4.0.

v1.2.0 — July 2026 | Mukesh Kund
         Playwright-based dropdown scraping. Aligned with
         scrape_approved_urls_updatedV4.py v4.5.0.

v1.2.1 — July 2026 | Mukesh Kund
         CDP mode for VDI. Aligned with scraper v4.5.2.

v1.2.2 — July 2026 | Mukesh Kund
         DevOps section fully documented.

v1.2.3 — July 2026 | Mukesh Kund
         Fix ValueError: I/O operation on closed pipe on Windows.

v1.2.4 — July 2026 | Mukesh Kund
         Suppress asyncio ProactorEventLoop GC noise on Windows.

v1.2.5 — July 2026 | Mukesh Kund
         Bulletproof dropdown scraping — contact signal validation
         + navigation guard + placeholder expansion.
         Aligned with scraper v4.5.5.

v1.3.0 — July 2026 | Mukesh Kund
         Content versioning and traceability. CHUNK_SIZE/CHUNK_OVERLAP/
         EMBEDDING_BATCH_SIZE externalised to env vars. save_run_manifest()
         added. chunk_page() stamped with versioning fields.
         All stale V3 references fixed.

v1.4.0 — July 2026 | Mukesh Kund
         Full HQA in freshness + dual-index auto mode.

         PROBLEM FIXED:
         Previously freshness re-indexed changed pages with empty
         augmented_questions and title_questions. This caused a
         quality mismatch — full-index chunks had HQA, freshness-
         delta chunks did not. Client correctly identified this as
         unacceptable. Every chunk in the index must have the same
         HQA quality regardless of whether it was indexed in the
         full run or a delta freshness run.

         SOLUTION — Full HQA pipeline in freshness (self-contained):
         All HQA logic (prompts, generation, validation, scoring,
         deduplication, embedding text building) from
         chunk_and_index_hqaV4.py duplicated here. No imports from
         other scripts. Freshness is now fully independent.

         DUAL-INDEX AUTO MODE:
         Every apply run automatically updates BOTH indexes:
           rlg-faq-index-v4           — full HQA + title_questions
           rlg-faq-index-v4-baseline  — no HQA (baseline mode)
         Scrape once, chunk once, HQA augment once, then:
           Upload HQA chunks  -> main index
           Upload same chunks (HQA cleared) -> baseline index
         No --no-hqa flag needed. Both indexes always stay in sync.

         FULL METADATA EXTRACTION:
         scrape_url_with_dropdowns() now calls extract_page_metadata()
         for rich metadata (has_video, content_type, product_category,
         description, thumbnail_url, publish_date, collection_name,
         read_time_mins). Matches full scraper quality.

         VERSIONING:
         SCRAPER_VERSION and METADATA_VERSION constants added.
         Stamped on all delta-scraped pages and chunks.
         Independent of scraper file constants — bump here when
         freshness scraping or metadata logic changes.

         BUMPING RULES:
           SCRAPER_VERSION  — bump when scrape_url_with_dropdowns()
                              or CDP/Playwright logic changes here
           METADATA_VERSION — bump when extract_page_metadata() or
                              CONTENT_TYPE_MAP / PRODUCT_CATEGORY_MAP
                              or EXCEL_CATEGORY_MAP changes here
           PIPELINE_VERSION — bump when chunking, HQA prompts/logic,
                              embedding, or schema changes here
           FRESHNESS_JOB_VERSION— bump when hash comparison, URL
                              classification, or apply logic changes

         NEW Key Vault secrets required:
           AZURE-SEARCH-BASELINE-INDEX-NAME  rlg-faq-index-v4-baseline
           AZURE-OPENAI-DEPLOYMENT-HQA       gpt-4o-mini
           HQA-QUESTIONS-FIRST-CHUNK         8
           HQA-QUESTIONS-OTHER-CHUNKS        5
           TITLE-QUESTIONS-COUNT             3
           TITLE-QUESTIONS-MAX-WORDS         12
           MAX-COLLISION-THRESHOLD           3

v1.5.0 — July 2026 | Mukesh Kund
         refresh_count auto-increment + chunk archive + versioning
         descriptions + FRESHNESS_VERSION renamed.

         RENAME:
         FRESHNESS_VERSION -> FRESHNESS_JOB_VERSION throughout.
         Prevents confusion with refresh_count (chunk-level int field).
         Old name appeared identical in style to chunk field names.
         New name makes clear this is a SCRIPT/JOB version constant,
         not a per-chunk data field.

         refresh_count AUTO-INCREMENT (new field on chunks):
         Before deleting old chunks for a changed URL, freshness now
         reads the existing refresh_count from the index. New chunks
         are stamped with refresh_count + 1.
           0 = indexed by chunk_and_index_hqaV4.py full run
           1 = changed once and re-indexed by freshness
           N = changed N times and re-indexed by freshness
         get_refresh_count_for_url() — new helper that reads existing
         chunk from main index and returns current refresh_count.
         chunk_page() and index_pages_dual() updated to accept and
         stamp refresh_count on all chunks.

         CHUNK ARCHIVE BEFORE DELETE (Point 2 from design discussion):
         Before deleting any chunks from either index, the existing
         chunk documents are serialised to JSON and saved to Blob at:
           freshness/archive/deleted_<url_hash>_<ts>.json
         Purpose: audit trail + manual rollback capability if a content
         change causes regression in model responses. Non-fatal — archive
         failure never blocks the delete/reindex.
         save_deleted_chunks_to_blob() — new helper function.
         BLOB_ARCHIVE_PREFIX constant added (default: freshness/archive/).

         VERSIONING DESCRIPTION BLOCK:
         All versioning fields documented with their meaning, who sets
         them, and bumping rules — as code comments near constants.

         FRESHNESS_JOB_VERSION meaning:
         Tracks WHICH VERSION OF THIS FRESHNESS SCRIPT ran a given job.
         Stored in run manifest only (NOT on chunks).
         Bump when: hash comparison strategy, URL classification,
         delta detection, or apply orchestration logic changes.

         refresh_count meaning (on chunks):
         Tracks HOW MANY TIMES this page's content has been REFRESHED
         by the nightly freshness job. Auto-incremented, never
         developer-bumped. 0 = full run, N = refreshed N times.

         New Blob path (Andy action):
           BLOB-ARCHIVE-PREFIX   freshness/archive/

v1.5.1 — July 2026 | Mukesh Kund
         GPT-5 compatibility: max_tokens → max_completion_tokens,
         temperature removed.

         generate_title_questions() [MODIFIED]:
         generate_hqa_questions()   [MODIFIED]:
         - GPT-5 models reject max_tokens with HTTP 400.
           Renamed to max_completion_tokens in both HQA call sites.
         - GPT-5 models reject temperature=0.3 with HTTP 400
           ("only default value 1 is supported"). Removed entirely
           from both HQA call sites.

         ── ROLLBACK TO gpt-4o-mini ──────────────────────────────
         If reverting HQA_DEPLOYMENT back to gpt-4o-mini:
         1. .env / Key Vault: set AZURE_OPENAI_DEPLOYMENT_HQA
            back to your gpt-4o-mini deployment name.
         2. generate_title_questions(): restore
              max_completion_tokens=200  →  max_tokens=200
              add back: temperature=0.3
         3. generate_hqa_questions(): restore
              max_completion_tokens=max_tokens  →  max_tokens=max_tokens
              add back: temperature=0.3
         No other changes needed — all other logic is model-agnostic.
         ─────────────────────────────────────────────────────────

v1.5.2 — July 2026 | Mukesh Kund
         GPT-5 compatibility: robust JSON array extraction.

         generate_title_questions() [MODIFIED]:
         generate_hqa_questions()   [MODIFIED]:
         - GPT-5 reasoning models return prose preamble or postamble
           around the JSON array, causing hqa_json_parse_failed on
           every chunk with the old startswith("```") fence-strip.
         - FIX: extract JSON by finding the outermost [ ... ] span
           in the raw response before attempting json.loads(). The
           markdown fence strip is kept as a secondary fallback.

         ── ROLLBACK TO gpt-4o-mini ──────────────────────────────
         If reverting HQA_DEPLOYMENT back to gpt-4o-mini, the JSON
         extraction change is safe to leave in place — gpt-4o-mini
         returns bare JSON arrays so find("[") / rfind("]") still
         works correctly. No code revert needed for this change.
         ─────────────────────────────────────────────────────────

v1.5.3 — July 2026 | Mukesh Kund
         GPT-5 compatibility: max_completion_tokens floor raised to 4000.

         generate_title_questions() [MODIFIED]:
         generate_hqa_questions()   [MODIFIED]:
         - finish_reason=length with content='' confirmed GPT-5-mini
           exhausts token budget on internal reasoning tokens before
           generating JSON output.
         - title_questions: 200 → 4000
         - hqa_questions: int(num_questions*55+50) → max(..., 4000)
         - gpt-4o-mini unaffected (non-reasoning; billed for actual
           output tokens only regardless of limit set).

         ── ROLLBACK TO gpt-4o-mini ──────────────────────────────
         Safe to leave 4000 in place. No revert required.
         ─────────────────────────────────────────────────────────

v1.5.4 — July 2026 | Mukesh Kund
         HQA cost tracking + model-aware estimates (mirrors indexer v5.8.5/5.8.6).

         MODEL_PRICING, CostTracker, _cost_tracker [NEW]:
         - Mirrors chunk_and_index_hqaV4.py v5.8.5 exactly.
           Keep in sync when adding new models or pricing changes.

         augment_chunks_with_hqa() [MODIFIED]:
         - Initialises CostTracker (run_mode="freshness-delta").
         - Model-aware time/cost estimate (gpt-5 vs gpt-4o-mini).
         - Prints cost summary after HQA completion.

         generate_title_questions() [MODIFIED]:
         generate_hqa_questions()   [MODIFIED]:
         - Call _cost_tracker.record(response.usage) on success.

v1.5.5 — July 2026 | Mukesh Kund
         Bugfix: read_time_mins int → str cast (upload failure).

         extract_metadata_from_html() [MODIFIED]:
         chunk_page() dropdown path [MODIFIED]:
         chunk_page() regular path [MODIFIED]:
         - Same root cause as indexer v5.8.7 — read_time_mins
           stored/passed as int but schema expects Edm.String.
         - Fixed at all 4 sites: default dict, metadata calc,
           dropdown scrape path, and page.get() pass-through.

v1.6.0 — July 2026 | Mukesh Kund
    FIX 1 — parent_url on dropdown state dicts in _scrape_dropdown_states_playwright().
    Mirrors scraper v4.8.0 fix. Dropdown state entries built here
    during delta re-index were also missing parent_url, so re-indexed
    dropdown chunks would still have dead #state= citation URLs.
    - Added "parent_url": url to every dropdown state dict.
    FIX 2 — parent_url pass-through in chunk_page().
    chunk_page() builds the final index document. parent_url must be
    carried from the page dict into the chunk dict (both atomic and
    regular paths) so the field reaches the index and retriever.py
    can resolve the clean citation URL. Mirrors indexer v5.9.0 fix.

v1.7.0 — August 2026 | Mukesh Kund
    RENAMED: content_freshness.py → content_freshnessV1.py.
    Targets rlg-faq-index-v5 / rlg-faq-index-v5-baseline — matches
    chunk_and_index_hqaV5.py. rlg-faq-index-v4 and v4-baseline are
    NOT touched by this script and remain live/untouched during the
    v5 rollout.

    ELEMENT-AWARE CHUNKING (inlined, not imported):
    chunk_page() now uses element-aware chunking — tables chunked
    atomically (row-capped), ##/### headers act as hard section
    boundaries. Same confirmed design as chunk_and_index_hqaV5.py
    v5.10.0 (see that file's changelog for the full Phase 0/1 audit
    trail this was built against).

    WHY INLINED, NOT IMPORTED FROM element_detector.py:
    This file's stated design principle (v1.4.0) is to be a SINGLE
    INDEPENDENT SCRIPT with zero dependency on other project code
    files — a deliberate choice so the nightly freshness job can
    never break because an unrelated script changed. Confirmed with
    project owner before building: that principle stands, so the
    element-aware chunking logic is duplicated here (prefixed _ea_)
    rather than imported. Same accepted tradeoff as the existing
    HQA/embedding duplication in this file — if element_detector.py
    is later fixed or improved, this block must be updated manually
    to match. No automated drift guard; this is a known, accepted
    risk, not a new category introduced by this change.

    NEW element_type FIELD:
    Every chunk (dropdown_state / prose / table) now carries
    element_type, matching the indexer's v5.10.0 schema addition.
    No schema changes needed in THIS file — chunk_and_index_hqaV5.py
    owns index creation; this file only writes documents into an
    index whose schema already includes the field.

    STRICTER CHECKPOINT — validate_chunking_preflight() [NEW]:
    New Step 7.5 in run_freshness_job(), between scraping (Step 7)
    and index mutation (Step 8). Chunks every scraped page and
    sanity-checks the output (no exceptions, no zero-chunk pages for
    non-trivial content, no oversized chunks, every chunk has
    element_type, total_chunks matches actual count) BEFORE Step 8
    deletes anything from the live index.

    WHY THIS MATTERS: Step 8 deletes existing chunks for a changed
    URL BEFORE indexing replacements. If chunking failed AFTER
    deletion, the URL would be left with nothing searchable until
    the next successful run. The new element-aware chunking logic
    (table parsing, section grouping) is new code with more edge
    cases than the old flat splitter, so this gate runs BEFORE any
    deletion — apply mode aborts with ZERO index writes if the
    pre-flight failure ratio exceeds 5% (_MAX_VALIDATION_FAILURE_
    RATIO). Report mode logs failures but continues (nothing to
    protect — no writes happen in report mode regardless).

v1.7.1 — August 2026 | Mukesh Kund
    BUGFIX: total_chunks/chunk_index computed before <50-char
    fragment filtering, not after — found by validate_chunking_
    preflight() during testing (see v1.7.0 entry above).

    chunk_page() [MODIFIED — regular (non-dropdown) path]:
    - OLD: total = len(pieces) computed BEFORE the filter loop that
      drops sub-50-char fragments. A page whose element-aware
      chunking produced e.g. 3 pieces, where 2 were filtered out for
      being too short, ended up with its 1 surviving chunk reporting
      total_chunks=3 — an internal inconsistency (declared count
      didn't match actual count in the index).
    - ROOT CAUSE: this exact pattern was inherited from
      chunk_and_index_hqaV4.py (present since v5.x, unchanged there
      per project decision to leave v4 untouched). Pre-existing, not
      introduced by v1.7.0 — but element-aware chunking's header-
      boundary splitting is MORE likely to produce short leading/
      trailing fragments (e.g. a one-line closing sentence after a
      table) than the old flat greedy splitter was, making this
      latent bug much more likely to actually trigger.
    - FIX: filter pieces first (valid_pieces = [p for p in pieces if
      len(p["text"].strip()) >= 50]), THEN compute
      total = len(valid_pieces) and enumerate over valid_pieces only.
      chunk_index and total_chunks now always agree with what's
      actually written to the index.
    - Same fix applied in parallel to chunk_and_index_hqaV5.py
      v5.10.1 (chunk_pages() regular path) — rlg-faq-index-v4 is
      NOT touched; the bug remains latent there by design (v4 is
      frozen/untouched per project decision).

v1.7.2 — August 2026 | Mukesh Kund
    PIPELINE_VERSION bumped 1.0.0 -> 1.1.0, matching
    chunk_and_index_hqaV5.py's PIPELINE_VERSION bump (v5.10.0).
    Was missed when v1.7.0 introduced element-aware chunking here —
    chunk_page() started using the new logic but the version constant
    stamped on chunks wasn't updated to reflect it. Without this fix,
    full-index chunks (pipeline_version=1.1.0) and freshness-delta
    chunks (pipeline_version=1.0.0) would report DIFFERENT versions
    despite both being produced by the same element-aware chunking
    logic — misleading for anyone auditing the index by this field.

v1.7.3 — August 2026 | Mukesh Kund
    ROOT-CAUSE FIX: chunk_id changed from random uuid.uuid4() to a
    deterministic SHA-256 hash of (source_url, chunk_index, content).
    Applied in parallel with chunk_and_index_hqaV5.py v5.10.2 — see
    that file's changelog for the full incident writeup (5,194
    duplicate chunks found in rlg-faq-index-v5 after its first
    --full build, root cause: random UUIDs don't survive accidental
    re-processing without creating duplicates).

    compute_chunk_id() [NEW] — byte-identical hash construction to
    chunk_and_index_hqaV5.py's version of the same function, so a
    chunk built by either script for the same (url, index, content)
    always resolves to the same ID — whether written by a --full
    run or a nightly freshness delta. Duplicated rather than
    imported, per this file's zero-cross-file-dependency principle
    (see module header) — same accepted tradeoff as every other
    piece of logic this file intentionally duplicates.

    chunk_page() [MODIFIED — both dropdown atomic and regular paths]:
    - "chunk_id": str(uuid.uuid4()) → compute_chunk_id(url, index, content)
    - Dropdown atomic chunks use index=0 (matches existing
      chunk_index=0 / total_chunks=1 convention).

    WHY THIS MATTERS FOR FRESHNESS SPECIFICALLY: the nightly job
    already deletes a changed URL's existing chunks before re-adding
    fresh ones (see Step 8 in run_freshness_job()) — deterministic
    IDs don't change that flow, but they DO mean if the delete step
    ever partially fails or a page is (re)processed twice within one
    run, re-adding the same content self-heals via overwrite instead
    of accumulating an orphaned duplicate.

v1.7.4 — August 2026 | Mukesh Kund
    BUGFIX: 100% false-positive rate on URL health check — every
    single approved URL (297/297) was being flagged as
    "internal_redirect: add new URL to Excel" on a --mode report run,
    despite the scraper (real Chrome via CDP) successfully fetching
    every one of those same URLs minutes earlier with no issue.

    ROOT CAUSE: check_single_url() used allow_redirects=False and
    inspected only the FIRST redirect hop's Location header,
    truncated at 80 characters for the report note. Royal London's
    site/CDN appears to issue a single transparent redirect on every
    URL (trailing slash / scheme canonicalisation) — exactly what a
    real browser follows silently without it being a "moved page" in
    any meaningful sense. The 80-char truncation was hiding that the
    redirect target and the source URL were IDENTICAL once you looked
    past the truncation point — confirmed by inspecting
    freshness_report_report_20260812_161133.xlsx directly: every
    "Reason" cell showed "Redirects to <same URL, cut off> — add new
    URL to Excel", with 0 actual content difference.

    FIX — check_single_url() [MODIFIED]:
    - allow_redirects=False → allow_redirects=True, max_redirects=5.
      Follows the full redirect chain to its actual destination,
      same as a real browser would.
    - Compares the FINAL resolved URL (resp.url after following all
      hops) against the original, both passed through
      normalise_url_path() (case/trailing-slash/query insensitive —
      already used elsewhere in this file for the same purpose).
    - If normalised final URL == normalised original URL: genuinely
      "live" — the redirect was transparent, no action needed.
    - Only flags internal_redirect / external_redirect when the
      final URL actually resolves to a DIFFERENT page or domain —
      meaning these flags can now be trusted as real signal, not
      noise from an artifact of the check method itself.

    NO DATA WAS LOST from the false-positive run: report mode makes
    zero index writes regardless of classification, confirmed via
    "Main Chunks Deleted: 0" on every row of the affected report.
    This was a signal-quality bug, not a data-safety incident.

v1.7.5 — August 2026 | Mukesh Kund
    DIAGNOSTIC FIX: scrape_failed log line only recorded the URL, not
    the actual failure reason — every scrape failure was undiagnosable
    from logs alone. Found during a --mode report run where 5-7 URLs
    failed with no visible cause (unlike playwright_option_error a few
    lines later in the same run, which DID include a full error
    message and clearly showed a real "Execution context was destroyed
    — navigation" race condition on a dropdown/filter click).

    scrape_page() [MODIFIED]: log.warning("scrape_failed", url=url) →
    now also logs error=result.error_message, success=result.success,
    has_markdown=bool(result.markdown). Same pattern already used in
    scrape_approved_urls_updatedV5.py's scrape_page() — this file just
    hadn't carried it over into its own independent copy.

    NOT YET FIXED: the underlying navigation-race-condition itself
    (confirmed real via playwright_option_error's error detail) is
    still open — root cause needs investigating with the new error
    detail from this fix before deciding on a code change. Logged
    here as a known follow-up, not resolved by this entry.

v1.7.6 — August 2026 | Mukesh Kund
    ROOT-CAUSE FIX: concurrent per-URL AsyncWebCrawler instances
    racing over the same shared CDP Chrome connection. Diagnosed
    using the error detail unlocked by v1.7.5's scrape_failed fix —
    a --mode report run showed:
      scrape_exception error="BrowserType.connect_over_cdp:
        Connection closed while reading from the driver"
      scrape_failed error="Page.goto: net::ERR_ABORTED ...
        Failed on navigating ACS-GOTO"

    ROOT CAUSE: scrape_url_with_dropdowns() created a BRAND NEW
    AsyncWebCrawler (async with AsyncWebCrawler(...) as crawler)
    for every single URL, internally. With SCRAPE_CONCURRENCY=3,
    three concurrent scrape tasks each independently opened and
    closed their own CDP connection to the SAME shared Chrome
    instance at nearly the same moment. One task's connection
    teardown could kill another concurrent task's in-flight CDP
    session or abort its navigation — a genuine race condition, not
    a site issue, not a WAF, not a timing fluke.

    CONFIRMED via direct comparison: scrape_approved_urls_updatedV5.py's
    scrape_page() has ALWAYS created exactly ONE AsyncWebCrawler
    outside its entire batch loop and passed that single instance
    into every scrape_page() call across all batches — the working,
    proven pattern this file's independent copy had drifted from.

    FIX:
    - scrape_url_with_dropdowns() [MODIFIED]: now takes `crawler` as
      a parameter instead of creating its own. No longer opens/closes
      a CDP connection per URL.
    - scrape_urls_batch() [MODIFIED]: creates ONE AsyncWebCrawler
      before the semaphore-gated concurrent tasks start, wraps the
      whole asyncio.gather() call inside that single `async with`
      block, passes the shared `crawler` into every
      scrape_url_with_dropdowns() call.

    NOT FIXED BY THIS ENTRY: playwright_option_error ("Execution
    context was destroyed... navigation") on dropdown/filter option
    clicks — that's a separate code path
    (_scrape_dropdown_states_playwright(), its own independent
    Playwright browser launch via thread-pool executor, not the
    shared CDP crawler touched here). Still open, tracked separately.

v1.7.7 — August 2026 | Mukesh Kund
    ROOT-CAUSE FIX: --mode report showed 293/297 approved URLs as
    "changed" — implausible given the short window since the index
    was last built. Investigated on user's reasonable skepticism
    rather than accepted at face value.

    TWO COMPOUNDING BUGS FOUND, both in scrape_url_with_dropdowns()'s
    content-hash computation:

    BUG 1 — wrong clean_content() used for hashing:
    scrape_approved_urls_updatedV5.py has ONE clean_content() that
    strips breadcrumbs, social-share sections, footer boilerplate,
    and duplicate article copy — used at SCRAPE time, and the
    content_hash stored in the index is computed DIRECTLY on its
    output (URL-stripping happens separately, later, at CHUNK time,
    inside chunk_and_index_hqaV5.py — AFTER the hash is already
    computed and stored). This file only ever ported THAT LATER,
    chunk-time clean_content() (URL-stripper only) and was using it
    — at scrape time — to compute content_hash. Every freshly-scraped
    page still had breadcrumb/social/footer noise in it when hashed,
    while the indexed hash was computed from properly-cleaned
    content. Comparing "still noisy" against "already clean"
    mismatches on nearly every page regardless of whether the real
    article text changed at all.

    BUG 2 — remove_duplicate_content() had drifted:
    This file's copy used a naive half-content character-position
    comparison (split at the midpoint, compare last/first 200 chars).
    The scraper's actual algorithm finds H1/H2 headings and compares
    WORD-SET overlap between the sections they bound — a completely
    different algorithm, producing different output on the same
    input. Even after fixing Bug 1, hashes would still have mismatched
    on any page where this function's behaviour diverged.

    FIX:
    - remove_duplicate_content() [REPLACED]: now an exact copy of
      the scraper's H1/H2 + word-overlap algorithm.
    - clean_scraped_content() [NEW]: exact port of the scraper's full
      clean_content() (breadcrumbs/social/footer/whitespace, calls
      the now-fixed remove_duplicate_content() internally). Named
      differently from this file's existing clean_content() (URL-
      stripper) since that name was already taken by the chunk-time
      function and both are legitimately needed for their own
      distinct purposes.
    - scrape_url_with_dropdowns() [MODIFIED]: page_content and
      content_hash both now computed via clean_scraped_content(),
      hash taken directly on its output with NO separate URL-
      stripping step in between — matching the scraper's exact
      hash-computation timing.

    IMPACT: also fixes the STORED content quality, not just the
    hash — previously, if apply mode had ever run, freshly re-scraped
    pages would have stored breadcrumb/footer noise INTO the index
    (the existing clean_content() URL-stripper does not remove that
    noise), silently degrading content quality below what the
    original --full build produced. No apply-mode run had happened
    yet with the buggy code, so no index content was actually
    polluted — caught in report mode before any write occurred.

v1.7.8 — August 2026 | Mukesh Kund
    ROOT-CAUSE FIX: 4 index-query functions missing pagination —
    found investigating a 10-URL test report showing "Main Chunks
    Before: 0" for EVERY row, including pages known to have real
    indexed chunks (tell-us-about-a-bereavement, confirmed 13
    dropdown states earlier this session). This was unrelated to the
    v1.7.7 content-hash fix — a completely separate bug class.

    ROOT CAUSE: get_chunk_ids_for_url(), get_all_urls_to_delete(),
    get_refresh_count_for_url(), and the archive-before-delete
    function each called client.search(top=N) ONCE with no skip
    loop — meaning only the FIRST N documents of an unsorted
    match-all query were ever checked, out of ~3,116+ total
    documents in the index post-cleanup:
      get_chunk_ids_for_url        top=1000
      get_all_urls_to_delete       top=5000 (accidentally "safe"
                                    only because index size is
                                    currently under 5000 — would
                                    silently break again after any
                                    rebuild that grows the index)
      get_refresh_count_for_url    top=1    (worst case — checks
                                    ONE arbitrary document)
      archive-before-delete        top=500

    IMPACT PER FUNCTION:
      get_chunk_ids_for_url: silently returned [] for most URLs —
        this is what "Main Chunks Before: 0" was actually showing.
      get_all_urls_to_delete: would silently miss dropdown #state=
        variants for deletion in apply mode once index size exceeds
        5000 documents.
      get_refresh_count_for_url: returned 0 for nearly every URL
        (top=1 against a multi-thousand-document index) — refresh_
        count would never correctly increment across freshness runs,
        every re-indexed chunk incorrectly stamped refresh_count=1
        regardless of true history.
      archive-before-delete: silently produced an EMPTY archive for
        most URLs right before their real chunks were deleted in
        apply mode — archive failure is designed to be non-fatal
        (see v1.5.0), so this failed completely silently, defeating
        the rollback-safety feature's entire purpose with no error
        surfaced anywhere.

    FIX: all 4 functions now use the same skip/page_sz pagination
    loop already used correctly in fetch_current_hashes_from_index()
    since this file's original version — that function was never
    affected by this bug class.

    NOTE: this fix does NOT confirm or deny whether v1.7.7's content-
    hash fix is working correctly — "Main Chunks Before" and the
    "changed" classification are computed by different code paths.
    The "changed" count on a fresh test run still needs independent
    verification against this fix.
"""

from __future__ import annotations

import argparse
import asyncio
import concurrent.futures
import copy
import hashlib
import io
import json
import os
import pickle
import re
import sys
import time
import traceback
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional
from urllib.parse import urlparse
import urllib.parse

import aiohttp
import structlog
from dotenv import find_dotenv, load_dotenv

_dotenv_path = find_dotenv(usecwd=False)
load_dotenv(_dotenv_path, override=True)
log = structlog.get_logger()

from azure.identity import DefaultAzureCredential, get_bearer_token_provider
from azure.search.documents import SearchClient
from openai import AzureOpenAI

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False

try:
    from crawl4ai import AsyncWebCrawler, CacheMode
    from crawl4ai.async_configs import BrowserConfig, CrawlerRunConfig
    from crawl4ai.markdown_generation_strategy import DefaultMarkdownGenerator
    from crawl4ai.content_filter_strategy import PruningContentFilter
    _CRAWL4AI_AVAILABLE = True
except ImportError:
    _CRAWL4AI_AVAILABLE = False

try:
    from langchain_text_splitters import RecursiveCharacterTextSplitter
    _LANGCHAIN_AVAILABLE = True
except ImportError:
    _LANGCHAIN_AVAILABLE = False

try:
    from bs4 import BeautifulSoup
    _BS4_AVAILABLE = True
except ImportError:
    _BS4_AVAILABLE = False


# ══════════════════════════════════════════════════════════════
# VERSIONING (v1.3.0 / v1.5.0)
# ══════════════════════════════════════════════════════════════
#
# COMPLETE VERSIONING FIELD REFERENCE — ALL FIELDS PRODUCED BY THIS SCRIPT:
#
# SCRAPER_VERSION ("1.0.0" -> "1.1.0" -> ...)
#   Tracks WHICH SCRAPING LOGIC in THIS FILE produced each page.
#   Independent of scrape_approved_urls_updatedV5.py — freshness
#   has its own scraping logic and own version constant.
#   Bump when: scrape_url_with_dropdowns(), CDP/Playwright logic,
#   content extraction, or crawl4ai config in THIS file changes.
#   Never auto-increments. Developer-bumped only.
#   Stored as: scraper_version (String) on every delta chunk.
#
# METADATA_VERSION ("1.0.0" -> "1.1.0" -> ...)
#   Tracks WHICH METADATA EXTRACTION LOGIC in THIS FILE produced pages.
#   Bump when: extract_page_metadata(), CONTENT_TYPE_MAP,
#   PRODUCT_CATEGORY_MAP, EXCEL_CATEGORY_MAP, or video detection
#   logic in THIS file changes.
#   Never auto-increments. Developer-bumped only.
#   Stored as: metadata_version (String) on every delta chunk.
#
# PIPELINE_VERSION ("1.0.0" -> "1.1.0" -> ...)
#   Tracks WHICH CHUNKING/HQA/EMBEDDING LOGIC produced each chunk.
#   Should logically match chunk_and_index_hqaV5.py PIPELINE_VERSION.
#   Bump when: CHUNK_SIZE, CHUNK_OVERLAP, HQA prompts/generation/
#   validation logic, embedding model/dimensions, or schema changes.
#   Never auto-increments. Developer-bumped only.
#   Stored as: pipeline_version (String) on every delta chunk.
#
# FRESHNESS_JOB_VERSION ("1.0.0" -> "1.1.0" -> ...)
#   Tracks WHICH VERSION OF THIS FRESHNESS JOB ran a given execution.
#   Renamed from FRESHNESS_VERSION in v1.5.0 to avoid confusion with
#   refresh_count (the chunk-level int field — see below).
#   Bump when: hash comparison strategy, URL classification logic,
#   delta detection, apply orchestration, or report logic changes.
#   Never auto-increments. Developer-bumped only.
#   STORED IN RUN MANIFEST ONLY — not stamped on chunks.
#
# refresh_count (Int32, auto-incremented, NOT a constant)
#   Tracks HOW MANY TIMES this page's content has been REFRESHED
#   by the nightly freshness job.
#     0 = indexed by chunk_and_index_hqaV5.py full run, never delta'd
#     1 = content changed once, re-indexed by freshness
#     N = content changed and re-indexed N times by freshness
#   get_refresh_count_for_url() reads existing value from main index
#   before delete. New chunks stamped with existing_count + 1.
#   Never developer-bumped — always auto-incremented.
#   Stored as: refresh_count (Int32, filterable, sortable) on chunks.
#   Useful query: filter=refresh_count gt 0 -> all freshness-touched chunks.
#
# scrape_run_id (UUID, auto per freshness apply run)
#   Groups ALL PAGES scraped in ONE freshness execution together.
#   Set to freshness_run_id UUID for all delta pages in this run.
#   Stored as: scrape_run_id (String) on every delta chunk.
#
# index_run_id (UUID, auto per freshness apply run)
#   Groups ALL CHUNKS indexed in ONE freshness apply execution.
#   Set to freshness_run_id UUID (same as scrape_run_id in freshness
#   since scrape + index happen in one execution here).
#   Stored as: index_run_id (String) on every delta chunk.
#
# indexed_at (auto ISO timestamp)
#   When the delta chunk was uploaded to Azure AI Search.
#   Stored as: indexed_at (String, sortable) on every chunk.
#
# scraped_at (auto ISO timestamp per page)
#   When the page was fetched from the Royal London website.
#   Set during scrape_url_with_dropdowns().
#   Stored as: scraped_at (String) on every delta chunk.
#
# BUMPING RULES SUMMARY:
#   Developer-bumped (manual — code logic changes only):
#     SCRAPER_VERSION, METADATA_VERSION, PIPELINE_VERSION,
#     FRESHNESS_JOB_VERSION
#   Auto-generated (no developer action needed):
#     scrape_run_id, index_run_id, indexed_at, scraped_at
#   Auto-incremented (freshness reads existing + adds 1):
#     refresh_count
#
SCRAPER_VERSION       = "1.0.0"
METADATA_VERSION      = "1.0.0"
# v1.7.0: bumped 1.0.0 -> 1.1.0 — element-aware chunking changed HOW
# chunk_page() splits standard pages (table atomicity, header
# boundaries). Matches chunk_and_index_hqaV5.py's PIPELINE_VERSION
# bump for the same reason, so full-index chunks and freshness-delta
# chunks report a consistent version for chunks produced by the same
# logic. Bump again if chunking/HQA/embedding/schema logic changes.
PIPELINE_VERSION      = "1.1.0"
FRESHNESS_JOB_VERSION = "1.0.0"


# ══════════════════════════════════════════════════════════════
# CONFIG
# ══════════════════════════════════════════════════════════════

EXPECTED_DOMAIN = "royallondon.com"

# Azure AI Search — main + baseline indexes (auto dual-index mode)
# v1.7.0: default changed to v5 — see chunk_and_index_hqaV5.py for
# the matching full-index build. rlg-faq-index-v4/v4-baseline are
# NOT touched by this script and remain live as fallback.
SEARCH_ENDPOINT     = os.getenv("AZURE_SEARCH_ENDPOINT", "").rstrip("/")
INDEX_NAME          = os.getenv("AZURE_SEARCH_INDEX_NAME", "rlg-faq-index-v5")
BASELINE_INDEX_NAME = os.getenv("AZURE_SEARCH_BASELINE_INDEX_NAME", "rlg-faq-index-v5-baseline")

# Azure OpenAI
AZURE_OPENAI_ENDPOINT = os.getenv("AZURE_OPENAI_ENDPOINT", "").rstrip("/")
EMBEDDING_DEPLOYMENT  = os.getenv("AZURE_OPENAI_EMBEDDING_DEPLOYMENT", "text-embedding-3-large")
EMBEDDING_DIMS        = int(os.getenv("AZURE_OPENAI_EMBEDDING_DIMENSIONS", "1536"))
HQA_DEPLOYMENT        = os.getenv("AZURE_OPENAI_DEPLOYMENT_HQA", "gpt-4o-mini")

# Azure Blob Storage
BLOB_STORAGE_CONNECTION  = os.getenv("AZURE_STORAGE_CONNECTION", "")
BLOB_CONTAINER_NAME      = os.getenv("BLOB_CONTAINER_NAME", "scraper-data")
BLOB_APPROVED_EXCEL_NAME = os.getenv("BLOB_APPROVED_EXCEL_NAME", "approved-urls/Approved_URLs.xlsx")
BLOB_HASH_STATE_NAME     = os.getenv("BLOB_HASH_STATE_NAME", "freshness/content_hashes.json")
BLOB_REPORT_PREFIX       = os.getenv("BLOB_REPORT_PREFIX", "freshness/reports/")
BLOB_MANIFEST_PREFIX     = os.getenv("BLOB_MANIFEST_PREFIX", "freshness/manifests/")
# Archive prefix: deleted chunks saved here before removal for audit/rollback.
# Format: freshness/archive/deleted_<url_hash>_<ts>.json
BLOB_ARCHIVE_PREFIX      = os.getenv("BLOB_ARCHIVE_PREFIX", "freshness/archive/")

# Redis
REDIS_URL              = os.getenv("REDIS_URL", "redis://localhost:6379")
REDIS_CACHE_KEY_PREFIX = "rlg:cache:"

# Chunking — must match chunk_and_index_hqaV5.py exactly.
# Changing these requires --full reindex of chunk_and_index_hqaV5.py.
CHUNK_SIZE    = int(os.getenv("CHUNK_SIZE", "1600"))
CHUNK_OVERLAP = int(os.getenv("CHUNK_OVERLAP", "200"))

# Batching
UPLOAD_BATCH_SIZE    = 100
EMBEDDING_BATCH_SIZE = int(os.getenv("EMBEDDING_BATCH_SIZE", "50"))

# HQA tuning — must match chunk_and_index_hqaV5.py exactly
HQA_QUESTIONS_FIRST_CHUNK  = int(os.getenv("HQA_QUESTIONS_FIRST_CHUNK", "8"))
HQA_QUESTIONS_OTHER_CHUNKS = int(os.getenv("HQA_QUESTIONS_OTHER_CHUNKS", "5"))
TITLE_QUESTIONS_COUNT      = int(os.getenv("TITLE_QUESTIONS_COUNT", "3"))
TITLE_QUESTIONS_MAX_WORDS  = int(os.getenv("TITLE_QUESTIONS_MAX_WORDS", "12"))
MAX_COLLISION_THRESHOLD    = int(os.getenv("MAX_COLLISION_THRESHOLD", "3"))

# ── v1.5.4: Cost tracking ─────────────────────────────────────
# Mirrors CostTracker from chunk_and_index_hqaV5.py v5.8.5.
# Keep in sync when adding new models or pricing changes.
MODEL_PRICING = {
    "gpt-5-mini":  (1.25,  5.00),
    "gpt-5-nano":  (0.50,  2.00),
    "gpt-5.1":     (2.00,  8.00),
    "gpt-5":       (2.00,  8.00),
    "gpt-4o-mini": (0.15,  0.60),
    "gpt-4o":      (2.50, 10.00),
    "gpt-4.1":     (2.00,  8.00),
}

# Look up $/1M token pricing for a model name; falls back to a safe default.
def _get_model_pricing(model_name: str) -> tuple[float, float]:
    name = model_name.lower()
    for key, prices in MODEL_PRICING.items():
        if key in name:
            return prices
    return (2.00, 8.00)


# Accumulates token usage and dollar cost across one HQA run.
class CostTracker:
    # Initialise token/cost counters for one HQA run.
    def __init__(self, model: str, run_mode: str):
        self.model            = model
        self.run_mode         = run_mode
        self.input_tokens     = 0
        self.output_tokens    = 0
        self.reasoning_tokens = 0
        self.calls            = 0
        self.input_price, self.output_price = _get_model_pricing(model)

    # Add one API call's token usage to the running totals.
    def record(self, usage) -> None:
        if usage is None:
            return
        self.input_tokens  += getattr(usage, "prompt_tokens", 0)
        self.output_tokens += getattr(usage, "completion_tokens", 0)
        details = getattr(usage, "completion_tokens_details", None)
        if details:
            self.reasoning_tokens += getattr(details, "reasoning_tokens", 0)
        self.calls += 1

    # Dollar cost of input tokens accumulated so far.
    @property
    def input_cost(self)  -> float:
        return self.input_tokens  / 1_000_000 * self.input_price

    # Dollar cost of output tokens accumulated so far.
    @property
    def output_cost(self) -> float:
        return self.output_tokens / 1_000_000 * self.output_price

    # Combined input + output dollar cost.
    @property
    def total_cost(self)  -> float:
        return self.input_cost + self.output_cost

    # Build a printable cost/token breakdown for this run.
    def summary(self) -> str:
        lines = [
            "",
            "=" * 60,
            "💰 HQA COST SUMMARY",
            "=" * 60,
            f"   Run mode:          {self.run_mode}",
            f"   Model:             {self.model}",
            f"   API calls:         {self.calls:,}",
            f"   Input tokens:      {self.input_tokens:,}   (${self.input_cost:.4f})",
            f"   Output tokens:     {self.output_tokens:,}   (${self.output_cost:.4f})",
        ]
        if self.reasoning_tokens:
            lines.append(
                f"   Reasoning tokens:  {self.reasoning_tokens:,}   "
                f"(included in output cost)"
            )
        lines += [
            f"   ─────────────────────────────────────────────────",
            f"   TOTAL ESTIMATED:   ${self.total_cost:.4f}",
            f"",
            f"   Pricing used:      ${self.input_price}/1M input, "
            f"${self.output_price}/1M output",
            f"   Note: Azure pricing may differ slightly from OpenAI "
            f"direct. Check Azure portal for actuals.",
            "=" * 60,
        ]
        return "\n".join(lines)


_cost_tracker: CostTracker | None = None

# HTTP health check
HTTP_CONCURRENCY     = 15
HTTP_TIMEOUT_SECONDS = 12

# Scraping concurrency — capped at 3 for container memory safety
SCRAPE_CONCURRENCY = 3

# Local output dir
LOCAL_DATA_DIR = Path("scraper/data")

# VDI Chrome CDP
_CDP_PORT    = int(os.getenv("PLAYWRIGHT_CDP_PORT", "9222"))
_CDP_URL     = f"http://localhost:{_CDP_PORT}"
_CHROME_PROC = None

# Playwright executable (VDI only — DO NOT add to Key Vault)
_PLAYWRIGHT_DEFAULT_WIN    = r"C:\Program Files\Google\Chrome\Application\chrome.exe"
PLAYWRIGHT_EXECUTABLE_PATH = os.getenv("PLAYWRIGHT_EXECUTABLE_PATH", _PLAYWRIGHT_DEFAULT_WIN)

# Section mapping — mirrors scraper v3.0.0+
SECTION_MAP = {
    "existing-customers":       "Existing Customers",
    "insurance":                "Insurance",
    "pensions":                 "Pensions",
    "guides-tools":             "Guides and Tools",
    "retirement-planning":      "Retirement Planning",
    "isa":                      "ISA",
    "profitshare":              "ProfitShare",
    "about-us":                 "About Us",
    "find-a-financial-adviser": "Find a Financial Adviser",
    "accessibility":            "Accessibility",
    "informational-pages":      "Information",
}

# Excel Category -> content_type — mirrors scraper v4.3.0+
EXCEL_CATEGORY_MAP = {
    "brand":    "article",
    "guidance": "guide",
    "other":    "article",
    "product":  "article",
    "tool":     "tool",
}

# Excel column header detection
URL_HEADERS      = {"url", "page url", "link", "webpage", "web page", "web url"}
TITLE_HEADERS    = {"title", "page title", "name"}
CATEGORY_HEADERS = {"category", "content category", "page category", "type"}

# Video detection signals — mirrors scraper v3.0.0+
VIDEO_CSS_SIGNALS = [
    "video-player", "webinar-player", "brightcove-player",
    "bc-player", "vjs-tech", "kaltura-player", "jwplayer",
    "data-video-id", "data-webinar-id", "data-brightcove",
]
VIDEO_URL_SIGNALS        = ["/webinars/", "/videos/", "/video/", "/webinar/"]
VIDEO_COLLECTION_SIGNALS = ["webinar", "video", "podcast"]

# Product category mapping — mirrors scraper v3.0.0+
PRODUCT_CATEGORY_MAP = [
    ("/pension",           "pensions"),
    ("/retirement",        "retirement"),
    ("/life-insurance",    "life_insurance"),
    ("/life-cover",        "life_insurance"),
    ("/whole-of-life",     "life_insurance"),
    ("/income-protection", "income_protection"),
    ("/critical-illness",  "critical_illness"),
    ("/illness-income",    "income_protection"),
    ("/isa",               "isa"),
    ("/investments",       "investments"),
    ("/investment",        "investments"),
    ("/fund",              "investments"),
    ("/funeral",           "funeral"),
    ("/profitshare",       "profitshare"),
    ("/financial-adviser", "financial_advice"),
    ("/find-a-financial",  "financial_advice"),
    ("/about-us",          "corporate"),
    ("/media",             "corporate"),
    ("/existing-customers","customer_support"),
]

# Content type mapping — mirrors scraper v3.0.0+
CONTENT_TYPE_MAP = [
    ("/webinars/",           "webinar"),
    ("/videos/",             "video"),
    ("/video/",              "video"),
    ("/guides-tools/",       "guide"),
    ("/pension-calculator",  "tool"),
    ("/retirement-planner",  "tool"),
    ("/lump-sum-calculator", "tool"),
    ("/risk-profiler",       "tool"),
    ("/calculator",          "tool"),
    ("/planner",             "tool"),
    ("/existing-customers/", "faq"),
    ("/help-and-support/",   "faq"),
    ("/pensions-explained",  "faq"),
    ("/about-us/",           "corporate"),
    ("/media/",              "news"),
    ("/press-release",       "news"),
    ("/news/",               "news"),
    ("/agm/",                "corporate"),
]

# Dropdown scraping — mirrors scraper v4.5.5+
_DROPDOWN_PLACEHOLDERS = {
    "select...", "select", "please select", "--", "choose...",
    "choose", "please choose", "-- select an option --",
    "- select -", "select an option", "select option",
    "none", "n/a", "0", "all",
}
_CONTACT_SIGNAL_PATTERNS = [
    re.compile(r'0\d{3,4}[\s\-]?\d{3,4}[\s\-]?\d{3,4}'),
    re.compile(r'1\d{3,4}[\s\-]?\d{3,4}[\s\-]?\d{3,4}'),
]
_CONTACT_SIGNAL_KEYWORDS = [
    "call us", "write to us", "lines are open",
    "excluding bank holidays", "fill in our online form",
    "tell us someone has died", "fill out our online form",
    "monday to friday", "8am to", "9am to",
]

# HQA — dedicated/generic page patterns for deduplication
DEDICATED_PAGE_PATTERNS = [
    "/pensions/", "/insurance/", "/life-insurance/",
    "/income-protection/", "/critical-illness/", "/isa/",
    "/investments/", "/funeral/", "/pension-guides/",
    "/life-insurance-guides/", "/isa-guides/", "/existing-customers/",
]
GENERIC_PAGE_PATTERNS = [
    "/about-us/", "/our-purpose/", "/our-performance/",
    "/agm/", "/life-events/", "/cost-of-living/",
    "/planning-ahead/", "/money-guides/",
]

# HQA blocked questions
BLOCKED_QUESTIONS = {
    "how can i save money on my energy bills",
    "what state benefits can i get if im struggling",
    "how do i save money on my household bills",
    "how can i create a budget to save money",
    "how can i pay off my credit cards faster",
    "what tips do you have for monitoring my spending",
    "what tips do you have for building an emergency fund",
    "how can i find guides on saving and budgeting",
}

# HQA Prompts — mirrors chunk_and_index_hqaV5.py v5.0.0+ exactly
HQA_SYSTEM_PROMPT_TEMPLATE = """You are an expert at generating realistic customer search questions.

Given a chunk of text from Royal London's insurance and pension FAQ pages, generate exactly {num_questions} questions that:
1. A real Royal London customer would type into a search box or chatbot
2. Are ONLY answerable using the provided text — not general knowledge
3. Use natural customer language — not technical or legal jargon
4. Are SPECIFIC to this exact chunk — not generic insurance questions
5. Vary in phrasing to cover different ways a customer might ask the same thing

PRIMARY TOPIC RULE:
Generate questions ONLY about the PRIMARY topic of this chunk.
The primary topic is what the majority of the chunk content discusses.
If a chunk about divorce mentions pensions in one sentence, do NOT generate pension questions.
A question must be fully answerable from this chunk alone — not require other pages.

CROSS-TOPIC RULE:
If this chunk covers multiple topics equally, generate questions for the FIRST topic only.

RULES:
- Include specific product names, numbers, or terms from the chunk
- Do NOT generate: "What is insurance?", "How do pensions work?" or any generic question
- Do NOT generate questions about topics mentioned only in passing
- Do NOT generate general money-saving or budgeting questions
- Keep questions under 15 words each
- Questions must feel like real customer queries

Return ONLY a valid JSON array of exactly {num_questions} strings.
No explanation, no preamble, no markdown formatting.
Example format: ["question 1", "question 2", ...] — must contain exactly {num_questions} items"""

HQA_CORPORATE_PROMPT_TEMPLATE = """You are an expert at generating customer search questions about Royal London as a company.

Given a chunk of text from Royal London's corporate pages (About Us, Our Purpose, Social Impact, AGM etc),
generate exactly {num_questions} questions that a customer or stakeholder would ask about Royal London as an organisation.

ALLOWED question topics:
- Royal London's mission, values, and purpose
- How Royal London is owned and run (mutual ownership, profitshare)
- Royal London's social impact and charity partnerships
- Royal London's financial performance and results
- Royal London's history and heritage
- Who leads Royal London (Board, executives)
- AGM, governance, member rights

STRICTLY FORBIDDEN question topics (do NOT generate these):
- Specific products (pensions, life insurance, ISAs, income protection)
- How to make a claim
- Policy terms, premiums, or benefits
- Product pricing or eligibility
- Customer account questions

RULES:
- Questions must be about Royal London as a company, not its products
- Use natural language a stakeholder or curious customer would use
- Keep questions under 15 words each

Return ONLY a valid JSON array of exactly {num_questions} strings.
No explanation, no preamble, no markdown formatting."""

TITLE_QUESTIONS_PROMPT = """You are an expert at generating broad ENTRY-POINT customer questions for Royal London's insurance and pension pages.

Given the first chunk of a page (its most overview-like content), generate exactly 3 questions that:
1. A customer would ask BEFORE knowing specific product detail
2. Are answerable using the provided text
3. Use natural customer language — not technical or legal jargon

Generate exactly these 3 questions, in this order:
Q1: The most natural "what is X" or "what are X" question for this page's topic
Q2: A "how does X work" or "what types of X" question for this page's topic
Q3: A Royal London specific question (e.g. "does Royal London offer X?")

RULES:
- Do NOT generate questions that are too specific
- Keep each question under 12 words
- Questions must feel like real, natural customer search queries

Return ONLY a valid JSON array of exactly 3 strings.
No explanation, no preamble, no markdown formatting.
Example format: ["question 1", "question 2", "question 3"]"""


# ══════════════════════════════════════════════════════════════
# SINGLETON CLIENTS
# ══════════════════════════════════════════════════════════════

_credential:    Optional[DefaultAzureCredential] = None
_openai_client: Optional[AzureOpenAI]            = None
_redis_client                                    = None


# Return a cached DefaultAzureCredential, creating it on first use.
def get_credential() -> DefaultAzureCredential:
    global _credential
    if _credential is None:
        _credential = DefaultAzureCredential()
    return _credential


# Return a cached Azure OpenAI client (managed-identity auth), creating it on first use.
def get_openai_client() -> AzureOpenAI:
    global _openai_client
    if _openai_client is None:
        if not AZURE_OPENAI_ENDPOINT:
            raise ValueError("AZURE_OPENAI_ENDPOINT is not set in .env")
        token_provider = get_bearer_token_provider(
            get_credential(),
            "https://cognitiveservices.azure.com/.default",
        )
        _openai_client = AzureOpenAI(
            azure_endpoint=AZURE_OPENAI_ENDPOINT,
            azure_ad_token_provider=token_provider,
            api_version="2024-12-01-preview",
        )
    return _openai_client


# Get SearchClient for the specified index. Supports dual-index writes.
def get_search_client(index_name: str = INDEX_NAME) -> SearchClient:
    """Get SearchClient for the specified index. Supports dual-index writes."""
    return SearchClient(
        endpoint=SEARCH_ENDPOINT,
        index_name=index_name,
        credential=get_credential(),
    )


# Return a cached Redis client, connecting (with ping test) on first use; None if unavailable.
def get_redis():
    global _redis_client
    if _redis_client is not None:
        return _redis_client
    try:
        import redis
        pool = redis.ConnectionPool.from_url(
            REDIS_URL,
            decode_responses=False,
            socket_connect_timeout=3,
            socket_timeout=3,
            max_connections=5,
        )
        client = redis.Redis(connection_pool=pool)
        client.ping()
        _redis_client = client
        log.info("redis_connected", url=REDIS_URL)
        return _redis_client
    except Exception as e:
        log.warning("redis_unavailable", error=str(e))
        return None


# ══════════════════════════════════════════════════════════════
# CDP / CHROME SUBPROCESS (VDI — mirrors scraper v4.5.2+)
# ══════════════════════════════════════════════════════════════

# Launch system Chrome with remote debugging for CDP mode.
def _cf_start_chrome_cdp() -> bool:
    """
    Launch system Chrome with remote debugging for CDP mode.
    VDI only — returns False on Linux/production (no-op).
    v1.2.3: CREATE_NO_WINDOW prevents I/O on closed pipe on Windows.
    """
    global _CHROME_PROC
    import socket
    import subprocess
    import sys as _sys
    import time as _t

    exec_path = PLAYWRIGHT_EXECUTABLE_PATH
    if not exec_path or not os.path.exists(exec_path):
        return False

    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        s.settimeout(1)
        if s.connect_ex(("localhost", _CDP_PORT)) == 0:
            log.info("cf_cdp_chrome_already_running", port=_CDP_PORT)
            return True

    try:
        creation_flags = 0
        if _sys.platform == "win32":
            creation_flags = subprocess.CREATE_NO_WINDOW

        _CHROME_PROC = subprocess.Popen(
            [exec_path,
             f"--remote-debugging-port={_CDP_PORT}",
             "--headless=new", "--no-sandbox",
             "--disable-dev-shm-usage", "--disable-gpu",
             "--no-first-run", "--no-default-browser-check"],
            stdin=subprocess.DEVNULL,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            creationflags=creation_flags,
        )
        for _ in range(15):
            _t.sleep(0.5)
            with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
                s.settimeout(0.5)
                if s.connect_ex(("localhost", _CDP_PORT)) == 0:
                    log.info("cf_cdp_chrome_started", pid=_CHROME_PROC.pid)
                    return True
        log.error("cf_cdp_chrome_start_timeout", port=_CDP_PORT)
        return False
    except Exception as e:
        log.error("cf_cdp_chrome_start_failed", error=str(e))
        return False


# Terminate Chrome CDP subprocess if we started it.
def _cf_stop_chrome_cdp() -> None:
    """Terminate Chrome CDP subprocess if we started it."""
    global _CHROME_PROC
    if _CHROME_PROC is not None:
        pid = _CHROME_PROC.pid
        try:
            _CHROME_PROC.terminate()
        except Exception:
            # Process may already be gone — cleanup path, safe to ignore.
            pass
        try:
            _CHROME_PROC.wait(timeout=5)
        except Exception:
            # Already exited or terminate() didn't finish in time —
            # nothing left to clean up either way.
            pass
        for pipe in [_CHROME_PROC.stdin, _CHROME_PROC.stdout, _CHROME_PROC.stderr]:
            if pipe is not None:
                try:
                    pipe.close()
                except Exception:
                    # Pipe may already be closed by process exit — safe to ignore.
                    pass
        _CHROME_PROC = None
        log.info("cf_cdp_chrome_stopped", pid=pid)


# CDP mode on VDI, normal Playwright on production.
def _cf_make_browser_config() -> "BrowserConfig":
    """CDP mode on VDI, normal Playwright on production."""
    use_cdp = _cf_start_chrome_cdp()
    if use_cdp:
        return BrowserConfig(headless=True, verbose=False, cdp_url=_CDP_URL)
    return BrowserConfig(headless=True, verbose=False)


# ══════════════════════════════════════════════════════════════
# BLOB STORAGE HELPERS
# ══════════════════════════════════════════════════════════════

# Build a BlobServiceClient from the configured connection string.
def get_blob_service():
    if not BLOB_STORAGE_CONNECTION:
        raise EnvironmentError(
            "AZURE_STORAGE_CONNECTION is not set. "
            "Set it in .env or use --file for local mode."
        )
    from azure.storage.blob import BlobServiceClient
    return BlobServiceClient.from_connection_string(BLOB_STORAGE_CONNECTION)


# Download the approved-URLs Excel file from Blob Storage as raw bytes.
def load_approved_excel_from_blob(blob_name: str | None = None) -> bytes:
    name   = blob_name or BLOB_APPROVED_EXCEL_NAME
    svc    = get_blob_service()
    client = svc.get_blob_client(container=BLOB_CONTAINER_NAME, blob=name)
    data   = client.download_blob().readall()
    log.info("excel_downloaded_from_blob", blob=name, size_bytes=len(data))
    return data


# Load previous run's hash state. Returns {} on first run.
def load_hash_state() -> dict[str, str]:
    """Load previous run's hash state. Returns {} on first run."""
    if BLOB_STORAGE_CONNECTION:
        try:
            svc    = get_blob_service()
            client = svc.get_blob_client(
                container=BLOB_CONTAINER_NAME, blob=BLOB_HASH_STATE_NAME,
            )
            data  = client.download_blob().readall()
            state = json.loads(data)
            log.info("hash_state_loaded", count=len(state), source="blob")
            return state
        except Exception as e:
            log.info("hash_state_not_found", error=str(e), reason="first_run_or_missing")
            return {}
    return {}


# Persist the URL->content_hash state to Blob (or local disk) for next run.
def save_hash_state(state: dict[str, str]) -> None:
    data = json.dumps(state, indent=2, ensure_ascii=False).encode("utf-8")
    if BLOB_STORAGE_CONNECTION:
        svc    = get_blob_service()
        client = svc.get_blob_client(
            container=BLOB_CONTAINER_NAME, blob=BLOB_HASH_STATE_NAME,
        )
        client.upload_blob(data, overwrite=True)
        log.info("hash_state_saved", count=len(state), dest="blob")
    else:
        dest = LOCAL_DATA_DIR / "content_hashes.json"
        dest.parent.mkdir(parents=True, exist_ok=True)
        dest.write_bytes(data)
        log.info("hash_state_saved", count=len(state), dest=str(dest))


# Write run manifest JSON after every freshness run.
def save_run_manifest(manifest: dict, ts_str: str) -> str | None:
    """
    Write run manifest JSON after every freshness run.
    Records versions, run ID, and stats for full audit trail.
    Non-fatal — never blocks the freshness run.
    """
    filename = f"run_manifest_{ts_str}.json"
    data     = json.dumps(manifest, indent=2, ensure_ascii=False).encode("utf-8")
    if BLOB_STORAGE_CONNECTION:
        try:
            svc    = get_blob_service()
            blob   = f"{BLOB_MANIFEST_PREFIX}{filename}"
            client = svc.get_blob_client(container=BLOB_CONTAINER_NAME, blob=blob)
            client.upload_blob(data, overwrite=True)
            log.info("run_manifest_saved", dest="blob", blob=blob)
            return blob
        except Exception as e:
            log.warning("run_manifest_blob_failed", error=str(e))
            return None
    else:
        try:
            dest = LOCAL_DATA_DIR / filename
            dest.parent.mkdir(parents=True, exist_ok=True)
            dest.write_bytes(data)
            log.info("run_manifest_saved", dest="local", path=str(dest))
            return str(dest)
        except Exception as e:
            log.warning("run_manifest_local_failed", error=str(e))
            return None


# Upload Excel report to Blob for archival. Non-fatal.
def upload_report_to_blob(report_path: Path) -> str | None:
    """Upload Excel report to Blob for archival. Non-fatal."""
    if not BLOB_STORAGE_CONNECTION:
        return None
    try:
        svc    = get_blob_service()
        blob   = f"{BLOB_REPORT_PREFIX}{report_path.name}"
        client = svc.get_blob_client(container=BLOB_CONTAINER_NAME, blob=blob)
        with open(report_path, "rb") as f:
            client.upload_blob(f, overwrite=True)
        log.info("report_uploaded", blob=blob)
        return blob
    except Exception as e:
        log.warning("report_upload_failed", error=str(e))
        return None


# ══════════════════════════════════════════════════════════════
# EXCEL LOADING
# ══════════════════════════════════════════════════════════════

# Scan header row to find which columns hold URL/title/category.
def _detect_columns(ws) -> tuple[int | None, int | None, int | None]:
    url_col = title_col = cat_col = None
    for cell in ws[1]:
        if not cell.value:
            continue
        h = str(cell.value).strip().lower()
        if h in URL_HEADERS and url_col is None:
            url_col = cell.column
        elif h in TITLE_HEADERS and title_col is None:
            title_col = cell.column
        elif h in CATEGORY_HEADERS and cat_col is None:
            cat_col = cell.column
    return url_col, title_col, cat_col


# Parse approved-URLs entries from an in-memory Excel file (Blob download).
def load_urls_from_excel_bytes(data: bytes) -> list[dict]:
    if not _OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl not installed.")
    wb = load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    url_col, title_col, cat_col = _detect_columns(ws)
    if url_col is None:
        raise ValueError(f"No URL column found. Headers: {[c.value for c in ws[1]]}")
    entries = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        url = str(row[url_col - 1] or "").strip()
        if not url or not url.startswith("http"):
            continue
        title    = str(row[title_col - 1] or "").strip() if title_col else ""
        category = str(row[cat_col - 1] or "").strip().lower() if cat_col else ""
        entries.append({"url": url, "title": title, "category": category})
    wb.close()
    return entries


# Parse approved-URLs entries from a local Excel file path.
def load_urls_from_excel_file(file_path: str) -> list[dict]:
    if not _OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl not installed.")
    wb = load_workbook(filename=file_path, read_only=True, data_only=True)
    ws = wb.active
    url_col, title_col, cat_col = _detect_columns(ws)
    if url_col is None:
        raise ValueError(f"No URL column found in {file_path}")
    entries = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        url = str(row[url_col - 1] or "").strip()
        if not url or not url.startswith("http"):
            continue
        title    = str(row[title_col - 1] or "").strip() if title_col else ""
        category = str(row[cat_col - 1] or "").strip().lower() if cat_col else ""
        entries.append({"url": url, "title": title, "category": category})
    wb.close()
    return entries


# ══════════════════════════════════════════════════════════════
# URL HELPERS
# ══════════════════════════════════════════════════════════════

# Lowercase, strip trailing slash — canonical form for URL comparison.
def normalise_url(url: str) -> str:
    try:
        parsed = urlparse(url.strip().rstrip("/").lower())
        return parsed.geturl()
    except Exception:
        return url.strip().rstrip("/").lower()


# True if this URL is a dropdown-state fragment (#state= or #policy=).
def is_dropdown_url(url: str) -> bool:
    return "#state=" in url or "#policy=" in url


# Strip any #fragment, returning the clean parent page URL.
def get_base_url(url: str) -> str:
    return url.split("#")[0]


# ══════════════════════════════════════════════════════════════
# CONTENT UTILITIES — mirrors scraper + indexer exactly
# ══════════════════════════════════════════════════════════════

# Guess the site section (e.g. "Pensions") from the URL's first path segment.
def derive_section(url: str) -> str:
    try:
        path          = url.split("://", 1)[-1]
        path          = path.split("/", 1)[-1]
        first_segment = path.split("/")[0].lower()
        return SECTION_MAP.get(first_segment, "General")
    except Exception:
        return "General"


# Guess content_type (guide/article/faq/tool) from URL pattern matching.
def derive_content_type(url: str) -> str:
    url_lower = url.lower()
    for pattern, ct in CONTENT_TYPE_MAP:
        if pattern in url_lower:
            return ct
    return "article"


# Guess product_category (pensions/isa/etc.) from URL pattern matching.
def derive_product_category(url: str) -> str:
    url_lower = url.lower()
    for pattern, cat in PRODUCT_CATEGORY_MAP:
        if pattern in url_lower:
            return cat
    return "general"


# Classify audience (adviser/employer/customer) from subdomain or path.
def derive_audience_from_url(url: str) -> str:
    url_lower = url.lower()
    if "adviser.royallondon.com" in url_lower or "/adviser/" in url_lower:
        return "adviser"
    if "employer.royallondon.com" in url_lower or "/employer/" in url_lower:
        return "employer"
    return "customer"


# Excel Category primary; URL-pattern fallback. Mirrors scraper v4.3.0+.
def map_excel_category_to_content_type(excel_category: str, url: str) -> str:
    """Excel Category primary; URL-pattern fallback. Mirrors scraper v4.3.0+."""
    if excel_category:
        mapped = EXCEL_CATEGORY_MAP.get(excel_category.strip().lower())
        if mapped:
            url_type = derive_content_type(url)
            if url_type in ("webinar", "video", "tool", "faq", "news"):
                return url_type
            return mapped
    return derive_content_type(url)


# Detect embedded video via URL pattern or meta tags/iframes in the HTML.
def detect_video_from_html(html: str, url: str) -> bool:
    url_lower = url.lower()
    for pattern in VIDEO_URL_SIGNALS:
        if pattern in url_lower:
            return True
    if not html or not _BS4_AVAILABLE:
        return False
    try:
        soup = BeautifulSoup(html, "html.parser")
        collection_meta = (
            soup.find("meta", attrs={"name": "Collection_name"}) or
            soup.find("meta", attrs={"name": "collection_name"})
        )
        if collection_meta:
            collection_val = (collection_meta.get("content", "") or "").lower()
            for signal in VIDEO_COLLECTION_SIGNALS:
                if signal in collection_val:
                    return True
        og_type = soup.find("meta", property="og:type")
        if og_type and "video" in (og_type.get("content", "") or "").lower():
            return True
        html_lower = html.lower()
        for signal in VIDEO_CSS_SIGNALS:
            if signal in html_lower:
                return True
    except Exception as e:
        log.warning("video_detection_error", error=str(e))
    return False


# Extract rich metadata from page HTML.
def extract_page_metadata(html: str, url: str) -> dict:
    """
    Extract rich metadata from page HTML.
    Mirrors extract_page_metadata() from scraper v3.0.0+ exactly.
    """
    metadata = {
        "has_video":        detect_video_from_html(html, url),
        "content_type":     derive_content_type(url),
        "product_category": derive_product_category(url),
        "audience":         derive_audience_from_url(url),
        "description":      "",
        "thumbnail_url":    "",
        "publish_date":     "",
        "collection_name":  "",
        "read_time_mins":   "5",
    }
    if not html or not _BS4_AVAILABLE:
        return metadata
    try:
        soup = BeautifulSoup(html, "html.parser")
        for attr, key in [
            ({"name": "description"}, "content"),
            ({"property": "og:description"}, "content"),
            ({"name": "st-description"}, "content"),
        ]:
            tag = soup.find("meta", attrs=attr)
            if tag and tag.get(key, "").strip():
                metadata["description"] = tag[key].strip()[:300]
                break
        teaser = soup.find("meta", attrs={"name": "teaser_image"})
        if teaser and teaser.get("content", "").strip():
            metadata["thumbnail_url"] = teaser["content"].strip()
        else:
            og_image = soup.find("meta", property="og:image")
            if og_image and og_image.get("content", "").strip():
                img_url = og_image["content"].strip()
                if "rl-logo-meta-image" not in img_url:
                    metadata["thumbnail_url"] = img_url
        pub_date_tag = soup.find("meta", attrs={"name": "st-publish-date"})
        if pub_date_tag and pub_date_tag.get("content", "").strip():
            raw_date = pub_date_tag["content"].strip()
            try:
                parsed = datetime.strptime(raw_date, "%d %B %Y")
                metadata["publish_date"] = parsed.strftime("%Y-%m-%d")
            except ValueError:
                metadata["publish_date"] = raw_date[:20]
        collection_tag = soup.find("meta", attrs={"name": "Collection_name"})
        if collection_tag and collection_tag.get("content", "").strip():
            metadata["collection_name"] = collection_tag["content"].strip()[:100]
        body_text = soup.get_text(separator=" ", strip=True)
        metadata["read_time_mins"] = str(max(1, round(len(body_text.split()) / 200)))
    except Exception as e:
        log.warning("metadata_extraction_error", url=url, error=str(e))
    return metadata


# Remove exact duplicate article content (crawl4ai sometimes scrapes a page twice).
def remove_duplicate_content(content: str) -> str:
    """
    Remove exact duplicate of article content.
    crawl4ai sometimes scrapes the page twice.

    v1.7.7 FIX: this function had DRIFTED from
    scrape_approved_urls_updatedV5.py's version — the previous
    implementation here split content into two raw character-count
    HALVES and compared the last/first 200 characters positionally.
    The scraper's actual algorithm finds H1/H2 HEADINGS and compares
    WORD-SET overlap between the two sections they bound. These are
    genuinely different algorithms producing genuinely different
    output on the same input — one contributing factor (alongside
    the missing full clean_content(), see clean_scraped_content()
    below) to a --mode report run showing 293/297 URLs as "changed"
    when it's implausible that much real site content shifted in
    such a short window. Replaced with an exact copy of the
    scraper's H1/H2 + word-overlap algorithm.

    Finds H1/H2 headings — if the second one starts after
    40% of content AND has >60% word overlap with the first
    chunk, it's a true duplicate → cut it.
    """
    h1_pattern = re.compile(r'^#{1,2}\s+\S', re.MULTILINE)
    matches = list(h1_pattern.finditer(content))

    if len(matches) < 2:
        return content

    second_pos  = matches[1].start()
    content_len = len(content)

    if second_pos < content_len * 0.40:
        return content

    first_chunk  = content[:second_pos].strip()
    second_chunk = content[second_pos:].strip()

    first_words  = set(first_chunk.split()[:200])
    second_words = set(second_chunk.split()[:200])

    if not first_words:
        return content

    overlap_pct = len(first_words & second_words) / len(first_words)

    if overlap_pct > 0.60:
        log.info(
            "duplicate_removed",
            overlap_pct=round(overlap_pct, 2),
            chars_removed=len(second_chunk),
        )
        return first_chunk

    return content


# v1.7.7 — full scrape-time content cleaner, mirrors scraper exactly.
def clean_scraped_content(content: str) -> str:
    """
    Safe content cleaning — removes pure UI noise (breadcrumbs,
    social-share sections, footer boilerplate, duplicate article
    copy). Exact port of scrape_approved_urls_updatedV5.py's
    clean_content() — same function name was already taken in THIS
    file by a different function (URL-stripper, used at chunk time
    inside chunk_page() — see that function's docstring: "Mirrors
    clean_content() from chunk_and_index_hqaV5.py exactly"), hence
    the distinct name here.

    v1.7.7 FIX — WHY THIS WAS MISSING:
    scrape_url_with_dropdowns() was previously computing both the
    STORED page content AND the content_hash using only
    remove_duplicate_content() + the WRONG clean_content() (the
    chunk-time URL-stripper). Neither breadcrumb navigation, social-
    share sections, nor footer boilerplate were ever stripped from
    freshly-scraped content before hashing or storing — while the
    indexed content_hash (and indexed content itself) WAS built from
    the scraper's fully-cleaned output. Comparing "still has
    breadcrumbs/footer noise" against "already clean" produces a
    hash mismatch on nearly every page, regardless of whether the
    actual article text changed — the direct cause of a --mode
    report run showing 293/297 URLs as "changed".

    NOTE ON HASH TIMING: the scraper computes content_hash directly
    on THIS function's output — URL-stripping (the other
    clean_content() in this file) is applied later, at CHUNK time,
    AFTER the hash is already computed and stored. This function
    must be used the same way here: page_content = clean_scraped_
    content(raw), content_hash = compute_content_hash(page_content) —
    with NO separate URL-stripping step in between.
    """
    # Step 1: Remove duplicate article copy
    content = remove_duplicate_content(content)

    # Step 2: Remove breadcrumb navigation
    content = re.sub(
        r'^\s*\d+\.\s*\[.*?\]\(.*?\)\s*>\s*$',
        '', content, flags=re.MULTILINE,
    )
    content = re.sub(
        r'^\s*\d+\.\s*\[.*?\]\(.*?\)\s*$',
        '', content, flags=re.MULTILINE,
    )
    content = re.sub(
        r'^\s*\d+\.\s+[A-Z][^\n]{3,60}$',
        '', content, flags=re.MULTILINE,
    )

    # Step 3: Remove social share sections
    content = re.sub(
        r'Share\s*\n(\s*\*\s*(\[?\s*\]?\([^\)]*\))?\s*\n)+',
        '', content,
    )
    content = re.sub(
        r'^\s*\*\s*\[?\s*\]?\(\s*[^\)]{0,10}\)\s*$',
        '', content, flags=re.MULTILINE,
    )
    content = re.sub(
        r'^Share\s*$',
        '', content, flags=re.MULTILINE,
    )

    # Step 4: Remove Twitter share links
    content = re.sub(
        r'\[?\s*\]?\(https://twitter\.com/intent/tweet[^\)]*\)\s*',
        '', content,
    )

    # Step 5: Remove other social media URLs
    content = re.sub(
        r'https://www\.(facebook|instagram|linkedin|x|youtube|twitter)\.com/\S+',
        '', content,
    )

    # Step 6: Remove empty markdown links
    content = re.sub(r'\[\s*\]\(\s*\)', '', content)
    content = re.sub(
        r'^\s*\*\s*\[\s*\]\s*$', '', content, flags=re.MULTILINE,
    )

    # Step 7: Remove Previous/Next Item labels
    content = re.sub(
        r'^(Previous Item|Next Item)\s*$',
        '', content, flags=re.MULTILINE | re.IGNORECASE,
    )

    # Step 8: Remove footer boilerplate
    content = re.sub(
        r'Your browser is not supported\..*?×\s*',
        '', content, flags=re.DOTALL,
    )
    content = re.sub(
        r'#{1,3}\s*Connect with us.*$',
        '', content, flags=re.DOTALL | re.MULTILINE,
    )
    content = re.sub(
        r'#{1,3}\s*Products and services.*$',
        '', content, flags=re.DOTALL | re.MULTILINE,
    )
    content = re.sub(
        r'#{1,3}\s*About Royal London.*$',
        '', content, flags=re.DOTALL | re.MULTILINE,
    )
    content = re.sub(
        r'#{1,3}\s*Useful links.*$',
        '', content, flags=re.DOTALL | re.MULTILINE,
    )
    content = re.sub(
        r'\*\*The Royal London Mutual Insurance.*$',
        '', content, flags=re.DOTALL | re.MULTILINE,
    )
    content = re.sub(
        r'©\s*Royal London \d{4}.*$',
        '', content, flags=re.DOTALL | re.MULTILINE,
    )
    content = re.sub(r'\[Back to top\].*?\n', '', content)

    # Step 9: Normalize whitespace
    content = re.sub(r'\n{3,}', '\n\n', content)
    content = re.sub(r'[ \t]+\n', '\n', content)
    content = re.sub(r'\n[ \t]+\n', '\n\n', content)

    return content.strip()


# Remove external URLs from page content before chunking.
def clean_content(text: str) -> str:
    """
    Remove external URLs from page content before chunking.
    Mirrors clean_content() from chunk_and_index_hqaV5.py exactly.
    royallondon.com URLs preserved for citation system.
    """
    # Strip non-RLG markdown links, keeping only the anchor text.
    def replace_markdown_link(match):
        anchor_text = match.group(1)
        url         = match.group(2)
        if "royallondon.com" in url:
            return match.group(0)
        return anchor_text

    text = re.sub(r'\[([^\]]+)\]\((https?://[^\)]+)\)', replace_markdown_link, text)

    # Strip bare non-RLG URLs entirely.
    def replace_raw_url(match):
        url = match.group(0)
        if "royallondon.com" in url:
            return url
        return ""

    text = re.sub(r'https?://[^\s\)\]"\'<>,]+', replace_raw_url, text)
    text = re.sub(r'  +', ' ', text)
    return text.strip()


# SHA-256 of content. Must match indexer and scraper exactly.
def compute_content_hash(content: str) -> str:
    """SHA-256 of content. Must match indexer and scraper exactly."""
    return hashlib.sha256(content.encode("utf-8")).hexdigest()


# v1.7.2 — deterministic chunk_id, replaces uuid.uuid4().
def compute_chunk_id(source_url: str, chunk_index: int, content: str) -> str:
    """
    Deterministic chunk_id — SHA-256 of (source_url, chunk_index,
    content). Replaces uuid.uuid4() (random, non-deterministic).

    Byte-identical hash construction to chunk_and_index_hqaV5.py's
    compute_chunk_id() — see that file's v5.10.2 changelog entry for
    the full root-cause writeup (5,194 duplicate chunks found in
    rlg-faq-index-v5 after its first --full build, caused by random
    UUIDs not surviving accidental re-processing without creating
    duplicates). Duplicated here rather than imported — this file's
    stated design principle is zero cross-file dependency (see
    module header) — but kept byte-identical so a chunk built by
    either script for the same (url, index, content) always resolves
    to the same ID, whether it was written by a --full run or a
    nightly freshness delta.

    chunk_index is part of the hash (not just url+content) so two
    genuinely different chunks containing identical text (e.g. a
    repeated boilerplate paragraph) don't collide into one ID.
    """
    key_material = f"{source_url}|{chunk_index}|{content}"
    return hashlib.sha256(key_material.encode("utf-8")).hexdigest()


# ══════════════════════════════════════════════════════════════
# ELEMENT-AWARE CHUNKING — v1.7.0
# ══════════════════════════════════════════════════════════════
# INLINED, not imported. This file's stated design principle (see
# module header, v1.4.0) is to be a SINGLE INDEPENDENT SCRIPT with
# no dependency on other project code files — chosen deliberately so
# the nightly freshness job can never break because an unrelated
# script changed. That principle is preserved here: this logic is a
# duplicate of element_detector.py (used by chunk_and_index_hqaV5.py),
# not an import of it.
#
# CONSEQUENCE — accepted tradeoff, same as the existing HQA/embedding
# duplication in this file: if element_detector.py's logic is fixed
# or improved, this block must be updated MANUALLY to match. There is
# no automated drift guard. This mirrors the exact situation already
# accepted for clean_content()/compute_content_hash()/chunk_page() in
# this file since v1.4.0 — one more duplicated block, same known
# tradeoff, not a new category of risk.
#
# DESIGN — confirmed via Phase 0/1 audits on the 297-URL approved
# list before this was built (see chunk_and_index_hqaV5.py v5.10.0
# changelog for the full audit trail):
#   - Tables: atomic per table, row-capped at TABLE_ROWS_PER_CHUNK
#     (safety net — current max observed table is 96 rows across 16
#     tables on one page, so this rarely/never fires in practice).
#   - ##/### headers: hard section boundaries. #### and deeper stay
#     as in-section formatting, not a boundary.
#   - Pages with no tables/headers: byte-identical output to the old
#     flat splitter (fast-path below) — zero behavioural change for
#     the vast majority of pages.

TABLE_ROWS_PER_CHUNK = 30

_BOUNDARY_HEADER_RE = re.compile(r'^(#{2,3})\s+(.+)$')


# Parse a contiguous block of |pipe| lines into header/rows.
def _ea_parse_table_block(lines: list) -> dict:
    """Parse a contiguous block of |pipe| lines into header/rows."""
    header_row    = None
    separator_row = None
    data_rows     = []
    for line in lines:
        stripped = line.strip()
        if not stripped.startswith("|"):
            continue
        cells = [c.strip() for c in stripped.strip("|").split("|")]
        if all(re.match(r'^[-:]+$', c) for c in cells if c):
            separator_row = cells
        elif header_row is None and separator_row is None:
            header_row = cells
        else:
            data_rows.append(cells)
    return {"header_row": header_row, "data_rows": data_rows}


# Parse raw markdown into ordered typed elements: header/table/
def _ea_parse_elements(content: str) -> list:
    """
    Parse raw markdown into ordered typed elements: header/table/
    prose/blank. Mirrors element_detector.py's parse_elements()
    exactly — see that module for full docstring detail.
    """
    lines    = content.splitlines()
    elements = []
    i = 0
    while i < len(lines):
        line     = lines[i]
        stripped = line.strip()

        header_match = re.match(r'^(#{1,6})\s+(.+)$', stripped)
        if header_match:
            level = len(header_match.group(1))
            elements.append({
                "type": "header", "level": level,
                "text": header_match.group(2).strip(), "lines": [line],
            })
            i += 1
            continue

        if stripped.startswith("|") and "|" in stripped[1:]:
            block = [line]
            i += 1
            while i < len(lines) and lines[i].strip().startswith("|"):
                block.append(lines[i])
                i += 1
            parsed = _ea_parse_table_block(block)
            elements.append({
                "type": "table", "lines": block,
                "header_row": parsed["header_row"],
                "data_rows":  parsed["data_rows"],
            })
            continue

        if not stripped:
            elements.append({"type": "blank", "lines": [line]})
            i += 1
            continue

        block = [line]
        i += 1
        while i < len(lines):
            ns = lines[i].strip()
            if ns.startswith("|") or re.match(r'^#{1,6}\s', ns) or not ns:
                break
            block.append(lines[i])
            i += 1
        elements.append({"type": "prose", "lines": block})

    return elements


# Group elements into header-bounded sections (## / ### only).
def _ea_group_into_sections(elements: list) -> list:
    """Group elements into header-bounded sections (## / ### only)."""
    sections = []
    current  = {"header_text": None, "header_level": None, "body": []}

    for el in elements:
        if el["type"] == "header" and el["level"] in (2, 3):
            if current["body"] or current["header_text"] is not None:
                sections.append(current)
            current = {
                "header_text": el["text"], "header_level": el["level"],
                "body": [],
            }
        else:
            current["body"].append(el)

    if current["body"] or current["header_text"] is not None:
        sections.append(current)

    return sections


# Render a (possibly batched) set of table rows to markdown.
def _ea_render_table_chunk(header_row: list, data_rows: list) -> str:
    """Render a (possibly batched) set of table rows to markdown."""
    if not header_row:
        return "\n".join("| " + " | ".join(r) + " |" for r in data_rows)
    sep = ["-" * max(3, len(h)) for h in header_row]
    lines = [
        "| " + " | ".join(header_row) + " |",
        "| " + " | ".join(sep) + " |",
    ]
    for row in data_rows:
        lines.append("| " + " | ".join(row) + " |")
    return "\n".join(lines)


# Convert a table element into atomic (row-capped) chunk texts.
def _ea_chunk_table_element(table_el: dict) -> list:
    """Convert a table element into atomic (row-capped) chunk texts."""
    header_row = table_el.get("header_row")
    data_rows  = table_el.get("data_rows") or []

    if not data_rows:
        return ["\n".join(table_el["lines"])]

    if len(data_rows) <= TABLE_ROWS_PER_CHUNK:
        return [_ea_render_table_chunk(header_row, data_rows)]

    batches = []
    for start in range(0, len(data_rows), TABLE_ROWS_PER_CHUNK):
        batch = data_rows[start:start + TABLE_ROWS_PER_CHUNK]
        batches.append(_ea_render_table_chunk(header_row, batch))
    return batches


# Convert a section's body elements into prose/table segments.
def _ea_section_to_text_segments(section: dict) -> list:
    """Convert a section's body elements into prose/table segments."""
    segments  = []
    prose_buf = []

    # Join buffered prose lines into one segment and reset the buffer.
    def _flush_prose():
        if prose_buf:
            text = "\n".join(prose_buf).strip()
            if text:
                segments.append({"text": text, "element_type": "prose"})
            prose_buf.clear()

    for el in section["body"]:
        if el["type"] == "table":
            _flush_prose()
            for chunk_text in _ea_chunk_table_element(el):
                segments.append({"text": chunk_text, "element_type": "table"})
        elif el["type"] == "blank":
            prose_buf.append("")
        else:
            prose_buf.extend(el["lines"])

    _flush_prose()
    return segments


# Guard against the splitter isolating a prepended '### <header>'
def _ea_merge_orphaned_header(pieces: list, header_text: str) -> list:
    """
    Guard against the splitter isolating a prepended '### <header>'
    line into its own useless first chunk when the section body is
    large and separator-sparse. See element_detector.py v1.1.0
    changelog for the full bug description this fixes.
    """
    if len(pieces) < 2:
        return pieces
    header_line = f"### {header_text}".strip()
    if pieces[0].strip() == header_line:
        return [pieces[0] + "\n" + pieces[1]] + pieces[2:]
    return pieces


# Main entry point — inlined duplicate of
def chunk_content_element_aware(
    content: str,
    chunk_size: int    = 1600,
    chunk_overlap: int = 200,
) -> list:
    """
    Main entry point — inlined duplicate of
    element_detector.chunk_content_element_aware(). See that module
    for full design rationale. Returns ordered list of
    {"text": str, "element_type": "prose"|"table"}.
    """
    splitter = RecursiveCharacterTextSplitter(
        chunk_size=chunk_size,
        chunk_overlap=chunk_overlap,
        separators=["\n\n", "\n", ". ", " ", ""],
    )

    elements = _ea_parse_elements(content)

    has_boundary_header = any(
        e["type"] == "header" and e["level"] in (2, 3) for e in elements
    )
    has_table = any(e["type"] == "table" for e in elements)

    if not has_boundary_header and not has_table:
        pieces = splitter.split_text(content)
        return [{"text": p, "element_type": "prose"} for p in pieces if p.strip()]

    sections = _ea_group_into_sections(elements)
    results  = []

    for section in sections:
        header_text = section.get("header_text")
        segments    = _ea_section_to_text_segments(section)

        for seg in segments:
            if seg["element_type"] == "table":
                text = seg["text"]
                if header_text:
                    text = f"### {header_text}\n{text}"
                results.append({"text": text, "element_type": "table"})
            else:
                section_text = seg["text"]
                if header_text:
                    section_text = f"### {header_text}\n{section_text}"
                pieces = splitter.split_text(section_text)
                if header_text:
                    pieces = _ea_merge_orphaned_header(pieces, header_text)
                for p in pieces:
                    if p.strip():
                        results.append({"text": p, "element_type": "prose"})

    return results


# Canonical URL — lowercase path, strip trailing slash/query.
def normalise_url_path(url: str) -> str:
    """Canonical URL — lowercase path, strip trailing slash/query."""
    try:
        parsed = urlparse(url)
        norm   = parsed._replace(
            path=parsed.path.lower().rstrip("/"),
            query="",
            fragment="",
        )
        return norm.geturl()
    except Exception:
        return url.strip().rstrip("/").lower()


# ══════════════════════════════════════════════════════════════
# DROPDOWN SCRAPING HELPERS — mirrors scraper v4.5.5+
# ══════════════════════════════════════════════════════════════

# True if text contains phone/contact keywords or patterns (routing-dropdown signal).
def _has_contact_signals(text: str) -> bool:
    text_lower = text.lower()
    for kw in _CONTACT_SIGNAL_KEYWORDS:
        if kw in text_lower:
            return True
    for pattern in _CONTACT_SIGNAL_PATTERNS:
        if pattern.search(text):
            return True
    return False


# Check rendered HTML for a genuine contact-routing <select> dropdown.
def _has_routing_dropdowns_in_html(html: str) -> bool:
    if not html or not _BS4_AVAILABLE:
        return False
    try:
        soup = BeautifulSoup(html, "html.parser")
        for select in soup.find_all("select"):
            valid_opts = [
                o for o in select.find_all("option")
                if o.get_text(strip=True).lower() not in _DROPDOWN_PLACEHOLDERS
                and o.get_text(strip=True)
            ]
            if len(valid_opts) > 1:
                return True
        return False
    except Exception:
        return False


# Scrape per-option content from routing dropdown page using Playwright.
def _scrape_dropdown_states_playwright(
    url:              str,
    base_title:       str,
    base_page_data:   dict,
    freshness_run_id: str = "",
) -> list[dict]:
    """
    Scrape per-option content from routing dropdown page using Playwright.
    Mirrors _scrape_dropdown_states_playwright() from scraper v4.5.5.
    3-layer filter: navigation guard + contact signal + placeholder check.
    SCRAPER_VERSION and METADATA_VERSION stamped on all results.
    """
    try:
        from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
    except ImportError:
        log.error("playwright_not_installed")
        return []

    results: list[dict] = []
    try:
        with sync_playwright() as pw:
            import os as _os
            _exec     = PLAYWRIGHT_EXECUTABLE_PATH
            _exec_arg = _exec if _exec and _os.path.exists(_exec) else None
            browser   = pw.chromium.launch(headless=True, executable_path=_exec_arg)
            try:
                page = browser.new_page()
                page.route(
                    "**/*.{png,jpg,jpeg,gif,svg,woff,woff2,ttf,eot}",
                    lambda route: route.abort(),
                )
                page.goto(url, wait_until="networkidle", timeout=45000)
                try:
                    page.wait_for_selector("main, article, [role='main']", timeout=10000)
                except PWTimeout:
                    # Selector never appeared — proceed anyway, the
                    # subsequent select/option scan just finds nothing.
                    pass

                selects           = page.query_selector_all("select")
                routing_dropdowns = []
                for select in selects:
                    options    = select.query_selector_all("option")
                    valid_opts = []
                    for opt in options:
                        text  = opt.inner_text().strip()
                        value = opt.get_attribute("value") or ""
                        if text and text.lower() not in _DROPDOWN_PLACEHOLDERS:
                            valid_opts.append({"value": value, "text": text})
                    if len(valid_opts) > 1:
                        routing_dropdowns.append({"select_element": select, "options": valid_opts})

                if not routing_dropdowns:
                    return []

                default_raw_text = page.inner_text("body")
                default_lines    = {
                    line.strip() for line in default_raw_text.split("\n") if line.strip()
                }

                for dropdown in routing_dropdowns:
                    select_el = dropdown["select_element"]
                    for opt in dropdown["options"]:
                        opt_text  = opt["text"]
                        opt_value = opt["value"]
                        try:
                            original_url = page.url
                            select_el.evaluate(
                                f"(el) => {{ el.value = {json.dumps(opt_value or opt_text)}; "
                                f"el.dispatchEvent(new Event('input', {{bubbles:true}})); "
                                f"el.dispatchEvent(new Event('change', {{bubbles:true}})); }}"
                            )
                            page.wait_for_timeout(1500)

                            # Layer 1: Navigation guard
                            if page.url != original_url:
                                try:
                                    page.goto(url, wait_until="networkidle", timeout=30000)
                                    default_raw_text = page.inner_text("body")
                                    default_lines = {
                                        l.strip() for l in default_raw_text.split("\n") if l.strip()
                                    }
                                    selects = page.query_selector_all("select")
                                except Exception as e:
                                    # Page reset after navigation failed — this
                                    # dropdown option's state is being skipped
                                    # rather than silently assumed unchanged.
                                    # Logged so a run with unexpectedly few
                                    # dropdown states can be traced back to
                                    # this cause instead of looking like a
                                    # site-content change.
                                    log.warning(
                                        "dropdown_page_reset_failed",
                                        url=url,
                                        error=str(e),
                                    )
                                continue

                            new_raw_text    = page.inner_text("body")
                            new_lines       = [l.strip() for l in new_raw_text.split("\n") if l.strip()]
                            changed_lines   = [l for l in new_lines if l not in default_lines]
                            dynamic_content = "\n".join(changed_lines)

                            # Layer 2a: Minimum content check
                            if not dynamic_content or len(dynamic_content.strip()) < 20:
                                continue

                            # Layer 2b: Contact signal validation
                            if not _has_contact_signals(dynamic_content):
                                continue

                            # All layers passed — build page dict
                            safe_value = opt_value if opt_value else opt_text
                            state_url  = f"{url}#state={urllib.parse.quote(safe_value)}"
                            content    = dynamic_content.strip()

                            results.append({
                                "url":              state_url,
                                # v1.6.0: clean parent URL (no #state= fragment)
                                # mirrors scraper v4.8.0 fix.
                                "parent_url":       url,
                                "title":            f"{base_title} — {opt_text}",
                                "section":          base_page_data["section"],
                                "content":          content,
                                "scraped_at":       datetime.now(timezone.utc).isoformat(),
                                "content_length":   len(content),
                                "content_hash":     compute_content_hash(content),
                                "scraper_version":  SCRAPER_VERSION,
                                "metadata_version": METADATA_VERSION,
                                "audience":         base_page_data["audience"],
                                "has_video":        base_page_data["has_video"],
                                "content_type":     base_page_data["content_type"],
                                "product_category": base_page_data["product_category"],
                                "description":      base_page_data["description"],
                                "thumbnail_url":    base_page_data["thumbnail_url"],
                                "publish_date":     base_page_data["publish_date"],
                                "collection_name":  base_page_data["collection_name"],
                                "read_time_mins":   str(max(1, len(content.split()) // 200)),
                                "dropdown_state":   opt_text,
                                "dropdown_value":   opt_value or "",
                            })
                            log.info("playwright_option_scraped", url=state_url, option=opt_text)

                        except Exception as e:
                            log.warning("playwright_option_error", url=url, option=opt_text, error=str(e))
                            continue
            finally:
                browser.close()
    except Exception as e:
        log.error("playwright_dropdown_scrape_error", url=url, error=str(e))
    return results


# ══════════════════════════════════════════════════════════════
# URL HEALTH CHECKING
# ══════════════════════════════════════════════════════════════

# HTTP HEAD one URL, classify as live/redirect/dead_404/dead_5xx/timeout.
async def check_single_url(session: aiohttp.ClientSession, entry: dict) -> dict:
    """
    v1.7.4 FIX: previously used allow_redirects=False and inspected only
    the FIRST hop's Location header, truncated at 80 chars. Royal
    London's site (or its WAF/CDN) appears to 301/302 EVERY URL on a
    single transparent hop (trailing slash / scheme canonicalisation —
    confirmed via freshness_report_report_20260812_161133.xlsx: all
    297 approved URLs showed "Redirects to <same-url-prefix, truncated>"
    — the truncation was hiding that source and destination were the
    same page). A real browser (and our crawl4ai scraper) follows that
    hop silently and lands on the live page with no issue — but this
    check flagged every single URL as "internal_redirect: add new URL
    to Excel", a 100% false-positive rate with zero real signal in it.

    FIX: follow redirects (allow_redirects=True, capped at 5 hops) and
    compare the FINAL resolved URL against the original, normalised
    (normalise_url_path — case/trailing-slash/query insensitive). If
    they match, this was a transparent single-hop redirect — genuinely
    "live", no action needed. Only if the final URL resolves to a
    DIFFERENT page (or a different domain) does this get flagged as
    internal_redirect / external_redirect — meaning the flag can now
    actually be trusted.
    """
    url    = entry["url"]
    result = {"url": url, "status": "unknown", "status_code": None, "redirect_note": ""}
    try:
        async with session.head(
            url,
            allow_redirects=True,
            max_redirects=5,
            timeout=aiohttp.ClientTimeout(total=HTTP_TIMEOUT_SECONDS),
        ) as resp:
            code      = resp.status
            final_url = str(resp.url)
            result["status_code"] = code

            if code < 300:
                if normalise_url_path(final_url) == normalise_url_path(url):
                    # Transparent redirect (trailing slash, scheme, etc.)
                    # — resolves to the same page. Genuinely live.
                    result["status"] = "live"
                elif EXPECTED_DOMAIN in final_url:
                    result["status"]        = "internal_redirect"
                    result["redirect_note"] = (
                        f"Redirects to {final_url} — add new URL to Excel."
                    )
                else:
                    result["status"]        = "external_redirect"
                    result["redirect_note"] = (
                        f"Redirects to {final_url} — removing."
                    )
            elif code < 500:
                result["status"] = "dead_404"
            else:
                result["status"] = "dead_5xx"
    except asyncio.TimeoutError:
        result["status"] = "timeout"
    except Exception as e:
        result["status"] = "error"
        log.warning("health_check_error", url=url, error=str(e))
    return result


# Health-check all approved URLs concurrently (bounded by HTTP_CONCURRENCY).
async def check_all_urls_health(entries: list[dict]) -> list[dict]:
    sem       = asyncio.Semaphore(HTTP_CONCURRENCY)
    connector = aiohttp.TCPConnector(ssl=False, limit=HTTP_CONCURRENCY)
    async with aiohttp.ClientSession(
        connector=connector,
        headers={"User-Agent": "DigitalAssistance-ContentFreshness/1.4"},
    ) as session:
        # Cap concurrent HEAD requests via the semaphore.
        async def _bounded(entry):
            async with sem:
                return await check_single_url(session, entry)
        return await asyncio.gather(*[_bounded(e) for e in entries])


# ══════════════════════════════════════════════════════════════
# INDEX MANAGEMENT (BOTH INDEXES)
# ══════════════════════════════════════════════════════════════

# Read content_hash for all documents in the specified index.
def fetch_current_hashes_from_index(index_name: str = INDEX_NAME) -> dict[str, str]:
    """Read content_hash for all documents in the specified index."""
    client  = get_search_client(index_name)
    hashes  = {}
    skip    = 0
    page_sz = 1000
    while True:
        try:
            results = client.search(
                search_text="*",
                select=["source_url", "content_hash"],
                top=page_sz,
                skip=skip,
            )
            batch = list(results)
            if not batch:
                break
            for r in batch:
                url = r.get("source_url", "")
                h   = r.get("content_hash", "")
                if url and h:
                    hashes[normalise_url(url)] = h
            if len(batch) < page_sz:
                break
            skip += page_sz
        except Exception as e:
            log.error("fetch_hashes_error", index=index_name, error=str(e))
            break
    log.info("hashes_fetched", index=index_name, count=len(hashes))
    return hashes


# Get all chunk IDs for a URL including dropdown #state= variants.
def get_chunk_ids_for_url(url: str, index_name: str = INDEX_NAME) -> list[str]:
    """
    Get all chunk IDs for a URL including dropdown #state= variants.

    v1.7.8 FIX: previously used a single client.search(top=1000) call
    with NO pagination — only the first 1000 documents (arbitrary
    order, unsorted match-all query) were ever checked. With ~3,116+
    documents in the index post-cleanup, most URLs' actual chunks
    simply weren't in that first arbitrary batch, so the function
    silently returned [] for the majority of genuinely-indexed pages.
    Confirmed via a 10-URL test report showing "Main Chunks Before: 0"
    for every row, including pages known to have real indexed chunks
    (e.g. tell-us-about-a-bereavement, confirmed 13 dropdown states
    earlier in this session). Fixed with the same skip/page_sz
    pagination loop already used correctly in
    fetch_current_hashes_from_index().
    """
    client = get_search_client(index_name)
    base   = get_base_url(url)
    norm   = normalise_url(base)
    ids    = []
    skip    = 0
    page_sz = 1000
    try:
        while True:
            results = client.search(
                search_text="*",
                select=["chunk_id", "source_url"],
                top=page_sz,
                skip=skip,
            )
            batch = list(results)
            if not batch:
                break
            for r in batch:
                src = r.get("source_url", "")
                if normalise_url(get_base_url(src)) == norm:
                    ids.append(r["chunk_id"])
            if len(batch) < page_sz:
                break
            skip += page_sz
    except Exception as e:
        log.error("get_chunk_ids_error", url=url, index=index_name, error=str(e))
    return ids


# Expand base URLs to include all #state= dropdown variants in the index.
def get_all_urls_to_delete(base_urls: list[str], index_name: str = INDEX_NAME) -> list[str]:
    """
    Expand base URLs to include all #state= dropdown variants in the index.

    v1.7.8 FIX: same missing-pagination bug as get_chunk_ids_for_url()
    — top=5000 with no skip loop. Currently "accidentally safe" only
    because the index (~3,116 docs post-cleanup) happens to be under
    5000 — would silently break again the moment a full rebuild
    produces more chunks than that. Fixed with the same pagination
    pattern used throughout this file for index-wide scans.
    """
    client   = get_search_client(index_name)
    all_urls = set(base_urls)
    skip     = 0
    page_sz  = 1000
    try:
        while True:
            results = client.search(
                search_text="*",
                select=["source_url"],
                top=page_sz,
                skip=skip,
            )
            batch = list(results)
            if not batch:
                break
            for r in batch:
                src = r.get("source_url", "")
                if src and is_dropdown_url(src):
                    base_norm = normalise_url(get_base_url(src))
                    if any(normalise_url(b) == base_norm for b in base_urls):
                        all_urls.add(src)
            if len(batch) < page_sz:
                break
            skip += page_sz
    except Exception as e:
        log.warning("get_all_urls_error", index=index_name, error=str(e))
    return list(all_urls)


# Delete all chunks for given URLs from the specified index.
def delete_chunks_for_urls(
    urls:       list[str],
    index_name: str  = INDEX_NAME,
    dry_run:    bool = False,
) -> dict[str, int]:
    """Delete all chunks for given URLs from the specified index."""
    client  = get_search_client(index_name)
    summary = {}
    for url in urls:
        ids = get_chunk_ids_for_url(url, index_name)
        if not ids:
            continue
        summary[url] = len(ids)
        if dry_run:
            log.info("dry_run_would_delete", url=url, count=len(ids), index=index_name)
            continue
        for i in range(0, len(ids), 100):
            batch = [{"chunk_id": cid} for cid in ids[i:i + 100]]
            try:
                client.delete_documents(documents=batch)
            except Exception as e:
                log.error("delete_error", url=url, index=index_name, error=str(e))
        log.info("chunks_deleted", url=url, count=len(ids), index=index_name)
    return summary


# Read the current refresh_count for a URL from the main index.
def get_refresh_count_for_url(url: str) -> int:
    """
    Read the current refresh_count for a URL from the main index.

    Returns the refresh_count from the FIRST chunk found for this URL.
    All chunks for a given URL share the same refresh_count (they were
    all indexed together in the same freshness run). Returns 0 if URL
    not found or field absent (backward compat with pre-v1.5.0 chunks).

    Called in Step 8 (apply mode) before deleting changed URL chunks
    so the new chunks can be stamped with refresh_count + 1.

    v1.7.8 FIX: same missing-pagination bug as get_chunk_ids_for_url()
    and get_all_urls_to_delete() — but worse here: top=1 meant this
    checked a SINGLE arbitrary document from the entire index and
    almost never matched the target URL. refresh_count would
    effectively never increment correctly across freshness runs —
    every re-indexed chunk always got refresh_count=1 (0+1) regardless
    of true history. Fixed with the same pagination pattern.
    """
    client  = get_search_client(INDEX_NAME)
    base    = get_base_url(url)
    norm    = normalise_url(base)
    skip    = 0
    page_sz = 1000
    try:
        while True:
            results = client.search(
                search_text="*",
                select=["source_url", "refresh_count"],
                top=page_sz,
                skip=skip,
            )
            batch = list(results)
            if not batch:
                break
            for r in batch:
                src = r.get("source_url", "")
                if normalise_url(get_base_url(src)) == norm:
                    return int(r.get("refresh_count") or 0)
            if len(batch) < page_sz:
                break
            skip += page_sz
    except Exception as e:
        log.warning("get_refresh_count_error", url=url, error=str(e))
    return 0


# Archive existing chunks for a URL to Blob before deletion.
def save_deleted_chunks_to_blob(
    url:        str,
    index_name: str,
    ts_str:     str,
) -> str | None:
    """
    Archive existing chunks for a URL to Blob before deletion.

    Called before delete_chunks_for_urls() in Step 8 (apply mode).
    Provides audit trail and manual rollback capability if a content
    change causes regression in model responses.

    Blob path: freshness/archive/deleted_<url_hash>_<index>_<ts>.json
    Contains: full chunk documents as they existed just before deletion.
    Non-fatal — archive failure never blocks the delete/reindex.

    HOW TO ROLL BACK:
      1. Download the archive JSON from Blob
      2. Re-upload documents to AI Search using SearchClient.upload_documents()
      3. Invalidate Redis cache for the affected URL

    v1.5.0 NEW — added in response to PO question:
    "What happens if web-page content change impacts model response quality?"
    Archive gives ops a point-in-time snapshot of what was in the index
    before the change, enabling manual comparison and rollback if needed.

    v1.7.8 FIX: same missing-pagination bug as get_chunk_ids_for_url(),
    get_all_urls_to_delete(), and get_refresh_count_for_url() — top=500
    with no skip loop meant only the first arbitrary 500 of ~3,116+
    documents were ever checked. For most URLs this silently produced
    an EMPTY archive (no chunks found → early return None) right
    before their real chunks were deleted in apply mode — defeating
    the rollback-safety purpose of this function entirely, with no
    error surfaced (archive failure is designed to be non-fatal, so
    this failed completely silently). Fixed with the same pagination
    pattern used throughout this file.
    """
    if not BLOB_STORAGE_CONNECTION:
        log.debug("archive_skipped_no_blob", url=url)
        return None

    client = get_search_client(index_name)
    base   = get_base_url(url)
    norm   = normalise_url(base)

    try:
        chunks  = []
        skip    = 0
        page_sz = 500
        while True:
            results = client.search(
                search_text="*",
                select=[
                    "chunk_id", "source_url", "title", "content",
                    "section", "chunk_index", "total_chunks",
                    "content_hash", "scraped_at", "indexed_at",
                    "scraper_version", "metadata_version",
                    "pipeline_version", "index_run_id", "scrape_run_id",
                    "refresh_count", "augmented_questions", "title_questions",
                    "has_video", "content_type", "product_category",
                ],
                top=page_sz,
                skip=skip,
            )
            batch = list(results)
            if not batch:
                break
            for r in batch:
                src = r.get("source_url", "")
                if normalise_url(get_base_url(src)) == norm:
                    chunks.append(dict(r))
            if len(batch) < page_sz:
                break
            skip += page_sz

        if not chunks:
            return None

        # url_hash: short identifier for the URL in the filename
        url_hash    = hashlib.sha256(base.encode()).hexdigest()[:12]
        index_short = "main" if index_name == INDEX_NAME else "baseline"
        filename    = f"deleted_{url_hash}_{index_short}_{ts_str}.json"
        blob_path   = f"{BLOB_ARCHIVE_PREFIX}{filename}"

        archive = {
            "archived_at":  datetime.now(timezone.utc).isoformat(),
            "url":          base,
            "index_name":   index_name,
            "chunk_count":  len(chunks),
            "chunks":       chunks,
        }
        data   = json.dumps(archive, indent=2, ensure_ascii=False).encode("utf-8")
        svc    = get_blob_service()
        bclient = svc.get_blob_client(container=BLOB_CONTAINER_NAME, blob=blob_path)
        bclient.upload_blob(data, overwrite=True)
        log.info("chunks_archived", url=base, count=len(chunks),
                 blob=blob_path, index=index_name)
        return blob_path

    except Exception as e:
        log.warning("archive_failed", url=url, index=index_name, error=str(e))
        return None


# Targeted Redis cache invalidation — only affected keys flushed.
def invalidate_cache_for_urls(urls: list[str], dry_run: bool = False) -> int:
    """Targeted Redis cache invalidation — only affected keys flushed."""
    redis = get_redis()
    if not redis:
        return 0
    invalidated = 0
    norm_urls   = {normalise_url(u) for u in urls}
    try:
        cursor  = 0
        pattern = f"{REDIS_CACHE_KEY_PREFIX}*".encode()
        while True:
            cursor, keys = redis.scan(cursor, match=pattern, count=200)
            for key in keys:
                try:
                    raw = redis.get(key)
                    if raw:
                        cached = pickle.loads(raw)
                        cached_urls = set()
                        if isinstance(cached, dict):
                            cached_urls = {normalise_url(u) for u in cached.get("source_urls", [])}
                        if cached_urls & norm_urls:
                            if not dry_run:
                                redis.delete(key)
                            invalidated += 1
                except Exception as e:
                    # One corrupted/incompatible cache entry shouldn't
                    # abort the whole invalidation scan — log at debug
                    # level so a systematic issue (e.g. pickle format
                    # change) is still traceable, without the noise of
                    # a warning for an expected occasional bad key.
                    log.debug("cache_key_skip", key=str(key), error=str(e))
            if cursor == 0:
                break
    except Exception as e:
        log.warning("cache_invalidation_error", error=str(e))
    log.info("cache_invalidated", count=invalidated, dry_run=dry_run)
    return invalidated


# ══════════════════════════════════════════════════════════════
# SCRAPING — self-contained, full metadata extraction
# ══════════════════════════════════════════════════════════════

# Scrape a URL and any routing dropdown states.
async def scrape_url_with_dropdowns(
    crawler,
    entry:            dict,
    freshness_run_id: str = "",
) -> list[dict] | None:
    """
    Scrape a URL and any routing dropdown states.
    Full metadata extraction via extract_page_metadata() — matches
    full scraper quality including has_video, content_type, etc.
    SCRAPER_VERSION and METADATA_VERSION stamped on every page dict.

    v1.7.6 FIX: now takes `crawler` (an already-open AsyncWebCrawler)
    as a parameter instead of creating a fresh one internally. Found
    via a --mode report run with real error detail (see v1.7.5's
    scrape_failed fix — this is what that fix was FOR): concurrent
    scrape tasks were EACH opening and closing their own
    AsyncWebCrawler (CDP connection) around a single URL, and with
    SCRAPE_CONCURRENCY=3 those 3 concurrent connect/disconnect
    cycles against the SAME shared Chrome CDP instance raced each
    other — one task's teardown could kill another's in-flight
    connection or navigation. Confirmed by comparison against
    scrape_approved_urls_updatedV5.py's scrape_page(), which has
    always created ONE AsyncWebCrawler outside its entire batch loop
    and passed it into every call — this function now does the same
    (the shared instance is created once in scrape_urls_batch() and
    passed down). Symptoms this fixes:
      - scrape_exception: "BrowserType.connect_over_cdp: Connection
        closed while reading from the driver"
      - scrape_failed: "Page.goto: net::ERR_ABORTED" /
        "Failed on navigating ACS-GOTO"

    NOTE: does NOT fix playwright_option_error ("Execution context
    was destroyed... navigation") seen on dropdown/filter option
    clicks (e.g. "Our first cohort" on the changemakers page) — that
    happens inside _scrape_dropdown_states_playwright(), which uses
    its OWN separate Playwright browser launch (thread pool executor,
    not the shared CDP crawler) and is a different code path. Left
    as a separate, still-open follow-up.

    Returns:
        list[dict] — [base_page, *dropdown_states] or [base_page]
        None       — scrape failed
    """
    if not _CRAWL4AI_AVAILABLE:
        raise RuntimeError("crawl4ai not installed.")

    url      = entry["url"]
    title    = entry.get("title", "")
    category = entry.get("category", "")

    run_cfg = CrawlerRunConfig(
        css_selector=(
            "main, article, .content, #content, "
            ".page-content, .main-content, [role='main']"
        ),
        markdown_generator=DefaultMarkdownGenerator(
            content_filter=PruningContentFilter(threshold=0.45, threshold_type="fixed"),
            options={"ignore_links": False, "ignore_images": True, "skip_internal_links": True},
        ),
        wait_until="domcontentloaded",
        page_timeout=30000,
        verbose=False,
        excluded_tags=["nav", "header", "footer", "aside", "script", "style", "noscript"],
    )

    try:
        result = await crawler.arun(url=url, config=run_cfg)

        if not result.success or not result.markdown:
            # v1.7.5: log the actual failure reason — previously
            # only the URL was logged, making every scrape_failed
            # occurrence undiagnosable from logs alone. See
            # scrape_approved_urls_updatedV5.py's scrape_page()
            # for the same result.error_message pattern.
            log.warning(
                "scrape_failed",
                url=url,
                error=result.error_message or "no error_message from crawl4ai",
                success=result.success,
                has_markdown=bool(result.markdown),
            )
            return None

        raw_content = result.markdown.raw_markdown
        if not raw_content or len(raw_content.strip()) < 100:
            log.warning("content_too_short", url=url)
            return None

        status_code = getattr(result, "status_code", None)
        if status_code and status_code >= 400:
            log.warning("scrape_http_error", url=url, status_code=status_code)
            return None

        # v1.7.7: use clean_scraped_content() (full boilerplate strip,
        # exact port of scraper's clean_content()) — NOT the other
        # clean_content() in this file (URL-stripper, chunk-time only).
        # Hash computed directly on this output, no separate
        # URL-stripping step in between — matches exactly how the
        # scraper computes the content_hash that's already stored
        # in the index.
        page_content = clean_scraped_content(raw_content.strip())
        content_hash = compute_content_hash(page_content.strip())

        raw_html = getattr(result, "html", "") or ""
        metadata = extract_page_metadata(raw_html, url)

        # Excel Category overrides URL-pattern for content_type
        content_type = map_excel_category_to_content_type(category, url)

        base_page = {
            "url":              normalise_url_path(url),
            "title":            title,
            "section":          derive_section(url),
            "content":          page_content,
            "scraped_at":       datetime.now(timezone.utc).isoformat(),
            "content_length":   len(page_content),
            "content_hash":     content_hash,
            # Versioning — stamped here, passed through to chunk_page()
            "scraper_version":  SCRAPER_VERSION,
            "metadata_version": METADATA_VERSION,
            # Rich metadata
            "audience":         metadata["audience"],
            "has_video":        metadata["has_video"],
            "content_type":     content_type,
            "product_category": metadata["product_category"],
            "description":      metadata["description"],
            "thumbnail_url":    metadata["thumbnail_url"],
            "publish_date":     metadata["publish_date"],
            "collection_name":  metadata["collection_name"],
            "read_time_mins":   str(metadata["read_time_mins"]),
            # Dropdown — empty for base page
            "dropdown_state":   "",
            "dropdown_value":   "",
        }

        # Dropdown detection and scraping
        dropdown_states: list[dict] = []
        if _has_routing_dropdowns_in_html(raw_html):
            log.info("dropdown_page_detected", url=url)
            try:
                loop     = asyncio.get_event_loop()
                executor = concurrent.futures.ThreadPoolExecutor(max_workers=1)
                dropdown_states = await loop.run_in_executor(
                    executor,
                    _scrape_dropdown_states_playwright,
                    url, title, base_page, freshness_run_id,
                )
            except Exception as e:
                log.warning("playwright_dropdown_skipped", url=url, error=str(e))
                dropdown_states = []

        pages = [base_page] + dropdown_states
        if dropdown_states:
            log.info("multi_state_page_scraped", url=url, states=len(dropdown_states))
        return pages

    except Exception as e:
        log.error("scrape_exception", url=url, error=str(e))
        return None


# Scrape multiple URLs with limited concurrency.
async def scrape_urls_batch(
    entries:          list[dict],
    concurrency:      int = SCRAPE_CONCURRENCY,
    freshness_run_id: str = "",
) -> list[list[dict] | None]:
    """
    Scrape multiple URLs with limited concurrency.

    v1.7.6 FIX: AsyncWebCrawler is now created ONCE here, before the
    semaphore-gated concurrent tasks start, and shared across all of
    them — see scrape_url_with_dropdowns() docstring for the full
    incident this fixes (concurrent per-URL AsyncWebCrawler instances
    racing over the same CDP connection). Mirrors
    scrape_approved_urls_updatedV5.py's proven batch-loop pattern:
    one AsyncWebCrawler for the whole run, not one per URL.
    """
    if not _CRAWL4AI_AVAILABLE:
        raise RuntimeError("crawl4ai not installed.")

    sem         = asyncio.Semaphore(concurrency)
    browser_cfg = _cf_make_browser_config()

    async with AsyncWebCrawler(config=browser_cfg) as crawler:
        # Cap concurrent scrape tasks via the semaphore.
        async def _bounded(entry):
            async with sem:
                return await scrape_url_with_dropdowns(
                    crawler, entry, freshness_run_id=freshness_run_id
                )

        return await asyncio.gather(*[_bounded(e) for e in entries])


# ══════════════════════════════════════════════════════════════
# CHUNKING — mirrors chunk_pages() from indexer v5.7.0 exactly
# ══════════════════════════════════════════════════════════════

# Split a page into index-ready chunks.
def chunk_page(
    page:             dict,
    freshness_run_id: str = "",
    indexed_at:       str = "",
    refresh_count:    int = 0,
) -> list[dict]:
    """
    Split a page into index-ready chunks.
    Mirrors chunk_pages() in chunk_and_index_hqaV5.py exactly.
    All versioning fields stamped on every chunk including refresh_count.
    Atomic chunking for dropdown state pages (v1.1.0).

    refresh_count: how many times this page has been refreshed by
    the nightly freshness job. Caller reads existing value from index
    via get_refresh_count_for_url() and passes existing_count + 1.
    0 on first freshness run for a given page (was 0 from full indexer
    run; freshness increments to 1 on first change).

    v1.7.0 — ELEMENT-AWARE CHUNKING (standard pages only):
    Standard pages now go through chunk_content_element_aware()
    (inlined above — see that section's header comment for the
    self-contained-vs-shared-module design decision) instead of a
    flat RecursiveCharacterTextSplitter call. Tables chunked
    atomically (row-capped); ##/### headers act as hard section
    boundaries. Pages with no tables/headers get byte-identical
    output to the old flat splitter.
    """
    if not _LANGCHAIN_AVAILABLE:
        raise RuntimeError("langchain_text_splitters not installed.")

    raw_content = page.get("content", "").strip()
    if not raw_content or len(raw_content) < 50:
        return []

    content = clean_content(raw_content)
    if len(content) < 50:
        return []

    title              = page.get("title", "")
    content_with_title = f"{title}\n\n{content}" if title else content
    _indexed_at        = indexed_at or datetime.now(timezone.utc).isoformat()

    # Build the version/traceability fields shared by every chunk dict.
    def _versioning() -> dict:
        return {
            "pipeline_version":  PIPELINE_VERSION,
            "index_run_id":      freshness_run_id,
            "indexed_at":        _indexed_at,
            "scraper_version":   page.get("scraper_version", SCRAPER_VERSION),
            "metadata_version":  page.get("metadata_version", METADATA_VERSION),
            # scrape_run_id: links all chunks from this freshness batch
            "scrape_run_id":     freshness_run_id,
            # refresh_count: how many times this URL has been delta re-indexed.
            # Caller reads existing value from index and passes + 1.
            "refresh_count":     refresh_count,
        }

    # Build the enrichment metadata fields shared by every chunk dict.
    def _enrichment() -> dict:
        return {
            "has_video":        page.get("has_video", False),
            "content_type":     page.get("content_type", "article"),
            "product_category": page.get("product_category", "general"),
            "description":      page.get("description", ""),
            "thumbnail_url":    page.get("thumbnail_url", ""),
            "publish_date":     page.get("publish_date", ""),
            "collection_name":  page.get("collection_name", ""),
            "read_time_mins":   str(page.get("read_time_mins", "5")),
        }

    # Atomic chunking for dropdown state pages
    is_dropdown_state = bool(page.get("dropdown_state", ""))
    if is_dropdown_state:
        if len(content_with_title.strip()) < 50:
            return []
        log.info("dropdown_atomic_chunk", url=page.get("url", ""), chars=len(content_with_title))
        return [{
            # v1.7.2: deterministic chunk_id — see compute_chunk_id()
            "chunk_id":            compute_chunk_id(page["url"], 0, content_with_title.strip()),
            "content":             content_with_title.strip(),
            "source_url":          page["url"],
            "title":               title,
            "section":             page.get("section", "General"),
            "audience":            page.get("audience", "customer"),
            "scraped_at":          page.get("scraped_at", ""),
            "chunk_index":         0,
            "total_chunks":        1,
            "content_hash":        page.get("content_hash", compute_content_hash(content)),
            "augmented_questions": "",
            "title_questions":     "",
            # v1.6.0: clean parent URL for dropdown state chunks
            "parent_url":          page.get("parent_url", ""),
            # v1.7.0: element_type — dropdown_state pages are always
            # atomic contact-detail content. Mirrors indexer v5.10.0.
            "element_type":        "dropdown_state",
            **_versioning(),
            **_enrichment(),
        }]

    # v1.7.0: element-aware chunking replaces the flat splitter call.
    # For pages with no headers/tables this produces IDENTICAL output
    # to the old flat splitter (fast-path in chunk_content_element_aware).
    #
    # v1.7.1 — BUGFIX: total_chunks must be computed AFTER filtering
    # out sub-50-char fragments, not before. The original pattern
    # (inherited from chunk_and_index_hqaV5.py, present since v5.x)
    # set total_chunks from the pre-filter piece count, so any small
    # fragment being dropped left surviving chunks with a total_chunks
    # value that didn't match their actual count. Rare with the old
    # flat greedy splitter; more likely with element-aware chunking,
    # since header-boundary splitting naturally produces short leading/
    # trailing fragments (e.g. a one-line closing sentence after a
    # table). Caught by validate_chunking_preflight() during testing —
    # fixed here rather than left for the checkpoint to keep flagging.
    pieces        = chunk_content_element_aware(content_with_title, CHUNK_SIZE, CHUNK_OVERLAP)
    valid_pieces  = [p for p in pieces if len(p["text"].strip()) >= 50]
    total         = len(valid_pieces)
    chunks        = []
    for idx, piece in enumerate(valid_pieces):
        split = piece["text"]
        chunks.append({
            # v1.7.2: deterministic chunk_id — see compute_chunk_id()
            "chunk_id":            compute_chunk_id(page["url"], idx, split.strip()),
            "content":             split.strip(),
            "source_url":          page["url"],
            "title":               title,
            "section":             page.get("section", "General"),
            "audience":            page.get("audience", "customer"),
            "scraped_at":          page.get("scraped_at", ""),
            "chunk_index":         idx,
            "total_chunks":        total,
            "content_hash":        page.get("content_hash", compute_content_hash(content)),
            "augmented_questions": "",
            "title_questions":     "",
            # v1.6.0: clean parent URL for dropdown state chunks
            "parent_url":          page.get("parent_url", ""),
            # v1.7.0: element_type — "prose" or "table". Mirrors
            # indexer v5.10.0's schema field for audit traceability.
            "element_type":        piece["element_type"],
            **_versioning(),
            **_enrichment(),
        })
    return chunks


# ══════════════════════════════════════════════════════════════
# STRICTER CHECKPOINTS — v1.7.0
# ══════════════════════════════════════════════════════════════
# WHY THIS EXISTS:
# Step 8 of run_freshness_job() deletes existing chunks for a changed
# URL BEFORE indexing the replacement chunks (see save_deleted_chunks_
# to_blob / delete_chunks_for_urls call sites). That ordering creates
# a live-index gap: if chunking or indexing fails AFTER deletion has
# already happened, the URL is left with nothing searchable until the
# next successful run. This risk existed before v1.7.0 too, but the
# element-aware chunking logic (table parsing, section grouping) is
# new code with more edge cases than the old flat splitter, so a
# pre-flight validation gate is added here — chunk (and sanity-check)
# every scraped page BEFORE Step 8 touches the index at all. If
# validation fails for any page, or fails above threshold, the whole
# apply run aborts with zero index writes, not a partial one.

# Safety threshold — max characters for a SINGLE chunk. Legitimate
# chunks should never approach this (CHUNK_SIZE=1600, largest atomic
# table chunk observed ~96 rows). A chunk this large signals a bug in
# section grouping (e.g. a header boundary failed to fire) rather
# than genuine content — better to abort than index a broken chunk.
_MAX_SANE_CHUNK_CHARS = 8000

# Max fraction of scraped pages allowed to fail validation before the
# whole apply run aborts. A single oddly-formatted page failing is
# tolerable (logged, skipped); widespread failure signals a real bug
# in the chunking logic itself and must not proceed to Step 8.
_MAX_VALIDATION_FAILURE_RATIO = 0.05  # 5%


# Pre-flight checkpoint — chunk every scraped page and sanity-check
def validate_chunking_preflight(scraped_pages: list) -> dict:
    """
    Pre-flight checkpoint — chunk every scraped page and sanity-check
    the output BEFORE run_freshness_job() proceeds to Step 8 (delete +
    index). Called once, after scraping, before any index mutation.

    Checks per page:
      1. chunk_page() completes without raising
      2. At least 1 chunk produced for non-trivial content
         (content >= 50 chars after cleaning)
      3. No single chunk exceeds _MAX_SANE_CHUNK_CHARS
      4. Every chunk has a non-empty element_type
      5. total_chunks matches len(chunks) actually produced

    Returns:
      {
        "passed":            bool — True if ratio of failures is
                              within _MAX_VALIDATION_FAILURE_RATIO
        "total_pages":       int,
        "failed_pages":      list of {"url", "reason"} dicts,
        "failure_ratio":     float,
      }

    Caller (run_freshness_job) MUST check result["passed"] and abort
    the entire apply run (zero index writes) if False.
    """
    failed_pages = []

    for page in scraped_pages:
        url = page.get("url", "unknown")
        try:
            chunks = chunk_page(page, freshness_run_id="preflight-check")
        except Exception as e:
            failed_pages.append({
                "url": url,
                "reason": f"chunk_page() raised: {type(e).__name__}: {e}",
            })
            continue

        raw_content = page.get("content", "").strip()
        if len(raw_content) >= 50 and not chunks:
            failed_pages.append({
                "url": url,
                "reason": "Non-trivial content produced ZERO chunks",
            })
            continue

        for c in chunks:
            content_len = len(c.get("content", ""))
            if content_len > _MAX_SANE_CHUNK_CHARS:
                failed_pages.append({
                    "url": url,
                    "reason": (
                        f"Chunk exceeds safety cap: {content_len} chars "
                        f"(cap={_MAX_SANE_CHUNK_CHARS}) — likely a section-"
                        f"grouping bug, not genuine content"
                    ),
                })
                break

            if not c.get("element_type"):
                failed_pages.append({
                    "url": url,
                    "reason": "Chunk missing element_type field",
                })
                break

        else:
            # No break — also verify total_chunks matches actual count
            if chunks and chunks[0].get("total_chunks") != len(chunks):
                # Only a genuine mismatch if all chunks share the same
                # total_chunks value (they should) and it disagrees
                # with the real count.
                declared = chunks[0].get("total_chunks")
                if declared != len(chunks):
                    failed_pages.append({
                        "url": url,
                        "reason": (
                            f"total_chunks mismatch: declared={declared}, "
                            f"actual={len(chunks)}"
                        ),
                    })

    total          = len(scraped_pages)
    failure_ratio  = (len(failed_pages) / total) if total else 0.0
    passed         = failure_ratio <= _MAX_VALIDATION_FAILURE_RATIO

    if failed_pages:
        log.warning(
            "chunking_preflight_failures",
            total_pages=total,
            failed_count=len(failed_pages),
            failure_ratio=round(failure_ratio, 4),
            passed=passed,
        )
        for f in failed_pages[:20]:  # cap logged detail, avoid log flood
            log.warning("chunking_preflight_page_failed",
                        url=f["url"], reason=f["reason"])
    else:
        log.info("chunking_preflight_passed", total_pages=total)

    return {
        "passed":        passed,
        "total_pages":   total,
        "failed_pages":  failed_pages,
        "failure_ratio": failure_ratio,
    }



# HQA — mirrors chunk_and_index_hqaV5.py v5.0.0+ exactly
# ══════════════════════════════════════════════════════════════

# True if the question shares at least 1 meaningful word with the chunk (anti-hallucination check).
def is_grounded(question: str, chunk_content: str) -> bool:
    stop_words = {
        "the", "a", "an", "is", "are", "was", "were", "in",
        "on", "at", "to", "for", "of", "and", "or", "but",
        "it", "this", "that", "with", "my", "your", "our",
        "i", "we", "you", "they", "be", "do", "have", "will",
        "can", "could", "would", "should", "may", "might",
        "from", "by", "as", "if", "when", "what", "how",
        "about", "which", "there", "their", "than", "then",
    }
    chunk_words = set(
        w.lower().strip(".,?!:;()[]\"'")
        for w in chunk_content.split()
        if len(w) > 3 and w.lower() not in stop_words
    )
    question_lower = question.lower()
    return sum(1 for w in chunk_words if w in question_lower) >= 1


# Reject generic/templated questions (too short or matching known generic patterns).
def is_specific_enough(question: str) -> bool:
    if len(question.split()) < 5:
        return False
    generic_patterns = [
        r"^what is (insurance|a pension|an isa)\??$",
        r"^how do (pensions|isas|policies) work\??$",
        r"^tell me about",
        r"^explain ",
        r"^what does royal london (do|offer|provide)\??$",
        r"^how can (i|you) help",
    ]
    question_lower = question.lower().strip()
    for pattern in generic_patterns:
        if re.search(pattern, question_lower):
            return False
    return True


# Score a question 0 (reject) / 1 (ok) / 2 (good — mentions RL domain terms).
def score_question(question: str, chunk_content: str) -> int:
    normalised = re.sub(r'[?!.,;:]+$', '', question.lower().strip()).strip()
    if normalised in BLOCKED_QUESTIONS:
        return 0
    if not is_grounded(question, chunk_content):
        return 0
    if not is_specific_enough(question):
        return 0
    rl_domain_terms = [
        "royal london", "pension", "isa", "insurance",
        "protection", "annuity", "drawdown", "premium",
        "policy", "benefit", "contribution", "allowance",
        "claim", "bereavement", "life cover", "critical illness",
        "income protection", "whole of life", "term insurance",
        "workplace", "personal pension", "sipp", "profitshare",
        "financial adviser", "retirement", "surrender",
    ]
    if any(term in question.lower() for term in rl_domain_terms):
        return 2
    return 1


# Validate a broad title-level question: grounded, not blocked, within word limit.
def is_valid_title_question(question: str, chunk_content: str) -> bool:
    normalised = re.sub(r'[?!.,;:]+$', '', question.lower().strip()).strip()
    if normalised in BLOCKED_QUESTIONS:
        return False
    if not is_grounded(question, chunk_content):
        return False
    word_count = len(question.split())
    if word_count == 0 or word_count > TITLE_QUESTIONS_MAX_WORDS:
        return False
    return True


# Generate broad entry-point questions for chunk_index == 0 only.
def generate_title_questions(chunk: dict, retry_count: int = 3) -> str:
    """Generate broad entry-point questions for chunk_index == 0 only."""
    from openai import RateLimitError
    if chunk.get("chunk_index", -1) != 0:
        return ""
    client          = get_openai_client()
    chunk_content   = chunk["content"]
    content_for_hqa = chunk_content[:2000]
    for attempt in range(retry_count):
        try:
            response = client.chat.completions.create(
                model=HQA_DEPLOYMENT,
                messages=[
                    {"role": "system", "content": TITLE_QUESTIONS_PROMPT},
                    {"role": "user", "content": (
                        f"Generate exactly {TITLE_QUESTIONS_COUNT} "
                        f"entry-point questions for this page's first chunk:\n\n{content_for_hqa}"
                    )},
                ],
                max_completion_tokens=4000,  # v1.5.3: GPT-5 needs reasoning headroom; 200 caused finish_reason=length with empty content
            )
            raw = response.choices[0].message.content or ""
            raw = raw.strip()
            # GPT-5 reasoning models may prepend/append prose or
            # wrap JSON in markdown fences. Extract the JSON array
            # directly by finding the outermost [ ... ] span.
            start = raw.find("[")
            end   = raw.rfind("]")
            if start != -1 and end != -1 and end > start:
                raw = raw[start:end + 1]
            elif raw.startswith("```"):
                raw = re.sub(r"```[a-z]*\n?", "", raw).strip().rstrip("`").strip()
            questions = json.loads(raw)
            if not isinstance(questions, list):
                continue
            accepted = [
                q.strip() for q in questions
                if isinstance(q, str) and q.strip()
                and is_valid_title_question(q.strip(), chunk_content)
            ]
            if _cost_tracker:
                _cost_tracker.record(response.usage)
            return "\n".join(accepted)
        except json.JSONDecodeError:
            if attempt < retry_count - 1:
                time.sleep(2)
            continue
        except RateLimitError:
            wait = 10 * (2 ** attempt)
            print(f"   ⚠️  Title questions rate limit. Waiting {wait}s...")
            time.sleep(wait)
            continue
        except Exception as e:
            log.error("title_questions_error", chunk_id=chunk.get("chunk_id", ""), error=str(e))
            if attempt < retry_count - 1:
                time.sleep(2)
            continue
    return ""


# Generate and validate HQA questions for a single chunk.
def generate_hqa_questions(
    chunk:         dict,
    num_questions: int = HQA_QUESTIONS_OTHER_CHUNKS,
    retry_count:   int = 3,
) -> list[str]:
    """Generate and validate HQA questions for a single chunk."""
    from openai import RateLimitError
    client          = get_openai_client()
    chunk_content   = chunk["content"]
    content_type    = chunk.get("content_type", "article")
    prompt_template = (
        HQA_CORPORATE_PROMPT_TEMPLATE
        if content_type == "corporate"
        else HQA_SYSTEM_PROMPT_TEMPLATE
    )
    system_prompt   = prompt_template.format(num_questions=num_questions)
    content_for_hqa = chunk_content[:2000]
    # v1.5.3: GPT-5 reasoning models consume internal reasoning tokens
    # before generating output. finish_reason=length with empty content
    # confirmed reasoning exhausted the budget. Floor raised to 4000.
    # gpt-4o-mini ignores extra budget (non-reasoning model).
    max_tokens      = max(int(num_questions * 55 + 50), 4000)

    for attempt in range(retry_count):
        try:
            response = client.chat.completions.create(
                model=HQA_DEPLOYMENT,
                messages=[
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": (
                        f"Generate {num_questions} questions for this chunk:\n\n{content_for_hqa}"
                    )},
                ],
                max_completion_tokens=max_tokens,
            )
            raw = response.choices[0].message.content or ""
            raw = raw.strip()
            # GPT-5 reasoning models may prepend/append prose or
            # wrap JSON in markdown fences. Extract the JSON array
            # directly by finding the outermost [ ... ] span.
            start = raw.find("[")
            end   = raw.rfind("]")
            if start != -1 and end != -1 and end > start:
                raw = raw[start:end + 1]
            elif raw.startswith("```"):
                raw = re.sub(r"```[a-z]*\n?", "", raw).strip().rstrip("`").strip()
            questions = json.loads(raw)
            if not isinstance(questions, list):
                continue
            accepted = []
            for q in questions:
                if isinstance(q, str) and q.strip() and score_question(q.strip(), chunk_content) >= 1:
                    accepted.append(q.strip())
            log.info("hqa_questions_generated",
                     chunk_id=chunk.get("chunk_id", "")[:8],
                     generated=len(questions), accepted=len(accepted))
            if _cost_tracker:
                _cost_tracker.record(response.usage)
            return accepted
        except json.JSONDecodeError:
            if attempt < retry_count - 1:
                time.sleep(2)
            continue
        except RateLimitError:
            wait = 10 * (2 ** attempt)
            print(f"   ⚠️  HQA rate limit. Waiting {wait}s...")
            time.sleep(wait)
            continue
        except Exception as e:
            log.error("hqa_error", chunk_id=chunk.get("chunk_id", ""), error=str(e))
            if attempt < retry_count - 1:
                time.sleep(2)
            continue
    return []


# True if URL matches a known dedicated (high-value) page pattern.
def _is_dedicated_page(source_url: str) -> bool:
    url_lower = source_url.lower()
    return any(p in url_lower for p in DEDICATED_PAGE_PATTERNS)


# True if URL matches a known generic/low-value page pattern.
def _is_generic_page(source_url: str) -> bool:
    url_lower = source_url.lower()
    return any(p in url_lower for p in GENERIC_PAGE_PATTERNS)


# Sort key for question dedup — prefer dedicated pages, chunk 0, lower index.
def _dedup_priority(chunk: dict) -> tuple:
    url        = chunk.get("source_url", "")
    idx        = chunk.get("chunk_index", 999)
    is_first   = 1 if idx == 0 else 2
    page_score = (
        0 if _is_dedicated_page(url) and not _is_generic_page(url)
        else 1 if not _is_generic_page(url)
        else 2
    )
    return (page_score, is_first, idx)


# Cross-chunk question deduplication. Mirrors indexer v5.0.0+ exactly.
def deduplicate_questions_across_chunks(chunks: list[dict]) -> list[dict]:
    """Cross-chunk question deduplication. Mirrors indexer v5.0.0+ exactly."""
    # Lowercase, strip trailing punctuation and extra whitespace so
    # near-duplicate questions collapse to the same key.
    def normalise(q: str) -> str:
        q = q.lower().strip()
        q = re.sub(r'[?!.,;:]+$', '', q)
        return ' '.join(q.split())

    QUESTION_FIELDS = ("title_questions", "augmented_questions")
    print(f"\n🔄 Deduplicating questions across {len(chunks):,} chunks...")

    question_to_chunks: dict[str, list[tuple]] = {}
    for ci, chunk in enumerate(chunks):
        for field in QUESTION_FIELDS:
            raw = chunk.get(field, "") or ""
            for q in [x.strip() for x in raw.split("\n") if x.strip()]:
                norm = normalise(q)
                if norm not in question_to_chunks:
                    question_to_chunks[norm] = []
                question_to_chunks[norm].append((ci, _dedup_priority(chunk), field, q))

    collisions = {
        norm: entries for norm, entries in question_to_chunks.items()
        if len(entries) > MAX_COLLISION_THRESHOLD
    }

    if not collisions:
        print("   ✅ No collisions above threshold")
        return chunks

    questions_removed = 0
    chunks_modified   = set()
    for norm_q, entries in collisions.items():
        sorted_entries = sorted(entries, key=lambda x: x[1])
        for ci, priority, field, original_q in sorted_entries[1:]:
            chunk  = chunks[ci]
            raw    = chunk.get(field, "") or ""
            qs     = [q.strip() for q in raw.split("\n") if q.strip()]
            new_qs = [q for q in qs if normalise(q) != norm_q]
            if len(new_qs) < len(qs):
                chunk[field] = "\n".join(new_qs)
                questions_removed += 1
                chunks_modified.add(ci)

    print(f"   ✅ Dedup: {len(collisions):,} collisions, {questions_removed:,} removed")
    return chunks


# Generate HQA + title_questions for all chunks.
def augment_chunks_with_hqa(chunks: list[dict]) -> list[dict]:
    """
    Generate HQA + title_questions for all chunks.
    Mirrors augment_chunks_with_hqa() from indexer v5.0.0+ exactly.
    Never raises — generation failure never stops indexing.
    """
    total             = len(chunks)
    first_chunk_count = sum(1 for c in chunks if c.get("chunk_index", -1) == 0)
    est_questions = (
        (total - first_chunk_count) * HQA_QUESTIONS_OTHER_CHUNKS
        + first_chunk_count * (HQA_QUESTIONS_FIRST_CHUNK + TITLE_QUESTIONS_COUNT)
    )
    # v1.5.4: model-aware time/cost estimate
    # gpt-4o-mini: $0.000116/question | gpt-5-mini: $0.547/100 chunks (pilot actuals)
    _is_gpt5 = "gpt-5" in HQA_DEPLOYMENT.lower()
    if _is_gpt5:
        est_cost_str = f"~${(total / 100) * 0.547:.2f} (gpt-5 actuals)"
        est_time_str = f"~{round(total * 12 / 60) + 1} min"
    else:
        est_cost_str = f"~${est_questions * 0.000116:.2f}"
        est_time_str = f"~{total // 30 + 1} min"

    # v1.5.4: initialise cost tracker
    global _cost_tracker
    _cost_tracker = CostTracker(model=HQA_DEPLOYMENT, run_mode="freshness-delta")

    print(f"\n🧠 HQA: Generating questions for {total:,} delta chunks...")
    print(f"   Model: {HQA_DEPLOYMENT} | Est: ~{est_questions:,} questions | "
          f"{est_cost_str} | {est_time_str}")

    accepted_total       = 0
    accepted_title_total = 0
    rejected_total       = 0
    failed_chunks        = 0

    for i, chunk in enumerate(chunks):
        is_first_chunk = chunk.get("chunk_index", -1) == 0
        num_questions  = HQA_QUESTIONS_FIRST_CHUNK if is_first_chunk else HQA_QUESTIONS_OTHER_CHUNKS

        if is_first_chunk:
            title_qs_str         = generate_title_questions(chunk)
            chunk["title_questions"] = title_qs_str
            if title_qs_str:
                accepted_title_total += len(title_qs_str.split("\n"))
        else:
            chunk["title_questions"] = ""

        questions = generate_hqa_questions(chunk, num_questions=num_questions)
        if questions:
            chunk["augmented_questions"] = "\n".join(questions)
            accepted_total += len(questions)
        else:
            chunk["augmented_questions"] = ""
            failed_chunks += 1

        rejected_total += (num_questions - len(questions))

        if (i + 1) % 50 == 0 or (i + 1) == total:
            pct = round((i + 1) / total * 100)
            print(f"   [{pct:3d}%] {i + 1:,}/{total:,} | "
                  f"HQA: {accepted_total:,} | Title: {accepted_title_total:,} | "
                  f"Failed: {failed_chunks}")

        time.sleep(0.1)

    print(f"\n   ✅ HQA done: {accepted_total:,} HQA accepted, "
          f"{accepted_title_total:,} title, {failed_chunks:,} failed chunks")

    chunks = deduplicate_questions_across_chunks(chunks)

    # v1.5.4: print cost summary for budgeting
    if _cost_tracker and _cost_tracker.calls > 0:
        print(_cost_tracker.summary())

    return chunks


# Build embedding text per chunk.
def build_embedding_texts(chunks: list[dict]) -> list[str]:
    """
    Build embedding text per chunk.
    Mirrors build_embedding_texts() from indexer exactly.
    Includes title_questions ahead of augmented_questions for chunk_index=0.
    """
    texts = []
    for chunk in chunks:
        content         = chunk["content"]
        title_questions = (chunk.get("title_questions", "") or "").strip()
        questions       = (chunk.get("augmented_questions", "") or "").strip()
        parts = [content]
        if title_questions:
            parts.append(f"Entry-point questions this page answers:\n{title_questions}")
        if questions:
            parts.append(f"Questions this answers:\n{questions}")
        texts.append("\n\n".join(parts) if len(parts) > 1 else content)
    return texts


# ══════════════════════════════════════════════════════════════
# EMBEDDING + DUAL-INDEX UPLOAD
# ══════════════════════════════════════════════════════════════

# Generate embeddings. Mirrors indexer — same model, batch size, backoff.
def get_embeddings_batch(texts: list[str]) -> list[list[float]]:
    """Generate embeddings. Mirrors indexer — same model, batch size, backoff."""
    from openai import RateLimitError
    client     = get_openai_client()
    all_embeds: list[list[float]] = []
    for i in range(0, len(texts), EMBEDDING_BATCH_SIZE):
        batch = texts[i:i + EMBEDDING_BATCH_SIZE]
        for attempt in range(6):
            try:
                response = client.embeddings.create(
                    input=batch,
                    model=EMBEDDING_DEPLOYMENT,
                    dimensions=EMBEDDING_DIMS,
                )
                all_embeds.extend([d.embedding for d in response.data])
                break
            except RateLimitError:
                wait = 30 * (2 ** attempt)
                log.warning("rate_limit_hit", attempt=attempt + 1, wait_seconds=wait)
                time.sleep(wait)
        else:
            raise RuntimeError(f"Embedding failed after 6 attempts (batch {i})")
        if i + EMBEDDING_BATCH_SIZE < len(texts):
            time.sleep(2)
    return all_embeds


# Upload chunks with embeddings to the specified index.
def upload_chunks_to_index(
    chunks:     list[dict],
    embeddings: list[list[float]],
    index_name: str,
) -> int:
    """Upload chunks with embeddings to the specified index."""
    client    = get_search_client(index_name)
    documents = [{**chunk, "embedding": emb} for chunk, emb in zip(chunks, embeddings)]
    total_uploaded = 0
    for i in range(0, len(documents), UPLOAD_BATCH_SIZE):
        batch  = documents[i:i + UPLOAD_BATCH_SIZE]
        result = client.upload_documents(documents=batch)
        succeeded = sum(1 for r in result if r.succeeded)
        total_uploaded += succeeded
        log.info("upload_batch_done",
                 uploaded=total_uploaded, total=len(documents), index=index_name)
    log.info("upload_complete", total=total_uploaded, index=index_name)
    return total_uploaded


# Chunk -> HQA augment -> embed -> upload to BOTH indexes automatically.
def index_pages_dual(
    pages:            list[dict],
    freshness_run_id: str            = "",
    dry_run:          bool           = False,
    refresh_counts:   dict | None    = None,
) -> tuple[int, int]:
    """
    Chunk -> HQA augment -> embed -> upload to BOTH indexes automatically.

    Strategy (scrape once, embed twice, index twice):
      1. Chunk all pages once, stamping refresh_count per URL
      2. Run full HQA augmentation once
      3. Build HQA embedding texts -> embed -> upload to main index
      4. Copy chunks, clear HQA fields -> baseline texts -> embed -> upload to baseline

    refresh_counts: dict mapping normalised base URL -> new refresh_count value.
    Built by run_freshness_job() Step 8: reads existing refresh_count from index
    before deletion, adds 1, passes here so new chunks carry correct value.
      refresh_count = 0  (set by full indexer run — never touched by freshness)
      refresh_count = 1  (changed once, re-indexed by freshness for first time)
      refresh_count = N  (changed and re-indexed N times by freshness)
    Defaults to 1 if URL not found in dict (first freshness re-index of a page).

    Returns (chunks_added_main, chunks_added_baseline).
    """
    if not pages:
        return 0, 0

    indexed_at     = datetime.now(timezone.utc).isoformat()
    refresh_counts = refresh_counts or {}

    # Step 1: Chunk all pages — stamp refresh_count per URL
    all_chunks: list[dict] = []
    for page in pages:
        url_norm     = normalise_url(get_base_url(page.get("url", "")))
        page_refresh = refresh_counts.get(url_norm, 1)
        all_chunks.extend(
            chunk_page(
                page,
                freshness_run_id=freshness_run_id,
                indexed_at=indexed_at,
                refresh_count=page_refresh,
            )
        )
    if not all_chunks:
        log.warning("no_chunks_produced", page_count=len(pages))
        return 0, 0

    print(f"\n   Chunked {len(all_chunks):,} chunks from {len(pages):,} pages")

    if dry_run:
        log.info("dry_run_dual_index", chunks=len(all_chunks),
                 main=INDEX_NAME, baseline=BASELINE_INDEX_NAME)
        return len(all_chunks), len(all_chunks)

    # Step 2: Full HQA augmentation
    hqa_chunks = augment_chunks_with_hqa(all_chunks)

    # Step 3: Main index — HQA embeddings
    print(f"\n📝 Building HQA embedding texts for main index...")
    hqa_texts  = build_embedding_texts(hqa_chunks)
    print(f"🔢 Generating {len(hqa_texts):,} embeddings for main index...")
    hqa_embeds = get_embeddings_batch(hqa_texts)
    print(f"📤 Uploading to '{INDEX_NAME}'...")
    main_count = upload_chunks_to_index(hqa_chunks, hqa_embeds, INDEX_NAME)
    print(f"   ✅ {main_count:,} chunks -> main index")

    # Step 4: Baseline index — content-only embeddings (no HQA)
    baseline_chunks = copy.deepcopy(hqa_chunks)
    for chunk in baseline_chunks:
        chunk["augmented_questions"] = ""
        chunk["title_questions"]     = ""

    print(f"\n📝 Building content-only embedding texts for baseline index...")
    base_texts  = build_embedding_texts(baseline_chunks)
    print(f"🔢 Generating {len(base_texts):,} embeddings for baseline index...")
    base_embeds = get_embeddings_batch(base_texts)
    print(f"📤 Uploading to '{BASELINE_INDEX_NAME}'...")
    base_count  = upload_chunks_to_index(baseline_chunks, base_embeds, BASELINE_INDEX_NAME)
    print(f"   ✅ {base_count:,} chunks -> baseline index")

    return main_count, base_count


# ══════════════════════════════════════════════════════════════
# REPORT GENERATION
# ══════════════════════════════════════════════════════════════

# Build the colour-coded Excel report summarising this run's scan results.
def build_report(
    scan_results: list[dict],
    run_summary:  dict,
    output_path:  Path,
) -> Path:
    if not _OPENPYXL_AVAILABLE:
        log.warning("openpyxl_unavailable", note="Report skipped.")
        return output_path

    C = {
        "unchanged":    "D4EDDA",
        "changed":      "FFF3CD",
        "new":          "CCE5FF",
        "removed":      "F8D7DA",
        "int_redirect": "FFE0B2",
        "ext_redirect": "E0E0E0",
        "header":       "2C3E50",
    }
    ACTION_COLOUR = {
        "unchanged":             C["unchanged"],
        "new":                   C["new"],
        "changed":               C["changed"],
        "removed_404":           C["removed"],
        "removed_5xx":           C["removed"],
        "removed_delisted":      C["removed"],
        "removed_int_redir":     C["int_redirect"],
        "removed_ext_redir":     C["ext_redirect"],
        "scrape_failed":         C["removed"],
        "pending_content_check": C["changed"],
    }
    HEADERS = [
        "URL", "Title", "Category", "Dropdown?",
        "HTTP Status", "Action",
        "Main Chunks Before", "Main Chunks After",
        "Baseline Chunks Before", "Baseline Chunks After",
        "Cache Keys Invalidated", "Notes",
    ]

    wb  = Workbook()
    ws1 = wb.active
    ws1.title = "Full Results"
    ws1.append(HEADERS)
    _style_header_row(ws1, 1, C["header"])
    for r in scan_results:
        colour = ACTION_COLOUR.get(r.get("action", ""), "FFFFFF")
        ws1.append([
            r.get("url", ""),
            r.get("title", ""),
            r.get("category", ""),
            "Yes" if r.get("is_dropdown") else "No",
            r.get("status_code", ""),
            r.get("action", ""),
            r.get("chunks_before_main", 0),
            r.get("chunks_after_main", 0),
            r.get("chunks_before_baseline", 0),
            r.get("chunks_after_baseline", 0),
            r.get("cache_invalidated", 0),
            r.get("notes", ""),
        ])
        _fill_row(ws1, ws1.max_row, colour)
    _auto_width(ws1)

    ws2 = wb.create_sheet("Action Required")
    ws2.append(HEADERS)
    _style_header_row(ws2, 1, C["header"])
    for r in scan_results:
        if r.get("action", "unchanged") != "unchanged":
            ws2.append([
                r.get("url", ""), r.get("title", ""), r.get("category", ""),
                "Yes" if r.get("is_dropdown") else "No",
                r.get("status_code", ""), r.get("action", ""),
                r.get("chunks_before_main", 0), r.get("chunks_after_main", 0),
                r.get("chunks_before_baseline", 0), r.get("chunks_after_baseline", 0),
                r.get("cache_invalidated", 0), r.get("notes", ""),
            ])
    _auto_width(ws2)

    ws3 = wb.create_sheet("Summary")
    for row_data in [
        ["Digital Assistance Content Freshness Report", ""],
        ["Run At (UTC)",           run_summary.get("run_at", "")],
        ["Mode",                   run_summary.get("mode", "").upper()],
        ["Main Index",             INDEX_NAME],
        ["Baseline Index",         BASELINE_INDEX_NAME],
        ["Freshness Version",      FRESHNESS_JOB_VERSION],
        ["Pipeline Version",       PIPELINE_VERSION],
        ["Scraper Version",        SCRAPER_VERSION],
        ["Metadata Version",       METADATA_VERSION],
        ["Run ID",                 run_summary.get("freshness_run_id", "")],
        ["", ""],
        ["Total Approved URLs",    run_summary.get("total_approved", 0)],
        ["Unchanged",              run_summary.get("live_unchanged", 0)],
        ["Changed (re-indexed)",   run_summary.get("changed", 0)],
        ["New (indexed)",          run_summary.get("new", 0)],
        ["Dead (404)",             run_summary.get("dead_404", 0)],
        ["Dead (5xx)",             run_summary.get("dead_5xx", 0)],
        ["Internal redirect",      run_summary.get("internal_redirect", 0)],
        ["External redirect",      run_summary.get("external_redirect", 0)],
        ["De-listed",              run_summary.get("delisted", 0)],
        ["Scrape failed",          run_summary.get("scrape_failed", 0)],
        ["", ""],
        ["Main index chunks added",        run_summary.get("chunks_added_main", 0)],
        ["Main index chunks deleted",      run_summary.get("chunks_deleted_main", 0)],
        ["Baseline index chunks added",    run_summary.get("chunks_added_baseline", 0)],
        ["Baseline index chunks deleted",  run_summary.get("chunks_deleted_baseline", 0)],
        ["Cache keys invalidated",         run_summary.get("cache_invalidated", 0)],
        ["", ""],
        ["POLICY NOTES", ""],
        ["Dual-index auto mode", "Both indexes updated automatically on every apply run."],
        ["HQA consistency",      "Full HQA generated for all delta pages — no quality mismatch."],
        ["Internal redirects",   "Treated as removed — new URL must be added to Excel."],
        ["De-listed URLs",       "Removed from both indexes — not in approved Excel."],
        ["Dropdown variants",    "All #state= URLs deleted when base URL changes/removed."],
    ]:
        ws3.append(row_data)
    _auto_width(ws3)

    ws4 = wb.create_sheet("Removed from Index")
    ws4.append(["URL", "Title", "Reason", "Main Chunks Deleted", "Baseline Chunks Deleted", "Cache Keys"])
    _style_header_row(ws4, 1, C["header"])
    removed_actions = {
        "removed_404", "removed_5xx", "removed_delisted",
        "removed_int_redir", "removed_ext_redir", "scrape_failed",
    }
    for r in scan_results:
        if r.get("action") in removed_actions:
            ws4.append([
                r.get("url", ""), r.get("title", ""), r.get("notes", ""),
                r.get("chunks_before_main", 0), r.get("chunks_before_baseline", 0),
                r.get("cache_invalidated", 0),
            ])
            _fill_row(ws4, ws4.max_row, C["removed"])
    _auto_width(ws4)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    log.info("report_saved", path=str(output_path))
    return output_path


# Bold white-on-colour styling for a header row.
def _style_header_row(ws, row_num: int, bg_hex: str):
    fill = PatternFill("solid", fgColor=bg_hex)
    font = Font(bold=True, color="FFFFFF")
    for cell in ws[row_num]:
        cell.fill      = fill
        cell.font      = font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


# Solid background fill for a data row (status colour-coding).
def _fill_row(ws, row_num: int, bg_hex: str):
    fill = PatternFill("solid", fgColor=bg_hex)
    for cell in ws[row_num]:
        cell.fill = fill


# Auto-size each column to fit its longest cell value (capped at max_width).
def _auto_width(ws, max_width: int = 70):
    for col in ws.columns:
        max_len    = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                # A single unmeasurable cell shouldn't break column sizing.
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, max_width)


# ══════════════════════════════════════════════════════════════
# MAIN ORCHESTRATOR
# ══════════════════════════════════════════════════════════════

# Full orchestration: scan -> classify -> act on BOTH indexes or report.
def run_freshness_job(
    mode:      str        = "report",
    file_path: str | None = None,
    blob_name: str | None = None,
    dry_run:   bool       = False,
) -> dict:
    """
    Full orchestration: scan -> classify -> act on BOTH indexes or report.

    In apply mode, BOTH rlg-faq-index-v5 (HQA) and
    rlg-faq-index-v5-baseline (no HQA) are updated automatically.
    No flags required for dual-index operation.
    """
    run_at           = datetime.now(timezone.utc)
    ts_str           = run_at.strftime("%Y%m%d_%H%M%S")
    freshness_run_id = str(uuid.uuid4())

    result = {
        "success":                 False,
        "mode":                    mode,
        "run_at":                  run_at.isoformat(),
        "freshness_job_version":   FRESHNESS_JOB_VERSION,
        "freshness_run_id":        freshness_run_id,
        "pipeline_version":        PIPELINE_VERSION,
        "scraper_version":         SCRAPER_VERSION,
        "metadata_version":        METADATA_VERSION,
        "main_index":              INDEX_NAME,
        "baseline_index":          BASELINE_INDEX_NAME,
        "total_approved":          0,
        "live_unchanged":          0,
        "changed":                 0,
        "new":                     0,
        "dead_404":                0,
        "dead_5xx":                0,
        "internal_redirect":       0,
        "external_redirect":       0,
        "delisted":                0,
        "scrape_failed":           0,
        "chunks_added_main":       0,
        "chunks_deleted_main":     0,
        "chunks_added_baseline":   0,
        "chunks_deleted_baseline": 0,
        "cache_invalidated":       0,
        "output_report":           "",
        "error":                   "",
    }

    print("\n" + "=" * 70)
    print(f"   Digital Assistance Content Freshness Manager — {mode.upper()}")
    print(f"   Run at:          {run_at.strftime('%Y-%m-%d %H:%M:%S UTC')}")
    print(f"   Run ID:          {freshness_run_id}")
    print(f"   Main index:      {INDEX_NAME}")
    print(f"   Baseline index:  {BASELINE_INDEX_NAME}")
    print(f"   Versions:        scraper={SCRAPER_VERSION} metadata={METADATA_VERSION} "
          f"pipeline={PIPELINE_VERSION} freshness={FRESHNESS_JOB_VERSION}")
    if dry_run:
        print("   ⚠️  DRY RUN — no index or cache writes")
    print("=" * 70 + "\n")

    try:
        # ── Step 1: Load approved URLs ─────────────────────────
        print("📋 Step 1: Loading approved URLs from Excel...")
        if file_path:
            entries = load_urls_from_excel_file(file_path)
            print(f"   {len(entries):,} URLs from local file: {file_path}")
        else:
            excel_bytes = load_approved_excel_from_blob(blob_name)
            entries     = load_urls_from_excel_bytes(excel_bytes)
            print(f"   {len(entries):,} URLs from Blob: {blob_name or BLOB_APPROVED_EXCEL_NAME}")

        if not entries:
            raise ValueError("No URLs found in approved Excel — aborting.")

        result["total_approved"] = len(entries)
        approved_norm_base = {normalise_url(get_base_url(e["url"])) for e in entries}

        # ── Step 2: Load previous hash state ──────────────────
        print("\n💾 Step 2: Loading previous hash state...")
        hash_state = load_hash_state()
        print(f"   {len(hash_state):,} URLs in hash state.")

        # ── Step 3: Fetch hashes from MAIN index only ─────────
        # Main index is source of truth for hash comparison.
        print(f"\n🔍 Step 3: Reading content_hash from '{INDEX_NAME}'...")
        index_hashes = fetch_current_hashes_from_index(INDEX_NAME)
        print(f"   {len(index_hashes):,} indexed URLs found.")
        if not index_hashes:
            print("   ⚠️  No hashes returned — check content_hash field is "
                  "retrievable=True (requires chunk_and_index_hqaV5.py v5.10.0+ schema).")

        # ── Step 4: Health check all URLs ─────────────────────
        print(f"\n🌐 Step 4: Health-checking {len(entries):,} URLs...")
        health_results = asyncio.run(check_all_urls_health(entries))
        health_by_norm = {normalise_url(r["url"]): r for r in health_results}
        live_count     = sum(1 for r in health_results if r["status"] == "live")
        print(f"   {live_count:,} live / {len(entries) - live_count:,} not-live.")

        # ── Step 5: Detect de-listed URLs ─────────────────────
        print("\n📊 Step 5: Detecting de-listed URLs...")
        all_indexed_norms = set(index_hashes.keys())
        delisted_norm = {
            normalise_url(get_base_url(u))
            for u in all_indexed_norms
            if normalise_url(get_base_url(u)) not in approved_norm_base
            and not is_dropdown_url(u)
        }
        print(f"   {len(delisted_norm):,} de-listed base URLs.")

        # ── Step 6: Classify each URL ─────────────────────────
        print("\n🗂️  Step 6: Classifying actions...")
        scan_results:        list[dict] = []
        urls_to_scrape:      list[dict] = []
        base_urls_to_delete: list[str]  = []

        for entry in entries:
            norm   = normalise_url(entry["url"])
            health = health_by_norm.get(norm, {"status": "unknown", "status_code": None})
            status = health.get("status", "unknown")

            base = {
                "url":                    entry["url"],
                "title":                  entry.get("title", ""),
                "category":               entry.get("category", ""),
                "is_dropdown":            is_dropdown_url(entry["url"]),
                "status_code":            health.get("status_code"),
                "chunks_before_main":     0,
                "chunks_after_main":      0,
                "chunks_before_baseline": 0,
                "chunks_after_baseline":  0,
                "cache_invalidated":      0,
                "notes":                  "",
            }

            if status in ("dead_404", "dead_5xx"):
                short = status.replace("dead_", "")
                base.update({
                    "action": f"removed_{short}",
                    "notes":  f"HTTP {health.get('status_code','N/A')} — removing from both indexes.",
                })
                base_urls_to_delete.append(get_base_url(entry["url"]))
                result[status] += 1

            elif status == "internal_redirect":
                base.update({
                    "action": "removed_int_redir",
                    "notes":  health.get("redirect_note", "Internal redirect — not followed."),
                })
                base_urls_to_delete.append(get_base_url(entry["url"]))
                result["internal_redirect"] += 1

            elif status == "external_redirect":
                base.update({
                    "action": "removed_ext_redir",
                    "notes":  health.get("redirect_note", "External redirect — removing."),
                })
                base_urls_to_delete.append(get_base_url(entry["url"]))
                result["external_redirect"] += 1

            elif status == "live":
                stored_hash = index_hashes.get(norm) or hash_state.get(norm, "")
                if norm not in index_hashes:
                    base.update({"action": "new", "notes": "New URL — scraped and indexed into both indexes."})
                    urls_to_scrape.append(entry)
                    result["new"] += 1
                else:
                    base.update({
                        "action":       "pending_content_check",
                        "notes":        "Live — content hash check pending.",
                        "_stored_hash": stored_hash,
                    })
                    urls_to_scrape.append(entry)
            else:
                base.update({
                    "action": "unchanged",
                    "notes":  f"Health: {status} — skipped.",
                })
                result["live_unchanged"] += 1

            scan_results.append(base)

        for norm_base in delisted_norm:
            scan_results.append({
                "url": norm_base, "title": "", "category": "",
                "is_dropdown": False, "status_code": "N/A",
                "action": "removed_delisted",
                "chunks_before_main": 0, "chunks_after_main": 0,
                "chunks_before_baseline": 0, "chunks_after_baseline": 0,
                "cache_invalidated": 0,
                "notes": "Removed from approved Excel — de-indexing both indexes.",
            })
            base_urls_to_delete.append(norm_base)
            result["delisted"] += 1

        # ── Step 7: Scrape pending URLs ────────────────────────
        scraped_pages:     list[dict] = []
        changed_base_urls: list[str]  = []

        if urls_to_scrape:
            print(f"\n🕷️  Step 7: Scraping {len(urls_to_scrape):,} URLs...")
            try:
                scrape_results = asyncio.run(
                    scrape_urls_batch(urls_to_scrape, freshness_run_id=freshness_run_id)
                )
            finally:
                _cf_stop_chrome_cdp()

            for entry, pages in zip(urls_to_scrape, scrape_results):
                norm   = normalise_url(entry["url"])
                record = next(
                    (r for r in scan_results if normalise_url(r["url"]) == norm), None
                )
                if pages is None:
                    if record:
                        record["action"] = "scrape_failed"
                        record["notes"]  = "Scrape returned no content — skipped."
                    result["scrape_failed"] += 1
                    continue

                base_page   = pages[0]
                live_hash   = base_page["content_hash"]
                stored_hash = (record or {}).get("_stored_hash", "")

                if stored_hash and live_hash == stored_hash:
                    if record:
                        record["action"] = "unchanged"
                        record["notes"]  = "Content hash matches — no change."
                    result["live_unchanged"] += 1
                else:
                    if record and record.get("action") == "pending_content_check":
                        record["action"] = "changed"
                        record["notes"]  = "Content hash mismatch — re-indexing both indexes."
                        result["changed"] += 1
                    scraped_pages.extend(pages)
                    if norm in index_hashes:
                        changed_base_urls.append(get_base_url(entry["url"]))

            print(f"   Scraped {len(urls_to_scrape)} | "
                  f"Changed/New: {len(scraped_pages)} pages | "
                  f"Unchanged: {result['live_unchanged']}")
        else:
            print("\n   Step 7: No URLs to scrape.")

        # ── Step 7.5: PRE-FLIGHT CHUNKING VALIDATION (v1.7.0) ──
        # Stricter checkpoint — chunk every scraped page and sanity-
        # check the output BEFORE Step 8 deletes anything from the
        # live index. See validate_chunking_preflight() docstring for
        # the full rationale (delete-before-reindex ordering risk).
        # Runs for BOTH modes (report and apply) so report mode also
        # surfaces chunking problems early, but only ABORTS in apply
        # mode — report mode logs and continues (nothing to protect).
        if scraped_pages:
            print(f"\n🛡️  Step 7.5: Pre-flight chunking validation "
                  f"({len(scraped_pages):,} pages)...")
            preflight = validate_chunking_preflight(scraped_pages)
            print(f"   {preflight['total_pages'] - len(preflight['failed_pages']):,} "
                  f"passed / {len(preflight['failed_pages']):,} failed "
                  f"(ratio={preflight['failure_ratio']:.1%}, "
                  f"threshold={_MAX_VALIDATION_FAILURE_RATIO:.0%})")

            if not preflight["passed"]:
                if mode == "apply":
                    print(f"\n   🛑 ABORTING — chunking validation failure "
                          f"ratio ({preflight['failure_ratio']:.1%}) exceeds "
                          f"threshold ({_MAX_VALIDATION_FAILURE_RATIO:.0%}).")
                    print(f"   NO index writes have occurred. Review failed "
                          f"pages below, fix, and re-run.")
                    for f in preflight["failed_pages"][:20]:
                        print(f"      ✗ {f['url']}: {f['reason']}")
                    result["error"] = (
                        f"Pre-flight chunking validation failed: "
                        f"{len(preflight['failed_pages'])}/"
                        f"{preflight['total_pages']} pages "
                        f"({preflight['failure_ratio']:.1%})"
                    )
                    result["success"] = False
                    return result
                else:
                    print(f"   ⚠️  Report mode — continuing despite failures "
                          f"(no index writes to protect).")
        else:
            print("\n   Step 7.5: No pages to validate (nothing scraped).")

        # ── Step 8: Apply changes to BOTH indexes ─────────────
        all_base_to_delete = list(set(base_urls_to_delete + changed_base_urls))

        if mode == "apply":
            print("\n⚡ Step 8: Applying changes to BOTH indexes...")

            if all_base_to_delete:
                # ── Read refresh_counts BEFORE delete (changed URLs only) ──
                # For each changed URL (not dead/delisted — those go to 0
                # anyway since they won't be re-indexed), read existing
                # refresh_count and build lookup dict for index_pages_dual().
                refresh_counts: dict[str, int] = {}
                for url in changed_base_urls:
                    norm    = normalise_url(get_base_url(url))
                    current = get_refresh_count_for_url(url)
                    refresh_counts[norm] = current + 1
                    log.info("refresh_count_incremented",
                             url=url, old=current, new=current + 1)

                # ── Archive before delete (audit trail + rollback) ────────
                if not dry_run:
                    print(f"   Archiving existing chunks before deletion...")
                    archived = 0
                    for url in all_base_to_delete:
                        archive_main = save_deleted_chunks_to_blob(url, INDEX_NAME, ts_str)
                        archive_base = save_deleted_chunks_to_blob(url, BASELINE_INDEX_NAME, ts_str)
                        if archive_main or archive_base:
                            archived += 1
                    print(f"   Archived {archived:,} URLs to Blob "
                          f"(path: {BLOB_ARCHIVE_PREFIX})")

                # ── Delete from MAIN index ────────────────────────────────
                main_del_urls = get_all_urls_to_delete(all_base_to_delete, INDEX_NAME)
                print(f"   [{INDEX_NAME}] Deleting {len(main_del_urls):,} URLs...")
                del_main       = delete_chunks_for_urls(main_del_urls, INDEX_NAME, dry_run=dry_run)
                total_del_main = sum(del_main.values())
                result["chunks_deleted_main"] += total_del_main
                print(f"   Deleted {total_del_main:,} chunks from main index.")

                # ── Delete from BASELINE index ────────────────────────────
                base_del_urls  = get_all_urls_to_delete(all_base_to_delete, BASELINE_INDEX_NAME)
                print(f"   [{BASELINE_INDEX_NAME}] Deleting {len(base_del_urls):,} URLs...")
                del_baseline   = delete_chunks_for_urls(base_del_urls, BASELINE_INDEX_NAME, dry_run=dry_run)
                total_del_base = sum(del_baseline.values())
                result["chunks_deleted_baseline"] += total_del_base
                print(f"   Deleted {total_del_base:,} chunks from baseline index.")

                # ── Cache invalidation ────────────────────────────────────
                all_del_urls = list(set(main_del_urls + base_del_urls))
                print(f"   Invalidating cache for {len(all_del_urls):,} URLs...")
                cache_count = invalidate_cache_for_urls(all_del_urls, dry_run=dry_run)
                result["cache_invalidated"] += cache_count
                print(f"   Invalidated {cache_count:,} cache keys.")

                # Update scan records with deletion counts
                for url, count in del_main.items():
                    base_norm = normalise_url(get_base_url(url))
                    for r in scan_results:
                        if normalise_url(get_base_url(r["url"])) == base_norm:
                            r["chunks_before_main"] = r.get("chunks_before_main", 0) + count
                for url, count in del_baseline.items():
                    base_norm = normalise_url(get_base_url(url))
                    for r in scan_results:
                        if normalise_url(get_base_url(r["url"])) == base_norm:
                            r["chunks_before_baseline"] = r.get("chunks_before_baseline", 0) + count
            else:
                refresh_counts = {}

            if scraped_pages:
                print(f"\n   Indexing {len(scraped_pages):,} pages into BOTH indexes...")
                main_added, base_added = index_pages_dual(
                    scraped_pages,
                    freshness_run_id=freshness_run_id,
                    dry_run=dry_run,
                    refresh_counts=refresh_counts,
                )
                result["chunks_added_main"]     += main_added
                result["chunks_added_baseline"] += base_added

                # Update scan records with after-chunk counts
                chunk_counts: dict[str, int] = {}
                for page in scraped_pages:
                    base_norm    = normalise_url(get_base_url(page["url"]))
                    page_refresh = refresh_counts.get(base_norm, 1)
                    chunk_counts[base_norm] = chunk_counts.get(base_norm, 0) + len(
                        chunk_page(page,
                                   freshness_run_id=freshness_run_id,
                                   refresh_count=page_refresh)
                    )
                for r in scan_results:
                    base_norm = normalise_url(get_base_url(r["url"]))
                    if base_norm in chunk_counts:
                        r["chunks_after_main"]     = chunk_counts[base_norm]
                        r["chunks_after_baseline"] = chunk_counts[base_norm]

            # Save updated hash state
            if not dry_run:
                new_hash_state = dict(hash_state)
                for page in scraped_pages:
                    new_hash_state[normalise_url(page["url"])] = page["content_hash"]
                for url in all_base_to_delete:
                    new_hash_state.pop(normalise_url(url), None)
                save_hash_state(new_hash_state)
                print(f"   Hash state saved ({len(new_hash_state):,} URLs).")
        else:
            print("\n   Step 8: Report mode — no index writes.")

        # ── Step 9: Generate report ────────────────────────────
        print("\n📊 Step 9: Generating Excel report...")
        report_name = f"freshness_report_{mode}_{ts_str}.xlsx"
        report_path = LOCAL_DATA_DIR / report_name
        build_report(
            scan_results=scan_results,
            run_summary={**result, "mode": mode},
            output_path=report_path,
        )
        result["output_report"] = str(report_path)
        print(f"   Report: {report_path}")

        blob_report = upload_report_to_blob(report_path)
        if blob_report:
            print(f"   Report uploaded: {blob_report}")

        # ── Step 10: Save run manifest ─────────────────────────
        manifest = {
            "freshness_job_version":   FRESHNESS_JOB_VERSION,
            "freshness_run_id":        freshness_run_id,
            "run_at":                  run_at.isoformat(),
            "mode":                    mode,
            "dry_run":                 dry_run,
            "main_index":              INDEX_NAME,
            "baseline_index":          BASELINE_INDEX_NAME,
            "pipeline_version":        PIPELINE_VERSION,
            "scraper_version":         SCRAPER_VERSION,
            "metadata_version":        METADATA_VERSION,
            "total_approved":          result["total_approved"],
            "live_unchanged":          result["live_unchanged"],
            "changed":                 result["changed"],
            "new":                     result["new"],
            "dead_404":                result["dead_404"],
            "dead_5xx":                result["dead_5xx"],
            "internal_redirect":       result["internal_redirect"],
            "external_redirect":       result["external_redirect"],
            "delisted":                result["delisted"],
            "scrape_failed":           result["scrape_failed"],
            "chunks_added_main":       result["chunks_added_main"],
            "chunks_deleted_main":     result["chunks_deleted_main"],
            "chunks_added_baseline":   result["chunks_added_baseline"],
            "chunks_deleted_baseline": result["chunks_deleted_baseline"],
            "cache_invalidated":       result["cache_invalidated"],
        }
        manifest_dest = save_run_manifest(manifest, ts_str)
        if manifest_dest:
            print(f"   Run manifest: {manifest_dest}")

        result["success"] = True

        print("\n" + "=" * 70)
        print("   RUN COMPLETE")
        print("=" * 70)
        print(f"   Total approved URLs  : {result['total_approved']:,}")
        print(f"   Unchanged            : {result['live_unchanged']:,}")
        print(f"   New                  : {result['new']:,}")
        print(f"   Changed              : {result['changed']:,}")
        print(f"   Removed (404/5xx)    : {result['dead_404'] + result['dead_5xx']:,}")
        print(f"   Redirected           : {result['internal_redirect'] + result['external_redirect']:,}")
        print(f"   De-listed            : {result['delisted']:,}")
        print(f"   Scrape failed        : {result['scrape_failed']:,}")
        if mode == "apply":
            print(f"   Main index added     : {result['chunks_added_main']:,}")
            print(f"   Main index deleted   : {result['chunks_deleted_main']:,}")
            print(f"   Baseline added       : {result['chunks_added_baseline']:,}")
            print(f"   Baseline deleted     : {result['chunks_deleted_baseline']:,}")
            print(f"   Cache invalidated    : {result['cache_invalidated']:,}")
        print("=" * 70 + "\n")

    except Exception as e:
        result["error"] = str(e)
        log.error("freshness_job_failed", error=str(e), traceback=traceback.format_exc())
        print(f"\n❌ FATAL ERROR: {e}\n")

    return result


# ══════════════════════════════════════════════════════════════
# CLI ENTRY POINT
# ══════════════════════════════════════════════════════════════

# CLI entry point — parses args and runs the freshness job.
def main():
    parser = argparse.ArgumentParser(
        description=(
            "Digital Assistance Content Freshness Manager\n"
            "Nightly job — detects changed/removed URLs and keeps\n"
            "BOTH Azure AI Search indexes + Redis cache in sync.\n\n"
            "  report mode: Read-only scan + Excel report (safe).\n"
            "  apply mode:  Scan + execute changes on BOTH indexes.\n"
            "               Dual-index is automatic — no flags needed."
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "--mode",
        choices=["report", "apply"],
        default=os.getenv("FRESHNESS_MODE", "report"),
        help="report = scan only (default). apply = execute changes.",
    )
    parser.add_argument(
        "--file",
        default=None,
        help="Local Excel path. Bypasses Blob (local dev / CI).",
    )
    parser.add_argument(
        "--blob-name",
        default=None,
        help="Override Blob Excel path for this run.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Validate config + connectivity. No writes (even in apply mode).",
    )
    args = parser.parse_args()

    if args.mode == "apply" and not args.dry_run:
        missing = []
        if not SEARCH_ENDPOINT:
            missing.append("AZURE_SEARCH_ENDPOINT")
        if not AZURE_OPENAI_ENDPOINT:
            missing.append("AZURE_OPENAI_ENDPOINT")
        if not args.file and not BLOB_STORAGE_CONNECTION:
            missing.append("AZURE_STORAGE_CONNECTION (or use --file for local)")
        if missing:
            print(f"\n❌ Missing env vars for apply mode:\n   {', '.join(missing)}\n")
            sys.exit(1)

    result = run_freshness_job(
        mode=args.mode,
        file_path=args.file,
        blob_name=args.blob_name,
        dry_run=args.dry_run,
    )
    sys.exit(0 if result.get("success") else 1)


if __name__ == "__main__":
    import sys as _sys
    if _sys.platform == "win32":
        import warnings as _warnings
        _warnings.filterwarnings(
            "ignore", message=".*I/O operation on closed pipe.*", category=ResourceWarning,
        )
        _warnings.filterwarnings(
            "ignore", message=".*unclosed transport.*", category=ResourceWarning,
        )
        _orig_unraisablehook = _sys.unraisablehook

        # Suppress the noisy Windows ProactorEventLoop GC warning.
        def _unraisablehook(unraisable):
            if "I/O operation on closed pipe" in str(unraisable.exc_value):
                return
            _orig_unraisablehook(unraisable)

        _sys.unraisablehook = _unraisablehook
    main()