"""
Royal London FAQ — Full Scrape (Customer Approved URLs Only)
================================================================
Builds the knowledge base from scratch using ONLY the customer's
approved URLs (status=200 in their verification Excel).

Customer shared 389 URLs total — 3 are dead (404), so this
scrapes the 384 unique approved URLs.

Pipeline (single pass, no dependency on previous JSON files):
  1. Read approved URLs + titles directly from customer Excel
     (status == 200 only, deduplicated)
  2. Scrape each URL with crawl4ai (main content only,
     nav/header/footer excluded via CSS selector)
  3. Clean content:
       - remove duplicate article copies (crawl4ai sometimes
         scrapes the same page twice)
       - remove breadcrumb navigation
       - remove social share buttons + Twitter links
       - remove Previous/Next Item labels
       - remove footer boilerplate (copyright, "Connect with
         us", "Products and services", etc.)
     NOTE: external (non-royallondon) URL stripping is handled
     separately by chunk_and_index.py's clean_content() at
     chunk time — not duplicated here.
  4. Attach title (from Excel) + section (from URL path)
  5. Save JSON ready for chunk_and_index.py

Output fields per page (matches chunk_and_index.py input format):
  url, title, section, audience, content, scraped_at, content_length

Input:
  scraper/data/royal_london_verification_retried.xlsx

Output:
  scraper/data/royal_london_faq_approved_<timestamp>.json

Usage (local):
    uv run python scraper/scrape_approved_urls.py

Usage (programmatic — DevOps / Function App):
    from scraper.scrape_approved_urls import run_scraper
    result = run_scraper()

═══════════════════════════════════════════════════════════════
CHANGE LOG
═══════════════════════════════════════════════════════════════

v1.0.0 — Initial version
         crawl4ai scraping from customer Excel file.
         Saves output JSON to scraper/data/ locally.

v2.0.0 — June 2026 | Mukesh Kund
         Production readiness: Blob Storage + entry point

         PRODUCTION GAP (pre v2.0.0):
         - Scraper saved output JSON to local disk only.
         - Azure Function App has no persistent local disk.
         - Output JSON must go to Azure Blob Storage so
           chunk_and_index.py can read it from a separate
           Function App invocation.
         - Excel URL source is developer-only. Production
           URL source (Blob JSON, SharePoint, CMS API) to
           be agreed with brand/marketing team and DevOps.

         save_scraped_pages() [NEW]:
         - Abstracts local vs Blob Storage output.
         - Local: saves to scraper/data/ as before.
         - Production: uploads to Azure Blob Storage.
         - Switched by AZURE_STORAGE_CONNECTION env var.
         - Not set → local mode, zero behaviour change.

         load_url_source() [NEW]:
         - Abstracts local Excel vs Blob Storage URL list.
         - Local: reads Excel file as before.
         - Production: reads JSON from Blob Storage.
         - Switched by AZURE_STORAGE_CONNECTION env var.
         - TODO: production URL source format to be agreed
           with brand/marketing team and DevOps.

         run_scraper() [NEW]:
         - Clean entry point for DevOps / Function App.
         - Returns structured result dict with stats.
         - TODO (DevOps): wrap in Function App trigger.

═══════════════════════════════════════════════════════════════
"""

import asyncio
import os
import json
import re
import sys
from pathlib import Path
from datetime import datetime, timezone

import structlog
from openpyxl import load_workbook
from crawl4ai import AsyncWebCrawler, BrowserConfig, CrawlerRunConfig
from crawl4ai.markdown_generation_strategy import DefaultMarkdownGenerator
from crawl4ai.content_filter_strategy import PruningContentFilter

log = structlog.get_logger()

# ── Config ────────────────────────────────────────────
APPROVED_EXCEL = "scraper/data/royal_london_verification_retried.xlsx"
BATCH_SIZE = 5
BATCH_DELAY_SECONDS = 2

# ── Production: Azure Blob Storage ─────────────────────────────
# TODO (DevOps): Set these in Azure Key Vault before go-live.
#
# AZURE_STORAGE_CONNECTION — Blob Storage connection string.
#   Not set (default) → local mode, saves to scraper/data/.
#   Set in production → uploads JSON to Blob Storage instead.
#
# BLOB_CONTAINER_NAME — container name in Blob Storage.
#   DevOps creates this in the Azure Storage Account.
#   Default: "scraper-data"
#
# BLOB_SCRAPED_FILENAME — filename written to Blob Storage.
#   chunk_and_index.py reads this exact filename.
#   Both scripts must agree on the same filename.
#   Default: "royal_london_faq_latest.json"
#
# TODO: Production URL source format TBD with client.
#   Currently reads approved URLs from Excel (dev-only workflow).
#   Production options (agree with brand/marketing + DevOps):
#     Option A: JSON file in Blob Storage (content team uploads)
#     Option B: SharePoint list via Microsoft Graph API
#     Option C: CMS API endpoint from Royal London website team
#   Until decided, Excel path works for local + manual prod runs.
BLOB_STORAGE_CONNECTION = os.getenv("AZURE_STORAGE_CONNECTION", "")
BLOB_CONTAINER_NAME     = os.getenv("BLOB_CONTAINER_NAME", "scraper-data")
BLOB_SCRAPED_FILENAME   = os.getenv(
    "BLOB_SCRAPED_FILENAME",
    "royal_london_faq_latest.json",
)

# Section mapping from first URL path segment
SECTION_MAP = {
    "existing-customers":       "Existing Customers",
    "insurance":                "Insurance",
    "pensions":                 "Pensions",
    "guides-tools":             "Guides and Tools",
    "retirement-planning":      "Retirement Planning",
    "isa":                      "ISA",
    "profitshare":              "ProfitShare",
    "about-us":                 "About Us",
    "find-a-financial-adviser": "Find a Financial Adviser",
    "accessibility":            "Accessibility",
    "informational-pages":      "Information",
}


# ── Step 1: Load approved URLs from customer Excel ────
def load_approved_pages(excel_path: str) -> list[dict]:
    """
    Reads customer Excel, returns list of dicts:
    [{"url": "...", "title": "..."}]

    Only rows with status == 200. Deduplicates by
    normalized URL (keeps first occurrence's title).
    """
    wb = load_workbook(excel_path, read_only=True)
    ws = wb.active

    seen = set()
    pages = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 5:
            continue

        raw_title = str(row[1]).strip() if row[1] else ""
        url       = str(row[2]).strip() if row[2] else ""
        status    = str(row[4]).strip() if row[4] else ""

        if status != "200" or not url.startswith("http"):
            continue

        normalized = url.rstrip("/").split("?")[0].rstrip("/")
        if normalized in seen:
            continue
        seen.add(normalized)

        # Clean title — strip " - Royal London" suffix
        title = raw_title
        for suffix in [
            " - Royal London", " | Royal London",
            "- Royal London", "| Royal London",
        ]:
            if title.endswith(suffix):
                title = title[: -len(suffix)].strip()

        pages.append({"url": normalized, "title": title})

    wb.close()
    return pages


def derive_section(url: str) -> str:
    """Derive section name from first URL path segment."""
    path = url.replace("https://www.royallondon.com", "").strip("/")
    if not path:
        return "General"
    first_segment = path.split("/")[0]
    return SECTION_MAP.get(
        first_segment,
        first_segment.replace("-", " ").title(),
    )


# ── Step 2: Content cleaning ───────────────────────────
def remove_duplicate_content(content: str) -> str:
    """
    Remove exact duplicate of article content.
    crawl4ai sometimes scrapes the page twice.

    Finds H1/H2 headings — if the second one starts after
    40% of content AND has >60% word overlap with the first
    chunk, it's a true duplicate → cut it.
    """
    h1_pattern = re.compile(r'^#{1,2}\s+\S', re.MULTILINE)
    matches = list(h1_pattern.finditer(content))

    if len(matches) < 2:
        return content

    second_pos = matches[1].start()
    content_len = len(content)

    if second_pos < content_len * 0.40:
        return content

    first_chunk = content[:second_pos].strip()
    second_chunk = content[second_pos:].strip()

    first_words = set(first_chunk.split()[:200])
    second_words = set(second_chunk.split()[:200])

    if not first_words:
        return content

    overlap_pct = len(first_words & second_words) / len(first_words)

    if overlap_pct > 0.60:
        log.info(
            "duplicate_removed",
            overlap_pct=round(overlap_pct, 2),
            chars_removed=len(second_chunk),
        )
        return first_chunk

    return content


def clean_content(content: str) -> str:
    """
    Safe content cleaning — removes only pure UI noise.
    No meaningful content removed.

    NOTE: External (non-royallondon) URL stripping is NOT
    done here — chunk_and_index.py handles that separately
    at chunk time.
    """

    # Step 1: Remove duplicate article copy
    content = remove_duplicate_content(content)

    # Step 2: Remove breadcrumb navigation
    # "1. [ Home ](url) >" or last item "5. Section Name"
    content = re.sub(
        r'^\s*\d+\.\s*\[.*?\]\(.*?\)\s*>\s*$',
        '', content, flags=re.MULTILINE,
    )
    content = re.sub(
        r'^\s*\d+\.\s*\[.*?\]\(.*?\)\s*$',
        '', content, flags=re.MULTILINE,
    )
    content = re.sub(
        r'^\s*\d+\.\s+[A-Z][^\n]{3,60}$',
        '', content, flags=re.MULTILINE,
    )

    # Step 3: Remove social share sections
    # "Share" heading + following empty bullets
    content = re.sub(
        r'Share\s*\n(\s*\*\s*(\[?\s*\]?\([^\)]*\))?\s*\n)+',
        '', content,
    )
    content = re.sub(
        r'^\s*\*\s*\[?\s*\]?\(\s*[^\)]{0,10}\)\s*$',
        '', content, flags=re.MULTILINE,
    )
    content = re.sub(
        r'^Share\s*$',
        '', content, flags=re.MULTILINE,
    )

    # Step 4: Remove Twitter share links
    content = re.sub(
        r'\[?\s*\]?\(https://twitter\.com/intent/tweet[^\)]*\)\s*',
        '', content,
    )

    # Step 5: Remove other social media URLs
    content = re.sub(
        r'https://www\.(facebook|instagram|linkedin|x|youtube|twitter)\.com/\S+',
        '', content,
    )

    # Step 6: Remove empty markdown links
    content = re.sub(r'\[\s*\]\(\s*\)', '', content)
    content = re.sub(
        r'^\s*\*\s*\[\s*\]\s*$', '', content, flags=re.MULTILINE,
    )

    # Step 7: Remove Previous/Next Item labels
    content = re.sub(
        r'^(Previous Item|Next Item)\s*$',
        '', content, flags=re.MULTILINE | re.IGNORECASE,
    )

    # Step 8: Remove footer boilerplate
    content = re.sub(
        r'Your browser is not supported\..*?×\s*',
        '', content, flags=re.DOTALL,
    )
    content = re.sub(
        r'#{1,3}\s*Connect with us.*$',
        '', content, flags=re.DOTALL | re.MULTILINE,
    )
    content = re.sub(
        r'#{1,3}\s*Products and services.*$',
        '', content, flags=re.DOTALL | re.MULTILINE,
    )
    content = re.sub(
        r'#{1,3}\s*About Royal London.*$',
        '', content, flags=re.DOTALL | re.MULTILINE,
    )
    content = re.sub(
        r'#{1,3}\s*Useful links.*$',
        '', content, flags=re.DOTALL | re.MULTILINE,
    )
    content = re.sub(
        r'\*\*The Royal London Mutual Insurance.*$',
        '', content, flags=re.DOTALL | re.MULTILINE,
    )
    content = re.sub(
        r'©\s*Royal London \d{4}.*$',
        '', content, flags=re.DOTALL | re.MULTILINE,
    )
    content = re.sub(r'\[Back to top\].*?\n', '', content)

    # Step 9: Normalize whitespace
    content = re.sub(r'\n{3,}', '\n\n', content)
    content = re.sub(r'[ \t]+\n', '\n', content)
    content = re.sub(r'\n[ \t]+\n', '\n\n', content)

    return content.strip()


# ── Step 3: Scrape a single page ───────────────────────
async def scrape_page(
    crawler: AsyncWebCrawler,
    page_info: dict,
    index: int,
    total: int,
) -> dict | None:
    """Scrape a single page and return cleaned page data."""
    url   = page_info["url"]
    title = page_info["title"]

    log.info("scraping_page", url=url, index=index, total=total)

    try:
        run_config = CrawlerRunConfig(
            css_selector=(
                "main, article, .content, #content, "
                ".page-content, .main-content, [role='main']"
            ),
            markdown_generator=DefaultMarkdownGenerator(
                content_filter=PruningContentFilter(
                    threshold=0.45,
                    threshold_type="fixed",
                ),
                options={
                    "ignore_links": False,
                    "ignore_images": True,
                    "skip_internal_links": True,
                }
            ),
            wait_until="domcontentloaded",
            page_timeout=30000,
            verbose=False,
            excluded_tags=[
                "nav", "header", "footer", "aside",
                "script", "style", "noscript",
            ],
        )

        result = await crawler.arun(url=url, config=run_config)

        if not result.success:
            log.error(
                "scrape_failed",
                url=url,
                error=result.error_message,
            )
            return None

        content = result.markdown.raw_markdown

        if not content or len(content.strip()) < 100:
            log.warning(
                "content_too_short",
                url=url,
                length=len(content or ""),
            )
            return None

        content = clean_content(content)

        if len(content.strip()) < 50:
            log.warning("content_too_short_after_cleaning", url=url)
            return None

        page_data = {
            "url":            url,
            "title":          title,
            "section":        derive_section(url),
            "audience":       "customer",
            "content":        content.strip(),
            "scraped_at":     datetime.now(timezone.utc).isoformat(),
            "content_length": len(content.strip()),
        }

        log.info(
            "scrape_success",
            url=url,
            index=index,
            total=total,
            content_length=page_data["content_length"],
        )
        return page_data

    except Exception as e:
        log.error("scrape_error", url=url, error=str(e))
        return None



# ── Production helpers (v2.0.0) ──────────────────────────────

def load_url_source(excel_path: str) -> list[dict]:
    """
    Load approved URLs from local Excel or Blob Storage.

    Local development (AZURE_STORAGE_CONNECTION not set):
        Reads from Excel file at excel_path — same as v1.0.0.
        No behaviour change for local development workflow.

    Production (AZURE_STORAGE_CONNECTION set):
        Reads from Azure Blob Storage JSON file instead.
        JSON format: [{"url": "...", "title": "..."}, ...]
        DevOps uploads this JSON to Blob Storage; brand/marketing
        team maintains the URL list in agreed format.

    TODO: Production URL source format to be agreed with client.
    Until decided, this falls back to Excel for all environments.
    """
    # ── Production: read from Blob Storage ───────────────────
    # TODO (DevOps): uncomment when AZURE_STORAGE_CONNECTION
    # is configured in Key Vault and URL source format agreed.
    #
    # if BLOB_STORAGE_CONNECTION:
    #     from azure.storage.blob import BlobServiceClient
    #     blob_service = BlobServiceClient.from_connection_string(
    #         BLOB_STORAGE_CONNECTION
    #     )
    #     blob_client = blob_service.get_blob_client(
    #         container=BLOB_CONTAINER_NAME,
    #         blob="approved_urls.json",   # DevOps agrees this name
    #     )
    #     data  = blob_client.download_blob().readall()
    #     pages = json.loads(data)
    #     log.info(
    #         "url_source_loaded_from_blob",
    #         container=BLOB_CONTAINER_NAME,
    #         count=len(pages),
    #     )
    #     return pages   # [{"url": ..., "title": ...}]

    # ── Local development: read from Excel ───────────────────
    # Current behaviour — unchanged from v1.0.0.
    return load_approved_pages(excel_path)


def save_scraped_pages(
    results: list[dict],
    output_file: "Path",
) -> str:
    """
    Save scraped pages to local file or Azure Blob Storage.

    Local development (AZURE_STORAGE_CONNECTION not set):
        Saves JSON to scraper/data/ as before.
        Returns local file path string.

    Production (AZURE_STORAGE_CONNECTION set):
        Uploads JSON to Azure Blob Storage.
        chunk_and_index.py reads from the same Blob container.
        Returns blob filename string.

    Args:
        results:     List of scraped page dicts
        output_file: Local Path object (used in local mode,
                     filename used as blob name in production)

    Returns:
        str — local file path or blob filename
    """
    # ── Production: upload to Blob Storage ───────────────────
    # TODO (DevOps): uncomment when AZURE_STORAGE_CONNECTION
    # is configured in Key Vault.
    #
    # if BLOB_STORAGE_CONNECTION:
    #     from azure.storage.blob import BlobServiceClient
    #     blob_service = BlobServiceClient.from_connection_string(
    #         BLOB_STORAGE_CONNECTION
    #     )
    #     blob_name   = BLOB_SCRAPED_FILENAME
    #     blob_client = blob_service.get_blob_client(
    #         container=BLOB_CONTAINER_NAME,
    #         blob=blob_name,
    #     )
    #     data = json.dumps(results, ensure_ascii=False, indent=2)
    #     blob_client.upload_blob(data, overwrite=True)
    #     log.info(
    #         "scraped_pages_saved_to_blob",
    #         container=BLOB_CONTAINER_NAME,
    #         blob=blob_name,
    #         pages=len(results),
    #     )
    #     print(f"   ✅ Uploaded to Blob Storage: {blob_name}")
    #     return blob_name

    # ── Local development: save to local file ────────────────
    output_file.parent.mkdir(parents=True, exist_ok=True)
    with open(output_file, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    log.info(
        "scraped_pages_saved_locally",
        file=str(output_file),
        pages=len(results),
    )
    return str(output_file)


def run_scraper(
    excel_path: str | None = None,
) -> dict:
    """
    Programmatic entry point for the scraping pipeline.
    Called by DevOps / Azure Function App trigger.

    Runs the full scraping pipeline synchronously by wrapping
    the async main() logic. Returns structured result dict.

    Args:
        excel_path: Path to approved URLs Excel file.
                    If None, uses APPROVED_EXCEL constant.

    Returns:
        dict with keys:
            success        (bool) — True if completed without error
            pages_scraped  (int)  — number of pages scraped
            pages_failed   (int)  — number of pages that failed
            output_path    (str)  — local path or blob filename
            error          (str)  — error message if success=False

    TODO (DevOps): Wrap in Azure Function App trigger:

        import azure.functions as func
        from scraper.scrape_approved_urls import run_scraper

        app = func.FunctionApp()

        # Monthly scheduled scrape (1st of month, 11pm —
        # runs before chunk_and_index at midnight)
        @app.timer_trigger(
            schedule="0 23 1 * *",
            arg_name="timer",
        )
        def monthly_scrape(timer: func.TimerRequest):
            result = run_scraper()
            logging.info(f"Scrape complete: {result}")
            # chunk_and_index Function App triggered separately
            # or chained here via Service Bus / Event Grid

        # On-demand HTTP trigger
        @app.route(route="scrape", methods=["POST"])
        def on_demand_scrape(req: func.HttpRequest):
            result = run_scraper()
            return func.HttpResponse(
                json.dumps(result),
                mimetype="application/json",
            )
    """
    import traceback

    result = {
        "success":       False,
        "pages_scraped": 0,
        "pages_failed":  0,
        "output_path":   "",
        "error":         "",
    }

    try:
        import asyncio

        excel = excel_path or APPROVED_EXCEL

        async def _run():
            pages_to_scrape = load_url_source(excel)
            browser_config  = BrowserConfig(
                browser_type="chromium",
                headless=True,
                verbose=False,
            )
            scraped      = []
            failed_urls  = []
            total        = len(pages_to_scrape)

            async with AsyncWebCrawler(config=browser_config) as crawler:
                for batch_start in range(0, total, BATCH_SIZE):
                    batch  = pages_to_scrape[
                        batch_start:batch_start + BATCH_SIZE
                    ]
                    tasks  = [
                        scrape_page(
                            crawler, page_info,
                            batch_start + i + 1, total,
                        )
                        for i, page_info in enumerate(batch)
                    ]
                    results = await asyncio.gather(*tasks)
                    for page_info, r in zip(batch, results):
                        if r:
                            scraped.append(r)
                        else:
                            failed_urls.append(page_info["url"])
                    if batch_start + BATCH_SIZE < total:
                        await asyncio.sleep(BATCH_DELAY_SECONDS)

            return scraped, failed_urls

        scraped, failed = asyncio.run(_run())

        # Save output
        timestamp   = datetime.now(timezone.utc).strftime(
            "%Y%m%d_%H%M%S"
        )
        output_file = (
            Path("scraper/data") /
            f"royal_london_faq_approved_{timestamp}.json"
        )
        output_path = save_scraped_pages(scraped, output_file)

        result["success"]       = True
        result["pages_scraped"] = len(scraped)
        result["pages_failed"]  = len(failed)
        result["output_path"]   = output_path
        return result

    except Exception as e:
        result["error"] = str(e)
        log.error(
            "scraper_pipeline_error",
            error=str(e),
            traceback=traceback.format_exc(),
        )
        return result

# ── Main ────────────────────────────────────────────────
async def main():
    print("\n" + "=" * 60)
    print("RLG FAQ SCRAPER — Customer Approved URLs Only")
    print("=" * 60)

    # ── Load approved URLs from Excel ──────────────────
    excel_path = Path(APPROVED_EXCEL)
    if not excel_path.exists():
        print(f"\nERROR: Customer Excel not found: {excel_path}")
        sys.exit(1)

    # v2.0.0: load_url_source() abstracts local Excel vs
    # Blob Storage. Local: reads Excel as before. Production:
    # reads JSON from Blob Storage when AZURE_STORAGE_CONNECTION set.
    pages_to_scrape = load_url_source(str(excel_path))
    print(f"\nApproved URLs to scrape: {len(pages_to_scrape)}")

    # ── Scrape in batches ───────────────────────────────
    browser_config = BrowserConfig(
        browser_type="chromium",
        headless=True,
        verbose=False,
    )

    results = []
    failed_urls = []
    total = len(pages_to_scrape)

    async with AsyncWebCrawler(config=browser_config) as crawler:
        for batch_start in range(0, total, BATCH_SIZE):
            batch = pages_to_scrape[batch_start:batch_start + BATCH_SIZE]
            batch_num = batch_start // BATCH_SIZE + 1
            total_batches = (total + BATCH_SIZE - 1) // BATCH_SIZE

            print(f"\nBatch {batch_num}/{total_batches}")

            tasks = [
                scrape_page(crawler, page_info, batch_start + i + 1, total)
                for i, page_info in enumerate(batch)
            ]
            batch_results = await asyncio.gather(*tasks)

            for page_info, result in zip(batch, batch_results):
                if result:
                    results.append(result)
                else:
                    failed_urls.append(page_info["url"])

            if batch_start + BATCH_SIZE < total:
                await asyncio.sleep(BATCH_DELAY_SECONDS)

    # ── Save output ─────────────────────────────────────
    # v2.0.0: save_scraped_pages() abstracts local file vs
    # Blob Storage. Local: saves to scraper/data/ as before.
    # Production: uploads to Azure Blob Storage when
    # AZURE_STORAGE_CONNECTION env var is set.
    timestamp   = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    output_file = (
        Path("scraper/data") /
        f"royal_london_faq_approved_{timestamp}.json"
    )
    output_path = save_scraped_pages(results, output_file)

    # ── Summary ───────────────────────────────────────────
    print("\n" + "=" * 60)
    print("SCRAPE SUMMARY")
    print("=" * 60)
    print(f"Approved URLs:    {total}")
    print(f"Scraped success:  {len(results)}")
    print(f"Failed:           {len(failed_urls)}")

    if failed_urls:
        print("\nFailed URLs:")
        for url in failed_urls:
            print(f"  - {url}")

    if results:
        total_chars = sum(r["content_length"] for r in results)
        lengths = [r["content_length"] for r in results]
        print(f"\nTotal content:    {total_chars:,} chars")
        print(f"Shortest page:    {min(lengths):,} chars")
        print(f"Longest page:     {max(lengths):,} chars")
        print(f"Average page:     {total_chars // len(results):,} chars")

    print(f"\nSaved to: {output_path}")
    print("=" * 60)
    print(f"\nNext: uv run python scraper/chunk_and_index.py --full "
          f"--file {output_path}")


if __name__ == "__main__":
    asyncio.run(main())