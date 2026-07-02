"""
Royal London FAQ — Full Scrape (Customer Approved URLs Only)
================================================================
Builds the knowledge base from scratch using ONLY the customer's
approved URLs (status=200 in their verification Excel).

Customer shared 389 URLs total — 3 are dead (404), so this
scrapes the 384 unique approved URLs.

Pipeline (single pass, no dependency on previous JSON files):
  1. Read approved URLs + titles directly from customer Excel
     (status == 200 only, deduplicated)
  2. Scrape each URL with crawl4ai (main content only,
     nav/header/footer excluded via CSS selector)
  3. Clean content:
       - remove duplicate article copies (crawl4ai sometimes
         scrapes the same page twice)
       - remove breadcrumb navigation
       - remove social share buttons + Twitter links
       - remove Previous/Next Item labels
       - remove footer boilerplate (copyright, "Connect with
         us", "Products and services", etc.)
     NOTE: external (non-royallondon) URL stripping is handled
     separately by chunk_and_index.py's clean_content() at
     chunk time — not duplicated here.
  4. Attach title (from Excel) + section (from URL path)
  5. Save JSON ready for chunk_and_index.py

Output fields per page (matches chunk_and_index.py input format):
  url, title, section, audience, content, scraped_at, content_length,
  has_video, content_type, product_category, description,
  thumbnail_url, publish_date, collection_name, read_time_mins

Input:
  scraper/data/royal_london_verification_retried.xlsx

Output:
  scraper/data/royal_london_faq_approved_<timestamp>.json

Usage (local):
    uv run python scraper/scrape_approved_urls.py

Usage (programmatic — DevOps / Function App):
    from scraper.scrape_approved_urls import run_scraper
    result = run_scraper()

═══════════════════════════════════════════════════════════════
CHANGE LOG
═══════════════════════════════════════════════════════════════

v1.0.0 — Initial version
         crawl4ai scraping from customer Excel file.
         Saves output JSON to scraper/data/ locally.

v2.0.0 — June 2026 | Mukesh Kund
         Production readiness: Blob Storage + entry point

         PRODUCTION GAP (pre v2.0.0):
         - Scraper saved output JSON to local disk only.
         - Azure Function App has no persistent local disk.
         - Output JSON must go to Azure Blob Storage so
           chunk_and_index.py can read it from a separate
           Function App invocation.
         - Excel URL source is developer-only. Production
           URL source (Blob JSON, SharePoint, CMS API) to
           be agreed with brand/marketing team and DevOps.

         save_scraped_pages() [NEW]:
         - Abstracts local vs Blob Storage output.
         - Local: saves to scraper/data/ as before.
         - Production: uploads to Azure Blob Storage.
         - Switched by AZURE_STORAGE_CONNECTION env var.
         - Not set → local mode, zero behaviour change.

         load_url_source() [NEW]:
         - Abstracts local Excel vs Blob Storage URL list.
         - Local: reads Excel file as before.
         - Production: reads JSON from Blob Storage.
         - Switched by AZURE_STORAGE_CONNECTION env var.
         - TODO: production URL source format to be agreed
           with brand/marketing team and DevOps.

         run_scraper() [NEW]:
         - Clean entry point for DevOps / Function App.
         - Returns structured result dict with stats.
         - TODO (DevOps): wrap in Function App trigger.

v3.0.0 — June 2026 | Mukesh Kund
         Rich metadata extraction — video detection, content type,
         product category, audience, publish date, thumbnail

         WHY:
         - Royal London pages contain videos (webinars, guides,
           product explainers) on ANY page type — not just /webinars/.
           Videos appear on pension pages, insurance pages, tools,
           existing customer pages etc.
         - UI/UX team will need metadata to render rich citations
           (video cards, product badges, publish dates, thumbnails)
           without requiring a re-index.
         - Principle: index once, serve many use cases. All metadata
           extracted at scrape time from HTML already fetched by
           crawl4ai — zero extra HTTP calls, zero extra cost.

         VIDEO DETECTION STRATEGY:
         - Royal London uses a proprietary/JS-rendered video player.
           The video embed URL is NOT reliably extractable from HTML.
         - Decision: source_url IS the video reference. UI team links
           to the page — the video is on that page.
         - Detection uses three independent signals (any one = True):
             Signal 1: URL pattern (/webinars/, /videos/, /video/)
             Signal 2: meta-Collection_name contains "webinar"/"video"
             Signal 3: rendered HTML contains video player CSS classes
                       or data attributes (fallback, less reliable)
         - has_video field: True/False stored per page and per chunk.
           UI team uses this to render a "📹 Watch video" indicator.

         METADATA EXTRACTION — extract_page_metadata() [NEW]:
         - Parses result.html (already fetched by crawl4ai, no extra
           HTTP call) using BeautifulSoup to extract:
             has_video        — bool: page contains video content
             content_type     — webinar/guide/article/faq/tool/news
             product_category — pensions/insurance/isa/retirement/general
             audience         — customer/adviser/employer (from URL)
             description      — from meta-description or og:description
             thumbnail_url    — from meta-teaser_image or og:image
             publish_date     — from meta-st-publish-date (ISO format)
             collection_name  — from meta-Collection_name (e.g. "Pension webinar")
             read_time_mins   — estimated from word count (200 wpm)

         OUTPUT FORMAT UPDATED:
         - Old: url, title, section, audience, content, scraped_at, content_length
         - New: + has_video, content_type, product_category, description,
                  thumbnail_url, publish_date, collection_name, read_time_mins
         - All new fields have safe defaults — if extraction fails,
           defaults are used and scraping continues normally.
         - chunk_and_index.py passes all new fields through to index.

         audience FIELD FIX:
         - Was hardcoded to "customer" regardless of URL.
         - Now derived from URL: adviser.royallondon.com → "adviser",
           employer.royallondon.com → "employer", else → "customer".
         - Matches the audience derivation in extract_page_metadata().


v3.1.0 — June 2026 | Mukesh Kund
         Attempted find_dotenv(usecwd=False) fix (INCOMPLETE)

         - Replaced plain load_dotenv() with
           find_dotenv(usecwd=False) + load_dotenv(_dotenv_path)
           so .env loads correctly when the script is run from
           a subdirectory (e.g. scraper/) rather than repo root.
         - NOT DOCUMENTED in this CHANGE LOG at the time — only
           a one-line inline comment was left above the call.
         - INCOMPLETE: the required
           `from dotenv import load_dotenv, find_dotenv` import
           was never added alongside this change. The script
           has been unable to run at all since this change
           landed (NameError on import) until fixed in v4.0.0
           below.

v4.0.0 — July 2026 | Mukesh Kund
         URL path lowercase normalisation + dotenv import fix

         BUG FOUND (carried over from v3.1.0, undocumented then):
         - This file already called find_dotenv(usecwd=False)
           and load_dotenv(_dotenv_path) at module level, but
           NEITHER find_dotenv NOR load_dotenv was imported from
           the dotenv package anywhere in this file. This raised
           NameError: name 'find_dotenv' is not defined the
           moment the module was imported — the script could not
           run at all in this state.
         - FIX: added `from dotenv import load_dotenv, find_dotenv`
           to the import block.
         - Also added override=True to load_dotenv() so a real
           .env value always wins over any stale environment
           variable already set in the shell/process.

         DUPLICATE URL BUG (case sensitivity):
         - load_approved_pages() normalised URLs by stripping
           trailing slash + query string only:
             normalized = url.rstrip("/").split("?")[0].rstrip("/")
           This does NOT lowercase the path, so
           .../should-I-consolidate-my-pensions and
           .../should-i-consolidate-my-pensions survived
           deduplication as two separate "approved" URLs even
           though they are the same page — confirmed live via a
           real duplicate found in the customer's approved-URL
           Excel (case difference only).
         - FIX: new normalize_url() helper lowercases the URL
           PATH only (scheme + domain are already consistently
           lowercase in the source Excel; query strings are
           already stripped before this point). Used in
           load_approved_pages() for the dedup key AND for the
           stored "url" value, so downstream scraping,
           section/audience/product-category derivation, and
           chunk_and_index.py's HQA indexing all see one
           canonical casing per page.
         - Also re-applied in scrape_page() to page_data["url"]
           (the field chunk_and_index.py / the search index
           treats as source_url) as defence-in-depth — in case a
           redirect ever resolves to a differently-cased final
           URL that didn't go through load_approved_pages().

═══════════════════════════════════════════════════════════════
"""

import asyncio
import os
import json
import re
import sys
from pathlib import Path
from datetime import datetime, timezone
from urllib.parse import urlparse, urlunparse

import structlog
from bs4 import BeautifulSoup       # v3.0.0: meta tag + video detection
from openpyxl import load_workbook
from crawl4ai import AsyncWebCrawler, BrowserConfig, CrawlerRunConfig
from crawl4ai.markdown_generation_strategy import DefaultMarkdownGenerator
from crawl4ai.content_filter_strategy import PruningContentFilter
# v4.0.0 FIX: find_dotenv/load_dotenv were called below but never
# imported — this raised NameError on module import, so the
# script could not run at all. See CHANGE LOG v4.0.0.
from dotenv import load_dotenv, find_dotenv

# v3.1.0 fix: load_dotenv() with no args loads from CWD.
# If script is run from scraper/ subfolder, .env is not found
# and all os.getenv() calls silently return hardcoded defaults.
# find_dotenv() walks UP the directory tree until it finds .env
# — works correctly regardless of which directory you run from.
# v4.0.0: override=True added — ensures a real .env value always
# wins over any stale environment variable already set in the
# shell/process. See CHANGE LOG v3.1.0 / v4.0.0 above.
_dotenv_path = find_dotenv(usecwd=False)
load_dotenv(_dotenv_path, override=True)
log = structlog.get_logger()

# ── Config ────────────────────────────────────────────
APPROVED_EXCEL = "scraper/data/royal_london_verification_retried.xlsx"
BATCH_SIZE = 5
BATCH_DELAY_SECONDS = 2

# ── Production: Azure Blob Storage ─────────────────────────────
# TODO (DevOps): Set these in Azure Key Vault before go-live.
#
# AZURE_STORAGE_CONNECTION — Blob Storage connection string.
#   Not set (default) → local mode, saves to scraper/data/.
#   Set in production → uploads JSON to Blob Storage instead.
#
# BLOB_CONTAINER_NAME — container name in Blob Storage.
#   DevOps creates this in the Azure Storage Account.
#   Default: "scraper-data"
#
# BLOB_SCRAPED_FILENAME — filename written to Blob Storage.
#   chunk_and_index.py reads this exact filename.
#   Both scripts must agree on the same filename.
#   Default: "royal_london_faq_latest.json"
#
# TODO: Production URL source format TBD with client.
#   Currently reads approved URLs from Excel (dev-only workflow).
#   Production options (agree with brand/marketing + DevOps):
#     Option A: JSON file in Blob Storage (content team uploads)
#     Option B: SharePoint list via Microsoft Graph API
#     Option C: CMS API endpoint from Royal London website team
#   Until decided, Excel path works for local + manual prod runs.
BLOB_STORAGE_CONNECTION = os.getenv("AZURE_STORAGE_CONNECTION", "")
BLOB_CONTAINER_NAME     = os.getenv("BLOB_CONTAINER_NAME", "scraper-data")
BLOB_SCRAPED_FILENAME   = os.getenv(
    "BLOB_SCRAPED_FILENAME",
    "royal_london_faq_latest.json",
)

# Section mapping from first URL path segment
SECTION_MAP = {
    "existing-customers":       "Existing Customers",
    "insurance":                "Insurance",
    "pensions":                 "Pensions",
    "guides-tools":             "Guides and Tools",
    "retirement-planning":      "Retirement Planning",
    "isa":                      "ISA",
    "profitshare":              "ProfitShare",
    "about-us":                 "About Us",
    "find-a-financial-adviser": "Find a Financial Adviser",
    "accessibility":            "Accessibility",
    "informational-pages":      "Information",
}



# ── Step 0: Metadata extraction (v3.0.0) ────────────────────

# CSS class names and data attributes that Royal London uses
# for their video player components. Detected in rendered HTML
# as a fallback signal when meta tags don't confirm video.
# These were identified from inspecting Royal London page source.
# Update this list if new player classes are discovered.
VIDEO_CSS_SIGNALS = [
    "video-player",
    "webinar-player",
    "brightcove-player",
    "bc-player",
    "vjs-tech",          # Video.js (common enterprise player)
    "kaltura-player",
    "jwplayer",
    "data-video-id",
    "data-webinar-id",
    "data-brightcove",
]

# URL path segments that reliably indicate video content.
# Any segment present in the page URL → has_video = True.
VIDEO_URL_SIGNALS = [
    "/webinars/",
    "/videos/",
    "/video/",
    "/webinar/",
]

# meta-Collection_name values that indicate video pages.
# crawl4ai renders these as text in the HTML head.
# Case-insensitive matching used.
VIDEO_COLLECTION_SIGNALS = [
    "webinar",
    "video",
    "podcast",          # audio/video content
]

# Product category keyword mapping.
# URL path segments → product_category value.
# Order matters — first match wins.
PRODUCT_CATEGORY_MAP = [
    ("/pension",              "pensions"),
    ("/retirement",           "retirement"),
    ("/life-insurance",       "life_insurance"),
    ("/life-cover",           "life_insurance"),
    ("/whole-of-life",        "life_insurance"),
    ("/income-protection",    "income_protection"),
    ("/critical-illness",     "critical_illness"),
    ("/illness-income",       "income_protection"),
    ("/isa",                  "isa"),
    ("/investments",          "investments"),
    ("/investment",           "investments"),
    ("/fund",                 "investments"),
    ("/funeral",              "funeral"),
    ("/profitshare",          "profitshare"),
    ("/financial-adviser",    "financial_advice"),
    ("/find-a-financial",     "financial_advice"),
    ("/about-us",             "corporate"),
    ("/media",                "corporate"),
    ("/existing-customers",   "customer_support"),
]

# Content type mapping from URL path segments.
# Order matters — more specific patterns first.
CONTENT_TYPE_MAP = [
    ("/webinars/",            "webinar"),
    ("/videos/",              "video"),
    ("/video/",               "video"),
    ("/guides-tools/",        "guide"),
    ("/pension-calculator",   "tool"),
    ("/retirement-planner",   "tool"),
    ("/lump-sum-calculator",  "tool"),
    ("/risk-profiler",        "tool"),
    ("/calculator",           "tool"),
    ("/planner",              "tool"),
    ("/existing-customers/",  "faq"),
    ("/help-and-support/",    "faq"),
    ("/pensions-explained",   "faq"),
    ("/about-us/",            "corporate"),
    ("/media/",               "news"),
    ("/press-release",        "news"),
    ("/news/",                "news"),
    ("/agm/",                 "corporate"),
]


def derive_content_type(url: str) -> str:
    """
    Derive content_type from URL path pattern.

    Uses CONTENT_TYPE_MAP — first match wins (most specific first).
    Falls back to "article" if no pattern matches.

    Returns one of:
        webinar, video, guide, tool, faq, corporate, news, article
    """
    url_lower = url.lower()
    for pattern, content_type in CONTENT_TYPE_MAP:
        if pattern in url_lower:
            return content_type
    return "article"


def derive_product_category(url: str) -> str:
    """
    Derive product_category from URL path pattern.

    Uses PRODUCT_CATEGORY_MAP — first match wins.
    Falls back to "general" if no pattern matches.

    Returns one of:
        pensions, retirement, life_insurance, income_protection,
        critical_illness, isa, investments, funeral, profitshare,
        financial_advice, corporate, customer_support, general
    """
    url_lower = url.lower()
    for pattern, category in PRODUCT_CATEGORY_MAP:
        if pattern in url_lower:
            return category
    return "general"


def derive_audience_from_url(url: str) -> str:
    """
    Derive audience from URL domain/path.

    v3.0.0 FIX: was hardcoded to "customer" in scrape_page().
    Now properly derived from URL:
        adviser.royallondon.com  → "adviser"
        employer.royallondon.com → "employer"
        /adviser/ in path        → "adviser"
        /employer/ in path       → "employer"
        everything else          → "customer"
    """
    url_lower = url.lower()
    if "adviser.royallondon.com" in url_lower or "/adviser/" in url_lower:
        return "adviser"
    if "employer.royallondon.com" in url_lower or "/employer/" in url_lower:
        return "employer"
    return "customer"


def detect_video_from_html(html: str, url: str) -> bool:
    """
    Detect whether a page contains video content.

    Uses three independent signals — any one True = has_video:

    Signal 1 — URL pattern (most reliable):
        URL contains /webinars/, /videos/, /video/, /webinar/

    Signal 2 — Meta tag (very reliable):
        meta-Collection_name contains "webinar" or "video"
        Parsed from HTML <head> — always present, not JS-dependent.

    Signal 3 — HTML class/attribute (fallback):
        Rendered HTML contains known video player CSS class names
        or data attributes (VIDEO_CSS_SIGNALS list).
        Less reliable as Royal London's player is JS-rendered and
        may not appear at domcontentloaded — but included as belt-
        and-suspenders.

    Args:
        html: Raw HTML string from crawl4ai result.html
        url:  Page URL for Signal 1 check

    Returns:
        bool — True if any signal indicates video content
    """
    url_lower = url.lower()

    # Signal 1: URL pattern
    for pattern in VIDEO_URL_SIGNALS:
        if pattern in url_lower:
            return True

    # Signal 2 + 3: parse HTML
    if not html:
        return False

    try:
        soup = BeautifulSoup(html, "html.parser")

        # Signal 2: meta-Collection_name tag
        # Royal London sets this in the page head as:
        # <meta name="Collection_name" content="Pension webinar">
        collection_meta = (
            soup.find("meta", attrs={"name": "Collection_name"}) or
            soup.find("meta", attrs={"name": "collection_name"}) or
            soup.find("meta", property="Collection_name")
        )
        if collection_meta:
            collection_val = (
                collection_meta.get("content", "") or ""
            ).lower()
            for signal in VIDEO_COLLECTION_SIGNALS:
                if signal in collection_val:
                    return True

        # Also check og:type = "video" or similar
        og_type = soup.find("meta", property="og:type")
        if og_type and "video" in (og_type.get("content", "") or "").lower():
            return True

        # Signal 3: HTML class/attribute scan
        # Check rendered HTML (not just head) for video player signals
        html_lower = html.lower()
        for signal in VIDEO_CSS_SIGNALS:
            if signal in html_lower:
                return True

    except Exception as e:
        log.warning(
            "video_detection_parse_error",
            url=url,
            error=str(e),
        )

    return False


def extract_page_metadata(html: str, url: str) -> dict:
    """
    Extract rich metadata from page HTML for index enrichment.

    Parses result.html (already fetched by crawl4ai — zero extra
    HTTP calls, zero extra cost) using BeautifulSoup.

    All extractions are defensive — any failure returns safe
    defaults so scraping continues normally.

    Args:
        html: Raw HTML string from crawl4ai result.html
        url:  Page URL for URL-based derivations

    Returns dict with keys:
        has_video        (bool) — page contains video content
        content_type     (str)  — webinar/guide/article/faq/tool/news
        product_category (str)  — pensions/insurance/isa/etc
        audience         (str)  — customer/adviser/employer
        description      (str)  — page description for UI previews
        thumbnail_url    (str)  — teaser/og image URL
        publish_date     (str)  — ISO date string or ""
        collection_name  (str)  — e.g. "Pension webinar"
        read_time_mins   (str)  — estimated reading time

    All string fields default to "" if not found.
    has_video defaults to False if detection fails.
    """
    # Safe defaults — used if any extraction fails
    metadata = {
        "has_video":        False,
        "content_type":     derive_content_type(url),
        "product_category": derive_product_category(url),
        "audience":         derive_audience_from_url(url),
        "description":      "",
        "thumbnail_url":    "",
        "publish_date":     "",
        "collection_name":  "",
        "read_time_mins":   "5",  # default 5 min if calculation fails
    }

    # Video detection — uses URL + HTML signals
    metadata["has_video"] = detect_video_from_html(html, url)

    if not html:
        # No HTML available — return URL-derived defaults
        return metadata

    try:
        soup = BeautifulSoup(html, "html.parser")

        # ── Description ──────────────────────────────────────
        # Priority: meta-description > og:description > st-description
        # UI team uses this for citation preview tooltips.
        for attr, key in [
            ({"name": "description"}, "content"),
            ({"property": "og:description"}, "content"),
            ({"name": "st-description"}, "content"),
        ]:
            tag = soup.find("meta", attrs=attr)
            if tag and tag.get(key, "").strip():
                metadata["description"] = tag[key].strip()[:300]
                break

        # ── Thumbnail URL ─────────────────────────────────────
        # Priority: meta-teaser_image > og:image
        # meta-teaser_image is Royal London's custom teaser image
        # (350x200px, page-specific). og:image is usually the RL logo.
        # UI team can use teaser_image for rich citation cards.
        teaser = soup.find("meta", attrs={"name": "teaser_image"})
        if teaser and teaser.get("content", "").strip():
            metadata["thumbnail_url"] = teaser["content"].strip()
        else:
            og_image = soup.find("meta", property="og:image")
            if og_image and og_image.get("content", "").strip():
                # Only store if it's Royal London hosted (not generic logo)
                img_url = og_image["content"].strip()
                if "rl-logo-meta-image" not in img_url:
                    metadata["thumbnail_url"] = img_url

        # ── Publish date ──────────────────────────────────────
        # Royal London sets meta-st-publish-date as human-readable
        # e.g. "13 March 2024". Parse to ISO format for indexing.
        pub_date_tag = soup.find(
            "meta", attrs={"name": "st-publish-date"}
        )
        if pub_date_tag and pub_date_tag.get("content", "").strip():
            raw_date = pub_date_tag["content"].strip()
            try:
                # Parse "13 March 2024" → "2024-03-13"
                parsed = datetime.strptime(raw_date, "%d %B %Y")
                metadata["publish_date"] = parsed.strftime("%Y-%m-%d")
            except ValueError:
                # If format differs, store raw value
                metadata["publish_date"] = raw_date[:20]

        # ── Collection name ───────────────────────────────────
        # e.g. "Pension webinar", "Life insurance guide"
        # UI team can use this for content category badges.
        collection_tag = soup.find(
            "meta", attrs={"name": "Collection_name"}
        )
        if collection_tag and collection_tag.get("content", "").strip():
            metadata["collection_name"] = (
                collection_tag["content"].strip()[:100]
            )

        # ── Read time estimate ────────────────────────────────
        # Estimate from visible text word count at 200 wpm.
        # For webinar pages: transcript is ~5,000-8,000 words →
        # 25-40 min. For articles: ~500-1,500 words → 2-7 min.
        # UI team can display "8 min read" or "30 min webinar".
        body_text = soup.get_text(separator=" ", strip=True)
        word_count = len(body_text.split())
        read_time = max(1, round(word_count / 200))
        metadata["read_time_mins"] = str(read_time)

    except Exception as e:
        log.warning(
            "metadata_extraction_error",
            url=url,
            error=str(e),
            note="Using safe defaults for this page",
        )

    return metadata


# ── Step 1: Load approved URLs from customer Excel ────



def normalize_url(url: str) -> str:
    """
    Canonicalise a URL for deduplication and index storage.

    v4.0.0 — uses urllib.parse for component-level handling.
    Replaces fragile string manipulation (.split("?")[0] etc)
    with explicit per-component processing.

    WHAT THIS DOES (in order):
    1. Lowercases scheme    — HTTPS:// -> https://
    2. Lowercases netloc    — WWW.RoyalLondon.COM -> www.royallondon.com
    3. Lowercases path      — /Should-I-... -> /should-i-...
    4. Strips trailing slash from path
    5. Drops query string   — ?utm_source=email discarded entirely
    6. Drops fragment       — #section discarded entirely

    WHY query strings are stripped (not kept):
    - All URLs in the approved Excel are canonical content pages.
      Query strings on Royal London URLs are tracking params only
      (utm_source, utm_medium etc) — never meaningful page selectors.
    - Keeping them would treat /pensions and /pensions?utm_source=email
      as two different approved pages: duplicate scrapes, duplicate
      chunks, duplicate HQA questions in the index.
    - Customers must never see tracking-tagged URLs in citation chips.

    WHY urllib.parse instead of .split("?")[0]:
    - .split("?")[0] on "https://example.com/page#section?query"
      returns "https://example.com/page#section" — fragment survives.
      urlparse handles all six components independently and correctly
      regardless of order or edge cases.

    Examples — all normalise to the same value:
        https://www.royallondon.com/Should-I-Consolidate/
        https://www.royallondon.com/should-i-consolidate
        HTTPS://WWW.ROYALLONDON.COM/should-i-consolidate
        https://www.royallondon.com/should-i-consolidate?utm_source=x#top
        -> https://www.royallondon.com/should-i-consolidate
    """
    if not url:
        return url

    try:
        parsed = urlparse(url.strip())
        return urlunparse((
            parsed.scheme.lower(),           # https
            parsed.netloc.lower(),           # www.royallondon.com
            parsed.path.lower().rstrip("/"), # /path/to/page
            "",                              # params  — always empty for RL
            "",                              # query   — stripped (see WHY above)
            "",                              # fragment — stripped
        ))
    except Exception:
        # Fallback for any malformed URL — lowercase everything
        # rather than risk leaving case variants undeduplicated.
        return url.strip().lower()


def load_approved_pages(excel_path: str) -> list[dict]:
    """
    Reads customer Excel, returns list of dicts:
    [{"url": "...", "title": "..."}]

    Only rows with status == 200. Deduplicates by
    normalized URL (keeps first occurrence's title).

    v4.0.0:
    - normalize_url() now handles scheme + domain + path
      lowercasing, query-string stripping, and fragment
      stripping via urllib.parse (production-safe).
    - Logs exact count of duplicates removed — audit trail
      for QA and canary for Excel data quality issues.
    """
    wb = load_workbook(excel_path, read_only=True)
    ws = wb.active

    seen          = set()
    pages         = []
    total_rows    = 0
    skipped_status = 0
    duplicates    = []   # track for audit log

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 5:
            continue

        raw_title = str(row[1]).strip() if row[1] else ""
        url       = str(row[2]).strip() if row[2] else ""
        status    = str(row[4]).strip() if row[4] else ""

        if not url.startswith("http"):
            continue

        total_rows += 1

        if status != "200":
            skipped_status += 1
            continue

        normalized = normalize_url(url)

        if normalized in seen:
            duplicates.append(url)
            log.debug(
                "url_duplicate_skipped",
                raw_url=url,
                normalized=normalized,
            )
            continue

        seen.add(normalized)

        # Clean title — strip " - Royal London" suffix variants
        title = raw_title
        for suffix in [
            " - Royal London", " | Royal London",
            "- Royal London",  "| Royal London",
        ]:
            if title.endswith(suffix):
                title = title[: -len(suffix)].strip()
                break

        pages.append({"url": normalized, "title": title})

    wb.close()

    # Audit log — production canary for Excel data quality
    log.info(
        "approved_pages_loaded",
        total_rows=total_rows,
        skipped_non_200=skipped_status,
        duplicates_removed=len(duplicates),
        unique_pages=len(pages),
    )
    if duplicates:
        log.warning(
            "duplicate_urls_in_excel",
            count=len(duplicates),
            examples=duplicates[:5],  # first 5 only to keep log readable
        )

    return pages


def derive_section(url: str) -> str:
    """Derive section name from first URL path segment."""
    path = url.replace("https://www.royallondon.com", "").strip("/")
    if not path:
        return "General"
    first_segment = path.split("/")[0]
    return SECTION_MAP.get(
        first_segment,
        first_segment.replace("-", " ").title(),
    )


# ── Step 2: Content cleaning ───────────────────────────
def remove_duplicate_content(content: str) -> str:
    """
    Remove exact duplicate of article content.
    crawl4ai sometimes scrapes the page twice.

    Finds H1/H2 headings — if the second one starts after
    40% of content AND has >60% word overlap with the first
    chunk, it's a true duplicate → cut it.
    """
    h1_pattern = re.compile(r'^#{1,2}\s+\S', re.MULTILINE)
    matches = list(h1_pattern.finditer(content))

    if len(matches) < 2:
        return content

    second_pos = matches[1].start()
    content_len = len(content)

    if second_pos < content_len * 0.40:
        return content

    first_chunk = content[:second_pos].strip()
    second_chunk = content[second_pos:].strip()

    first_words = set(first_chunk.split()[:200])
    second_words = set(second_chunk.split()[:200])

    if not first_words:
        return content

    overlap_pct = len(first_words & second_words) / len(first_words)

    if overlap_pct > 0.60:
        log.info(
            "duplicate_removed",
            overlap_pct=round(overlap_pct, 2),
            chars_removed=len(second_chunk),
        )
        return first_chunk

    return content


def clean_content(content: str) -> str:
    """
    Safe content cleaning — removes only pure UI noise.
    No meaningful content removed.

    NOTE: External (non-royallondon) URL stripping is NOT
    done here — chunk_and_index.py handles that separately
    at chunk time.
    """

    # Step 1: Remove duplicate article copy
    content = remove_duplicate_content(content)

    # Step 2: Remove breadcrumb navigation
    # "1. [ Home ](url) >" or last item "5. Section Name"
    content = re.sub(
        r'^\s*\d+\.\s*\[.*?\]\(.*?\)\s*>\s*$',
        '', content, flags=re.MULTILINE,
    )
    content = re.sub(
        r'^\s*\d+\.\s*\[.*?\]\(.*?\)\s*$',
        '', content, flags=re.MULTILINE,
    )
    content = re.sub(
        r'^\s*\d+\.\s+[A-Z][^\n]{3,60}$',
        '', content, flags=re.MULTILINE,
    )

    # Step 3: Remove social share sections
    # "Share" heading + following empty bullets
    content = re.sub(
        r'Share\s*\n(\s*\*\s*(\[?\s*\]?\([^\)]*\))?\s*\n)+',
        '', content,
    )
    content = re.sub(
        r'^\s*\*\s*\[?\s*\]?\(\s*[^\)]{0,10}\)\s*$',
        '', content, flags=re.MULTILINE,
    )
    content = re.sub(
        r'^Share\s*$',
        '', content, flags=re.MULTILINE,
    )

    # Step 4: Remove Twitter share links
    content = re.sub(
        r'\[?\s*\]?\(https://twitter\.com/intent/tweet[^\)]*\)\s*',
        '', content,
    )

    # Step 5: Remove other social media URLs
    content = re.sub(
        r'https://www\.(facebook|instagram|linkedin|x|youtube|twitter)\.com/\S+',
        '', content,
    )

    # Step 6: Remove empty markdown links
    content = re.sub(r'\[\s*\]\(\s*\)', '', content)
    content = re.sub(
        r'^\s*\*\s*\[\s*\]\s*$', '', content, flags=re.MULTILINE,
    )

    # Step 7: Remove Previous/Next Item labels
    content = re.sub(
        r'^(Previous Item|Next Item)\s*$',
        '', content, flags=re.MULTILINE | re.IGNORECASE,
    )

    # Step 8: Remove footer boilerplate
    content = re.sub(
        r'Your browser is not supported\..*?×\s*',
        '', content, flags=re.DOTALL,
    )
    content = re.sub(
        r'#{1,3}\s*Connect with us.*$',
        '', content, flags=re.DOTALL | re.MULTILINE,
    )
    content = re.sub(
        r'#{1,3}\s*Products and services.*$',
        '', content, flags=re.DOTALL | re.MULTILINE,
    )
    content = re.sub(
        r'#{1,3}\s*About Royal London.*$',
        '', content, flags=re.DOTALL | re.MULTILINE,
    )
    content = re.sub(
        r'#{1,3}\s*Useful links.*$',
        '', content, flags=re.DOTALL | re.MULTILINE,
    )
    content = re.sub(
        r'\*\*The Royal London Mutual Insurance.*$',
        '', content, flags=re.DOTALL | re.MULTILINE,
    )
    content = re.sub(
        r'©\s*Royal London \d{4}.*$',
        '', content, flags=re.DOTALL | re.MULTILINE,
    )
    content = re.sub(r'\[Back to top\].*?\n', '', content)

    # Step 9: Normalize whitespace
    content = re.sub(r'\n{3,}', '\n\n', content)
    content = re.sub(r'[ \t]+\n', '\n', content)
    content = re.sub(r'\n[ \t]+\n', '\n\n', content)

    return content.strip()


# ── Step 3: Scrape a single page ───────────────────────
async def scrape_page(
    crawler: AsyncWebCrawler,
    page_info: dict,
    index: int,
    total: int,
) -> dict | None:
    """Scrape a single page and return cleaned page data."""
    url   = page_info["url"]
    title = page_info["title"]

    log.info("scraping_page", url=url, index=index, total=total)

    try:
        run_config = CrawlerRunConfig(
            css_selector=(
                "main, article, .content, #content, "
                ".page-content, .main-content, [role='main']"
            ),
            markdown_generator=DefaultMarkdownGenerator(
                content_filter=PruningContentFilter(
                    threshold=0.45,
                    threshold_type="fixed",
                ),
                options={
                    "ignore_links": False,
                    "ignore_images": True,
                    "skip_internal_links": True,
                }
            ),
            wait_until="domcontentloaded",
            page_timeout=30000,
            verbose=False,
            excluded_tags=[
                "nav", "header", "footer", "aside",
                "script", "style", "noscript",
            ],
        )

        result = await crawler.arun(url=url, config=run_config)

        if not result.success:
            log.error(
                "scrape_failed",
                url=url,
                error=result.error_message,
            )
            return None

        page_content = result.markdown.raw_markdown

        if not page_content or len(page_content.strip()) < 100:
            log.warning(
                "content_too_short",
                url=url,
                length=len(page_content or ""),
            )
            return None

        page_content = clean_content(page_content)

        if len(page_content.strip()) < 50:
            log.warning("content_too_short_after_cleaning", url=url)
            return None

        # v3.0.0 — extract rich metadata from raw HTML.
        # result.html is the full rendered HTML already fetched
        # by crawl4ai — no extra HTTP call needed.
        # Falls back to safe defaults if HTML unavailable.
        raw_html = getattr(result, "html", "") or ""
        metadata = extract_page_metadata(raw_html, url)

        page_data = {
            # ── Core fields (unchanged from v1/v2) ────────────
            # v4.0.0: normalize_url() applied here too as
            # defence-in-depth. If a redirect resolves to a
            # differently-cased or tracking-tagged final URL,
            # scheme + domain + path are still canonicalised
            # and query strings stripped before the record
            # reaches chunk_and_index.py / the search index.
            # See CHANGE LOG v4.0.0.
            "url":            normalize_url(url),
            "title":          title,
            "section":        derive_section(url),
            "content":        page_content.strip(),
            "scraped_at":     datetime.now(timezone.utc).isoformat(),
            "content_length": len(page_content.strip()),

            # ── Enrichment fields (v3.0.0) ────────────────────
            # All extracted from HTML already fetched by crawl4ai.
            # Zero extra HTTP calls. Safe defaults if extraction fails.
            "audience":         metadata["audience"],
            "has_video":        metadata["has_video"],
            "content_type":     metadata["content_type"],
            "product_category": metadata["product_category"],
            "description":      metadata["description"],
            "thumbnail_url":    metadata["thumbnail_url"],
            "publish_date":     metadata["publish_date"],
            "collection_name":  metadata["collection_name"],
            "read_time_mins":   metadata["read_time_mins"],
        }

        log.info(
            "scrape_success",
            url=url,
            index=index,
            total=total,
            content_length=page_data["content_length"],
            has_video=metadata["has_video"],
            content_type=metadata["content_type"],
            product_category=metadata["product_category"],
        )
        return page_data

    except Exception as e:
        log.error("scrape_error", url=url, error=str(e))
        return None



# ── Production helpers (v2.0.0) ──────────────────────────────

def load_url_source(excel_path: str) -> list[dict]:
    """
    Load approved URLs from local Excel or Blob Storage.

    Local development (AZURE_STORAGE_CONNECTION not set):
        Reads from Excel file at excel_path — same as v1.0.0.
        No behaviour change for local development workflow.

    Production (AZURE_STORAGE_CONNECTION set):
        Reads from Azure Blob Storage JSON file instead.
        JSON format: [{"url": "...", "title": "..."}, ...]
        DevOps uploads this JSON to Blob Storage; brand/marketing
        team maintains the URL list in agreed format.

    TODO: Production URL source format to be agreed with client.
    Until decided, this falls back to Excel for all environments.
    """
    # ── Production: read from Blob Storage ───────────────────
    # TODO (DevOps): uncomment when AZURE_STORAGE_CONNECTION
    # is configured in Key Vault and URL source format agreed.
    #
    # if BLOB_STORAGE_CONNECTION:
    #     from azure.storage.blob import BlobServiceClient
    #     blob_service = BlobServiceClient.from_connection_string(
    #         BLOB_STORAGE_CONNECTION
    #     )
    #     blob_client = blob_service.get_blob_client(
    #         container=BLOB_CONTAINER_NAME,
    #         blob="approved_urls.json",   # DevOps agrees this name
    #     )
    #     data  = blob_client.download_blob().readall()
    #     pages = json.loads(data)
    #     log.info(
    #         "url_source_loaded_from_blob",
    #         container=BLOB_CONTAINER_NAME,
    #         count=len(pages),
    #     )
    #     return pages   # [{"url": ..., "title": ...}]

    # ── Local development: read from Excel ───────────────────
    # Current behaviour — unchanged from v1.0.0.
    return load_approved_pages(excel_path)


def save_scraped_pages(
    results: list[dict],
    output_file: "Path",
) -> str:
    """
    Save scraped pages to local file or Azure Blob Storage.

    Local development (AZURE_STORAGE_CONNECTION not set):
        Saves JSON to scraper/data/ as before.
        Returns local file path string.

    Production (AZURE_STORAGE_CONNECTION set):
        Uploads JSON to Azure Blob Storage.
        chunk_and_index.py reads from the same Blob container.
        Returns blob filename string.

    Args:
        results:     List of scraped page dicts
        output_file: Local Path object (used in local mode,
                     filename used as blob name in production)

    Returns:
        str — local file path or blob filename
    """
    # ── Production: upload to Blob Storage ───────────────────
    # TODO (DevOps): uncomment when AZURE_STORAGE_CONNECTION
    # is configured in Key Vault.
    #
    # if BLOB_STORAGE_CONNECTION:
    #     from azure.storage.blob import BlobServiceClient
    #     blob_service = BlobServiceClient.from_connection_string(
    #         BLOB_STORAGE_CONNECTION
    #     )
    #     blob_name   = BLOB_SCRAPED_FILENAME
    #     blob_client = blob_service.get_blob_client(
    #         container=BLOB_CONTAINER_NAME,
    #         blob=blob_name,
    #     )
    #     data = json.dumps(results, ensure_ascii=False, indent=2)
    #     blob_client.upload_blob(data, overwrite=True)
    #     log.info(
    #         "scraped_pages_saved_to_blob",
    #         container=BLOB_CONTAINER_NAME,
    #         blob=blob_name,
    #         pages=len(results),
    #     )
    #     print(f"   ✅ Uploaded to Blob Storage: {blob_name}")
    #     return blob_name

    # ── Local development: save to local file ────────────────
    output_file.parent.mkdir(parents=True, exist_ok=True)
    with open(output_file, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    log.info(
        "scraped_pages_saved_locally",
        file=str(output_file),
        pages=len(results),
    )
    return str(output_file)


def run_scraper(
    excel_path: str | None = None,
) -> dict:
    """
    Programmatic entry point for the scraping pipeline.
    Called by DevOps / Azure Function App trigger.

    Runs the full scraping pipeline synchronously by wrapping
    the async main() logic. Returns structured result dict.

    Args:
        excel_path: Path to approved URLs Excel file.
                    If None, uses APPROVED_EXCEL constant.

    Returns:
        dict with keys:
            success        (bool) — True if completed without error
            pages_scraped  (int)  — number of pages scraped
            pages_failed   (int)  — number of pages that failed
            output_path    (str)  — local path or blob filename
            error          (str)  — error message if success=False

    TODO (DevOps): Wrap in Azure Function App trigger:

        import azure.functions as func
        from scraper.scrape_approved_urls import run_scraper

        app = func.FunctionApp()

        # Monthly scheduled scrape (1st of month, 11pm —
        # runs before chunk_and_index at midnight)
        @app.timer_trigger(
            schedule="0 23 1 * *",
            arg_name="timer",
        )
        def monthly_scrape(timer: func.TimerRequest):
            result = run_scraper()
            logging.info(f"Scrape complete: {result}")
            # chunk_and_index Function App triggered separately
            # or chained here via Service Bus / Event Grid

        # On-demand HTTP trigger
        @app.route(route="scrape", methods=["POST"])
        def on_demand_scrape(req: func.HttpRequest):
            result = run_scraper()
            return func.HttpResponse(
                json.dumps(result),
                mimetype="application/json",
            )
    """
    import traceback

    result = {
        "success":       False,
        "pages_scraped": 0,
        "pages_failed":  0,
        "output_path":   "",
        "error":         "",
    }

    try:
        import asyncio

        excel = excel_path or APPROVED_EXCEL

        async def _run():
            pages_to_scrape = load_url_source(excel)
            browser_config  = BrowserConfig(
                browser_type="chromium",
                headless=True,
                verbose=False,
            )
            scraped      = []
            failed_urls  = []
            total        = len(pages_to_scrape)

            async with AsyncWebCrawler(config=browser_config) as crawler:
                for batch_start in range(0, total, BATCH_SIZE):
                    batch  = pages_to_scrape[
                        batch_start:batch_start + BATCH_SIZE
                    ]
                    tasks  = [
                        scrape_page(
                            crawler, page_info,
                            batch_start + i + 1, total,
                        )
                        for i, page_info in enumerate(batch)
                    ]
                    results = await asyncio.gather(*tasks)
                    for page_info, r in zip(batch, results):
                        if r:
                            scraped.append(r)
                        else:
                            failed_urls.append(page_info["url"])
                    if batch_start + BATCH_SIZE < total:
                        await asyncio.sleep(BATCH_DELAY_SECONDS)

            return scraped, failed_urls

        scraped, failed = asyncio.run(_run())

        # Save output
        timestamp   = datetime.now(timezone.utc).strftime(
            "%Y%m%d_%H%M%S"
        )
        output_file = (
            Path("scraper/data") /
            f"royal_london_faq_approved_{timestamp}.json"
        )
        output_path = save_scraped_pages(scraped, output_file)

        result["success"]       = True
        result["pages_scraped"] = len(scraped)
        result["pages_failed"]  = len(failed)
        result["output_path"]   = output_path
        return result

    except Exception as e:
        result["error"] = str(e)
        log.error(
            "scraper_pipeline_error",
            error=str(e),
            traceback=traceback.format_exc(),
        )
        return result

# ── Main ────────────────────────────────────────────────
async def main():
    print("\n" + "=" * 60)
    print("RLG FAQ SCRAPER — Customer Approved URLs Only")
    print("=" * 60)

    # ── Load approved URLs from Excel ──────────────────
    excel_path = Path(APPROVED_EXCEL)
    if not excel_path.exists():
        print(f"\nERROR: Customer Excel not found: {excel_path}")
        sys.exit(1)

    # v2.0.0: load_url_source() abstracts local Excel vs
    # Blob Storage. Local: reads Excel as before. Production:
    # reads JSON from Blob Storage when AZURE_STORAGE_CONNECTION set.
    pages_to_scrape = load_url_source(str(excel_path))
    print(f"\nApproved URLs to scrape: {len(pages_to_scrape)}")

    # ── Scrape in batches ───────────────────────────────
    browser_config = BrowserConfig(
        browser_type="chromium",
        headless=True,
        verbose=False,
    )

    results = []
    failed_urls = []
    total = len(pages_to_scrape)

    async with AsyncWebCrawler(config=browser_config) as crawler:
        for batch_start in range(0, total, BATCH_SIZE):
            batch = pages_to_scrape[batch_start:batch_start + BATCH_SIZE]
            batch_num = batch_start // BATCH_SIZE + 1
            total_batches = (total + BATCH_SIZE - 1) // BATCH_SIZE

            print(f"\nBatch {batch_num}/{total_batches}")

            tasks = [
                scrape_page(crawler, page_info, batch_start + i + 1, total)
                for i, page_info in enumerate(batch)
            ]
            batch_results = await asyncio.gather(*tasks)

            for page_info, result in zip(batch, batch_results):
                if result:
                    results.append(result)
                else:
                    failed_urls.append(page_info["url"])

            if batch_start + BATCH_SIZE < total:
                await asyncio.sleep(BATCH_DELAY_SECONDS)

    # ── Save output ─────────────────────────────────────
    # v2.0.0: save_scraped_pages() abstracts local file vs
    # Blob Storage. Local: saves to scraper/data/ as before.
    # Production: uploads to Azure Blob Storage when
    # AZURE_STORAGE_CONNECTION env var is set.
    timestamp   = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    output_file = (
        Path("scraper/data") /
        f"royal_london_faq_approved_{timestamp}.json"
    )
    output_path = save_scraped_pages(results, output_file)

    # ── Summary ───────────────────────────────────────────
    print("\n" + "=" * 60)
    print("SCRAPE SUMMARY")
    print("=" * 60)
    print(f"Approved URLs:    {total}")
    print(f"Scraped success:  {len(results)}")
    print(f"Failed:           {len(failed_urls)}")

    if failed_urls:
        print("\nFailed URLs:")
        for url in failed_urls:
            print(f"  - {url}")

    if results:
        total_chars = sum(r["content_length"] for r in results)
        lengths     = [r["content_length"] for r in results]
        print(f"\nTotal content:    {total_chars:,} chars")
        print(f"Shortest page:    {min(lengths):,} chars")
        print(f"Longest page:     {max(lengths):,} chars")
        print(f"Average page:     {total_chars // len(results):,} chars")

        # v3.0.0 — enrichment summary
        video_pages = sum(1 for r in results if r.get("has_video"))
        type_counts = {}
        cat_counts  = {}
        for r in results:
            ct = r.get("content_type", "article")
            pc = r.get("product_category", "general")
            type_counts[ct] = type_counts.get(ct, 0) + 1
            cat_counts[pc]  = cat_counts.get(pc, 0) + 1

        print(f"\nPages with video: {video_pages}")
        print("\nContent types:")
        for ct, count in sorted(type_counts.items(), key=lambda x: -x[1]):
            print(f"  {ct:<20} {count}")
        print("\nProduct categories:")
        for pc, count in sorted(cat_counts.items(), key=lambda x: -x[1]):
            print(f"  {pc:<25} {count}")

    print(f"\nSaved to: {output_path}")
    print("=" * 60)
    print(f"\nNext: uv run python scraper/chunk_and_indexV3.py --full "
          f"--file {output_path}")


if __name__ == "__main__":
    asyncio.run(main())