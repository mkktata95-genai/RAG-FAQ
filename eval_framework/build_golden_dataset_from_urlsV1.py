"""
build_golden_dataset_from_urls.py — Golden Dataset Generator (page-level,
multi-question, category-tagged, generate-then-grade)

═══════════════════════════════════════════════════════════════
PURPOSE
═══════════════════════════════════════════════════════════════
Standalone. Does NOT import anything from the RLG pipeline — deliberately
self-contained so this script can be handed off, forked, or run
independently by anyone who wants to tweak the prompts or question count
without touching pipeline code.

v2 CHANGE OF APPROACH (see v2.0.0 changelog entry below for the full
reasoning): earlier versions generated 2 competing candidate questions per
URL and picked 1 winner. This version instead generates UP TO
TARGET_QUESTIONS_PER_URL (default 5) questions per URL, spread across
three categories a customer's follow-up questions typically fall into,
grades EACH one individually (not pick-one-of-N), and drafts an answer
for every accepted question. This matches a proposal to reduce SME
authorship workload to review workload instead — see the "REVIEW WORKFLOW
CHANGE" note below, which is the most important thing to understand
about this version before running it.

THE THREE QUESTION CATEGORIES
  1. on_page              — answerable purely from this page's content
  2. related_not_on_page  — a natural follow-up a customer would ask
                             after reading this page, but this page does
                             NOT cover it
  3. clarification         — asks for further explanation of a concept/
                             term that IS named on the page but may need
                             unpacking for a typical customer

  NOTE — category 2 has an open, unresolved question: when a customer's
  likely follow-up isn't covered by this page, should the "correct"
  answer draw on OTHER approved content, or should it be a correct
  refusal ("I don't have that information")? This script does NOT
  resolve that — it drafts a best-effort answer for related_not_on_page
  questions and explicitly flags it via the sourcing_note field so
  whoever reviews it knows this is provisional pending that decision.
  Do not treat related_not_on_page draft answers as authoritative until
  that's settled.

REVIEW WORKFLOW CHANGE — READ BEFORE USING THE HANDOFF FILE
  Earlier versions of this script deliberately withheld the draft answer
  from the PO/SME handoff file, so an SME's independent answer couldn't
  be biased by seeing a draft first. THIS VERSION INCLUDES THE DRAFT
  ANSWER IN THE HANDOFF FILE ON PURPOSE — the review model has changed
  from "SME authors from scratch" to "SME validates an AI draft against
  the source content." This is a materially different (weaker) tier of
  ground truth than fully independent authorship — see team discussion
  on "golden" vs "silver standard" terminology before presenting this
  dataset as equivalent to earlier fully-SME-authored work.
  The review column defaults to correct / needs_edit / wrong (not a
  plain yes/no) plus a notes field — a plain yes/no is more vulnerable
  to reviewers rubber-stamping fluent-sounding AI text without close
  reading. Change REVIEW_OPTIONS below if a simpler yes/no is preferred.

APPROACH (per URL, not per chunk)
  1. Read the approved URL list from the customer Excel.
  2. Pull that URL's FULL base-page content from the search index —
     concatenates all non-dropdown chunks (chunk_id has no #state=
     fragment) in chunk_index order. Dropdown-state content is
     deliberately excluded — base page only.
  3. Generation LLM produces up to TARGET_QUESTIONS_PER_URL questions,
     each tagged with a category and a one-line rationale for why a
     customer would plausibly ask it.
  4. Grading LLM (a DIFFERENT model from generation, to avoid
     self-evaluation bias) scores EACH question individually against a
     rubric including whether it genuinely fits its claimed category.
     Any question scoring below GRADE_THRESHOLD is dropped — never
     forced into the output to hit a target count.
  5. Generation LLM drafts an answer for every accepted question —
     on_page/clarification answers are grounded strictly in the page
     content; related_not_on_page answers are explicitly flagged as
     provisional (see note above).

THREE OUTPUT FILES
  - *_internal.csv  : question, category, source_url, draft_expected_answer,
                       grade scores/rationale — for your own eval-framework
                       testing. One row per accepted question (0-5 per URL).
  - *_handoff.csv    : question, category, source_url, draft_expected_answer,
                       sourcing_note, review, review_notes (review/
                       review_notes blank — send this to PO/SME). Draft
                       answer IS included here — see REVIEW WORKFLOW
                       CHANGE above.
  - *_audit.csv      : EVERY url attempted, how many questions were
                       generated/accepted/rejected per category, and why
                       — proof of full coverage attempt even where little
                       or nothing was accepted.

TWEAKABLE CONFIG (all in one block below)
  - GENERATION_MODEL / GRADING_MODEL deployment names
  - TARGET_QUESTIONS_PER_URL
  - GRADE_THRESHOLD
  - REVIEW_OPTIONS
  - GENERATION_PROMPT / GRADING_PROMPT / ANSWER_DRAFT_PROMPT templates

USAGE
    python build_golden_dataset_from_urls.py --excel approved_urls.xlsx --dry-run --limit 5
    python build_golden_dataset_from_urls.py --excel approved_urls.xlsx --out-prefix golden_dataset

REQUIRES
    .env with AZURE_SEARCH_ENDPOINT and AZURE_OPENAI_ENDPOINT set
    pip install azure-search-documents azure-identity azure-ai-inference openai python-dotenv openpyxl
    DefaultAzureCredential auth (RLG policy — no API keys)

═══════════════════════════════════════════════════════════════
CHANGE LOG
═══════════════════════════════════════════════════════════════
v2.0.0 — Aug 2026 | Mukesh Kund
         Full redesign per senior-member proposal to reduce SME workload
         from authorship to review. Changed from "generate 2 candidates,
         pick 1 winner, 1 output row per URL" to "generate up to 5
         category-tagged questions, grade each individually, output up
         to 5 rows per URL." Added the 3-category prompt structure
         (on_page / related_not_on_page / clarification). Draft answers
         are now included in the handoff file (previously deliberately
         withheld) since the review model itself changed — see
         "REVIEW WORKFLOW CHANGE" note in the module docstring. Added
         sourcing_note field to flag related_not_on_page answers as
         provisional pending an unresolved team decision (cross-page
         sourcing vs. correct refusal). Review column defaults to
         correct/needs_edit/wrong + notes, not a plain yes/no.
         ROLLBACK: this is a structural change, not a patch — revert to
         the prior single-file version if the pick-one-winner approach
         is still wanted (see git history / prior version in chat).

v1.2.0 — Aug 2026 | Mukesh Kund
         Fix (more complete): added standalone _build_create_kwargs()
         mirroring generator.py/classifier_node.py's own GPT-4 vs
         GPT-5 family handling — GPT-5 family needs
         max_completion_tokens, not max_tokens, in addition to
         omitting temperature.
         ROLLBACK: revert to plain max_tokens-only kwargs (breaks on
         GPT-5-family deployments).

v1.1.0 — Aug 2026 | Mukesh Kund
         Fix: removed explicit temperature parameter from both LLM
         calls. gpt-5.6-luna (reasoning-tier model) rejects any
         temperature value other than the default (1) with a 400.
         ROLLBACK: re-add temperature=0.3/0.2 — only valid on a
         non-reasoning-tier model.

v1.0.0 — Aug 2026 | Mukesh Kund
         Initial version. Standalone, no imports from pipeline scripts.
         ROLLBACK: n/a — new standalone script.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import time
from dataclasses import dataclass, field
from typing import Optional

from dotenv import load_dotenv, find_dotenv
from azure.identity import DefaultAzureCredential, get_bearer_token_provider
from azure.search.documents import SearchClient
from openai import AzureOpenAI
from openpyxl import load_workbook

load_dotenv(find_dotenv(usecwd=False))

# ═══════════════════════════════════════════════════════════════
# TWEAKABLE CONFIG — edit here, nothing else needs touching
# ═══════════════════════════════════════════════════════════════

AZURE_SEARCH_ENDPOINT = os.getenv("AZURE_SEARCH_ENDPOINT", "").rstrip("/")
SEARCH_INDEX_NAME     = os.getenv("AZURE_SEARCH_INDEX_NAME", "rlg-faq-index-v5")

AZURE_OPENAI_ENDPOINT = os.getenv("AZURE_OPENAI_ENDPOINT", "").rstrip("/")
AZURE_OPENAI_API_VERSION = os.getenv("AZURE_OPENAI_API_VERSION", "2024-10-21")

# Deployment names — set these to whatever your Azure OpenAI resource
# actually has deployed. Generation and grading MUST be different
# deployments/models to avoid self-evaluation bias.
GENERATION_MODEL = os.getenv("GOLDEN_GEN_MODEL", "gpt-5.6-luna")
GRADING_MODEL    = os.getenv("GOLDEN_GRADE_MODEL", "gpt-5-mini")

TARGET_QUESTIONS_PER_URL = 5  # upper bound — pages with thin content may yield fewer
GRADE_THRESHOLD = 7  # 0-10 scale; a question scoring below this is dropped, not forced in
VALID_CATEGORIES = {"on_page", "related_not_on_page", "clarification"}
REVIEW_OPTIONS = "correct / needs_edit / wrong"  # shown in handoff CSV header comment

REQUEST_SLEEP_SECONDS = 0.5  # small pause between URLs to be gentle on rate limits
MAX_RETRIES = 2
MAX_OUTPUT_TOKENS = 3000  # multi-question JSON responses need more headroom than v1's single-question ones


def _build_create_kwargs(model: str, max_tokens: int) -> dict:
    """Model-family-compatible kwargs, mirroring generator.py's
    _build_create_kwargs / classifier_node.py's equivalent helper.

    GPT-4 family (gpt-4*): supports max_tokens.
    GPT-5 family (gpt-5*, o*, etc.): uses max_completion_tokens; a
    custom temperature is REJECTED (400 error) — only the default (1)
    is supported, so temperature is never set for this family.
    Standalone copy, not imported, per this script's no-pipeline-
    dependency design — keep in sync manually if generator.py's
    version changes.
    """
    is_gpt4 = "gpt-4" in model.lower()
    kwargs: dict = {}
    if is_gpt4:
        kwargs["max_tokens"] = max_tokens
    else:
        kwargs["max_completion_tokens"] = max_tokens
    return kwargs


GENERATION_PROMPT = """You are helping build a test set for a customer-facing pensions/ISA/protection/equity-release chatbot for Royal London.

Below is the full content of one webpage. A real customer reading this page will often have follow-up questions. Generate realistic follow-up questions across THREE categories:

1. "on_page" — a question about something SPECIFICALLY mentioned on this page. Should be fully answerable using only this page's content.
2. "related_not_on_page" — a natural follow-up question about a CLOSELY RELATED topic that a customer would plausibly ask after reading this page, but which this page does NOT itself cover.
3. "clarification" — a question asking for further explanation of a CONCEPT OR TERM that IS named on this page but may need unpacking for a typical customer to fully understand.

Generate up to {target_count} questions total, spread across these three categories based on what's realistic for THIS page — do not force all three categories or the full count if the page is short/simple and doesn't genuinely support that many good questions. Quality over hitting the count.

For each question return:
- "question": phrased naturally, like a real customer typing into a chat box (not formal/robotic)
- "category": exactly one of "on_page", "related_not_on_page", "clarification"
- "rationale": one sentence — why would a customer plausibly ask this after reading this page?

Respond ONLY with a JSON array of objects in this exact shape, no other text:
[{{"question": "...", "category": "...", "rationale": "..."}}]

PAGE CONTENT:
{content}
"""

GRADING_PROMPT = """You are grading ONE candidate customer-support question for a golden test dataset for a Royal London pensions/ISA/protection/equity-release chatbot.

PAGE CONTENT:
{content}

CANDIDATE QUESTION: {question}
CLAIMED CATEGORY: {category}

Score 0-10 on each of:
- specificity: not generic — tests something concrete, not a vague "what is X" that could apply to any page
- natural_phrasing: sounds like a real customer, not a robotic/templated question
- product_relevance: relevant to a pensions/ISA/protection/equity-release/bereavement customer support chatbot (NOT internal IT topics like passwords/MFA, NOT generic corporate/accessibility-statement content)
- category_fit: does the question genuinely match its claimed category?
    - on_page must be fully answerable from the page content alone
    - related_not_on_page must be a plausible follow-up the page does NOT cover (if the page actually does cover it, this is a category mismatch, mark it down)
    - clarification must be about a concept/term actually named on the page

Respond ONLY with JSON in this exact shape, no other text:
{{"specificity": 0, "natural_phrasing": 0, "product_relevance": 0, "category_fit": 0, "overall": 0, "accept": true, "rationale": "one sentence"}}

"overall" is your single 0-10 judgement (not required to be an average of the four sub-scores — use your judgement). Set "accept" to true only if this question is good enough to include in the golden dataset as-is.
"""

ANSWER_DRAFT_PROMPT = """Draft an answer to the following question for a Royal London customer support chatbot test dataset.

PAGE CONTENT:
{content}

QUESTION: {question}
CATEGORY: {category}

Instructions depend on category:
- If category is "on_page" or "clarification": answer using ONLY the page content above. Be concise (2-4 sentences), factual, do not add information not present in the content.
- If category is "related_not_on_page": the page content above does NOT fully answer this question. Give your best good-faith brief answer based on general knowledge of Royal London's public products/policies if you are reasonably confident, but keep it short. This draft is PROVISIONAL — the correct approach for this category (whether the chatbot should draw on other approved content, or should refuse and say it doesn't have that information) has not been finalized by the team yet.

Respond ONLY with JSON in this exact shape, no other text:
{{"answer": "...", "sourcing_note": "..."}}

"sourcing_note" must be "" (empty string) for on_page/clarification questions. For related_not_on_page questions, sourcing_note must briefly state what the draft answer relies on (e.g. "Based on general knowledge of RL ISA products, not sourced from this page — needs team decision on cross-page sourcing vs. refusal approach before treating as final.") so a reviewer knows to treat it differently from the other two categories.
"""


# ═══════════════════════════════════════════════════════════════
# CLIENTS
# ═══════════════════════════════════════════════════════════════

def get_search_client() -> SearchClient:
    if not AZURE_SEARCH_ENDPOINT:
        raise SystemExit("AZURE_SEARCH_ENDPOINT not set in .env")
    return SearchClient(
        endpoint=AZURE_SEARCH_ENDPOINT,
        index_name=SEARCH_INDEX_NAME,
        credential=DefaultAzureCredential(),
    )


def get_openai_client() -> AzureOpenAI:
    if not AZURE_OPENAI_ENDPOINT:
        raise SystemExit("AZURE_OPENAI_ENDPOINT not set in .env")
    token_provider = get_bearer_token_provider(
        DefaultAzureCredential(), "https://cognitiveservices.azure.com/.default"
    )
    return AzureOpenAI(
        azure_endpoint=AZURE_OPENAI_ENDPOINT,
        azure_ad_token_provider=token_provider,
        api_version=AZURE_OPENAI_API_VERSION,
    )


# ═══════════════════════════════════════════════════════════════
# APPROVED URL LIST (standalone — mirrors the header-detection
# behaviour of the pipeline's loader but does not import it)
# ═══════════════════════════════════════════════════════════════

URL_HEADERS      = {"url", "page url", "link", "webpage", "web page", "web url"}
TITLE_HEADERS    = {"title", "page title", "name"}
CATEGORY_HEADERS = {"category", "content category", "page category", "type"}


def load_approved_urls(excel_path: str) -> list[dict]:
    wb = load_workbook(excel_path, read_only=True)
    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    headers = [str(h).strip().lower() if h is not None else "" for h in header_row]

    def find_col(candidates: set) -> Optional[int]:
        for idx, h in enumerate(headers):
            if h in candidates:
                return idx
        return None

    url_idx = find_col(URL_HEADERS)
    title_idx = find_col(TITLE_HEADERS)
    category_idx = find_col(CATEGORY_HEADERS)

    if url_idx is None:
        wb.close()
        raise ValueError(
            f"No URL column found in {excel_path!r}. Expected one of "
            f"{sorted(URL_HEADERS)}. Headers found: {header_row!r}"
        )

    rows = []
    seen = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or url_idx >= len(row) or not row[url_idx]:
            continue
        url = str(row[url_idx]).strip()
        if not url or url in seen:
            continue
        seen.add(url)
        rows.append({
            "url": url,
            "title": str(row[title_idx]).strip() if title_idx is not None and title_idx < len(row) and row[title_idx] else "",
            "category": str(row[category_idx]).strip() if category_idx is not None and category_idx < len(row) and row[category_idx] else "",
        })
    wb.close()
    return rows


# ═══════════════════════════════════════════════════════════════
# PAGE CONTENT RETRIEVAL (base page only — excludes #state= dropdown chunks)
# ═══════════════════════════════════════════════════════════════

def fetch_base_page_content(search_client: SearchClient, url: str) -> str:
    results = list(search_client.search(
        search_text="*",
        filter=f"source_url eq '{url}'",
        select=["chunk_index", "content"],
        top=200,
    ))
    if not results:
        return ""
    results.sort(key=lambda r: r.get("chunk_index", 0))
    return "\n\n".join(r.get("content", "") for r in results if r.get("content"))


# ═══════════════════════════════════════════════════════════════
# LLM CALLS
# ═══════════════════════════════════════════════════════════════

def _chat_json(client: AzureOpenAI, model: str, prompt: str) -> Optional[dict | list]:
    for attempt in range(MAX_RETRIES + 1):
        try:
            resp = client.chat.completions.create(
                model=model,
                messages=[{"role": "user", "content": prompt}],
                **_build_create_kwargs(model, MAX_OUTPUT_TOKENS),
            )
            raw = (resp.choices[0].message.content or "").strip()
            # Brace-position extraction, not naive backtick-strip — GPT-5
            # reasoning models often prepend prose before the JSON object.
            # Mirrors the established pattern in classifier_node.py's
            # classify_intent() (v1.2.0 FIX 2). Handles both object {}
            # and array [] responses — picks whichever bracket appears
            # first in the raw text.
            first_curly, first_square = raw.find("{"), raw.find("[")
            if first_square != -1 and (first_curly == -1 or first_square < first_curly):
                start, end = first_square, raw.rfind("]") + 1
            else:
                start, end = first_curly, raw.rfind("}") + 1
            if start == -1 or end == 0:
                raise ValueError(f"No JSON found in response: {raw[:150]!r}")
            return json.loads(raw[start:end])
        except Exception as exc:
            if attempt == MAX_RETRIES:
                print(f"    [LLM call failed after {MAX_RETRIES + 1} attempts]: {exc}")
                return None
            time.sleep(1.5 * (attempt + 1))
    return None


def generate_candidate_questions(client: AzureOpenAI, content: str) -> list[dict]:
    """Returns list of {"question", "category", "rationale"} dicts,
    filtered to valid categories only."""
    prompt = GENERATION_PROMPT.format(target_count=TARGET_QUESTIONS_PER_URL, content=content[:12000])
    result = _chat_json(client, GENERATION_MODEL, prompt)
    if not isinstance(result, list):
        return []
    out = []
    for item in result:
        if not isinstance(item, dict):
            continue
        q = str(item.get("question", "")).strip()
        cat = str(item.get("category", "")).strip()
        if q and cat in VALID_CATEGORIES:
            out.append({"question": q, "category": cat, "rationale": str(item.get("rationale", "")).strip()})
    return out


def grade_question(client: AzureOpenAI, content: str, question: str, category: str) -> dict:
    """Grades ONE question. Returns dict with overall score, accept bool, rationale."""
    prompt = GRADING_PROMPT.format(content=content[:12000], question=question, category=category)
    result = _chat_json(client, GRADING_MODEL, prompt)
    if not isinstance(result, dict):
        return {"overall": 0, "accept": False, "rationale": "grading call failed or malformed response"}
    return {
        "overall": result.get("overall", 0),
        "accept": bool(result.get("accept", False)),
        "rationale": result.get("rationale", ""),
    }


def draft_answer(client: AzureOpenAI, content: str, question: str, category: str) -> dict:
    """Returns {"answer": str, "sourcing_note": str}."""
    prompt = ANSWER_DRAFT_PROMPT.format(content=content[:12000], question=question, category=category)
    result = _chat_json(client, GENERATION_MODEL, prompt)
    if isinstance(result, dict) and result.get("answer"):
        return {"answer": str(result["answer"]).strip(), "sourcing_note": str(result.get("sourcing_note", "")).strip()}
    return {"answer": "[draft answer generation failed]", "sourcing_note": ""}


# ═══════════════════════════════════════════════════════════════
# MAIN PIPELINE
# ═══════════════════════════════════════════════════════════════

@dataclass
class QAItem:
    """One accepted question+answer for one URL."""
    question: str
    category: str
    draft_answer: str
    sourcing_note: str
    grade_score: float
    grade_rationale: str


@dataclass
class PageResult:
    url: str
    title: str
    page_category: str  # from the Excel's category column, if present — NOT the question category
    status: str  # "generated" | "no_questions_accepted" | "skipped_no_content" | "error"
    reason: str = ""
    items: list[QAItem] = field(default_factory=list)
    num_candidates_generated: int = 0
    num_accepted: int = 0
    num_rejected: int = 0


def process_url(search_client, openai_client, url: str, title: str, page_category: str, dry_run: bool) -> PageResult:
    content = fetch_base_page_content(search_client, url)
    if not content or len(content.split()) < 30:
        return PageResult(url=url, title=title, page_category=page_category, status="skipped_no_content",
                           reason=f"base page content too short/empty ({len(content.split())} words)")

    if dry_run:
        return PageResult(url=url, title=title, page_category=page_category, status="generated",
                           reason=f"[DRY RUN — no LLM calls] content_words={len(content.split())}")

    candidates = generate_candidate_questions(openai_client, content)
    if not candidates:
        return PageResult(url=url, title=title, page_category=page_category, status="error",
                           reason="question generation returned nothing valid")

    items: list[QAItem] = []
    num_rejected = 0
    for cand in candidates:
        grading = grade_question(openai_client, content, cand["question"], cand["category"])
        if not grading["accept"] or grading["overall"] < GRADE_THRESHOLD:
            num_rejected += 1
            continue
        drafted = draft_answer(openai_client, content, cand["question"], cand["category"])
        items.append(QAItem(
            question=cand["question"], category=cand["category"],
            draft_answer=drafted["answer"], sourcing_note=drafted["sourcing_note"],
            grade_score=grading["overall"], grade_rationale=grading["rationale"],
        ))

    if not items:
        return PageResult(url=url, title=title, page_category=page_category, status="no_questions_accepted",
                           reason=f"{len(candidates)} candidate(s) generated, none cleared threshold={GRADE_THRESHOLD}",
                           num_candidates_generated=len(candidates), num_rejected=num_rejected)

    return PageResult(url=url, title=title, page_category=page_category, status="generated",
                       items=items, num_candidates_generated=len(candidates),
                       num_accepted=len(items), num_rejected=num_rejected)


def main():
    parser = argparse.ArgumentParser(description="Generate golden dataset questions — up to 5 category-tagged questions per approved URL")
    parser.add_argument("--excel", required=True, help="Path to approved URLs Excel file")
    parser.add_argument("--out-prefix", default="golden_dataset", help="Prefix for the three output CSVs")
    parser.add_argument("--dry-run", action="store_true", help="Skip LLM calls — just verify URL list + content retrieval works")
    parser.add_argument("--limit", type=int, default=None, help="Only process the first N URLs (for testing)")
    args = parser.parse_args()

    pages = load_approved_urls(args.excel)
    if args.limit:
        pages = pages[:args.limit]
    print(f"Loaded {len(pages)} approved URL(s) from {args.excel}"
          f"{' (dry run — no LLM calls)' if args.dry_run else ''}")

    search_client = get_search_client()
    openai_client = None if args.dry_run else get_openai_client()

    results: list[PageResult] = []
    for i, page in enumerate(pages, 1):
        print(f"[{i}/{len(pages)}] {page['url']}")
        result = process_url(search_client, openai_client, page["url"], page["title"], page["category"], args.dry_run)
        results.append(result)
        if result.status == "generated" and not args.dry_run:
            print(f"    -> generated {result.num_accepted}/{result.num_candidates_generated} accepted "
                  f"({result.num_rejected} rejected)")
        else:
            print(f"    -> {result.status}" + (f" ({result.reason})" if result.reason else ""))
        if not args.dry_run:
            time.sleep(REQUEST_SLEEP_SECONDS)

    generated_pages = [r for r in results if r.status == "generated"]
    no_accept_pages = [r for r in results if r.status == "no_questions_accepted"]
    skipped_pages = [r for r in results if r.status == "skipped_no_content"]
    errored_pages = [r for r in results if r.status == "error"]
    total_items = sum(len(r.items) for r in generated_pages)

    # ── Internal file (framework testing) — one row per accepted question ──
    internal_path = f"{args.out_prefix}_internal.csv"
    with open(internal_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "id", "product_category", "question_category", "question", "draft_expected_answer",
            "sourcing_note", "source_url", "grade_score", "grade_rationale",
        ])
        writer.writeheader()
        row_id = 1
        for r in generated_pages:
            for item in r.items:
                writer.writerow({
                    "id": f"GOLD-{row_id:04d}", "product_category": r.page_category or "general",
                    "question_category": item.category, "question": item.question,
                    "draft_expected_answer": item.draft_answer, "sourcing_note": item.sourcing_note,
                    "source_url": r.url, "grade_score": item.grade_score, "grade_rationale": item.grade_rationale,
                })
                row_id += 1

    # ── PO/SME handoff file — DRAFT ANSWER INCLUDED, see module docstring ──
    handoff_path = f"{args.out_prefix}_handoff.csv"
    with open(handoff_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "id", "product_category", "question_category", "question", "draft_expected_answer",
            "sourcing_note", "source_url", "review", "review_notes",
        ])
        writer.writeheader()
        row_id = 1
        for r in generated_pages:
            for item in r.items:
                writer.writerow({
                    "id": f"GOLD-{row_id:04d}", "product_category": r.page_category or "general",
                    "question_category": item.category, "question": item.question,
                    "draft_expected_answer": item.draft_answer, "sourcing_note": item.sourcing_note,
                    "source_url": r.url, "review": "", "review_notes": "",
                })
                row_id += 1

    # ── Audit log — every URL attempted, coverage proof ──
    audit_path = f"{args.out_prefix}_audit.csv"
    with open(audit_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "url", "status", "reason", "num_candidates_generated", "num_accepted", "num_rejected",
        ])
        writer.writeheader()
        for r in results:
            writer.writerow({
                "url": r.url, "status": r.status, "reason": r.reason,
                "num_candidates_generated": r.num_candidates_generated,
                "num_accepted": r.num_accepted, "num_rejected": r.num_rejected,
            })

    print(f"\n{'='*60}")
    print(f"Total URLs processed: {len(results)}")
    print(f"  URLs with >=1 question accepted: {len(generated_pages)}")
    print(f"  URLs with 0 questions accepted: {len(no_accept_pages)}")
    print(f"  Skipped (no/insufficient content): {len(skipped_pages)}")
    print(f"  Errored: {len(errored_pages)}")
    print(f"  Total accepted questions across all URLs: {total_items}")
    print(f"\nOutputs:")
    print(f"  {handoff_path}  <- send this to PO/SME (includes draft answers — see module docstring)")
    print(f"  {audit_path}    <- send this too, as coverage proof")
    print(f"  {internal_path} <- keep for your own framework testing")


if __name__ == "__main__":
    main()