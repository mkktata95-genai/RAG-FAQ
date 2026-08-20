"""
build_golden_dataset_from_urls.py — Golden Dataset Generator (page-level, generate-then-grade)

═══════════════════════════════════════════════════════════════
PURPOSE
═══════════════════════════════════════════════════════════════
Standalone replacement for the HQA-chunk-based seed approach. Does NOT
import anything from the RLG pipeline (chunk_and_index_hqaV5.py,
scrape_approved_urls_updatedV5.py, etc.) — deliberately self-contained
so this script can be handed off, forked, or run independently by
anyone who wants to tweak the prompts or question count without
touching pipeline code.

APPROACH (per URL, not per chunk)
  1. Read the approved URL list from the customer Excel.
  2. Pull that URL's FULL base-page content from the search index —
     concatenates all non-dropdown chunks (chunk_id has no #state=
     fragment) in chunk_index order. Dropdown-state content is
     deliberately excluded — base page only.
  3. Generation LLM produces NUM_CANDIDATE_QUESTIONS candidate
     customer-style questions answerable purely from that content.
  4. Grading LLM (a DIFFERENT model from generation, to avoid
     self-evaluation bias) scores each candidate 0-10 against a
     rubric and picks the best one. If neither clears GRADE_THRESHOLD,
     the URL is skipped — no forced/weak question is ever output.
  5. Generation LLM also drafts an "expected_answer" from the same
     content, explicitly marked unverified/draft — for internal
     framework testing only. This draft is NEVER included in the
     PO/SME handoff file, so it can't bias an independent SME answer.

THREE OUTPUT FILES
  - *_internal.csv  : question, source_url, category, draft_expected_answer
                       (unverified), grade, grading_rationale — for your
                       own eval-framework testing.
  - *_handoff.csv    : question, source_url, category, expected_answer
                       (blank) — send this to PO/SME. No draft answer
                       included on purpose.
  - *_audit.csv      : EVERY url attempted, status (generated / rejected
                       / skipped_no_content), reason — proof of full
                       coverage attempt even for URLs that didn't make
                       the cut. Give this to PO alongside the handoff
                       file so they can see nothing was silently dropped.

TWEAKABLE CONFIG (all in one block below — no need to read the rest
of the file to adjust question count, models, or prompts)
  - GENERATION_MODEL / GRADING_MODEL deployment names
  - NUM_CANDIDATE_QUESTIONS
  - GRADE_THRESHOLD
  - GENERATION_PROMPT / GRADING_PROMPT / ANSWER_DRAFT_PROMPT templates

USAGE
    python build_golden_dataset_from_urls.py --excel approved_urls.xlsx --dry-run --limit 5
    python build_golden_dataset_from_urls.py --excel approved_urls.xlsx --out-prefix golden_dataset

REQUIRES
    .env with AZURE_SEARCH_ENDPOINT and AZURE_OPENAI_ENDPOINT set
    pip install azure-search-documents azure-identity azure-ai-inference openai python-dotenv openpyxl
    DefaultAzureCredential auth (RLG policy — no API keys)

═══════════════════════════════════════════════════════════════
CHANGE LOG
═══════════════════════════════════════════════════════════════
v1.0.0 — Aug 2026 | Mukesh Kund
         Initial version. Standalone, no imports from pipeline scripts.
         ROLLBACK: n/a — new standalone script.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import time
from dataclasses import dataclass, field
from typing import Optional

from dotenv import load_dotenv, find_dotenv
from azure.identity import DefaultAzureCredential, get_bearer_token_provider
from azure.search.documents import SearchClient
from openai import AzureOpenAI
from openpyxl import load_workbook

load_dotenv(find_dotenv(usecwd=False))

# ═══════════════════════════════════════════════════════════════
# TWEAKABLE CONFIG — edit here, nothing else needs touching
# ═══════════════════════════════════════════════════════════════

AZURE_SEARCH_ENDPOINT = os.getenv("AZURE_SEARCH_ENDPOINT", "").rstrip("/")
SEARCH_INDEX_NAME     = os.getenv("AZURE_SEARCH_INDEX_NAME", "rlg-faq-index-v5")

AZURE_OPENAI_ENDPOINT = os.getenv("AZURE_OPENAI_ENDPOINT", "").rstrip("/")
AZURE_OPENAI_API_VERSION = os.getenv("AZURE_OPENAI_API_VERSION", "2024-10-21")

# Deployment names — set these to whatever your Azure OpenAI resource
# actually has deployed. Generation and grading MUST be different
# deployments/models to avoid self-evaluation bias.
GENERATION_MODEL = os.getenv("GOLDEN_GEN_MODEL", "gpt-5.6-luna")
GRADING_MODEL    = os.getenv("GOLDEN_GRADE_MODEL", "gpt-5-mini")

NUM_CANDIDATE_QUESTIONS = 2
GRADE_THRESHOLD = 7  # 0-10 scale; neither candidate scoring >= this means the URL is skipped

REQUEST_SLEEP_SECONDS = 0.5  # small pause between URLs to be gentle on rate limits
MAX_RETRIES = 2

GENERATION_PROMPT = """You are helping build a test set for a customer-facing pensions/ISA/protection/equity-release chatbot.

Below is the full content of one webpage. Write {num_questions} DIFFERENT questions that:
- A real customer might plausibly type into a chat box (natural phrasing, not formal/robotic)
- Can be answered COMPLETELY using ONLY the content below (no outside knowledge needed)
- Are specific to this page's content — not generic ("What is a pension?" is too generic if the page is about something more specific)
- Do not overlap with each other in what they ask about

Respond ONLY with a JSON array of strings, e.g. ["question 1", "question 2"]. No other text.

PAGE CONTENT:
{content}
"""

GRADING_PROMPT = """You are grading candidate customer-support questions for a golden test dataset for a pensions/ISA/protection/equity-release chatbot.

PAGE CONTENT:
{content}

CANDIDATE QUESTIONS:
{candidates_numbered}

For EACH candidate, score 0-10 on:
- specificity: not generic, actually tests something concrete from this page
- natural_phrasing: sounds like a real customer, not a robotic/templated question
- answerable_from_content: can be fully answered using ONLY the page content above
- product_relevance: relevant to a pensions/ISA/protection/equity-release/bereavement customer support chatbot (NOT internal IT topics like passwords/MFA, NOT generic corporate/accessibility statement content)

Respond ONLY with JSON in this exact shape, no other text:
{{
  "scores": [
    {{"index": 1, "specificity": 0, "natural_phrasing": 0, "answerable_from_content": 0, "product_relevance": 0, "overall": 0}},
    {{"index": 2, "specificity": 0, "natural_phrasing": 0, "answerable_from_content": 0, "product_relevance": 0, "overall": 0}}
  ],
  "best_index": 1,
  "rationale": "one sentence why this one was picked over the other"
}}
"overall" is your single 0-10 judgement per candidate (not required to be an average of the four sub-scores — use your judgement).
"""

ANSWER_DRAFT_PROMPT = """Answer the following question using ONLY the page content below. Be concise (2-4 sentences), factual, and do not add information not present in the content.

PAGE CONTENT:
{content}

QUESTION: {question}

Respond with ONLY the answer text, no preamble.
"""


# ═══════════════════════════════════════════════════════════════
# CLIENTS
# ═══════════════════════════════════════════════════════════════

def get_search_client() -> SearchClient:
    if not AZURE_SEARCH_ENDPOINT:
        raise SystemExit("AZURE_SEARCH_ENDPOINT not set in .env")
    return SearchClient(
        endpoint=AZURE_SEARCH_ENDPOINT,
        index_name=SEARCH_INDEX_NAME,
        credential=DefaultAzureCredential(),
    )


def get_openai_client() -> AzureOpenAI:
    if not AZURE_OPENAI_ENDPOINT:
        raise SystemExit("AZURE_OPENAI_ENDPOINT not set in .env")
    token_provider = get_bearer_token_provider(
        DefaultAzureCredential(), "https://cognitiveservices.azure.com/.default"
    )
    return AzureOpenAI(
        azure_endpoint=AZURE_OPENAI_ENDPOINT,
        azure_ad_token_provider=token_provider,
        api_version=AZURE_OPENAI_API_VERSION,
    )


# ═══════════════════════════════════════════════════════════════
# APPROVED URL LIST (standalone — mirrors the header-detection
# behaviour of the pipeline's loader but does not import it)
# ═══════════════════════════════════════════════════════════════

URL_HEADERS      = {"url", "page url", "link", "webpage", "web page", "web url"}
TITLE_HEADERS    = {"title", "page title", "name"}
CATEGORY_HEADERS = {"category", "content category", "page category", "type"}


def load_approved_urls(excel_path: str) -> list[dict]:
    wb = load_workbook(excel_path, read_only=True)
    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    headers = [str(h).strip().lower() if h is not None else "" for h in header_row]

    def find_col(candidates: set) -> Optional[int]:
        for idx, h in enumerate(headers):
            if h in candidates:
                return idx
        return None

    url_idx = find_col(URL_HEADERS)
    title_idx = find_col(TITLE_HEADERS)
    category_idx = find_col(CATEGORY_HEADERS)

    if url_idx is None:
        wb.close()
        raise ValueError(
            f"No URL column found in {excel_path!r}. Expected one of "
            f"{sorted(URL_HEADERS)}. Headers found: {header_row!r}"
        )

    rows = []
    seen = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or url_idx >= len(row) or not row[url_idx]:
            continue
        url = str(row[url_idx]).strip()
        if not url or url in seen:
            continue
        seen.add(url)
        rows.append({
            "url": url,
            "title": str(row[title_idx]).strip() if title_idx is not None and title_idx < len(row) and row[title_idx] else "",
            "category": str(row[category_idx]).strip() if category_idx is not None and category_idx < len(row) and row[category_idx] else "",
        })
    wb.close()
    return rows


# ═══════════════════════════════════════════════════════════════
# PAGE CONTENT RETRIEVAL (base page only — excludes #state= dropdown chunks)
# ═══════════════════════════════════════════════════════════════

def fetch_base_page_content(search_client: SearchClient, url: str) -> str:
    results = list(search_client.search(
        search_text="*",
        filter=f"source_url eq '{url}'",
        select=["chunk_index", "content"],
        top=200,
    ))
    if not results:
        return ""
    results.sort(key=lambda r: r.get("chunk_index", 0))
    return "\n\n".join(r.get("content", "") for r in results if r.get("content"))


# ═══════════════════════════════════════════════════════════════
# LLM CALLS
# ═══════════════════════════════════════════════════════════════

def _chat_json(client: AzureOpenAI, model: str, prompt: str) -> Optional[dict | list]:
    for attempt in range(MAX_RETRIES + 1):
        try:
            resp = client.chat.completions.create(
                model=model,
                messages=[{"role": "user", "content": prompt}],
                temperature=0.3,
            )
            raw = resp.choices[0].message.content.strip()
            cleaned = raw.removeprefix("```json").removeprefix("```").removesuffix("```").strip()
            return json.loads(cleaned)
        except Exception as exc:
            if attempt == MAX_RETRIES:
                print(f"    [LLM call failed after {MAX_RETRIES + 1} attempts]: {exc}")
                return None
            time.sleep(1.5 * (attempt + 1))
    return None


def generate_candidate_questions(client: AzureOpenAI, content: str) -> list[str]:
    prompt = GENERATION_PROMPT.format(num_questions=NUM_CANDIDATE_QUESTIONS, content=content[:12000])
    result = _chat_json(client, GENERATION_MODEL, prompt)
    if isinstance(result, list):
        return [str(q).strip() for q in result if str(q).strip()]
    return []


def grade_questions(client: AzureOpenAI, content: str, candidates: list[str]) -> dict:
    numbered = "\n".join(f"{i+1}. {q}" for i, q in enumerate(candidates))
    prompt = GRADING_PROMPT.format(content=content[:12000], candidates_numbered=numbered)
    result = _chat_json(client, GRADING_MODEL, prompt)
    if not isinstance(result, dict) or "scores" not in result:
        return {"best_index": None, "best_score": 0, "rationale": "grading call failed or malformed response"}

    scores = {s["index"]: s.get("overall", 0) for s in result.get("scores", [])}
    best_index = result.get("best_index")
    best_score = scores.get(best_index, 0)
    return {"best_index": best_index, "best_score": best_score, "rationale": result.get("rationale", ""), "all_scores": scores}


def draft_answer(client: AzureOpenAI, content: str, question: str) -> str:
    prompt = ANSWER_DRAFT_PROMPT.format(content=content[:12000], question=question)
    for attempt in range(MAX_RETRIES + 1):
        try:
            resp = client.chat.completions.create(
                model=GENERATION_MODEL,
                messages=[{"role": "user", "content": prompt}],
                temperature=0.2,
            )
            return resp.choices[0].message.content.strip()
        except Exception as exc:
            if attempt == MAX_RETRIES:
                return f"[draft answer generation failed: {exc}]"
            time.sleep(1.5 * (attempt + 1))
    return ""


# ═══════════════════════════════════════════════════════════════
# MAIN PIPELINE
# ═══════════════════════════════════════════════════════════════

@dataclass
class PageResult:
    url: str
    title: str
    category: str
    status: str  # "generated" | "rejected" | "skipped_no_content" | "error"
    reason: str = ""
    question: str = ""
    draft_expected_answer: str = ""
    grade_score: Optional[int] = None
    grade_rationale: str = ""


def process_url(search_client, openai_client, url: str, title: str, category: str, dry_run: bool) -> PageResult:
    content = fetch_base_page_content(search_client, url)
    if not content or len(content.split()) < 30:
        return PageResult(url=url, title=title, category=category, status="skipped_no_content",
                           reason=f"base page content too short/empty ({len(content.split())} words)")

    if dry_run:
        return PageResult(url=url, title=title, category=category, status="generated",
                           question="[DRY RUN — no LLM called]", reason=f"content_words={len(content.split())}")

    candidates = generate_candidate_questions(openai_client, content)
    if not candidates:
        return PageResult(url=url, title=title, category=category, status="error",
                           reason="question generation returned nothing")

    grading = grade_questions(openai_client, content, candidates)
    best_index = grading.get("best_index")
    best_score = grading.get("best_score", 0)

    if not best_index or best_score < GRADE_THRESHOLD:
        return PageResult(url=url, title=title, category=category, status="rejected",
                           reason=f"best_score={best_score} below threshold={GRADE_THRESHOLD}",
                           grade_score=best_score, grade_rationale=grading.get("rationale", ""))

    chosen_question = candidates[best_index - 1]
    answer = draft_answer(openai_client, content, chosen_question)

    return PageResult(url=url, title=title, category=category, status="generated",
                       question=chosen_question, draft_expected_answer=answer,
                       grade_score=best_score, grade_rationale=grading.get("rationale", ""))


def main():
    parser = argparse.ArgumentParser(description="Generate golden dataset questions, one per approved URL")
    parser.add_argument("--excel", required=True, help="Path to approved URLs Excel file")
    parser.add_argument("--out-prefix", default="golden_dataset", help="Prefix for the three output CSVs")
    parser.add_argument("--dry-run", action="store_true", help="Skip LLM calls — just verify URL list + content retrieval works")
    parser.add_argument("--limit", type=int, default=None, help="Only process the first N URLs (for testing)")
    args = parser.parse_args()

    pages = load_approved_urls(args.excel)
    if args.limit:
        pages = pages[:args.limit]
    print(f"Loaded {len(pages)} approved URL(s) from {args.excel}"
          f"{' (dry run — no LLM calls)' if args.dry_run else ''}")

    search_client = get_search_client()
    openai_client = None if args.dry_run else get_openai_client()

    results: list[PageResult] = []
    for i, page in enumerate(pages, 1):
        print(f"[{i}/{len(pages)}] {page['url']}")
        result = process_url(search_client, openai_client, page["url"], page["title"], page["category"], args.dry_run)
        results.append(result)
        print(f"    -> {result.status}" + (f" ({result.reason})" if result.reason else ""))
        if not args.dry_run:
            time.sleep(REQUEST_SLEEP_SECONDS)

    generated = [r for r in results if r.status == "generated"]
    rejected = [r for r in results if r.status == "rejected"]
    skipped = [r for r in results if r.status == "skipped_no_content"]
    errored = [r for r in results if r.status == "error"]

    # ── Internal file (with draft answers, for framework testing) ──
    internal_path = f"{args.out_prefix}_internal.csv"
    with open(internal_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "id", "product_category", "question", "draft_expected_answer",
            "source_url", "grade_score", "grade_rationale",
        ])
        writer.writeheader()
        for i, r in enumerate(generated, 1):
            writer.writerow({
                "id": f"GOLD-{i:04d}", "product_category": r.category or "general",
                "question": r.question, "draft_expected_answer": r.draft_expected_answer,
                "source_url": r.url, "grade_score": r.grade_score, "grade_rationale": r.grade_rationale,
            })

    # ── PO/SME handoff file (NO draft answer — blank for independent review) ──
    handoff_path = f"{args.out_prefix}_handoff.csv"
    with open(handoff_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "id", "product_category", "question", "expected_answer", "source_url", "po_decision", "po_notes",
        ])
        writer.writeheader()
        for i, r in enumerate(generated, 1):
            writer.writerow({
                "id": f"GOLD-{i:04d}", "product_category": r.category or "general",
                "question": r.question, "expected_answer": "", "source_url": r.url,
                "po_decision": "", "po_notes": "",
            })

    # ── Audit log (every URL attempted, proof of coverage) ──
    audit_path = f"{args.out_prefix}_audit.csv"
    with open(audit_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=["url", "category", "status", "reason", "grade_score"])
        writer.writeheader()
        for r in results:
            writer.writerow({"url": r.url, "category": r.category, "status": r.status,
                              "reason": r.reason, "grade_score": r.grade_score})

    print(f"\n{'='*60}")
    print(f"Total URLs processed: {len(results)}")
    print(f"  Generated: {len(generated)}")
    print(f"  Rejected (below grade threshold): {len(rejected)}")
    print(f"  Skipped (no/insufficient content): {len(skipped)}")
    print(f"  Errored: {len(errored)}")
    print(f"\nOutputs:")
    print(f"  {handoff_path}  <- send this to PO/SME")
    print(f"  {audit_path}    <- send this too, as coverage proof")
    print(f"  {internal_path} <- keep for your own framework testing (has draft answers)")


if __name__ == "__main__":
    main()